#!/usr/bin/env python3
"""
Topologie Packing List Extractor — v9
Output: 2 sheets — Packing List (matches the official PL template layout,
plain black grid, no bold/no color fill) + Match_Status (internal QC log,
color-coded on purpose so mismatches stay easy to spot).
Updates: No/Product Name split, robust HS Code lookup, revised Origin logic,
Origin/HS Code repeated per item, Packing List sheet now mirrors the real
PL template columns 1:1 (Item#, OR No., SO No., Product Name in
English, SKU#, BarCode/UPC, UOM, Quantity, Carton#, Packaging code, Carton
Dimensions L/W/H, Weight, CBM, Origin Country, Origin Country's HTSCODE,
Shipping Mark, PORT, 中国标签名称) with a plain black-bordered grid
(no bold, no blue/red/green fills) instead of the old merged-diagnostic
style. PORT is auto-filled for CN-factory cartons using the same
store/port rule already used for the CN split (pl_group_export.py).
v9: Shipping Mark is now auto-filled with the source PDF's filename (no
.pdf extension) — no manual typing needed. CNEE / Notify Party for non-CN
factories can now be typed once on the app.html page (MANUAL_CONSIGNEE /
MANUAL_NOTIFY_PARTY) instead of edited by hand in every exported file; CN
shipments keep auto-filling from STORE_MASTER regardless of what's typed
there. OR No. / SO No. are auto-filled from a matched OR List when one is
uploaded (v13); 中国标签名称 is still intentionally left blank — fill it in
manually.
"""
from __future__ import annotations
import re, sys, logging, unicodedata, difflib, json
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import pandas as pd
import pdfplumber
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter

# ── Version / build identity ────────────────────────────────────────────────
# Shown in the app.html run summary so it's possible to verify the browser is
# actually running this revision (not a stale cached copy) and to tag every
# exported Audit_Summary sheet with the exact code that produced it.
PARSER_VERSION = "v11.0-sublist-2026-08"
GIT_COMMIT = __GIT_COMMIT__
LAST_RUN_META: Optional[dict] = None  # populated by run_pipeline(), read by app.html for the UI summary
LAST_OR_LIST_RESULT = None  # pl_or_list_import.OrListImportResult, populated by run_pipeline() -- read by
                             # the RUN_SUMMARY block below so the UI can distinguish "no file uploaded" from
                             # "file uploaded but header not recognized" (Turn 12 fix) instead of both looking
                             # identical (or_index empty either way).

logging.basicConfig(level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)])
log = logging.getLogger(__name__)

# =========================================================
# 1) CONFIG — filled in by app.html from browser uploads (Pyodide virtual FS).
#    Only this block differs from the original notebook; everything below is
#    byte-for-byte identical to the tested OCR_Packing List.ipynb / v7 logic,
#    except the Package dataclass (2 new optional fields) and the Excel
#    writer section (Packing List sheet layout), which are the parts this
#    revision (v8) intentionally changes.
# =========================================================
PL_FOLDER = Path("/work/pdfs")
OUTPUT_XLSX = Path("/work/PL_Total.xlsx")
DIM_WEIGHT_FILE = Path("/work/dim.xlsx")
DIM_WEIGHT_SHEET = __DIM_WEIGHT_SHEET__
MASTER_DATA_FILE = Path("/work/master.xlsx")
MASTER_DATA_SHEET = __MASTER_DATA_SHEET__
RECURSIVE = __RECURSIVE__
# v12: OR List (optional, spec section 4/9) -- None means "no file", and the
# whole pipeline must run fully in that case (never a fatal error, Run/
# Export never disabled). Wired from app.html's OR List upload input in a
# later step (Task 20); defaults to None until then.
OR_LIST_FILE = __OR_LIST_FILE__
# v9: optional manual CNEE / Notify Party, typed in on the app.html page —
# only used for non-CN factories (CN always auto-fills from STORE_MASTER).
# None / "" when the user left the field blank.
MANUAL_CONSIGNEE = __MANUAL_CONSIGNEE__
MANUAL_NOTIFY_PARTY = __MANUAL_NOTIFY_PARTY__
# v11: Sublist (Excel) generation -- OFF by default in v12+ (spec turn 5:
# the mandatory Sublist output is now the A5 PDF below; this Excel version
# is kept only as an optional SECONDARY output). Never required, never
# blocks Run/Export if OFF or if it fails to import (see AUTO SPLIT
# section at the bottom).
GENERATE_SUBLIST = __GENERATE_SUBLIST__
# v12: A5 carton Sublist PDF -- ON by default (spec: "Generate A5 carton
# Sublist PDF" checkbox, default checked). This is now the primary/
# mandatory Sublist deliverable (PL_SPLIT_OUTPUT/05_SUBLIST/SUBLIST_TOTAL.
# pdf). Its failure must NEVER block the legacy ZIP/export -- see
# SUBLIST_PDF_STATUS handling in the AUTO SPLIT section at the bottom.
GENERATE_SUBLIST_PDF = __GENERATE_SUBLIST_PDF__


# ── Constants ──────────────────────────────────────────────────────────────
VALID_UNITS  = {"PCS", "SET", "CARTON", "CTN", "BOX", "PACK"}
UNIT_PAT     = "|".join(VALID_UNITS)
STATUS_WORDS = {"MOI", "NEW", "USED", "CU"}
# Canonical display form for each condition word we recognise (used when a
# "Tinh trang" cell is split out of a merged "Tinh trang So luong" cell —
# see RE_COND_QTY below — so the exported value is stable regardless of
# accented/unaccented spelling in the source PDF).
CONDITION_CANON = {
    "MOI": "Mới", "NEW": "New", "CU": "Cũ", "USED": "Used",
    "REFURBISHED": "Refurbished",
}
VN_KEYWORDS  = {"POP", "JION", "QIFENG", "SBGEAR", "SB_GEAR"}
TABLE_HDR_KW = {"stt","barcode","ma vach","ma hang","ten hang",
                "don vi","so luong","tinh trang","condition","quantity"}

RE_PKG_HEADER = re.compile(
    r'(?:M[aã]\s*ki[eệ]n\s*h[aà]ng\s*[:\-]?\s*)?'
    r'(PGKEC[A-Z0-9]{5,})'
    r'(?:\s+(\d+\s*/\s*\d+))?',
    re.IGNORECASE | re.UNICODE)

RE_TOTAL = re.compile(
    r'T[oôồốổỗộòóỏõọ]ng\s+c[oôồốổỗộòóỏõọ]ng\s*:?\s*([\d,\.]+)',
    re.IGNORECASE | re.UNICODE)

RE_BARCODE   = re.compile(r'(?<!\d)(\d{8,14})(?!\d)')
RE_GTIN13    = re.compile(r'^\d{13}$')
RE_PROD_CODE = re.compile(r'(TP-[A-Z0-9]{2,}(?:-[A-Z0-9]+)*-?)', re.IGNORECASE)
RE_TERMINAL  = re.compile(rf'({UNIT_PAT})\s+[^\d\n]+?\s+([\d,]+)\s*$',
                          re.IGNORECASE | re.UNICODE)
# Bug #9 in the audit: pdfplumber sometimes merges the "Tinh trang" (condition)
# and "So luong" (quantity) columns into a single cell, e.g. "Moi 12". Left
# unsplit, that cell fails every classifier below (not a barcode, not a SKU,
# not a bare unit, not a bare integer) and silently falls through to the
# product-name bucket -- the item is then dropped for quantity==0. This
# pattern recognises and splits that merged cell.
RE_COND_QTY  = re.compile(
    r'^\s*(M[oớ]i|C[uũ]|New|Used|Refurbished)\s+([\d][\d,\.]*)\s*$',
    re.IGNORECASE | re.UNICODE)
# Bug #6: a SKU suffix like "-BOX" that lands on its own on the next physical
# row/line (continuation of the previous item's code) instead of staying
# attached to the code cell.
RE_SKU_SUFFIX_ROW = re.compile(r'^-?(BOX|SET|PACK|CTN|CARTON)$', re.IGNORECASE)
# Header/footer noise that must never be parsed as an item row (bug #12):
# browser tab title ("about:blank"), print timestamp, page numbers, and the
# "# Barcode  Ma san pham ..." table header repeated as plain text on every
# page once the table border detector fails to catch it as a real table.
RE_NOISE = re.compile(
    r'^(STT\b|No\.\s*$|PACKING\s*LIST|DANH\s*S[AÁ]CH|Page\s*\d+\s*(/|of)?\s*\d*\s*$'
    r'|about:blank'
    r'|\d{1,2}/\d{1,2}/\d{2,4},?\s*\d{1,2}:\d{2}\s*(AM|PM)?'
    r'|#?\s*Barcode\s+M[aã]\s*s[aả]n\s*ph[aẩ]m'
    r'|^Kho\s|^Hotline|^\u0110[iị]a\s*ch[iỉ]|^S[oố]\s*\u0111i[eệ]n\s*tho[aạ]i'
    r'|^Th[oô]ng\s*tin\s)',
    re.IGNORECASE | re.UNICODE)

# ── Helpers ────────────────────────────────────────────────────────────────
# Unicode artifacts observed in real PL PDFs (bug #13 in the audit): the text
# layer sometimes contains U+FFFE / U+FFFF ("noncharacter" codepoints) or a
# Unicode dash variant (en dash, em dash, non-breaking hyphen, ...) in place
# of a plain ASCII '-'; zero-width characters (ZWSP/ZWNJ/ZWJ/word joiner/BOM/
# soft hyphen) and NBSP-family spaces also show up inside otherwise-
# contiguous codes. All normalization below funnels through one function so
# every caller (DIM ref/pkg keys, SKU keys, display codes) handles these the
# same way instead of each reimplementing a partial fix.
_ZERO_WIDTH_RE    = re.compile('[\u200B\u200C\u200D\u2060\uFEFF\xad]')
_DASH_VARIANTS_RE = re.compile('[\u2010\u2011\u2012\u2013\u2014\u2015\uFFFE\uFFFF]')
_NBSP_RE          = re.compile('[\xa0\u202f]')

def fix_unicode_artifacts(s) -> str:
    """Canonical Unicode cleanup shared by every code/reference normalizer:
    NFKC-fold, convert PDF-extraction dash artifacts (U+FFFE/U+FFFF, en/em
    dash, non-breaking hyphen, ...) to a plain '-', drop zero-width
    characters, convert NBSP-like spaces to a normal space, then collapse any
    whitespace (including a literal newline) that sits directly around a '-'
    so a code split across a line ("...-56-\nBOX") re-joins into one token
    ("...-56-BOX")."""
    if s is None:
        return ""
    s = unicodedata.normalize("NFKC", str(s))
    s = _DASH_VARIANTS_RE.sub("-", s)
    s = _ZERO_WIDTH_RE.sub("", s)
    s = _NBSP_RE.sub(" ", s)
    s = re.sub(r"\s*-\s*", "-", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()

def normalize_code(text) -> str:
    """Display-safe code normalizer for SKU / package / reference codes:
    fixes Unicode artifacts via fix_unicode_artifacts(), then strips every
    remaining whitespace (a code never legitimately contains one) and
    upper-cases. Keeps hyphens intact, unlike normalize_sku_key() below,
    which is a stricter matching key. Never fuzzy-corrects characters (no
    O<->0 / I<->1 substitution) — only Unicode-artifact cleanup."""
    if text is None:
        return ""
    text = fix_unicode_artifacts(text)
    text = re.sub(r"\s+", "", text)
    return text.upper()

def strip_accents(s: str) -> str:
    nfd = unicodedata.normalize('NFD', str(s).strip())
    return ''.join(c for c in nfd if unicodedata.category(c) != 'Mn').upper()

def normalize(s: str) -> str:
    s = fix_unicode_artifacts(s)
    s = strip_accents(s)
    s = re.sub(r'[^\x20-\x7E]', '', s)
    s = re.sub(r'\s*[\(\[]\d+[\)\]]', '', s)
    return s.strip()

def is_valid_gtin13(s: str) -> bool:
    """Strict GTIN validation per spec: exactly 13 digits, exact match only
    -- never fuzzy-corrected (no O->0 / I->1 substitution)."""
    return bool(RE_GTIN13.fullmatch(str(s or '').strip()))

def parse_qty(s: str) -> int:
    return int(re.sub(r'[,\.]', '', s.strip()))

def get_origin(source_file: str, reference_code: str) -> str:
    """
    Origin logic:
    - If reference/source explicitly contains TOPOLOGIE or ends with _CN / -CN / space CN => CN.
      Example: CN-2659_SH_CN => CN.
    - Otherwise JION, POP, SBGEAR, QIFENG, VN => VN.
    - Default remains CN for safety.
    """
    raw = f"{source_file} {reference_code}"
    text = strip_accents(raw)
    ref = strip_accents(reference_code)

    # Strong CN markers
    if "TOPOLOGIE" in text:
        return "CN"
    if re.search(r'(?:^|[_\-\s])CN(?:$|[_\-\s.])', ref):
        # The final business code is CN, e.g. CN-2659_SH_CN
        if ref.endswith("_CN") or ref.endswith("-CN") or ref.endswith(" CN") or ref == "CN":
            return "CN"

    # VN markers
    for kw in {"JION", "POP", "SBGEAR", "SB_GEAR", "QI FENG", "QIFENG", "VN"}:
        if kw in text:
            return "VN"

    return "CN"

def join_split_product_code(text: str) -> str:
    return re.sub(r'(TP-[A-Z0-9-]+?-)\s+([A-Z0-9])', r'\1\2',
                  text, flags=re.IGNORECASE)

def is_table_hdr(cells: List[str]) -> bool:
    joined = strip_accents(" ".join(cells)).lower()
    return sum(1 for kw in TABLE_HDR_KW if kw in joined) >= 2

def is_noise(line: str) -> bool:
    return bool(RE_NOISE.match(line.strip()))

def is_noise_row(cells: List[str]) -> bool:
    """Same header/footer-noise filter as is_noise(), applied to a table row
    (list of cells) instead of a single text line -- covers repeated table
    headers and about:blank/date/page-number rows that pdfplumber sometimes
    hands back as an extra 1-row 'table' rather than as plain text."""
    joined = " ".join(c for c in cells if c).strip()
    return bool(joined) and bool(RE_NOISE.match(joined))

def safe_float(v) -> Optional[float]:
    try:
        f = float(v)
        return None if pd.isna(f) else f
    except Exception:
        return None

def normalize_sku_key(text: str) -> str:
    """Robust key for SKU/EAN lookup: remove hidden chars, spaces and
    punctuation. Bug fix: the previous version called
    .replace("\ufffe", "-") twice and never handled U+FFFF at all (both
    show up in real PL PDFs in place of a plain '-' inside a SKU) --
    fix_unicode_artifacts() now handles both, plus the other dash variants /
    zero-width characters, in one place shared with normalize_code()."""
    text = clean_excel_key(text) if 'clean_excel_key' in globals() else str(text or '').strip()
    text = fix_unicode_artifacts(text)
    return re.sub(r"[^A-Z0-9]", "", text.upper())

_INVISIBLE_CHARS_RE = re.compile('[\u200b\u200c\u200d\u2060\ufeff\xad]')  # bug fix: a literal ASCII space used to sit in this class (typo'd in place
# of NBSP U+00A0), silently deleting every space from every table cell
# ("14mm Rope Loop" -> "14mmRopeLoop") -- confirmed via real-file test on
# CN-1286-Kerry_PVG-CN.pdf before this fix.

def sanitize_ocr_cell(s: str) -> str:
    """Strip invisible/zero-width unicode artifacts (NBSP, zero-width space,
    soft hyphen, BOM, ...) that PDF text extraction sometimes inserts inside
    otherwise-contiguous codes. Left uncleaned, these can silently break the
    barcode/SKU regexes or leave a stray character inside a code that should
    read as one unbroken token."""
    if s is None:
        return s
    return _INVISIBLE_CHARS_RE.sub('', str(s))

def dequarantine_code(value: str, label: str, context: str = "") -> str:
    """SKU/product codes and EAN/barcodes never legitimately contain blanks —
    if OCR produced one anyway (stray space from a misread character), strip
    it and log a warning so the anomaly is visible for a quick sanity check,
    rather than silently mismatching or silently 'fixing' with no trace."""
    if not value:
        return value
    cleaned = re.sub(r'\s+', '', value)
    if cleaned != value:
        log.warning(f"{label} contained unexpected blank(s), auto-fixed: "
                     f"{value!r} -> {cleaned!r}" + (f" ({context})" if context else ""))
    return cleaned

def split_leading_no(product_name: str, current_no: str = "") -> Tuple[str, str]:
    """Split cases like '1 10mm Rope Loop' into No='1', Product Name='10mm Rope Loop'."""
    name = re.sub(r"\s+", " ", str(product_name or "")).strip()
    no = str(current_no or "").strip()
    m = re.match(r"^(\d{1,4})\s+(.+)$", name)
    if m and not no:
        no = m.group(1)
        name = m.group(2).strip()
    return no, name

# ── Data models ────────────────────────────────────────────────────────────
@dataclass
class Item:
    no: str
    product_name: str
    product_code: str
    barcode: str
    unit: str
    quantity: int
    hs_code: str = ""
    parse_method: str = "table"
    # -- audit/diagnostic fields (v10 audit) -- additive only; every field
    # above this line is unchanged so pl_group_export.py / the Excel writer
    # keep working exactly as before (backward compatible).
    condition: str = ""
    remark: str = ""
    sku_raw: str = ""
    gtin_raw: str = ""
    source_page: Optional[int] = None
    gtin_valid: bool = False
    dedup_key: Tuple = field(default=(), repr=False, compare=False)

    @property
    def sku(self) -> str:
        """Spec vocabulary alias for product_code (normalized SKU)."""
        return self.product_code
    @property
    def gtin(self) -> str:
        """Spec vocabulary alias for barcode."""
        return self.barcode

@dataclass
class Package:
    package_code: str
    source_file: str
    reference_code: str
    pdf_package_seq: str
    items: List[Item] = field(default_factory=list)
    declared_total_qty: Optional[int] = None
    header_count: int = 1
    global_carton_num: str = ""
    length:  Optional[float] = None
    width:   Optional[float] = None
    height:  Optional[float] = None
    weight:  Optional[float] = None
    cbm:     Optional[float] = None
    dim_matched: bool = False
    hs_code: str = ""
    # v8: CN store/port classification, filled in by classify_packages_for_port()
    # (same rule as pl_group_export.py's CN split — "quy luật chia port/store").
    # Stays "" for non-CN-factory packages; fill manually if needed.
    port: str = ""
    store: str = ""
    # v10 audit: first page the package header was seen on, and the internal
    # dedup set used while items are being fed in (see Parser._add_item).
    first_page: Optional[int] = None
    dim_source_method: str = ""
    # v11: OR / SO / Shipping Mark -- package-level metadata, parsed from a
    # "LABEL: value" line/cell inside the PL itself (see parse_document_sequence
    # helpers and Parser._maybe_capture_metadata below). Additive-only, every
    # field defaults to "" so any existing caller/test that doesn't know these
    # fields exist keeps working unmodified (dataclass field order stays the
    # same for every pre-existing field above this block).
    #
    # IMPORTANT -- these are NOT the same thing:
    #   or_number      = Customer OR (e.g. "OR1016"), a business/order reference
    #   shipping_mark   = Warehouse "OR" (e.g. "CN-1666-PVG-KERRY-POP"), today
    #                      defaulting to the source PDF filename (reference_code)
    #                      exactly like before this change -- see run_pipeline().
    # Never overwrite one with the other.
    or_number: str = ""
    or_source: str = ""             # PL_STRUCTURED_FIELD | PL_TEXT | ""
    so_number: str = ""
    so_source: str = ""             # PL_STRUCTURED_FIELD | PL_TEXT | ""
    shipping_mark: str = ""
    shipping_mark_source: str = ""  # PL_STRUCTURED_FIELD | PL_TEXT | FILENAME_REFERENCE_CODE | ""
    # v12: GW parsed directly from a "GW:" label line in the PL text. Kept
    # SEPARATE from `weight` (exclusively DIM/Weight-file-sourced, existing
    # field, unchanged behavior) -- this is only a fallback used by the
    # Sublist when no DIM match exists. Stored as the raw string (e.g.
    # "35.68 KG" or "35.68") since the label's own unit formatting should
    # not be guessed/reformatted here.
    pl_gross_weight: str = ""
    pl_gross_weight_source: str = ""
    # v11: structured carton-numbering fields backing the existing
    # global_carton_num string (kept as-is, unchanged shape, for 100% backward
    # compatibility with every current reader of that field) so callers never
    # have to re-parse "1/10" back into numbers.
    carton_sequence: int = 0
    carton_total: int = 0
    carton_display: str = ""
    # v12: counting SCOPE this package's carton number is relative to (spec
    # section 9) -- "" (the default/unset value) means "one implicit scope
    # containing every package", which is exactly today's flat global
    # numbering, so every existing caller/test that never sets this field
    # keeps behaving identically. Format when set: "<shipment_key>|
    # <normalized_store>", e.g. "OR:OR1016|KERRY" or "UPLOAD_BATCH|KERRY".
    counting_scope_key: str = ""
    counting_scope_source: str = ""  # OR_LIST_GROUPING | UPLOAD_BATCH_DEFAULT | ""
    # v12: Store/OR/SO resolved against an (optional) OR List -- kept
    # separate from `store` (existing field, CN-only, filled by
    # classify_packages_for_port/match_store) and from or_number/so_number
    # (PL-text-parsed, v11) so every source stays independently traceable.
    # or_number/so_number are ENRICHED from these (fills in only when still
    # blank -- PL-text-parsed value always wins, spec section 4.2/4.3 +
    # the original priority list: structured PL field > OCR text > OR List).
    or_list_store: str = ""
    or_list_match_source: str = ""   # SHIPMARK_TOKEN_EXACT | FILENAME_TOKEN_EXACT | RECEIVER_TEXT_EXACT | FUZZY | ""
    or_list_match_status: str = ""   # OK | REVIEW | NO_OR_LIST | ""
    or_list_review_reason: str = ""
    or_list_candidate_store: str = ""
    or_list_candidate_score: float = 0.0
    _seen_item_keys: set = field(default_factory=set, repr=False, compare=False)

    @property
    def calc_qty(self) -> int:
        return sum(i.quantity for i in self.items)
    @property
    def item_count(self) -> int:
        return len(self.items)
    @property
    def origin(self) -> str:
        return get_origin(self.source_file, self.reference_code)
    @property
    def master_match_status(self) -> str:
        """v14 (spec section 8/14): explicit spec-vocabulary status for
        Master Data enrichment (DIM/weight file + HS Code file). Master
        Data is ENRICHMENT-ONLY -- a package/item is NEVER removed from any
        output because Master Data didn't have a matching row for it; this
        property just reports that fact under the spec's own status name
        instead of leaving it implicit in dim_matched / a blank hs_code
        (both of which already behave this way and are unchanged by this
        property -- it derives from them, it doesn't drive any new
        filtering/dropping behavior anywhere)."""
        if not self.dim_matched:
            return "MASTER_UNMATCHED"
        if self.items and any(not (i.hs_code or "").strip() for i in self.items):
            return "MASTER_UNMATCHED"
        return "MASTER_MATCHED"

# ── v11: OR # / SO Order # / Shipping Mark -- label-value line detection ───
# These three fields are package-level metadata that MAY appear printed on
# the PL itself as a "LABEL: value" line (its own text line, or a 2-cell
# table row of [label, value]) -- e.g. "OR #: OR1016" / "SO Order #  so402064"
# / "Shipping Mark  CN-1666-PVG-KERRY-POP". No real production PL sample with
# this exact layout was available while writing this (see README/commit
# message) -- the matcher below is intentionally conservative: bare 2-letter
# aliases ("OR", "SO") only match when immediately followed by a punctuation
# separator (#, :, .), never by whitespace alone, so an ordinary sentence
# that happens to start with the English word "or"/"so" is never misread as
# a label. Multi-word aliases ("SO Order #", "Shipping Mark", ...) are
# unambiguous enough to also accept a plain space separator. Never widened
# beyond the alias lists below without a real-file confirmation (spec:
# "khong match qua rong dan den nhan nham field").
OR_ALIASES_RAW = ["OR #", "OR NO.", "OR NO", "OR NUMBER", "OR CODE", "OUTBOUND REQUEST", "OR"]
SO_ALIASES_RAW = ["SO ORDER #", "SO ORDER", "SO #", "SO NO.", "SO NO", "SO NUMBER",
                   "SALES ORDER #", "SALES ORDER", "SO"]
SHIPPING_MARK_ALIASES_RAW = ["SHIPPING MARKS", "SHIPPING MARK", "SHIPMARK #", "SHIPMARK",
                              "SHIP MARK", "MARKS & NOS", "MARKS AND NUMBERS", "WAREHOUSE OR",
                              # v14 (spec section 1): "Ma tham chieu" (Vietnamese
                              # PL template field) is the HIGHEST-priority Shipmark
                              # source per spec, but this codebase has no real PDF
                              # sample containing it to validate a separate priority
                              # tier against (documented limitation, same caveat as
                              # the rest of this alias-matcher) -- added here as
                              # ANOTHER shipping_mark alias (both accented and
                              # unaccented spellings, since OCR may strip diacritics)
                              # so it is at least recognised and captured from PL
                              # text/table cells, additive and safe. Both spellings
                              # normalize identically in match_metadata_label_cells()
                              # (accent-insensitive there); RE_SHIPPING_MARK_LABEL_LINE
                              # needs both forms listed explicitly since that path
                              # matches raw text.
                              "MA THAM CHIEU", "MÃ THAM CHIẾU", "THAM CHIEU", "THAM CHIẾU"]
# v12: GW (gross weight) as a package-level label line, same convention as
# OR#/SO#/Shipping Mark. NOTE this is intentionally kept SEPARATE from
# Package.weight (which is exclusively DIM/Weight-file-sourced, unchanged
# existing behavior) -- see pl_weight_from_label / _METADATA_SOURCE_FIELD
# and resolve_sublist_metadata() in pl_sublist_pdf_export.py for the
# resolution order (DIM file wins when present; this is a fallback only).
GW_ALIASES_RAW = ["GROSS WEIGHT", "G.W.", "G.W", "GW"]
# "GW" is NOT in the strict-separator set below -- unlike bare "OR"/"SO",
# which collide with common English sentence-starter words, "GW" has no
# such natural-language collision risk, so a plain space separator (the
# real spec example: "GW  35.68 KG", no punctuation) is accepted.
_BARE_2LETTER_ALIASES = {"OR", "SO"}


def _alias_word_pattern(alias: str) -> str:
    words = alias.strip().split()
    return r'[\s#\.\-_]*'.join(re.escape(w) for w in words)


def _build_label_matcher(aliases_raw):
    # Longest-alias-first so "SO Order #" is tried before the bare "SO".
    ordered = sorted(set(aliases_raw), key=len, reverse=True)
    alts = []
    for alias in ordered:
        pat = _alias_word_pattern(alias)
        sep = r'[:#\.]\s*' if alias.upper() in _BARE_2LETTER_ALIASES else r'[\s#\.\-_:]*\s*'
        alts.append(pat + sep)
    combined = r'^\s*(?:' + '|'.join(alts) + r')(\S.*?)\s*$'
    return re.compile(combined, re.IGNORECASE | re.UNICODE)


RE_OR_LABEL_LINE = _build_label_matcher(OR_ALIASES_RAW)
RE_SO_LABEL_LINE = _build_label_matcher(SO_ALIASES_RAW)
RE_SHIPPING_MARK_LABEL_LINE = _build_label_matcher(SHIPPING_MARK_ALIASES_RAW)
RE_GW_LABEL_LINE = _build_label_matcher(GW_ALIASES_RAW)

_ALL_LABEL_ALIASES_NORM = {
    re.sub(r'[^A-Z0-9]', '', a.upper())
    for group in (OR_ALIASES_RAW, SO_ALIASES_RAW, SHIPPING_MARK_ALIASES_RAW, GW_ALIASES_RAW)
    for a in group
}

_METADATA_SOURCE_FIELD = {
    "or_number": "or_source",
    "so_number": "so_source",
    "shipping_mark": "shipping_mark_source",
    "pl_gross_weight": "pl_gross_weight_source",
}


# v14 (spec section 1): ordinal confidence per Shipmark source -- a PL
# table field ("PL_STRUCTURED_FIELD") is the most reliable, a PL text-line
# regex match ("PL_TEXT") is somewhat less reliable (more prone to a
# stray/unrelated line matching), the filename fallback
# ("FILENAME_REFERENCE_CODE") is the weakest signal (it's not from the PL's
# own content at all). Never blank/"" -- an unresolved package still gets
# 0.0, never None, so every diagnostics consumer can sort/filter on it
# without a None-check.
_SHIPPING_MARK_SOURCE_CONFIDENCE = {
    "PL_STRUCTURED_FIELD": 1.0,
    "PL_TEXT": 0.85,
    "FILENAME_REFERENCE_CODE": 0.5,
    "": 0.0,
}


def resolve_shipping_mark_confidence(source: str) -> float:
    return _SHIPPING_MARK_SOURCE_CONFIDENCE.get(source or "", 0.0)


# v14 (spec sections 2-3): country detection from the resolved Shipmark's
# prefix -- "^(CN|KR|JP|BE|US|TW)" per spec, anchored at the very start of
# the string and requiring the next character (if any) to NOT be another
# letter, so "US-1234" matches "US" but "USER-1234" or "USA-1234" do not
# (never a guess beyond the literal 6-code list the spec names). Distinct
# from any store token embedded elsewhere in the Shipmark (e.g. the "KR" in
# "CN-1529_KR_PVG_POP" is a STORE token, not a country -- this only ever
# looks at the leading prefix).
RE_SHIPMARK_COUNTRY_PREFIX = re.compile(r'^(CN|KR|JP|BE|US|TW)(?![A-Z])', re.IGNORECASE)


def detect_shipment_country(shipmark: str) -> str:
    """Returns one of CN/KR/JP/BE/US/TW, or "" if the Shipmark's prefix
    doesn't match any known country code (never guessed)."""
    text = (shipmark or "").strip()
    m = RE_SHIPMARK_COUNTRY_PREFIX.match(text)
    return m.group(1).upper() if m else ""


def _norm_label_cell(s: str) -> str:
    return re.sub(r'[^A-Z0-9]', '', strip_accents(s).upper())


def match_metadata_label_line(line: str):
    """Try OR -> SO -> Shipping Mark -> GW (in that priority order) against
    one flattened text line. Returns (field, value) or (None, None). Used
    for the text-fallback parsing path (source=PL_TEXT)."""
    line = (line or "").strip()
    if not line:
        return None, None
    m = RE_OR_LABEL_LINE.match(line)
    if m:
        return "or_number", m.group(1).strip()
    m = RE_SO_LABEL_LINE.match(line)
    if m:
        return "so_number", m.group(1).strip()
    m = RE_SHIPPING_MARK_LABEL_LINE.match(line)
    if m:
        return "shipping_mark", m.group(1).strip()
    m = RE_GW_LABEL_LINE.match(line)
    if m:
        return "pl_gross_weight", m.group(1).strip()
    return None, None


def match_metadata_label_cells(cells):
    """2-cell table row [label, value] case (source=PL_STRUCTURED_FIELD) --
    safer than the line-prefix regex because the WHOLE first cell must
    normalize to a known alias, not just a line prefix. Returns (field,
    value) or (None, None)."""
    non_empty = [c.strip() for c in cells if c and str(c).strip()]
    if len(non_empty) != 2:
        return None, None
    label, value = non_empty
    norm = _norm_label_cell(label)
    if norm not in _ALL_LABEL_ALIASES_NORM:
        return None, None
    if norm in {"OR", "SO", "GW"} and len(value) < 1:
        return None, None
    if any(_norm_label_cell(a) == norm for a in OR_ALIASES_RAW):
        return "or_number", value.strip()
    if any(_norm_label_cell(a) == norm for a in SO_ALIASES_RAW):
        return "so_number", value.strip()
    if any(_norm_label_cell(a) == norm for a in SHIPPING_MARK_ALIASES_RAW):
        return "shipping_mark", value.strip()
    if any(_norm_label_cell(a) == norm for a in GW_ALIASES_RAW):
        return "pl_gross_weight", value.strip()
    return None, None


# ── v11: document (PDF file) sequence -- original vs _1 / _2 / ... copies ──
# Copy-suffix is ONLY recognised when it is a trailing pattern anchored at
# the very end of the reference_code/filename stem -- "CN-1666-PVG-KERRY-POP"
# must NOT be read as sequence=1666 or copy=POP; there is no trailing
# _N/-N/(N)/COPY-N pattern there at all, so it correctly falls through as an
# ordinary "no suffix" (original, sequence 0) document.
_SEQ_SUFFIX_RE = re.compile(
    r'^(?P<base>.+?)(?:[_\-]\s*(?P<n1>\d{1,4})|\((?P<n2>\d{1,4})\)|\s+COPY\s+(?P<n3>\d{1,4}))$',
    re.IGNORECASE)


class DocumentSequence:
    __slots__ = ("base_document_key", "document_sequence", "is_original", "sequence_source")

    def __init__(self, base_document_key, document_sequence, is_original, sequence_source):
        self.base_document_key = base_document_key
        self.document_sequence = document_sequence
        self.is_original = is_original
        self.sequence_source = sequence_source

    def __repr__(self):
        return (f"DocumentSequence(base_document_key={self.base_document_key!r}, "
                f"document_sequence={self.document_sequence!r}, is_original={self.is_original!r}, "
                f"sequence_source={self.sequence_source!r})")

    def __eq__(self, other):
        return isinstance(other, DocumentSequence) and (
            self.base_document_key, self.document_sequence, self.is_original, self.sequence_source
        ) == (other.base_document_key, other.document_sequence, other.is_original, other.sequence_source)


def parse_document_sequence(reference_code: str, source_file: str = "") -> DocumentSequence:
    """Pure function: is this PDF the ORIGINAL of a document, or a numbered
    copy (Kerry_POP -> sequence 0/original; Kerry_POP_1 -> sequence 1; ...)?

    Only a trailing _N / -N / (N) / "COPY N" pattern counts as a copy
    sequence (spec section 6.2) -- a number anywhere else in the code (e.g.
    a shipment number in the middle of a Shipping Mark) is never mistaken
    for one, because the pattern is anchored to the end of the string.
    """
    stem = (reference_code or "").strip()
    if not stem:
        stem = Path(source_file).stem if source_file else ""
    m = _SEQ_SUFFIX_RE.match(stem)
    if not m or not m.group("base"):
        return DocumentSequence(stem, 0, True, "NO_SUFFIX_ORIGINAL")
    n = m.group("n1") or m.group("n2") or m.group("n3")
    return DocumentSequence(m.group("base"), int(n), False, "TRAILING_SUFFIX_PATTERN")


# Spec-vocabulary status codes for the Raw_Data / Audit_Summary sheets and
# the app.html run summary (kept separate from overall_status() below, which
# stays byte-for-byte backward compatible for the existing Match_Status
# sheet's color-coding).
def audit_status(pkg: "Package") -> str:
    if pkg.item_count == 0:
        return "ZERO_ITEMS" if pkg.declared_total_qty is not None else "PARSE_REVIEW_REQUIRED"
    if pkg.declared_total_qty is None:
        return "MISSING_TOTAL"
    if pkg.declared_total_qty != pkg.calc_qty:
        return "QUANTITY_MISMATCH"
    return "OK"

# ── Item parsers ───────────────────────────────────────────────────────────
def parse_item_cells(cells: List[str], source_page: Optional[int] = None) -> Optional[Item]:
    """Parse one packing-list item row from a table's cells. Column-count
    agnostic (works for the 6/7/8-column shapes pdfplumber hands back — bug
    #8/#10) because every cell is classified by *content*, not a fixed
    index. RE_COND_QTY below is the fix for bug #9: a merged "Tinh trang So
    luong" cell ("Moi 12") used to fail every classifier and silently fall
    into the product-name bucket, dropping the whole item (quantity stayed
    0). Confirmed against CN-1286-CNWorld_PEK-CN.pdf p.1 table 1: rows 1-4
    all used this merged shape and were previously lost (23 units missing
    from package PGKECISZTEPU3490002's declared total)."""
    line_no = ""
    barcode = prod_code = unit = condition = ""
    sku_raw = gtin_raw = ""
    quantity = 0
    name_parts: List[str] = []
    unit_idx = -1
    barcode_idx = -1
    prod_idx = -1

    # Pass 1: Unicode-normalize every cell (bug #13: U+FFFE/U+FFFF/zero-width
    # artifacts) and re-join a SKU/product code pdfplumber split across two
    # adjacent cells (trailing '-' + continuation cell in the next column).
    merged: List[str] = []
    i = 0
    while i < len(cells):
        cell = fix_unicode_artifacts(sanitize_ocr_cell(str(cells[i])))
        if cell.endswith('-') and RE_PROD_CODE.fullmatch(cell) and i + 1 < len(cells):
            nxt = fix_unicode_artifacts(sanitize_ocr_cell(str(cells[i + 1])))
            merged.append(cell + nxt)
            i += 2
        else:
            merged.append(cell)
            i += 1

    for idx, raw in enumerate(merged):
        cell = str(raw).strip()
        if not cell:
            continue

        # Bug #9 fix: merged "Tinh trang So luong" cell, e.g. "Moi 12".
        m_cq = RE_COND_QTY.match(cell)
        if m_cq:
            cond_key = strip_accents(m_cq.group(1))
            condition = CONDITION_CANON.get(cond_key, m_cq.group(1))
            quantity = parse_qty(m_cq.group(2))
            continue

        # STT / No column usually appears before barcode or SKU.
        if not line_no and re.fullmatch(r'\d{1,4}', cell) and not barcode and not prod_code and idx <= 2:
            line_no = cell
            continue

        if RE_BARCODE.fullmatch(cell) and len(cell) in (8, 12, 13, 14):
            barcode = cell
            gtin_raw = cell
            barcode_idx = idx
            continue

        candidate = join_split_product_code(cell)
        m = RE_PROD_CODE.search(candidate)
        if m:
            pc = m.group(1).rstrip('-')
            if pc.count('-') >= 2:
                sku_raw = candidate
                prod_code = normalize_code(pc)
                prod_idx = idx
                rest = RE_PROD_CODE.sub('', candidate).strip()
                if rest and strip_accents(rest) not in STATUS_WORDS:
                    no_from_name, clean_name = split_leading_no(rest, line_no)
                    if no_from_name:
                        line_no = no_from_name
                    if clean_name:
                        name_parts.append(clean_name)
                continue

        if cell.upper() in VALID_UNITS:
            unit = cell.upper()
            unit_idx = idx
            continue

        # Bare condition word in its own cell (7/8-column layout, condition
        # and quantity NOT merged).
        cond_key = strip_accents(cell)
        if cond_key in STATUS_WORDS:
            condition = CONDITION_CANON.get(cond_key, cell)
            continue

        if re.fullmatch(r'\d{1,5}', cell) and idx > unit_idx >= 0:
            quantity = int(cell)
            continue

        if RE_PKG_HEADER.search(cell):
            continue

        # Avoid putting STT in product name if it appears as a separate numeric cell.
        if re.fullmatch(r'\d{1,4}', cell) and not name_parts and (idx < barcode_idx or idx < prod_idx or idx <= 2):
            if not line_no:
                line_no = cell
            continue

        name_parts.append(cell)

    if not (barcode or prod_code) or quantity == 0:
        return None

    product_name = re.sub(r'\s+', ' ', ' '.join(name_parts)).strip()
    line_no, product_name = split_leading_no(product_name, line_no)
    prod_code = dequarantine_code(prod_code, "SKU/product_code", "table row")
    barcode   = dequarantine_code(barcode, "EAN/barcode", "table row")
    return Item(no=line_no, product_name=product_name, product_code=prod_code,
                barcode=barcode, unit=unit or "PCS", quantity=quantity,
                parse_method="table", condition=condition, sku_raw=sku_raw,
                gtin_raw=gtin_raw, source_page=source_page,
                gtin_valid=is_valid_gtin13(barcode))

def parse_item_text(accumulated: str, source_page: Optional[int] = None) -> Optional[Item]:
    text = fix_unicode_artifacts(join_split_product_code(sanitize_ocr_cell(accumulated)))

    # Bug #9, text-fallback path: a merged "Tinh trang So luong" fragment can
    # also show up in the flattened text line (e.g. "... PCS Moi 12"). Try
    # the merged terminal first; RE_TERMINAL (unit + bare qty) still covers
    # the already-split case.
    m_cq_term = re.search(rf'({UNIT_PAT})\s+(M[oớ]i|C[uũ]|New|Used|Refurbished)\s+([\d,\.]+)\s*$',
                           text, re.IGNORECASE | re.UNICODE)
    condition = ""
    if m_cq_term:
        unit = m_cq_term.group(1).upper()
        cond_key = strip_accents(m_cq_term.group(2))
        condition = CONDITION_CANON.get(cond_key, m_cq_term.group(2))
        quantity = parse_qty(m_cq_term.group(3))
        term_start = m_cq_term.start()
    else:
        m_term = RE_TERMINAL.search(text)
        if not m_term:
            return None
        unit     = m_term.group(1).upper()
        quantity = parse_qty(m_term.group(2))
        term_start = m_term.start()

    line_no = ""
    m_no = re.match(r"^\s*(\d{1,4})\s+", text)
    if m_no:
        line_no = m_no.group(1)

    barcode = ""
    for m in RE_BARCODE.finditer(text):
        if len(m.group(1)) in (8, 12, 13, 14):
            barcode = m.group(1)
            break
    prod_code = ""
    sku_raw = ""
    pc_end = 0
    for m in RE_PROD_CODE.finditer(text):
        pc = m.group(1).rstrip('-')
        if pc.count('-') >= 2:
            sku_raw = m.group(1)
            prod_code = normalize_code(pc)
            pc_end = m.end()
            break
    product_name = ""
    if prod_code and pc_end < term_start:
        region = text[pc_end:term_start].strip()
        region = RE_BARCODE.sub('', region).strip()
        parts = [w for w in region.split() if strip_accents(w) not in STATUS_WORDS]
        product_name = re.sub(r'\s+', ' ', ' '.join(parts)).strip()
    if not (barcode or prod_code) or quantity == 0:
        return None
    line_no, product_name = split_leading_no(product_name, line_no)
    prod_code = dequarantine_code(prod_code, "SKU/product_code", "text line")
    barcode   = dequarantine_code(barcode, "EAN/barcode", "text line")
    return Item(no=line_no, product_name=product_name, product_code=prod_code,
                barcode=barcode, unit=unit, quantity=quantity,
                parse_method="text", condition=condition, sku_raw=sku_raw,
                gtin_raw=barcode, source_page=source_page,
                gtin_valid=is_valid_gtin13(barcode))

# ── Parser state machine ───────────────────────────────────────────────────
class Parser:
    """Package state machine. Identity is the *normalized* package code
    (bug #3/#17: continuation pages repeat the same "Ma kien hang" code —
    the code must be reused, never treated as a new package, never
    overwritten/reset). declared_total_qty is only ever set by an actual
    "Tong cong" line (bug #4: a package that starts at the bottom of one
    page and finishes on the next must keep accumulating across the page
    boundary, which this class does simply by not closing `_cur` until a
    real total line, or a different package header, or end-of-document is
    seen)."""

    def __init__(self):
        self.packages: List[Package] = []
        self._cur: Optional[Package] = None
        self._buf: List[str] = []
        self._source_file    = ""
        self._reference_code = ""
        self._page: Optional[int] = None
        # Counters exposed to run_pipeline() so it can implement the
        # per-page "table produced nothing new -> fall back to text" rule
        # (spec section 5) without depending on internal state layout.
        self.duplicate_items_skipped = 0
        # v12: OR#/SO#/Shipping Mark/GW label lines that appear BEFORE the
        # package that owns them opens (real PL layout: metadata block
        # printed above "Ma kien hang"/Packing Code) -- held here until the
        # next _on_pkg_header() call, then applied and immediately cleared.
        # Never forward-filled past that one package (spec section 6).
        self._pending_package_metadata: Dict[str, Tuple[str, str]] = {}

    def set_file(self, pdf_path: Path):
        self._source_file    = pdf_path.name
        self._reference_code = pdf_path.stem

    def set_page(self, page_no: int):
        self._page = page_no

    def total_item_count(self) -> int:
        """Snapshot used by run_pipeline() to detect whether a parsing pass
        added any new item (see run_pipeline's per-page fallback logic)."""
        n = sum(p.item_count for p in self.packages)
        if self._cur is not None:
            n += self._cur.item_count
        return n

    def total_package_count(self) -> int:
        n = len(self.packages)
        if self._cur is not None:
            n += 1
        return n

    def feed_table_row(self, cells: List[str]):
        joined = " ".join(c for c in cells if c)
        if not joined.strip() or is_table_hdr(cells) or is_noise_row(cells):
            return
        m = RE_PKG_HEADER.search(joined)
        if m:
            self._on_pkg_header(m.group(1).upper(), (m.group(2) or "").replace(' ',''))
            return
        m = RE_TOTAL.search(joined)
        if m:
            self._on_total(parse_qty(m.group(1)))
            return
        # v12: package-level OR#/SO#/Shipping Mark/GW as a 2-cell
        # [label,value] table row -- checked BEFORE the "no open package"
        # guard below, because this metadata commonly appears ABOVE the
        # package's own header line in the real PL layout (spec section 6);
        # _capture_metadata() itself queues it as pending when no package is
        # open yet, and _on_pkg_header() applies it to the next package.
        field, value = match_metadata_label_cells(cells)
        if field:
            self._capture_metadata(field, value, "PL_STRUCTURED_FIELD")
            return
        if self._cur is None:
            return
        # Bug #6: a lone "-BOX"/"BOX" row is the continuation of the
        # PREVIOUS row's SKU suffix, split onto its own line by the PDF
        # layout -- not a new item.
        non_empty = [c.strip() for c in cells if c and c.strip()]
        if len(non_empty) == 1 and RE_SKU_SUFFIX_ROW.match(non_empty[0]) and self._cur.items:
            self._append_sku_suffix(non_empty[0])
            return
        item = parse_item_cells(cells, source_page=self._page)
        if item:
            self._add_item(item)

    def feed_text_line(self, line: str):
        line = line.strip()
        if not line or is_noise(line):
            return
        m = RE_PKG_HEADER.search(line)
        if m:
            self._flush_buf()
            self._on_pkg_header(m.group(1).upper(), (m.group(2) or "").replace(' ',''))
            return
        m = RE_TOTAL.search(line)
        if m:
            self._flush_buf()
            self._on_total(parse_qty(m.group(1)))
            return
        # v12: same reasoning as the table-row case above -- checked BEFORE
        # the "no open package" guard so metadata printed above the package
        # header is queued as pending instead of silently dropped.
        field, value = match_metadata_label_line(line)
        if field:
            self._capture_metadata(field, value, "PL_TEXT")
            return
        if self._cur is None:
            return
        if RE_SKU_SUFFIX_ROW.match(line) and self._cur.items and not self._buf:
            self._append_sku_suffix(line)
            return
        if RE_BARCODE.search(line):
            if self._buf:
                item = parse_item_text(' '.join(self._buf), source_page=self._page)
                if item:
                    self._add_item(item)
                self._buf = []
            self._buf.append(line)
            item = parse_item_text(' '.join(self._buf), source_page=self._page)
            if item:
                self._add_item(item)
                self._buf = []
            return
        if self._buf:
            self._buf.append(line)
            item = parse_item_text(' '.join(self._buf), source_page=self._page)
            if item:
                self._add_item(item)
                self._buf = []

    def end_of_pdf(self):
        self._flush_buf()

    def finalise(self):
        self._flush_buf()
        if self._cur is not None:
            log.warning(f"EOF: {self._cur.package_code} never saw Tong cong")
            self._force_close()

    def _capture_metadata(self, field: str, value: str, source: str):
        """First match wins -- never overwrite an already-captured
        OR#/SO#/Shipping Mark/GW with a later, possibly-spurious match (same
        'never silently overwrite' philosophy as declared_total_qty / DIM
        duplicate-key handling elsewhere in this file).

        v12: when NO package is currently open (metadata block printed
        above "Ma kien hang"/Packing Code -- the common real layout), the
        value is held in _pending_package_metadata instead of being
        dropped, and gets applied to the NEXT package opened by
        _on_pkg_header() (then immediately cleared -- spec section 6:
        "khong duoc forward-fill vao mot package khong lien quan sau do").
        If a package IS already open, capture goes straight onto it, scoped
        to that package only, exactly as before."""
        value = (value or "").strip()
        if not value:
            return
        if self._cur is None:
            if field not in self._pending_package_metadata:
                self._pending_package_metadata[field] = (value, source)
            return
        if getattr(self._cur, field):
            return
        setattr(self._cur, field, value)
        setattr(self._cur, _METADATA_SOURCE_FIELD[field], source)

    def _apply_pending_metadata(self, pkg: "Package"):
        """Apply whatever metadata was captured before `pkg` opened, then
        clear it unconditionally so it can never leak into a LATER package
        (spec section 6, item 3-4). First-match-wins is preserved: if `pkg`
        already has a value for a field (e.g. a re-opened package that had
        already captured it directly), the pending value is discarded for
        that field rather than overwriting it."""
        if pkg is None:
            self._pending_package_metadata = {}
            return
        for field, (value, source) in self._pending_package_metadata.items():
            if not getattr(pkg, field):
                setattr(pkg, field, value)
                setattr(pkg, _METADATA_SOURCE_FIELD[field], source)
        self._pending_package_metadata = {}

    def _append_sku_suffix(self, suffix_cell: str):
        """Bug #6 fix: join a stray '-BOX' style continuation row onto the
        most recently added item's SKU instead of dropping it or treating it
        as a bogus new item."""
        suffix = suffix_cell.lstrip('-').upper()
        last = self._cur.items[-1]
        joined = last.product_code if last.product_code.endswith('-' + suffix) else                  (last.product_code.rstrip('-') + '-' + suffix)
        log.info(f"SKU suffix continuation: {last.product_code!r} -> {joined!r}")
        last.product_code = joined
        last.sku_raw = (last.sku_raw or '') + suffix_cell

    def _on_pkg_header(self, pkg_code: str, seq: str):
        if self._cur is not None and self._cur.package_code == pkg_code:
            self._cur.header_count += 1
            # v12: a continuation-page repeat of the SAME package's header
            # can still be preceded by its own stray metadata line (e.g. a
            # re-printed "GW:" on the continuation page) -- apply/clear any
            # pending metadata onto the (unchanged) current package too.
            self._apply_pending_metadata(self._cur)
            return
        if self._cur is not None:
            log.warning(f"INTERRUPTED: {self._cur.package_code} -> {pkg_code}")
            self._force_close()
        # Bug #17 defensive fix: if this code was already finalized earlier
        # (e.g. a premature/duplicate "Tong cong" closed it, then more rows
        # for the same package follow) re-open the SAME package object
        # instead of creating a duplicate. declared_total_qty is cleared so
        # it only gets set again by a real "Tong cong" line, per spec
        # section 7 ("declared total chi cap nhat khi gap dong Tong cong
        # hop le").
        if self.packages and self.packages[-1].package_code == pkg_code:
            log.warning(f"RE-OPENED package {pkg_code} after premature close "
                        f"(more rows followed the earlier 'Tong cong')")
            self._cur = self.packages.pop()
            self._cur.declared_total_qty = None
            self._cur.header_count += 1
            self._apply_pending_metadata(self._cur)
            return
        self._cur = Package(package_code=pkg_code,
                            source_file=self._source_file,
                            reference_code=self._reference_code,
                            pdf_package_seq=seq,
                            first_page=self._page)
        # v12: apply whatever OR#/SO#/Shipping Mark/GW label lines were seen
        # BEFORE this package's own header (the common real layout: metadata
        # block printed above "Ma kien hang"/Packing Code) -- then clear
        # them so they can never leak into a later, unrelated package.
        self._apply_pending_metadata(self._cur)

    def _on_total(self, declared: int):
        if self._cur is None:
            log.warning(f"Orphan Tong cong={declared}")
            return
        self._cur.declared_total_qty = declared
        self.packages.append(self._cur)
        self._cur = None

    def _force_close(self):
        if self._cur is not None:
            self.packages.append(self._cur)
            self._cur = None

    def _flush_buf(self):
        if self._buf and self._cur is not None:
            item = parse_item_text(' '.join(self._buf), source_page=self._page)
            if item:
                self._add_item(item)
        self._buf = []

    def _add_item(self, item: Item):
        """Append an item to the current package, deduplicating on full
        context (spec section 5): package + page + line_no/position + GTIN
        + normalized SKU + quantity + condition -- NOT just GTIN+quantity,
        since the same SKU can legitimately repeat within a package (e.g.
        the same color restocked in two different cartons of one package
        is rare but the same SKU CAN legitimately appear twice on
        genuinely different rows, which is why line_no/page is part of the
        key rather than being ignored)."""
        key = (self._page, item.no, normalize_sku_key(item.product_code),
               item.barcode, item.quantity, item.condition)
        if key in self._cur._seen_item_keys:
            self.duplicate_items_skipped += 1
            log.warning(f"DUPLICATE_ITEM_SUSPECTED skipped in {self._cur.package_code}: "
                        f"sku={item.product_code!r} gtin={item.barcode!r} qty={item.quantity} "
                        f"page={self._page} line_no={item.no!r}")
            return
        self._cur._seen_item_keys.add(key)
        item.dedup_key = key
        self._cur.items.append(item)

# ── DIM mapper ─────────────────────────────────────────────────────────────
RE_PACKAGE_CODE_DIM = re.compile(r'^PGKEC[A-Z0-9]+$', re.IGNORECASE)


class DimMapper:
    """DIM/weight lookup, keyed by normalized "reference|package_code".

    Detection runs a strict priority order and never silently guesses past
    it (spec section 10):
      1) HEADER_MAPPING              -- every column found by header alias
      2) PARTIAL_HEADER_POSITIONAL   -- ref+pkg found by header, L/W/H/Wt/CBM
                                         filled positionally (pkg_col+1..+5)
      3) PACKAGE_CODE_POSITIONAL_FALLBACK -- header detection failed entirely;
                                         anchor on a cell matching ^PGKEC...$
                                         and read the fixed 5 cells after it
                                         (Length, Width, Height, Weight, CBM
                                         -- this exact order, never changed)
      4) FAIL_WITH_DIAGNOSTIC        -- neither worked; sheet contributes 0
                                         rows and every reason is logged.
    """
    _ALIASES: Dict[str, List[str]] = {
        "ref":    ["lo","lot","lohang","reference_code","reference","ref","job","shipment"],
        "pkg":    ["tracking","package_code","package","pkg","carton_code","carton","kien","makien"],
        "length": ["dai","length","l","len","d","chieudai"],
        "width":  ["rong","width","w","wid","r","chieurong"],
        "height": ["cao","height","h","hei","high","c","chieucao"],
        "weight": ["kg","weight","wt","gross","gw","nang"],
        "cbm":    ["cbm","volume","vol","cubic","m3"],
    }
    _POSITIONAL_ORDER = ["length", "width", "height", "weight", "cbm"]

    def __init__(self, xlsx_path: Path, sheet_name: Optional[str] = None):
        self._data: Dict[str, dict] = {}
        self.diagnostics: List[dict] = []          # one row per DIM row processed (spec section 11)
        self.detection_method = "FAIL_WITH_DIAGNOSTIC"
        self.selected_sheet: Optional[str] = None
        self.header_row: Optional[int] = None
        self.headers_detected: List[str] = []
        self.canonical_mapping: Dict[str, str] = {}
        self.rows_scanned = 0
        self.valid_rows = 0
        self.duplicate_keys = 0
        self.malformed_rows = 0
        self.review_required = 0
        self._known_refs: set = set()
        self._known_pkgs: set = set()
        self._load(xlsx_path, sheet_name)

    # ── loading ──────────────────────────────────────────────────────────
    def _load(self, path: Path, sheet_name: Optional[str] = None):
        log.info(f"Loading DIM <- {path.name}")
        try:
            wb = openpyxl.load_workbook(str(path), data_only=True)
        except Exception as e:
            log.error(f"Cannot open DIM: {e}")
            return

        if sheet_name and sheet_name in wb.sheetnames:
            sheets_to_try = [sheet_name]
        else:
            if sheet_name:
                log.warning(f"DIM sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}. Auto-detect instead.")
            sheets_to_try = wb.sheetnames

        for sheet in sheets_to_try:
            ws = wb[sheet]
            if self._load_sheet(ws, sheet):
                self.selected_sheet = sheet
                break

        for key in self._data:
            ref, pkg = key.split("|", 1)
            self._known_refs.add(ref)
            self._known_pkgs.add(pkg)

        log.info(f"  DIM detection method: {self.detection_method} (sheet={self.selected_sheet!r})")
        log.info(f"  Rows scanned={self.rows_scanned} valid={self.valid_rows} "
                 f"malformed={self.malformed_rows} duplicate_keys={self.duplicate_keys} "
                 f"review_required={self.review_required}")
        log.info(f"  Total DIM records: {len(self._data)}")

    @staticmethod
    def _merge_map(ws) -> Dict[Tuple[int, int], object]:
        """Every cell coordinate covered by a merged range maps to the
        top-left cell's value, so reading a merged-away cell (e.g. a
        reference column merged down 2 rows) returns the real value instead
        of None."""
        m: Dict[Tuple[int, int], object] = {}
        for rng in ws.merged_cells.ranges:
            top_left = ws.cell(row=rng.min_row, column=rng.min_col).value
            for r in range(rng.min_row, rng.max_row + 1):
                for c in range(rng.min_col, rng.max_col + 1):
                    m[(r, c)] = top_left
        return m

    @staticmethod
    def _cell(ws, r: int, c: int, merge_map) -> object:
        v = ws.cell(row=r, column=c).value
        if v is None:
            v = merge_map.get((r, c))
        return v

    def _load_sheet(self, ws, sheet_name: str) -> bool:
        max_row, max_col = ws.max_row, ws.max_column
        if max_row < 2 or max_col < 1:
            return False
        merge_map = self._merge_map(ws)

        header_row_idx, cm = self._find_header_row(ws, merge_map, max_row, max_col)
        if header_row_idx and cm and "ref" in cm and "pkg" in cm:
            method = ("HEADER_MAPPING"
                      if all(f in cm for f in self._ALIASES)
                      else "PARTIAL_HEADER_POSITIONAL")
            n = self._extract_with_header(ws, merge_map, header_row_idx, cm,
                                           sheet_name, method, max_row, max_col)
            if n:
                self.detection_method = method
                self.header_row = header_row_idx
                self.headers_detected = [self._cell(ws, header_row_idx, c, merge_map)
                                          for c in range(1, max_col + 1)]
                self.canonical_mapping = dict(cm)
                return True
            log.warning(f"  Sheet '{sheet_name}': header row found (row {header_row_idx}) "
                        f"but 0 valid rows extracted -- trying PACKAGE_CODE_POSITIONAL_FALLBACK.")

        n = self._extract_positional_by_package_code(ws, merge_map, sheet_name, max_row, max_col)
        if n:
            self.detection_method = "PACKAGE_CODE_POSITIONAL_FALLBACK"
            return True

        log.error(f"  Sheet '{sheet_name}': FAIL_WITH_DIAGNOSTIC -- no header mapping "
                  f"(ref+pkg columns) and no cell matching ^PGKEC...$ found anywhere "
                  f"in {max_row}x{max_col} cells.")
        return False

    # ── tier 1 / 2: header-based ─────────────────────────────────────────
    def _find_header_row(self, ws, merge_map, max_row, max_col):
        """Scan the first few rows for one containing at least ref+pkg
        header aliases (exact or substring match)."""
        for r in range(1, min(6, max_row) + 1):
            cols = [str(self._cell(ws, r, c, merge_map) or "").strip() for c in range(1, max_col + 1)]
            if not any(cols):
                continue
            cm = self._detect(cols)
            if cm and "ref" in cm and "pkg" in cm:
                # cm values are column header strings; convert to 1-based column index
                col_idx = {name: cols.index(hdr) + 1 for name, hdr in cm.items()}
                return r, col_idx
        return None, None

    def _extract_with_header(self, ws, merge_map, header_row_idx, col_idx,
                              sheet_name, method, max_row, max_col) -> int:
        # `loaded` must count rows *successfully committed* to self._data,
        # not rows merely attempted -- a bug found while writing the
        # PARTIAL_HEADER_POSITIONAL regression test: the old code did
        # `loaded += 1` unconditionally after calling _commit_row(), so a
        # header-detected sheet where every row was malformed (missing
        # dimension, non-positive, duplicate key) still made `if n:` in
        # _load_sheet() true, silently reporting HEADER_MAPPING /
        # PARTIAL_HEADER_POSITIONAL as "successful" with 0 usable DIM
        # records AND never falling through to try
        # PACKAGE_CODE_POSITIONAL_FALLBACK, which might have recovered the
        # row. _commit_row() now returns bool so the caller's count reflects
        # reality.
        loaded = 0
        last_ref = None
        pkg_col = col_idx.get("pkg")
        for r in range(header_row_idx + 1, max_row + 1):
            self.rows_scanned += 1
            raw_ref = self._cell(ws, r, col_idx["ref"], merge_map)
            raw_pkg = self._cell(ws, r, col_idx["pkg"], merge_map)
            ref_norm = normalize(str(raw_ref)) if raw_ref not in (None, "") else ""
            if not ref_norm or ref_norm == "NAN":
                # forward-fill a blank reference from the nearest row above
                # (merged "Lo hang" cell) -- never fabricate a ref that was
                # never actually present.
                ref_norm = last_ref or ""
            else:
                last_ref = ref_norm
            pkg_norm = normalize(str(raw_pkg)) if raw_pkg not in (None, "") else ""
            if not ref_norm or not pkg_norm or pkg_norm == "NAN":
                continue  # blank row / spacer row, not malformed
            if method == "PARTIAL_HEADER_POSITIONAL" and pkg_col is not None:
                # Header only found ref+pkg (and maybe some, but not all, of
                # L/W/H/Wt/CBM) -- per spec, tier 2 reads ALL FIVE dimension
                # cells positionally (pkg_col+1..+5), the same fixed order
                # as tier 3, rather than trusting a partial/ambiguous header
                # match for some dimensions and not others.
                vals = {f: safe_float(self._cell(ws, r, pkg_col + i, merge_map))
                        for i, f in enumerate(self._POSITIONAL_ORDER, start=1)}
            else:
                vals = {f: safe_float(self._cell(ws, r, col_idx[f], merge_map)) if f in col_idx else None
                        for f in self._POSITIONAL_ORDER}
            if self._commit_row(sheet_name, r, method, col_idx.get("pkg"),
                                 raw_ref, ref_norm, raw_pkg, pkg_norm, vals):
                loaded += 1
        return loaded

    # ── tier 3: PACKAGE_CODE_POSITIONAL_FALLBACK ─────────────────────────
    def _extract_positional_by_package_code(self, ws, merge_map, sheet_name, max_row, max_col) -> int:
        loaded = 0
        last_ref = None
        for r in range(1, max_row + 1):
            pkg_col = None
            raw_pkg = None
            for c in range(1, max_col + 1):
                v = self._cell(ws, r, c, merge_map)
                if v is None:
                    continue
                candidate = normalize_code(str(v))
                if RE_PACKAGE_CODE_DIM.match(candidate):
                    pkg_col, raw_pkg = c, v
                    break
            if pkg_col is None:
                continue
            self.rows_scanned += 1
            if pkg_col + 5 > max_col:
                self.malformed_rows += 1
                self.diagnostics.append(self._diag(
                    sheet_name, r, "PACKAGE_CODE_POSITIONAL_FALLBACK", pkg_col,
                    None, None, raw_pkg, normalize_code(str(raw_pkg)),
                    {}, None, None, "INSUFFICIENT_POSITIONAL_FIELDS"))
                log.warning(f"  Row {r}: INSUFFICIENT_POSITIONAL_FIELDS "
                            f"(package code at col {pkg_col}, needs 5 cells after it, "
                            f"sheet only has {max_col} columns)")
                continue
            # reference: nearest non-blank, non-purely-numeric cell to the
            # LEFT of the package-code column on this row (spec: "khong lay
            # nham STT hoac text header lam reference"); else forward-fill
            # from the previous row that had one.
            raw_ref = None
            for c in range(pkg_col - 1, 0, -1):
                v = self._cell(ws, r, c, merge_map)
                if v is None or str(v).strip() == "":
                    continue
                if re.fullmatch(r'\d{1,3}', str(v).strip()):
                    continue  # looks like an STT / row-index column, skip it
                raw_ref = v
                break
            if raw_ref is not None:
                ref_norm = normalize(str(raw_ref))
                last_ref = ref_norm
            else:
                ref_norm = last_ref or ""
            pkg_norm = normalize_code(str(raw_pkg))
            if not ref_norm:
                self.malformed_rows += 1
                self.diagnostics.append(self._diag(
                    sheet_name, r, "PACKAGE_CODE_POSITIONAL_FALLBACK", pkg_col,
                    raw_ref, "", raw_pkg, pkg_norm, {}, None, None, "MISSING_REFERENCE"))
                continue
            vals = {}
            for i, field_name in enumerate(self._POSITIONAL_ORDER, start=1):
                vals[field_name] = safe_float(self._cell(ws, r, pkg_col + i, merge_map))
            if self._commit_row(sheet_name, r, "PACKAGE_CODE_POSITIONAL_FALLBACK", pkg_col,
                                 raw_ref, ref_norm, raw_pkg, pkg_norm, vals):
                loaded += 1
        return loaded

    # ── shared row commit (validation, CBM cross-check, dedup) ──────────
    def _commit_row(self, sheet_name, excel_row, method, pkg_col,
                     raw_ref, ref_norm, raw_pkg, pkg_norm, vals: Dict[str, Optional[float]]):
        length, width, height, weight, cbm = (vals.get(f) for f in self._POSITIONAL_ORDER)
        missing = [f for f in self._POSITIONAL_ORDER if vals.get(f) is None]
        non_positive = [f for f in self._POSITIONAL_ORDER
                         if vals.get(f) is not None and vals[f] <= 0]
        if missing or non_positive:
            self.malformed_rows += 1
            reason = "NON_NUMERIC_OR_MISSING_DIMENSION" if missing else "NON_POSITIVE_DIMENSION"
            self.diagnostics.append(self._diag(sheet_name, excel_row, method, pkg_col,
                                                raw_ref, ref_norm, raw_pkg, pkg_norm,
                                                vals, None, None, reason))
            log.warning(f"  Row {excel_row} ({ref_norm}|{pkg_norm}): {reason} "
                        f"missing={missing} non_positive={non_positive}")
            return False

        expected_cbm = (length * width * height) / 1_000_000 if length and width and height else None
        cbm_diff = abs(cbm - expected_cbm) if (cbm is not None and expected_cbm is not None) else None
        tolerance = max(0.001, expected_cbm * 0.05) if expected_cbm is not None else None
        status = "OK"
        if cbm_diff is not None and tolerance is not None and cbm_diff > tolerance:
            status = "DIM_REVIEW_REQUIRED"
            self.review_required += 1
            log.warning(f"  Row {excel_row} ({ref_norm}|{pkg_norm}): DIM_REVIEW_REQUIRED "
                        f"declared_cbm={cbm} expected_cbm={round(expected_cbm,6)} "
                        f"diff={round(cbm_diff,6)} tolerance={round(tolerance,6)} "
                        f"-- original CBM kept as-is, NOT overwritten.")

        key = f"{ref_norm}|{pkg_norm}"
        if key in self._data:
            self.duplicate_keys += 1
            log.warning(f"  Row {excel_row}: DUPLICATE DIM key {key!r} -- keeping the "
                        f"first occurrence, this row is NOT applied (no silent overwrite).")
            self.diagnostics.append(self._diag(sheet_name, excel_row, method, pkg_col,
                                                raw_ref, ref_norm, raw_pkg, pkg_norm, vals,
                                                expected_cbm, cbm_diff, "DUPLICATE_KEY"))
            return False

        self._data[key] = {"length": length, "width": width, "height": height,
                            "weight": weight, "cbm": cbm}
        self.valid_rows += 1
        self.diagnostics.append(self._diag(sheet_name, excel_row, method, pkg_col,
                                            raw_ref, ref_norm, raw_pkg, pkg_norm, vals,
                                            expected_cbm, cbm_diff, status, key=key))
        return True

    @staticmethod
    def _diag(sheet, excel_row, method, pkg_col, raw_ref, ref_norm, raw_pkg, pkg_norm,
              vals, expected_cbm, cbm_diff, status, key=None):
        return {
            "sheet": sheet, "excel_row": excel_row, "detection_method": method,
            "package_column_index": pkg_col,
            "raw_reference": raw_ref, "normalized_reference": ref_norm,
            "raw_package_code": raw_pkg, "normalized_package_code": pkg_norm,
            "length": vals.get("length"), "width": vals.get("width"),
            "height": vals.get("height"), "weight": vals.get("weight"), "cbm": vals.get("cbm"),
            "expected_cbm": round(expected_cbm, 6) if expected_cbm is not None else None,
            "cbm_diff": round(cbm_diff, 6) if cbm_diff is not None else None,
            "validation_status": status,
            "exact_dim_key": key,
        }

    # ── header alias detection (unchanged logic, kept as a static method) ─
    def _detect(self, cols: List[str]) -> Optional[Dict[str, str]]:
        def norm(s):
            return re.sub(r'[^a-z0-9]', '', strip_accents(s).lower())
        nc = {norm(c): c for c in cols if c}
        mapping: Dict[str, str] = {}
        claimed: set = set()
        for field_name, aliases in self._ALIASES.items():
            found = None
            for alias in aliases:
                a = norm(alias)
                if a in nc and nc[a] not in claimed:
                    found = nc[a]
                    break
            if found is None:
                for alias in aliases:
                    a = norm(alias)
                    if len(a) < 2:
                        continue
                    for key, orig_col in nc.items():
                        if orig_col in claimed:
                            continue
                        if a in key:
                            found = orig_col
                            break
                    if found:
                        break
            if found:
                mapping[field_name] = found
                claimed.add(found)
        if "ref" not in mapping or "pkg" not in mapping:
            return None
        return mapping

    # ── lookup (used by run_pipeline) ────────────────────────────────────
    def lookup(self, ref: str, pkg: str) -> Optional[dict]:
        key = f"{normalize(ref)}|{normalize(pkg)}"
        return self._data.get(key)

    def diagnose_miss(self, source_file: str, raw_ref: str, raw_pkg: str) -> dict:
        """Read-only diagnostic for a PDF package that found no exact DIM
        match: nearest reference/package candidates are for REVIEW ONLY --
        never used to auto-assign a DIM record (spec: no fuzzy auto-match)."""
        ref_norm, pkg_norm = normalize(raw_ref), normalize(raw_pkg)
        key = f"{ref_norm}|{pkg_norm}"
        nearest_ref = difflib.get_close_matches(ref_norm, self._known_refs, n=1, cutoff=0.6)
        nearest_pkg = difflib.get_close_matches(pkg_norm, self._known_pkgs, n=1, cutoff=0.6)
        reason = ("NO_REFERENCE_IN_DIM" if ref_norm not in self._known_refs
                  else "NO_PACKAGE_CODE_IN_DIM" if pkg_norm not in self._known_pkgs
                  else "KEY_COMBINATION_NOT_FOUND")
        return {
            "source_file": source_file, "raw_pdf_reference": raw_ref,
            "normalized_pdf_reference": ref_norm, "raw_pdf_package_code": raw_pkg,
            "normalized_pdf_package_code": pkg_norm, "exact_lookup_key": key,
            "nearest_reference_candidate": nearest_ref[0] if nearest_ref else "",
            "nearest_package_candidate": nearest_pkg[0] if nearest_pkg else "",
            "mismatch_reason": reason,
        }

# ── HS Code mapper ─────────────────────────────────────────────────────────
def clean_excel_key(text: str) -> str:
    """Clean key for exact Excel-like SKU matching while preserving hyphens."""
    if text is None:
        return ""
    try:
        if pd.isna(text):
            return ""
    except Exception:
        pass
    return re.sub(r"\s+", " ", str(text)).strip()

def norm_col_name(text: str) -> str:
    text = clean_excel_key(text)
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = text.upper()
    return re.sub(r"[^A-Z0-9]+", "", text)

class HsCodeMapper:
    def __init__(self, master_path: Optional[Path] = None, sheet_name: Optional[str] = None):
        self._data_exact: Dict[str, str] = {}
        self._data_norm: Dict[str, str] = {}
        self._data_barcode: Dict[str, str] = {}
        if master_path:
            self._load(master_path, sheet_name)

    def _load(self, path: Path, sheet_name: Optional[str] = None):
        if not path or not path.exists():
            log.warning(f"Master data file not found: {path}")
            return
        log.info(f"Loading HS Code <- {path.name}")
        try:
            xl = pd.ExcelFile(str(path))
        except Exception as e:
            log.error(f"Cannot open Master Data: {e}")
            return

        sheets_to_try = [sheet_name] if sheet_name and sheet_name in xl.sheet_names else xl.sheet_names
        if sheet_name and sheet_name not in xl.sheet_names:
            log.warning(f"Master data sheet '{sheet_name}' not found. Auto-detect instead.")

        for sheet in sheets_to_try:
            try:
                raw = pd.read_excel(path, sheet_name=sheet, header=None, dtype=str)
            except Exception as e:
                log.warning(f"  Sheet '{sheet}' error: {e}")
                continue
            if raw.empty:
                continue

            header_row_idx = None
            for i in range(min(40, len(raw))):
                row_join = " | ".join(clean_excel_key(x) for x in raw.iloc[i].tolist()).upper()
                row_norm = norm_col_name(row_join)
                if ("SKU" in row_norm or "PRODUCTCODE" in row_norm) and "HSCODE" in row_norm:
                    header_row_idx = i
                    break
            if header_row_idx is None:
                continue

            df = pd.read_excel(path, sheet_name=sheet, header=header_row_idx, dtype=str)
            df.columns = [clean_excel_key(c) for c in df.columns]

            sku_col = hs_col = None
            barcode_cols: List[str] = []
            for c in df.columns:
                n = norm_col_name(c)
                if n in {"SKUPRODUCTCODE", "SKU", "PRODUCTCODE"} or ("SKU" in n and "PRODUCT" in n and "CODE" in n):
                    sku_col = c
                if n == "HSCODE" or ("HS" in n and "CODE" in n):
                    hs_col = c
                if n in {"EAN", "BARCODE", "UPC"} or "EAN" in n or "BARCODE" in n:
                    barcode_cols.append(c)

            if not sku_col or not hs_col:
                log.warning(f"  Sheet '{sheet}': cannot find SKU / HS Code columns")
                continue

            loaded = 0
            for _, row in df.iterrows():
                sku = clean_excel_key(row.get(sku_col))
                hs = clean_excel_key(row.get(hs_col))
                if not sku or not hs:
                    continue
                self._data_exact[sku] = hs
                self._data_exact[sku.upper()] = hs
                sku_norm = normalize_sku_key(sku)
                if sku_norm:
                    self._data_norm[sku_norm] = hs
                for bc in barcode_cols:
                    barcode = re.sub(r"\D", "", clean_excel_key(row.get(bc)))
                    if barcode:
                        self._data_barcode[barcode] = hs
                loaded += 1
            log.info(f"  Sheet '{sheet}': loaded {loaded} HS Code records")

        log.info(
            f"  Total HS Code records: exact={len(self._data_exact)}, "
            f"normalized={len(self._data_norm)}, barcode={len(self._data_barcode)}"
        )

    def lookup(self, sku: str, barcode: str = "") -> str:
        sku_clean = clean_excel_key(sku)
        if sku_clean in self._data_exact:
            return self._data_exact[sku_clean]
        if sku_clean.upper() in self._data_exact:
            return self._data_exact[sku_clean.upper()]
        sku_norm = normalize_sku_key(sku_clean)
        if sku_norm in self._data_norm:
            return self._data_norm[sku_norm]
        barcode_clean = re.sub(r"\D", "", clean_excel_key(barcode))
        if barcode_clean in self._data_barcode:
            return self._data_barcode[barcode_clean]
        return ""

# ── v11: business document/carton ordering ─────────────────────────────────
# Fixes the pre-v11 bug where PDFs (and therefore cartons) were ordered by
# plain lexical filename sort -- "Kerry_POP_1" < "Kerry_POP_10" < "Kerry_POP_2"
# alphabetically, which is wrong (spec section 4/7). Order is now:
#   1) factory business order (CARTON_FACTORY_ORDER_WITH_CO in
#      pl_group_export.py: POP -> SBGEAR -> QIFENG -> JION -> CN; unknown/
#      REVIEW factories sort last, never crash the sort)
#   2) within a factory, by base_document_key (e.g. "Kerry_POP") so an
#      original PDF and all its "_1"/"_2" copies stay adjacent, never
#      interleaved with a different base document
#   3) within a base_document_key, by numeric document_sequence (original=0
#      first, then _1, _2, ... _10 in NUMERIC order, never lexical)
#   4) within one PDF, Python's sort is stable, so packages keep the exact
#      relative order the parser produced them in (== pdf_package_seq order)
#      without needing a separate explicit key.
#
# KNOWN LIMITATION (documented, not silently decided): full "CO" grouping
# (spec 7.2 -- one retail Store spanning several Factories, carton numbers
# continuous across the whole Store) needs a Store assigned to EVERY
# package, including non-CN factories (POP/SBGEAR/QIFENG/JION). Today, Store
# resolution (classify_packages_for_port / match_store) only runs for
# factory=CN packages -- non-CN packages have no Store source at all in this
# codebase yet (that requires the OR List Store-enrichment feature, not yet
# wired in). So this implementation orders by FACTORY BUSINESS ORDER only
# (which already satisfies "VN factories before CN" for the no-CO case, since
# POP/SBGEAR/QIFENG/JION all precede CN in CARTON_FACTORY_ORDER_WITH_CO) --
# it does not yet interleave-by-Store across factories for a true multi-
# factory CO shipment. Numbering stays global/continuous either way (no
# reset between factories/PDFs), which is the part of 7.2 this DOES fully
# satisfy today.
def business_sort_packages(packages: List[Package]) -> List[Package]:
    try:
        import pl_group_export as pge
        factory_order = list(pge.carton_factory_rank_table())
        detect_factory = pge.detect_factory
    except ImportError:
        log.warning("pl_group_export not importable -- carton business sort "
                    "falls back to input order (parse/filename order) unchanged.")
        return list(packages)

    rank = {f: i for i, f in enumerate(factory_order)}
    review_rank = len(factory_order)  # unknown/REVIEW factories sort last
    log.info(f"Sorting {len(packages)} carton(s) by business order "
             f"(factory {factory_order} -> document -> sequence)...")

    def key(pkg: Package):
        # detect_factory() now strips a trailing copy-suffix (_1/_2/...)
        # internally (v12 fix, single source of truth in pl_group_export.py)
        # so this can call it directly on the raw reference_code/source_file
        # exactly like every other caller does -- no local workaround needed
        # here any more.
        factory = detect_factory(pkg.reference_code, pkg.source_file)
        seq = parse_document_sequence(pkg.reference_code, pkg.source_file)
        return (rank.get(factory, review_rank), seq.base_document_key, seq.document_sequence)

    return sorted(packages, key=key)  # sorted() is stable -> ties keep parse order


# ── v12: per-Store counting scope (spec section 9) ──────────────────────────
def compute_counting_scope_key(pkg: Package) -> Tuple[str, str]:
    """(counting_scope_key, scope_source). Store comes from an OK OR List
    match first, else the existing CN-only `pkg.store` (classify_packages_
    for_port), else "UNRESOLVED" (never crashes/blocks on an unknown store).

    shipment_key: the OR List's own OR value is the strongest available
    signal ("OR List grouping" per spec's list of possible signals) --
    packages sharing the same Store AND the same OR are almost certainly the
    same customer order/shipment. Absent that, this tool's own existing
    workflow is already scoped to "1 upload = 1 lo hang / 1 shipment" (see
    app.html's own long-standing instructions: "Chon thu muc chua toan bo
    PDF Packing List cua 1 lo hang") -- so "the whole upload is one
    shipment" is not an arbitrary guess, it's the tool's existing, already-
    documented assumption, used here as the fallback shipment_key when no
    OR List disambiguates further. KNOWN LIMITATION: two independent
    shipments for the SAME Store combined into one upload, with no OR List
    to tell them apart, will still share one counting scope -- there is no
    other reliable shipment-boundary signal available in the PL PDFs' own
    structured data today. Flagged explicitly rather than guessed at
    silently (spec: "neu khong the xac dinh tin cay, danh dau REVIEW" --
    the ambiguity itself is documented here rather than resolved by
    invention)."""
    try:
        import pl_group_export as pge
        store_identity_fn = pge.store_identity
    except ImportError:
        store_identity_fn = lambda s: str(s or "").strip().upper()

    # v14 (spec sections 3/11): non-China countries (KR/JP/BE/US/TW) are
    # SINGLE_DESTINATION -- never apply the China multi-Store split/
    # numbering logic to them, even if an OR List or the CN-only classifier
    # happens to produce a store-looking match. `pkg.country` is only ever
    # a known non-"" value once detect_shipment_country() has matched one
    # of the 6 explicit country codes (never a guess), and "" (unknown/not
    # yet detected -- e.g. a caller/test that never touched this v14
    # field) intentionally falls through to the pre-v14 behavior below
    # unchanged, so every existing caller keeps working exactly as before.
    _NON_CN_COUNTRIES = {"KR", "JP", "BE", "US", "TW"}
    if pkg.country in _NON_CN_COUNTRIES:
        store = "UNRESOLVED"
    elif pkg.or_list_match_status == "OK" and pkg.or_list_store:
        store = pkg.or_list_store
    elif pkg.store and pkg.store != "REVIEW":
        store = pkg.store
    else:
        store = "UNRESOLVED"

    if pkg.or_list_match_status == "OK" and pkg.or_number:
        shipment_key = f"OR:{store_identity_fn(pkg.or_number)}"
        source = "OR_LIST_GROUPING"
    else:
        shipment_key = "UPLOAD_BATCH"
        source = "UPLOAD_BATCH_DEFAULT"

    return f"{shipment_key}|{store_identity_fn(store)}", source


def assign_counting_scope_keys(packages: List[Package]):
    for pkg in packages:
        pkg.counting_scope_key, pkg.counting_scope_source = compute_counting_scope_key(pkg)


# ── Carton numbers (per counting scope) ─────────────────────────────────────
def assign_global_numbers(packages: List[Package]):
    """Numbers each counting_scope_key group independently as 1/N .. N/N
    (spec section 9) -- NOT one denominator across the whole `packages`
    list. `counting_scope_key` defaults to "" on every Package (v12), so
    any caller that never touches counting_scope_key (every pre-v12 test,
    and the whole pipeline when no OR List is uploaded and no package
    resolves a Store at all) puts every package into that one shared ""
    scope -- i.e. exactly today's flat global "1/total ... total/total"
    numbering, byte-for-byte. The name `assign_global_numbers` and the
    `global_carton_num` field it sets are both kept unchanged so every
    existing caller keeps working without modification."""
    scopes: Dict[str, List[Package]] = {}
    for pkg in packages:
        scopes.setdefault(pkg.counting_scope_key, []).append(pkg)
    for scope_key, scoped_pkgs in scopes.items():
        total = len(scoped_pkgs)
        for i, pkg in enumerate(scoped_pkgs, start=1):
            pkg.carton_sequence = i
            pkg.carton_total = total
            pkg.carton_display = f"{i}/{total}"
            pkg.global_carton_num = pkg.carton_display  # unchanged shape/field, kept in sync
        label = scope_key if scope_key else "(single implicit scope)"
        log.info(f"Carton numbers for scope {label!r}: 1/{total} ... {total}/{total}")

# ── v8: CN store/port classification (same rule as pl_group_export.py) ─────
def classify_packages_for_port(packages: List[Package], pdf_folder: Path, recursive: bool):
    """Fill pkg.port / pkg.store for every CN-factory package, using the exact
    same detect_factory() + match_store() rule already used to split CN
    shipments by store/port (pl_group_export.py) — "quy luật chia port và
    store" the warehouse team already uses. Non-CN-factory packages (POP,
    SBGEAR, QIFENG, JION, or unclassifiable) are left with port="" / store=""
    (blank) — fill in manually if needed, same as OR No. / SO No. when no
    OR List match is available."""
    try:
        import pl_group_export as pge
    except ImportError:
        log.warning("pl_group_export not importable — PORT column will stay blank for all packages.")
        return
    cache: Dict[str, str] = {}
    n_cn = n_matched = 0
    for pkg in packages:
        factory = pge.detect_factory(pkg.reference_code, pkg.source_file)
        if factory != "CN":
            continue
        n_cn += 1
        signal = pge._collect_cn_signal(pkg, pdf_folder, recursive, cache)
        store, confidence, suggestion = pge.match_store(signal)
        if store in pge.STORE_MASTER:
            pkg.store = store
            pkg.port = str(pge.STORE_MASTER[store]["port"])
            n_matched += 1
        else:
            pkg.store = "REVIEW"
            pkg.port = ""
    log.info(f"CN store/port classification: {n_matched}/{n_cn} CN package(s) matched to a store+port.")

# ── Status ─────────────────────────────────────────────────────────────────
def overall_status(pkg: Package) -> Tuple[str, str]:
    decl = pkg.declared_total_qty
    calc = pkg.calc_qty
    if decl is not None and decl > 0 and pkg.item_count == 0:
        return "CRITICAL_ZERO_ITEMS", f"declared={decl} but 0 items parsed"
    if decl is None:
        return "MISMATCH_NO_TOTAL", "Tong cong not found"
    if decl != calc:
        return "MISMATCH_QTY", f"declared={decl} calc={calc} diff={decl - calc}"
    if not pkg.dim_matched:
        return "MISMATCH_DIM", "No row in Final_dim weight"
    return "MATCHED", ""

# ── Excel styles ───────────────────────────────────────────────────────────
# v8: the "Packing List" sheet (the customer-facing deliverable) is now a
# plain black-grid table — no bold, no header fill, no red/green/yellow
# status colors — matching the real PL template exactly. The Match_Status
# sheet is an internal QC log only (never sent to the customer), so it keeps
# its color-coding on purpose: that's what makes mismatches jump out.
THIN_BLACK = Side(style="thin", color="000000")
PLAIN_BORDER = Border(left=THIN_BLACK, right=THIN_BLACK, top=THIN_BLACK, bottom=THIN_BLACK)
PLAIN_FONT = Font(bold=False, size=10)

STATUS_FILL = {
    "MATCHED":             PatternFill("solid", fgColor="C6EFCE"),
    "MISMATCH_QTY":        PatternFill("solid", fgColor="FFCCCC"),
    "MISMATCH_DIM":        PatternFill("solid", fgColor="FFF2CC"),
    "MISMATCH_NO_TOTAL":   PatternFill("solid", fgColor="FFD966"),
    "CRITICAL_ZERO_ITEMS": PatternFill("solid", fgColor="FF4444"),
}
CRIT_FONT = Font(bold=True, color="FFFFFF", size=10)
MS_HDR_FILL = PatternFill("solid", fgColor="1F4E79")
MS_HDR_FONT = Font(bold=True, color="FFFFFF", size=10)

# Packing List column layout (1-based), matching the real PL template. The
# item table itself starts at row 14 (rows 1-11 = document header block,
# rows 12-13 = the bilingual table header) — same row numbers as the real
# template, so this sheet lines up with it exactly:
#  A Item#            B PO No.           C Invoice No.
#  D Product Name in English             E SKU#            F BarCode/UPC
#  G UOM              H Quantity         I Carton#         J Packaging code
#  K/L/M Carton Dimensions (Length/Width/Height, cm)
#  N Weight (KG)      O CBM
#  P Origin Country   Q Origin Country's HTSCODE            R Shipping Mark
#  S PORT             T 中国标签名称
# v13 (FIX6): the real production PL template's columns B/C are labelled
# "OR No." / "SO No." (confirmed against a real uploaded PL_Total.xlsx --
# NOT "PO No." / "Invoice No." as this file previously assumed/hardcoded).
# Now that pl_or_list_import.py + pl_group_export.match_store_and_or can
# resolve a package's OR/SO from an uploaded OR List (v12 feature, extended
# in v13 to work for every factory, not just CN), these columns are filled
# from pkg.or_number / pkg.so_number instead of being left permanently
# blank for manual entry -- "manual" is now only the fallback when no OR
# List is uploaded or a package doesn't resolve (both leave the cell "",
# same as before, never a guess).
PL_HEADERS_EN = [
    "Item#", "OR No.", "SO No.", "Product Name\nin English", "SKU#",
    "BarCode/UPC", "UOM", "Quantity", "Carton#", "Packaging code",
    "Carton Dimensions (cm)\n(Length*Weight*Height)", "", "",
    "Weight (KG)", "CBM", "Origin Country", "Origin Country's HTSCODE",
    "Shipping Mark", "PORT", "中国标签名称",
]
PL_HEADERS_CN = [
    "项目", "OR 编码", "SO 编码", "货品名称", "SKU编码",
    "条形码", "单位", "数量", "箱号", "包装条形码",
    "箱子尺寸", "", "",
    "", "", "原产国", "原产国",
    "", "", "",
]
NCOLS = 20
TABLE_HDR_ROW1 = 12  # English header row (matches the real template exactly)
TABLE_HDR_ROW2 = 13  # Chinese header row
FIRST_ITEM_ROW = 14
# Package/carton-level columns — same physical carton, so merged across all
# item rows of that carton (matches the real template exactly): Carton#,
# Packaging code, L/W/H, Weight, CBM. Everything else (incl. Origin Country,
# HTS Code, Shipping Mark, PORT) is left un-merged / repeated per row, same
# as the template.
_MERGE_COLS = [9, 10, 11, 12, 13, 14, 15]  # I,J,K,L,M,N,O

# v8: document header block (rows 1-11) — SHIPPER / CONSIGNEE are the same
# entity on every CN shipment (confirmed against 2 real PL samples), so they
# are filled in automatically. WPIC Purchase Order#, Seller's EIN#, Date,
# Invoice#, Remark (SO#) and Trade term vary per shipment and are NOT in the
# OCR/DIM/Master data, so they stay blank — fill in manually, same as before.
SHIPPER_BLOCK = (
    "SHIPPER:\n"
    "TOPOLOGIE GLOBAL LIMITED\n"
    "RM G, 9/F, King Palace Plaza\n"
    "55 King Yip Street, Kwun Tong, Hong Kong\n"
    "EMAIL: supplychainhk@topologie.com\n"
    "TEL: 852 3955 9963"
)
CONSIGNEE_BLOCK = (
    "CONSIGNEE:\n"
    "WORKING UNIT SHANGHAI TRADING CO LTD\n"
    "Room 301, No. 47, Branch Lane 51, Lane 2000, Beizhai Road, Minhang District,\n"
    "Shanghai, China\n"
    "13817762730"
)

# Column widths / row heights copied from the real PL template so this sheet
# looks identical when opened in Excel.
PL_COL_WIDTHS = {
    "A": 16.875, "B": 14.25, "C": 13, "D": 27.8125, "E": 25.8125, "F": 17.5,
    "G": 10, "H": 14.25, "I": 11.8125, "J": 23.1875, "K": 8.3125, "L": 8.0625,
    "M": 6.9375, "N": 9.6875, "O": 13, "P": 15.1875, "Q": 17.0625, "R": 27.125,
    "S": 9, "T": 25.75,
}
PL_ROW_HEIGHTS = {1: 33.75, 2: 31.9, 3: 21, 4: 100.9, 5: 158.65, 6: 21.4,
                   7: 21.4, 8: 15, 9: 15, 10: 15, 11: 15, 12: 37.15, 13: 15.75}


def _style_cell(cell, *, bold=False, align="center", wrap=True, size=10):
    cell.font = Font(bold=bold, size=size)
    cell.border = PLAIN_BORDER
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)


_NO_SIDE = Side(style=None)


def _hdr_border(*, top=False, bottom=False, left=False, right=False):
    return Border(top=THIN_BLACK if top else _NO_SIDE,
                   bottom=THIN_BLACK if bottom else _NO_SIDE,
                   left=THIN_BLACK if left else _NO_SIDE,
                   right=THIN_BLACK if right else _NO_SIDE)


def _style_text(cell, *, bold=False, align="center", wrap=True, size=10):
    """Font + alignment only — border is handled separately by
    _write_pl_doc_header's border pass (see below), so header borders match
    the real template's box exactly instead of a blanket grid."""
    cell.font = Font(bold=bold, size=size)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)


def _write_pl_doc_header(ws, notify_party_text: str, is_cn: bool = True):
    """Rows 1-11: PACKING LIST title, WPIC PO# / Date, Seller's EIN# /
    Invoice#, Shipper / Remark(SO#), Consignee / Notify Party, Trade term,
    then the Package/Quantity/Weight/CBM total block.

    Border layout copied cell-by-cell from the real template (checked
    against 3 real PL files): a single box around columns A:Q for rows 2-6
    (a horizontal divider between every field row, one vertical divider
    between the K/L columns splitting left field / right field), the
    "Trade term" row (6) only boxed on the left side (A:K) — matches the
    template exactly. Row 1 (title) and rows 7-11 (blank spacer + the
    Package/Qty/Weight/CBM total labels) only keep the K/L divider line
    continuing down, nothing else. Columns R/S/T are outside the header
    block in the real template and are left completely untouched (no
    border, no fill) — that's what was over-applied last time."""
    for letter, width in PL_COL_WIDTHS.items():
        ws.column_dimensions[letter].width = width
    for r, h in PL_ROW_HEIGHTS.items():
        ws.row_dimensions[r].height = h

    ws.cell(row=1, column=1, value="PACKING LIST（装箱单）")
    ws.merge_cells("A1:Q1")
    _style_text(ws.cell(row=1, column=1), bold=True, align="center", size=16)

    ws.cell(row=2, column=1, value="WPIC Purchase Order#/箱单编号：")
    ws.merge_cells("A2:D2")
    ws.cell(row=2, column=12, value=" 日期/Date：")
    ws.merge_cells("L2:M2")
    ws.merge_cells("N2:Q2")

    ws.cell(row=3, column=1, value="Seller's EIN#：")
    ws.merge_cells("A3:B3")
    ws.cell(row=3, column=12, value="Invoice#:")
    ws.merge_cells("L3:M3")
    ws.merge_cells("N3:Q3")

    ws.cell(row=4, column=1, value=SHIPPER_BLOCK)
    ws.merge_cells("A4:E4")
    ws.cell(row=4, column=12, value="Remark (SO#):")
    ws.merge_cells("L4:M4")
    ws.merge_cells("N4:Q4")

    ws.cell(row=5, column=1, value=_resolve_consignee(is_cn))
    ws.merge_cells("A5:E5")
    ws.cell(row=5, column=12, value=notify_party_text or "NOTIFY PARTY:\nDELIVERY ADDRESS:")
    ws.merge_cells("L5:Q5")

    ws.cell(row=6, column=1, value="成交方式/Trade term：")
    ws.merge_cells("A6:E6")

    ws.cell(row=8, column=1, value="Package Total: ")
    ws.cell(row=9, column=1, value="Quantity Total:")
    ws.cell(row=10, column=1, value="Gross Weight (KG):")
    ws.cell(row=11, column=1, value="CBM")
    # column B values (formulas referencing the TOTAL row) are filled in by
    # write_workbook() once the item table's TOTAL row number is known.

    # ── Font / alignment (all of A:Q, rows 1-11 — text only, no border) ────
    bold_label_cells = {"A1", "A2", "L2", "A3", "L3", "L4", "A6",
                         "A8", "A9", "A10", "A11"}
    for r in range(1, 7):
        for c in range(1, 18):  # A..Q only — R:T are outside the header box
            cell = ws.cell(row=r, column=c)
            align = "left" if r in (4, 5) and c in (1, 12) else "center"
            _style_text(cell, bold=(cell.coordinate in bold_label_cells or r == 1),
                        align=align, wrap=True, size=16 if r == 1 else 10)
    for r in (8, 9, 10, 11):
        _style_text(ws.cell(row=r, column=1), bold=True, align="left", wrap=False)

    # ── Border (matches the real template's box exactly) ───────────────────
    # Row 1: no border at all (floating title).
    ws.cell(row=1, column=1).border = Border()
    # Rows 2-5: full A:Q box, horizontal divider under every row, one
    # vertical divider between K (col 11) and L (col 12).
    for r in (2, 3, 4, 5):
        for c in range(1, 18):
            ws.cell(row=r, column=c).border = _hdr_border(
                top=True, bottom=True, left=(c == 1), right=(c == 17 or c == 11))
        ws.cell(row=r, column=12).border = _hdr_border(top=True, bottom=True, left=True)
    # Row 6 (Trade term): only the left half (A:K) is boxed — matches the
    # template, where the right/Notify-Party box stops at row 5.
    for c in range(1, 12):  # A..K
        ws.cell(row=6, column=c).border = _hdr_border(top=True, left=(c == 1), right=(c == 11))
    # Rows 7-11: blank spacer + totals — only the K/L divider continues.
    for r in range(7, 12):
        ws.cell(row=r, column=11).border = _hdr_border(right=True)
        ws.cell(row=r, column=12).border = _hdr_border(left=True)


def _write_pl_table_header(ws):
    for c, val in enumerate(PL_HEADERS_EN, start=1):
        ws.cell(row=TABLE_HDR_ROW1, column=c, value=val or None)
    for c, val in enumerate(PL_HEADERS_CN, start=1):
        ws.cell(row=TABLE_HDR_ROW2, column=c, value=val or None)
    ws.merge_cells(start_row=TABLE_HDR_ROW1, start_column=11, end_row=TABLE_HDR_ROW1, end_column=13)
    ws.merge_cells(start_row=TABLE_HDR_ROW2, start_column=11, end_row=TABLE_HDR_ROW2, end_column=13)
    for r in (TABLE_HDR_ROW1, TABLE_HDR_ROW2):
        for c in range(1, NCOLS + 1):
            _style_cell(ws.cell(row=r, column=c), bold=False, align="center", wrap=True)


def _auto_w(ws, cap=40):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        w = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[letter].width = min(w + 3, cap)


def _apply_pl_merge(ws, start_row: int, end_row: int):
    """Merge the carton-level columns across all item rows of one package —
    plain style, no fill, no bold (matches the requested no-color grid)."""
    if end_row <= start_row:
        return
    for col in _MERGE_COLS:
        letter = get_column_letter(col)
        ws.merge_cells(f"{letter}{start_row}:{letter}{end_row}")


def _is_all_cn_factory(packages: List[Package]) -> bool:
    """True only if every package in this file is factory=CN (the known CN
    retail network this template's hardcoded CONSIGNEE / STORE_MASTER data
    applies to). Non-CN factories (POP/SBGEAR/QIFENG/JION/REVIEW) go to
    different countries/consignees this tool has no data for — CNEE and
    NOTIFY PARTY must stay blank and be filled in manually for those."""
    if not packages:
        return False
    try:
        import pl_group_export as pge
    except ImportError:
        return False
    return all(pge.detect_factory(p.reference_code, p.source_file) == "CN" for p in packages)


_MANUAL_FILL_NOTE = "(Ngoài CN — vui lòng tự điền chính xác / non-CN: fill in manually)"


def _resolve_consignee(is_cn: bool) -> str:
    """CN: always the fixed CONSIGNEE_BLOCK. Non-CN: use whatever the user
    typed into the "CNEE / Consignee" box on the app.html page (MANUAL_
    CONSIGNEE) — filled in once on the page instead of editing the exported
    file by hand every time. Falls back to a manual-fill note only if that
    box was left empty."""
    if is_cn:
        return CONSIGNEE_BLOCK
    if MANUAL_CONSIGNEE:
        return "CONSIGNEE:\n" + MANUAL_CONSIGNEE
    return "CONSIGNEE:\n" + _MANUAL_FILL_NOTE


def _resolve_notify_party(packages: List[Package], is_cn: bool) -> str:
    """Auto-fill Notify Party / Delivery Address ONLY when every package in
    this file belongs to the exact same known CN store (e.g. a
    04_CN_BY_STORE split file) — never guessed/blended when a file mixes
    stores (e.g. PL_Total, a factory file, or a CN-by-port file with several
    stores sharing one port). For non-CN factories, use whatever the user
    typed into the "NOTIFY PARTY / Delivery Address" box on the app.html
    page (MANUAL_NOTIFY_PARTY); falls back to a manual-fill note if left
    empty. Ambiguous CN cases are left blank without the note (still CN,
    just needs the specific store filled in)."""
    if not is_cn:
        if MANUAL_NOTIFY_PARTY:
            return "NOTIFY PARTY:\nDELIVERY ADDRESS:\n" + MANUAL_NOTIFY_PARTY
        return "NOTIFY PARTY:\nDELIVERY ADDRESS:\n" + _MANUAL_FILL_NOTE
    stores = {p.store for p in packages if p.store and p.store != "REVIEW"}
    if len(stores) != 1:
        return ""
    try:
        import pl_group_export as pge
    except ImportError:
        return ""
    return pge.notify_party_block(next(iter(stores)))


# ── Workbook writer ────────────────────────────────────────────────────────
def write_workbook(output_path: Path, packages: List[Package], run_meta: Optional[dict] = None,
    carton_display_field: str = "global_carton_num"):
    wb = Workbook()
    wb.remove(wb.active)

    # ── Packing List (customer-facing, plain black grid) ────────────────────
    ws1 = wb.create_sheet("Packing List")
    is_cn = _is_all_cn_factory(packages)
    notify_party_text = _resolve_notify_party(packages, is_cn)
    _write_pl_doc_header(ws1, notify_party_text, is_cn)
    _write_pl_table_header(ws1)

    row_idx = FIRST_ITEM_ROW
    item_no = 0
    for pkg in packages:
        origin    = pkg.origin
        pkg_start = row_idx

        # v13 (FIX6): OR No. / SO No. -- filled from a successful OR List
        # match when available, left "" otherwise (never a guess; matches
        # the same "manual" fallback the sheet has always had for anything
        # this tool can't determine on its own).
        pkg_or = pkg.or_number or ""
        pkg_so = pkg.so_number or ""
        if not pkg.items:
            item_no += 1
            ws1.append([
                item_no, pkg_or, pkg_so,                # Item# / OR No. / SO No.
                "", "", "", "", "",                     # Product/SKU/Barcode/UOM/Qty (no items)
                getattr(pkg, carton_display_field), pkg.package_code,
                pkg.length, pkg.width, pkg.height,
                pkg.weight, pkg.cbm,
                origin, "", (pkg.shipping_mark or pkg.reference_code),  # HTS (manual) / Shipping Mark (v11: parsed label, else PDF filename)
                pkg.port, "",                             # PORT (auto for CN) / 中国标签名称 (manual)
            ])
            row_idx += 1
        else:
            for item in pkg.items:
                item_no += 1
                ws1.append([
                    item_no, pkg_or, pkg_so,                          # Item# / OR No. / SO No.
                    item.product_name, item.product_code, item.barcode,
                    item.unit, item.quantity,
                    getattr(pkg, carton_display_field), pkg.package_code,
                    pkg.length, pkg.width, pkg.height,
                    pkg.weight, pkg.cbm,
                    origin, item.hs_code, (pkg.shipping_mark or pkg.reference_code),  # Shipping Mark (v11: parsed label, else PDF filename)
                    pkg.port, "",                                       # PORT / 中国标签名称
                ])
                row_idx += 1

        end_row = row_idx - 1
        for r in range(pkg_start, end_row + 1):
            align_by_col = {4: "left"}  # Product Name left-aligned, rest centered
            for c in range(1, NCOLS + 1):
                _style_cell(ws1.cell(row=r, column=c), bold=False,
                            align=align_by_col.get(c, "center"), wrap=True)
        _apply_pl_merge(ws1, pkg_start, end_row)

    # ── TOTAL row ────────────────────────────────────────────────────────────
    total_qty = sum(p.calc_qty for p in packages)
    total_cartons = len(packages)
    total_weight = sum(p.weight for p in packages if p.weight is not None)
    total_cbm = sum(p.cbm for p in packages if p.cbm is not None)
    total_row = row_idx
    ws1.cell(row=total_row, column=1, value="TOTAL")
    ws1.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=7)
    ws1.cell(row=total_row, column=8, value=total_qty)
    ws1.cell(row=total_row, column=9, value=f"{total_cartons} Cartons")
    ws1.cell(row=total_row, column=14, value=round(total_weight, 3) if total_weight else 0)
    ws1.cell(row=total_row, column=15, value=round(total_cbm, 6) if total_cbm else 0)
    for c in range(1, NCOLS + 1):
        _style_cell(ws1.cell(row=total_row, column=c), bold=False, align="center", wrap=True)

    # Rows 8-11 (Package/Quantity/Weight/CBM totals, above the table) — live
    # formulas referencing the TOTAL row, same as the real template.
    ws1.cell(row=8, column=2, value=f"={get_column_letter(9)}{total_row}")
    ws1.cell(row=9, column=2, value=f"={get_column_letter(8)}{total_row}")
    ws1.cell(row=10, column=2, value=f"={get_column_letter(14)}{total_row}")
    ws1.cell(row=11, column=2, value=f"={get_column_letter(15)}{total_row}")
    for r in (8, 9, 10, 11):
        _style_text(ws1.cell(row=r, column=2), bold=False, align="left", wrap=False)

    for r in range(FIRST_ITEM_ROW, total_row + 1):
        ws1.row_dimensions[r].height = ws1.row_dimensions[r].height or 18
    for letter, width in PL_COL_WIDTHS.items():
        ws1.column_dimensions[letter].width = width
    ws1.freeze_panes = f"A{FIRST_ITEM_ROW}"

    # ── Match_Status (internal QC only — keeps color-coding on purpose) ─────
    ws2 = wb.create_sheet("Match_Status")
    ws2.append([
        "source_file", "reference_code", "package_code",
        "pdf_package_seq", "Carton number",
        "item_count", "calculated_total_qty", "declared_total_qty",
        "qty_match", "dim_match", "overall_status", "remark",
    ])
    for cell in ws2[1]:
        cell.fill = MS_HDR_FILL
        cell.font = MS_HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws2.row_dimensions[1].height = 28
    for i, pkg in enumerate(packages, start=2):
        decl = pkg.declared_total_qty
        calc = pkg.calc_qty
        q_ok = "YES" if (decl is not None and decl == calc) else "NO"
        d_ok = "YES" if pkg.dim_matched else "NO"
        status, remark = overall_status(pkg)
        ws2.append([
            pkg.source_file, pkg.reference_code, pkg.package_code,
            pkg.pdf_package_seq, getattr(pkg, carton_display_field),
            pkg.item_count, calc, decl,
            q_ok, d_ok, status, remark,
        ])
        fill = STATUS_FILL.get(status, STATUS_FILL["MISMATCH_DIM"])
        fnt  = CRIT_FONT if status == "CRITICAL_ZERO_ITEMS" else None
        for c in range(1, 13):
            ws2.cell(row=i, column=c).fill = fill
            if fnt:
                ws2.cell(row=i, column=c).font = fnt
    _auto_w(ws2)
    ws2.freeze_panes = "A2"

    # ── Raw_Data (spec section 12): one un-merged row per item, full audit
    # trail. Unlike the "Packing List" sheet above, nothing here is merged,
    # so every row is independently machine-readable / filterable, and DIM
    # values are NOT solely carried by merged cells.
    ws3 = wb.create_sheet("Raw_Data")
    raw_headers = [
        "source_filename", "source_page", "reference_raw", "reference_normalized",
        "package_code_raw", "package_code_normalized", "package_sequence",
        "gtin", "gtin_valid_13digit", "sku_raw", "sku_normalized", "description",
        "uom", "condition", "quantity", "declared_total", "calculated_total",
        "package_validation_status", "dim_match_status", "dim_source_method",
        "length", "width", "height", "weight", "cbm", "diagnostic_reason",
        # v14 (spec section 14): full per-package diagnostics -- additive
        # trailing columns only, every column above is unchanged (position
        # and meaning) so no existing reader of this sheet breaks.
        "master_match_status", "country", "country_source",
        "shipping_mark", "shipping_mark_source", "shipping_mark_confidence",
        "filename_reference", "or_list_match_status", "or_list_store",
        "counting_scope_key", "store_carton_display", "global_carton_display",
    ]
    ws3.append(raw_headers)
    for cell in ws3[1]:
        cell.fill = MS_HDR_FILL
        cell.font = MS_HDR_FONT
    for pkg in packages:
        pstatus = audit_status(pkg)
        dim_status = "EXACT_MATCH" if pkg.dim_matched else "MISMATCH"
        rows = pkg.items if pkg.items else [None]
        for item in rows:
            diag_tail = [
                pkg.master_match_status, pkg.country, pkg.country_source,
                pkg.shipping_mark, pkg.shipping_mark_source, pkg.shipping_mark_confidence,
                pkg.filename_reference, pkg.or_list_match_status, pkg.or_list_store,
                pkg.counting_scope_key, pkg.carton_display, pkg.global_carton_display,
            ]
            if item is None:
                ws3.append([
                    pkg.source_file, pkg.first_page, pkg.reference_code, normalize(pkg.reference_code),
                    pkg.package_code, normalize_code(pkg.package_code), pkg.pdf_package_seq,
                    "", "", "", "", "", "", "", "", pkg.declared_total_qty, pkg.calc_qty,
                    pstatus, dim_status, pkg.dim_source_method,
                    pkg.length, pkg.width, pkg.height, pkg.weight, pkg.cbm,
                    "ZERO_ITEMS_IN_PACKAGE",
                ] + diag_tail)
            else:
                ws3.append([
                    pkg.source_file, item.source_page, pkg.reference_code, normalize(pkg.reference_code),
                    pkg.package_code, normalize_code(pkg.package_code), pkg.pdf_package_seq,
                    item.barcode, item.gtin_valid, item.sku_raw, item.product_code, item.product_name,
                    item.unit, item.condition, item.quantity,
                    pkg.declared_total_qty, pkg.calc_qty,
                    pstatus, dim_status, pkg.dim_source_method,
                    pkg.length, pkg.width, pkg.height, pkg.weight, pkg.cbm,
                    item.parse_method,
                ] + diag_tail)
    _auto_w(ws3)
    ws3.freeze_panes = "A2"

    # ── Audit_Summary (spec section 12) ──────────────────────────────────
    ws4 = wb.create_sheet("Audit_Summary")
    meta = run_meta or {}
    dim_obj = meta.get("dim")
    summary_rows = [
        ("Parser/core version", PARSER_VERSION),
        ("Git commit SHA", GIT_COMMIT),
        ("Run timestamp (UTC)", meta.get("run_started_at", "")),
        ("Files processed", len({p.source_file for p in packages})),
        ("Pages processed", sum(d["page_count"] for d in meta.get("pdf_diagnostics", []))),
        ("Package count", len(packages)),
        ("Item count", sum(p.item_count for p in packages)),
        ("Duplicate items skipped (dedup)", meta.get("duplicate_items_skipped", 0)),
        ("Total checks passed (audit_status=OK)", meta.get("audit_counts", {}).get("OK", 0)),
        ("Total checks failed (audit_status!=OK)",
         sum(n for st, n in meta.get("audit_counts", {}).items() if st != "OK")),
        ("DIM detection method", dim_obj.detection_method if dim_obj else ""),
        ("DIM sheet selected", dim_obj.selected_sheet if dim_obj else ""),
        ("DIM rows loaded (valid)", dim_obj.valid_rows if dim_obj else 0),
        ("DIM rows scanned", dim_obj.rows_scanned if dim_obj else 0),
        ("DIM malformed rows", dim_obj.malformed_rows if dim_obj else 0),
        ("DIM duplicate keys", dim_obj.duplicate_keys if dim_obj else 0),
        ("DIM review-required (CBM tolerance)", dim_obj.review_required if dim_obj else 0),
        ("DIM exact matches (packages)", sum(1 for p in packages if p.dim_matched)),
        ("DIM mismatches (packages)", sum(1 for p in packages if not p.dim_matched)),
        ("Warning count (duplicates + malformed + review)",
         (dim_obj.malformed_rows + dim_obj.duplicate_keys + dim_obj.review_required
          if dim_obj else 0) + meta.get("duplicate_items_skipped", 0)),
        ("Error count (DIM FAIL_WITH_DIAGNOSTIC sheets)",
         1 if (dim_obj and dim_obj.detection_method == "FAIL_WITH_DIAGNOSTIC") else 0),
    ]
    ws4.append(["Metric", "Value"])
    for cell in ws4[1]:
        cell.fill = MS_HDR_FILL
        cell.font = MS_HDR_FONT
    for label, value in summary_rows:
        ws4.append([label, value])
    _auto_w(ws4)

    if meta.get("pdf_diagnostics"):
        ws4.append([])
        ws4.append(["Per-PDF reproduction report"])
        ws4.append(["filename", "page_count", "text_layer_usable"])
        for d in meta["pdf_diagnostics"]:
            ws4.append([d["filename"], d["page_count"], d["text_layer_usable"]])

    if meta.get("dim_mismatches"):
        ws4.append([])
        ws4.append(["DIM mismatches (review only -- nothing auto-assigned)"])
        ws4.append(["source_file", "normalized_pdf_reference", "normalized_pdf_package_code",
                    "nearest_reference_candidate", "nearest_package_candidate", "mismatch_reason"])
        for m in meta["dim_mismatches"]:
            ws4.append([m["source_file"], m["normalized_pdf_reference"], m["normalized_pdf_package_code"],
                        m["nearest_reference_candidate"], m["nearest_package_candidate"], m["mismatch_reason"]])

    output_path = Path(output_path)
    tmp_path = output_path.with_suffix(output_path.suffix + ".tmp")
    try:
        wb.save(str(tmp_path))
    except PermissionError as e:
        raise PermissionError(
            f"Cannot create temp file for '{output_path.name}': {e}. "
            f"Close any program locking that folder and re-run."
        ) from e
    try:
        if output_path.exists():
            output_path.unlink()
        tmp_path.replace(output_path)
    except PermissionError as e:
        try:
            tmp_path.unlink(missing_ok=True)
        except Exception:
            pass
        raise PermissionError(
            f"Cannot overwrite '{output_path}': the file appears to be open in Excel. "
            f"Close it and re-run."
        ) from e
    log.info(f"Saved -> {output_path}")

# ── Pipeline ───────────────────────────────────────────────────────────────
TABLE_CFG = {
    "vertical_strategy": "lines", "horizontal_strategy": "lines",
    "snap_tolerance": 5, "join_tolerance": 3,
    "edge_min_length": 3, "min_words_vertical": 1,
}

def run_pipeline(pl_folder: Path, dim_xlsx: Path,
                 output_path: Optional[Path] = None,
                 dim_sheet: Optional[str] = None,
                 master_data_file: Optional[Path] = None,
                 master_data_sheet: Optional[str] = None,
                 recursive: bool = False,
                 or_list_file: Optional[Path] = None):
    run_started_at = datetime.now(timezone.utc)
    if output_path is None:
        output_path = pl_folder / "PL_Output_v6_HS_DIM.xlsx"

    log.info(f"pl_ocr_core {PARSER_VERSION} (commit {GIT_COMMIT})")
    dim = DimMapper(dim_xlsx, sheet_name=dim_sheet)
    hs_mapper = HsCodeMapper(master_data_file, sheet_name=master_data_sheet)
    pdf_iter = pl_folder.rglob("*.pdf") if recursive else pl_folder.glob("*.pdf")
    pdf_files = sorted(pdf_iter, key=lambda p: p.name.upper())
    if not pdf_files:
        log.error(f"No PDFs in {pl_folder}")
        return []
    log.info(f"PDFs ({len(pdf_files)}): {[f.name for f in pdf_files]}")

    # Reproduction-report evidence (spec section 3), collected while parsing
    # runs so before/after comparisons don't need a second pass.
    pdf_diagnostics: List[dict] = []

    parser = Parser()
    for pdf_path in pdf_files:
        log.info(f"Parsing  {pdf_path.name}")
        parser.set_file(pdf_path)
        page_diag = []
        with pdfplumber.open(str(pdf_path)) as pdf:
            text_layer_chars = 0
            for page_idx, page in enumerate(pdf.pages, start=1):
                parser.set_page(page_idx)
                page_text = page.extract_text() or ""
                text_layer_chars += len(page_text.strip())
                tables = page.extract_tables(TABLE_CFG)

                items_before = parser.total_item_count()
                closed_before = len(parser.packages)
                for table in tables:
                    for row in table:
                        if row is None or all(c is None for c in row):
                            continue
                        cells = [str(c).strip() if c else "" for c in row]
                        if not any(cells):
                            continue
                        parser.feed_table_row(cells)
                items_after_table = parser.total_item_count()
                closed_after_table = len(parser.packages)

                # Bug #16 fix: the old rule was `if tables: used_table = True`
                # -- ANY non-empty tables list disabled the text fallback for
                # the rest of the page, even if every row in those tables
                # failed to produce a single valid item (exactly what
                # happened with the merged "Tinh trang So luong" cells, bug
                # #9). The rule now falls back to the text-line pass only
                # when the table pass produced NEITHER a new item NOR closed
                # a package (a real "Tong cong" match) on this page -- a
                # continuation page whose only content is a repeated header
                # + its "Tong cong" line legitimately adds 0 items but DID
                # finalize a package, and must NOT re-run the weaker text
                # pass over the same page (that re-triggers the package
                # header/total lines a second time for no benefit).
                fell_back_to_text = False
                if items_after_table == items_before and closed_after_table == closed_before:
                    for line in page_text.splitlines():
                        parser.feed_text_line(line)
                    fell_back_to_text = True

                page_diag.append({
                    "page": page_idx,
                    "text_layer_chars": len(page_text.strip()),
                    "tables_detected": len(tables),
                    "items_from_table_pass": items_after_table - items_before,
                    "used_text_fallback": fell_back_to_text,
                })
            pdf_diagnostics.append({
                "filename": pdf_path.name,
                "page_count": len(pdf.pages),
                "text_layer_usable": text_layer_chars > 0,
                "pages": page_diag,
            })
        parser.end_of_pdf()

    parser.finalise()
    packages = parser.packages
    log.info(f"Total packages: {len(packages)}")
    log.info(f"Duplicate items skipped (dedup): {parser.duplicate_items_skipped}")

    # v11: resolve Shipping Mark BEFORE sorting/numbering (sorting itself
    # doesn't need it, but this keeps every downstream consumer -- sort,
    # numbering, write_workbook, Sublist -- looking at the same fully-
    # resolved field instead of each doing its own "pkg.shipping_mark or
    # pkg.reference_code" fallback). Never overwrites an explicit label
    # already captured by the parser (Parser._capture_metadata / v11).
    log.info("Resolving OR# / SO# / Shipping Mark for every package...")
    for pkg in packages:
        # v14: filename_reference is ALWAYS recorded, independent of which
        # source ultimately wins for shipping_mark itself (spec section 14
        # diagnostics -- the filename signal must stay auditable even when
        # a PDF-sourced value was used).
        pkg.filename_reference = pkg.reference_code
        if not pkg.shipping_mark:
            pkg.shipping_mark = pkg.reference_code
            pkg.shipping_mark_source = "FILENAME_REFERENCE_CODE"
        pkg.shipping_mark_raw = pkg.shipping_mark
        pkg.shipping_mark_confidence = resolve_shipping_mark_confidence(pkg.shipping_mark_source)
        # v14 (spec sections 2-3): country from the resolved Shipmark's
        # prefix. Falls back to reference_code's own prefix when the
        # resolved shipping_mark itself doesn't carry a recognisable
        # country code (e.g. it resolved from PL text/table content that
        # doesn't start with the shipment code) -- filename_reference is
        # still Shipmark-derived data, just from a lower-priority source,
        # so this fallback stays within the spec's own priority list
        # rather than inventing a new signal.
        pkg.country = detect_shipment_country(pkg.shipping_mark) or detect_shipment_country(pkg.filename_reference)
        pkg.country_source = "SHIPPING_MARK_PREFIX" if pkg.country else ""
    log.info(f"  OR# found (from PL text): {sum(1 for p in packages if p.or_number)}/{len(packages)}")
    log.info(f"  SO# found (from PL text): {sum(1 for p in packages if p.so_number)}/{len(packages)}")
    log.info(f"  Shipping Mark from PL text (not filename fallback): "
             f"{sum(1 for p in packages if p.shipping_mark_source not in ('', 'FILENAME_REFERENCE_CODE'))}/{len(packages)}")
    n_country = sum(1 for p in packages if p.country)
    log.info(f"  Country detected (Shipmark prefix): {n_country}/{len(packages)} "
             f"({sorted({p.country for p in packages if p.country})})")

    # v11: business sort (natural _1/_2/... order + factory order) BEFORE
    # carton numbers are assigned -- numbering must never be computed first
    # and then have the sort order changed under it (spec section 6: "Carton
    # numbering phai duoc thuc hien sau khi hoan tat business sorting").
    packages = business_sort_packages(packages)
    parser.packages = packages

    # v8: CN store/port classification — fills pkg.port/pkg.store for the
    # Packing List sheet's PORT column, using the exact same rule as the CN
    # split step. v12: moved up from below DIM/HS matching -- it now must
    # run BEFORE OR List matching and carton numbering, since both need
    # pkg.store resolved for CN packages that have no OR List coverage.
    classify_packages_for_port(packages, pl_folder, recursive)

    # v12: OR List (optional) — Store/OR/SO matching hierarchy (spec section
    # 5). Never a fatal error, never disables Run/Export: an unusable/absent
    # OR List just leaves every package at or_list_match_status="NO_OR_LIST"
    # and the pipeline continues exactly as it did before this feature.
    log.info("Matching packages against OR List (optional)...")
    or_index: Dict[str, list] = {}
    pge_mod = None
    try:
        import pl_or_list_import as oli
        import pl_group_export as pge_mod
        or_list_result = oli.load_or_list(or_list_file)
        global LAST_OR_LIST_RESULT
        LAST_OR_LIST_RESULT = or_list_result
        if or_list_result.ok:
            or_index = oli.build_or_index(or_list_result)
            log.info(f"OR List loaded: {len(or_list_result.rows)} row(s) from "
                     f"sheet {or_list_result.sheet_used!r}.")
        elif or_list_result.status != "NO_FILE":
            log.warning(f"OR List not usable (status={or_list_result.status}): "
                        f"{'; '.join(or_list_result.errors) or '(no details)'} "
                        f"-- continuing without it, Run/Export not affected.")
    except ImportError:
        log.warning("pl_or_list_import/pl_group_export not importable — "
                    "OR List matching skipped, pipeline continues unaffected.")

    if or_index and pge_mod is not None:
        receiver_cache: Dict[str, str] = {}
        for pkg in packages:
            receiver_text = pge_mod._collect_cn_signal(pkg, pl_folder, recursive, receiver_cache)
            m = pge_mod.match_store_and_or(pkg, or_index, receiver_text=receiver_text)
            pkg.or_list_store = m.matched_store
            pkg.or_list_match_source = m.match_source
            pkg.or_list_match_status = m.status
            pkg.or_list_review_reason = m.review_reason
            pkg.or_list_candidate_store = m.candidate_store
            pkg.or_list_candidate_score = m.candidate_score
            if m.status == "OK":
                # PL-text-parsed OR/SO (captured earlier by the Parser)
                # always takes priority when present -- OR List only fills
                # in what the PL text itself didn't already give us.
                if not pkg.or_number:
                    pkg.or_number = m.matched_or
                    pkg.or_source = "OR_LIST"
                if not pkg.so_number:
                    pkg.so_number = m.matched_so
                    pkg.so_source = "OR_LIST"
        n_ok = sum(1 for p in packages if p.or_list_match_status == "OK")
        n_review = sum(1 for p in packages if p.or_list_match_status == "REVIEW")
        n_other = len(packages) - n_ok - n_review
        log.info(f"OR List match: {n_ok} OK, {n_review} REVIEW, {n_other} other, "
                 f"out of {len(packages)}")
    else:
        for pkg in packages:
            pkg.or_list_match_status = "NO_OR_LIST"
        log.info("No usable OR List -- every package left at "
                 "or_list_match_status='NO_OR_LIST' (tool still runs fully; "
                 "Run/Export never disabled).")

    # v12: carton numbers are now computed per counting_scope_key (spec
    # section 9), not as one flat sequence across the whole upload.
    assign_counting_scope_keys(packages)
    assign_global_numbers(packages)

    matched = 0
    dim_mismatches: List[dict] = []
    for pkg in packages:
        d = dim.lookup(pkg.reference_code, pkg.package_code)
        if d:
            pkg.length, pkg.width, pkg.height = d["length"], d["width"], d["height"]
            pkg.weight, pkg.cbm = d["weight"], d["cbm"]
            pkg.dim_matched = True
            pkg.dim_source_method = dim.detection_method
            matched += 1
        else:
            dim_mismatches.append(dim.diagnose_miss(pkg.source_file, pkg.reference_code, pkg.package_code))
    log.info(f"DIM matched: {matched}/{len(packages)}")

    hs_matched = 0
    hs_total = 0
    for pkg in packages:
        pkg_hs_codes = []
        for item in pkg.items:
            hs_total += 1
            item.hs_code = hs_mapper.lookup(item.product_code, item.barcode)
            if item.hs_code:
                hs_matched += 1
                pkg_hs_codes.append(item.hs_code)
        # For zero-item packages only, keep a carton-level fallback blank.
        pkg.hs_code = ", ".join(sorted(set(pkg_hs_codes))) if pkg_hs_codes else ""
    log.info(f"HS Code matched: {hs_matched}/{hs_total}")

    counts: Dict[str, int] = defaultdict(int)
    audit_counts: Dict[str, int] = defaultdict(int)
    for pkg in packages:
        counts[overall_status(pkg)[0]] += 1
        audit_counts[audit_status(pkg)] += 1
    for st, n in sorted(counts.items()):
        log.info(f"  {st:<30} {n}")
    log.info(f"Audit-vocabulary status: {dict(audit_counts)}")

    run_meta = {
        "run_started_at": run_started_at.isoformat(),
        "parser_version": PARSER_VERSION,
        "git_commit": GIT_COMMIT,
        "pdf_diagnostics": pdf_diagnostics,
        "dim_mismatches": dim_mismatches,
        "duplicate_items_skipped": parser.duplicate_items_skipped,
        "audit_counts": dict(audit_counts),
        "dim": dim,
    }
    global LAST_RUN_META
    LAST_RUN_META = run_meta
    write_workbook(output_path, packages, run_meta=run_meta,
        carton_display_field="global_carton_display")
    return packages

# ── Entry point ────────────────────────────────────────────────────────────
# ── Entry point (called directly instead of __main__ guard, since this
#    module is exec'd inside Pyodide rather than run as a script) ─────────
packages = run_pipeline(
    pl_folder=PL_FOLDER,
    dim_xlsx=DIM_WEIGHT_FILE,
    output_path=OUTPUT_XLSX,
    dim_sheet=DIM_WEIGHT_SHEET,
    master_data_file=MASTER_DATA_FILE,
    master_data_sheet=MASTER_DATA_SHEET,
    recursive=RECURSIVE,
    or_list_file=OR_LIST_FILE,
)

# ── UI summary (spec section 13) ─────────────────────────────────────────
# Plain dict of JSON-serializable values, read by app.html via
# `pyodide.globals.get('RUN_SUMMARY').toJs(...)` after this script finishes,
# and also echoed to stdout (captured into the on-page log box) so it is
# visible even if that JS read ever fails.
_dim_obj = LAST_RUN_META.get("dim") if LAST_RUN_META else None
RUN_SUMMARY = {
    "parser_version": PARSER_VERSION,
    "git_commit": GIT_COMMIT,
    "pdf_files_processed": len({p.source_file for p in packages}) if packages else 0,
    "packages_found": len(packages) if packages else 0,
    "items_found": sum(p.item_count for p in packages) if packages else 0,
    "qty_checks_passed": sum(1 for p in packages
                              if p.declared_total_qty is not None and p.declared_total_qty == p.calc_qty) if packages else 0,
    "qty_checks_failed": sum(1 for p in packages
                              if p.declared_total_qty is not None and p.declared_total_qty != p.calc_qty) if packages else 0,
    "dim_rows_loaded": _dim_obj.valid_rows if _dim_obj else 0,
    "dim_detection_method": _dim_obj.detection_method if _dim_obj else "",
    "dim_exact_matched": sum(1 for p in packages if p.dim_matched) if packages else 0,
    "dim_review_required": _dim_obj.review_required if _dim_obj else 0,
    "dim_mismatched": sum(1 for p in packages if not p.dim_matched) if packages else 0,
    "duplicate_items_skipped": LAST_RUN_META.get("duplicate_items_skipped", 0) if LAST_RUN_META else 0,
    "errors": sum(1 for p in packages if audit_status(p) != "OK") if packages else 0,
    # OR List status (Turn 12 fix): distinguishes "no file uploaded" (NO_FILE,
    # the normal/expected case for most runs -- OR List is optional) from "a
    # file WAS uploaded but couldn't be parsed" (HEADER_NOT_FOUND /
    # REQUIRED_FIELD_MISSING / LOAD_ERROR) -- previously both looked identical
    # to the UI (or_index just came back empty either way), so a real upload
    # mistake silently behaved as if no file had been given at all.
    "or_list_status": LAST_OR_LIST_RESULT.status if LAST_OR_LIST_RESULT else "NO_FILE",
    "or_list_sheet_used": LAST_OR_LIST_RESULT.sheet_used if LAST_OR_LIST_RESULT else None,
    "or_list_rows_loaded": len(LAST_OR_LIST_RESULT.rows) if LAST_OR_LIST_RESULT else 0,
    "or_list_error": ("; ".join(LAST_OR_LIST_RESULT.errors) if LAST_OR_LIST_RESULT and LAST_OR_LIST_RESULT.errors else ""),
}
print("RUN_SUMMARY_JSON=" + json.dumps(RUN_SUMMARY))

# =========================================================
# AUTO SPLIT: TOTAL -> FACTORY -> CN PORT -> CN STORE
# Requires: pl_group_export.py in the same folder as this notebook.
# Run this cell AFTER the cell above has produced `packages` via run_pipeline().
# =========================================================
import importlib
import pl_group_export
importlib.reload(pl_group_export)  # pick up edits to pl_group_export.py without restarting the kernel
from pl_group_export import export_grouped_pl

if not packages:
    raise RuntimeError(
        "`packages` is empty — the OCR pipeline above found no PDFs or produced "
        "no packages. Fix PL_FOLDER / the PDFs first, then re-run the cell above "
        "before running this split step."
    )

SPLIT_OUTPUT_DIR = PL_FOLDER / 'PL_SPLIT_OUTPUT'

try:
    control_file = export_grouped_pl(
        packages=packages,
        output_dir=SPLIT_OUTPUT_DIR,
        write_workbook=write_workbook,
        total_workbook=OUTPUT_XLSX,
        pdf_folder=PL_FOLDER,
        recursive=RECURSIVE,
    )
except RuntimeError as e:
    print("XXXX SPLIT FAILED — reconciliation mismatch, nothing was silently swallowed XXXX")
    print(e)
    raise
except PermissionError as e:
    print("XXXX SPLIT FAILED — a target .xlsx is locked/open in Excel XXXX")
    print(e)
    raise
else:
    print(f'Completed: {SPLIT_OUTPUT_DIR}')
    print(f'Control file: {control_file}')

# =========================================================
# v11/v12: SUBLIST generation -- Excel (optional secondary output) + A5 PDF
# (mandatory/default-on primary output, spec turn 5 section 8).
#
# NON-BLOCKING BY DESIGN (spec: "The current patch claims Sublist is
# optional, but re-raises RuntimeError and PermissionError. Fix this."):
# NEITHER Sublist output may ever raise out of this block. A failure here
# is recorded into SUBLIST_XLSX_STATUS / SUBLIST_PDF_STATUS
# (SUCCESS|FAILED|DISABLED), logged, and the run still completes with
# every other output (Packing List / Match_Status / Raw_Data /
# Audit_Summary / PL_SPLIT_OUTPUT / legacy ZIP) fully intact -- exactly as
# if this feature didn't exist at all.
# =========================================================
SUBLIST_XLSX_STATUS = "DISABLED"
SUBLIST_XLSX_PATH = None
SUBLIST_XLSX_ERROR = ""
SUBLIST_PDF_STATUS = "DISABLED"
SUBLIST_PDF_PATH = None
SUBLIST_PDF_ERROR = ""

if GENERATE_SUBLIST:
    try:
        import pl_sublist_export
        importlib.reload(pl_sublist_export)
        sublist_dir = SPLIT_OUTPUT_DIR / '05_SUBLIST'
        sublist_path = sublist_dir / 'SUBLIST_TOTAL.xlsx'
        # `packages` is already in the exact same order write_workbook() used
        # for PL_Total.xlsx (business_sort_packages() ran once, inside
        # run_pipeline(), before assign_global_numbers() -- see above) --
        # passing it straight through is what keeps Sublist order identical
        # to PL_TOTAL (spec requirement).
        log.info("Generating Sublist (Excel, optional secondary output)...")
        sublist_result = pl_sublist_export.generate_sublist_workbook(packages, sublist_path)
        sublist_ok, sublist_report = pl_sublist_export.validate_sublist(packages, sublist_result)
        print("\n" + "=" * 70)
        print("SUBLIST (XLSX) VALIDATION REPORT")
        print("=" * 70)
        print(sublist_report)
        print("=" * 70)
        if not sublist_ok:
            SUBLIST_XLSX_STATUS = "FAILED"
            SUBLIST_XLSX_ERROR = "Reconciliation FAILED -- see the report printed above."
            log.warning(f"Sublist (Excel) reconciliation FAILED (non-blocking, every other "
                        f"output is unaffected): {SUBLIST_XLSX_ERROR}")
        else:
            SUBLIST_XLSX_STATUS = "SUCCESS"
            SUBLIST_XLSX_PATH = str(sublist_path)
            print(f'Sublist (Excel) completed: {sublist_path}')
    except ImportError as e:
        SUBLIST_XLSX_STATUS = "FAILED"
        SUBLIST_XLSX_ERROR = f"pl_sublist_export not importable: {e}"
        log.warning(f"Sublist (Excel) NOT generated this run (non-blocking, every other "
                    f"output is unaffected): {SUBLIST_XLSX_ERROR}")
    except Exception as e:
        # Deliberately broad (was RuntimeError/PermissionError re-raised
        # before -- both, and anything else, must now be caught here):
        # an optional secondary output must NEVER take down the run.
        SUBLIST_XLSX_STATUS = "FAILED"
        SUBLIST_XLSX_ERROR = f"{type(e).__name__}: {e}"
        log.warning(f"Sublist (Excel) generation FAILED (non-blocking, every other output "
                    f"is unaffected): {SUBLIST_XLSX_ERROR}")
else:
    log.info("GENERATE_SUBLIST (Excel) is off -- Excel Sublist not generated this run.")

if GENERATE_SUBLIST_PDF:
    try:
        import pl_sublist_pdf_export
        importlib.reload(pl_sublist_pdf_export)
        pdf_dir = SPLIT_OUTPUT_DIR / '05_SUBLIST'
        pdf_path = pdf_dir / 'SUBLIST_TOTAL.pdf'
        log.info("Generating A5 carton Sublist PDF...")
        pdf_result = pl_sublist_pdf_export.generate_sublist_pdf(packages, pdf_path)
        pdf_problems = pl_sublist_pdf_export.validate_sublist_pdf(packages, pdf_result)
        print("\n" + "=" * 70)
        print("SUBLIST (PDF) VALIDATION REPORT")
        print("=" * 70)
        print(f"status: {pdf_result.status}")
        print(f"cartons_written: {pdf_result.cartons_written}  pages_written: {pdf_result.pages_written}  "
              f"items_written: {pdf_result.items_written}")
        if pdf_problems:
            print("problems:")
            for p in pdf_problems:
                print(f"  - {p}")
        print("=" * 70)
        if pdf_result.status == "SUCCESS" and not pdf_problems:
            SUBLIST_PDF_STATUS = "SUCCESS"
            SUBLIST_PDF_PATH = str(pdf_path)
            print(f'Sublist (PDF) completed: {pdf_path}')
        elif pdf_result.status == "SUCCESS" and pdf_problems:
            SUBLIST_PDF_STATUS = "FAILED"
            SUBLIST_PDF_ERROR = "; ".join(pdf_problems)
            log.warning(f"Sublist (PDF) reconciliation FAILED (non-blocking, every other "
                        f"output is unaffected): {SUBLIST_PDF_ERROR}")
        else:
            SUBLIST_PDF_STATUS = pdf_result.status  # FAILED or DISABLED
            SUBLIST_PDF_ERROR = pdf_result.error
            log.warning(f"Sublist (PDF) NOT generated this run (non-blocking, every other "
                        f"output is unaffected): status={pdf_result.status} error={pdf_result.error}")
    except ImportError as e:
        SUBLIST_PDF_STATUS = "FAILED"
        SUBLIST_PDF_ERROR = f"pl_sublist_pdf_export not importable: {e}"
        log.warning(f"Sublist (PDF) NOT generated this run (non-blocking, every other output "
                    f"is unaffected): {SUBLIST_PDF_ERROR}")
    except Exception as e:
        # generate_sublist_pdf() itself never raises (spec: non-blocking by
        # design) -- this outer catch is defense in depth for anything
        # unexpected in this wiring block itself (e.g. validate_sublist_pdf).
        SUBLIST_PDF_STATUS = "FAILED"
        SUBLIST_PDF_ERROR = f"{type(e).__name__}: {e}"
        log.warning(f"Sublist (PDF) generation FAILED (non-blocking, every other output is "
                    f"unaffected): {SUBLIST_PDF_ERROR}")
else:
    log.info("GENERATE_SUBLIST_PDF is off -- A5 Sublist PDF not generated this run.")

RUN_SUMMARY["sublist_xlsx_status"] = SUBLIST_XLSX_STATUS
RUN_SUMMARY["sublist_xlsx_path"] = SUBLIST_XLSX_PATH
RUN_SUMMARY["sublist_xlsx_error"] = SUBLIST_XLSX_ERROR
RUN_SUMMARY["sublist_pdf_status"] = SUBLIST_PDF_STATUS
RUN_SUMMARY["sublist_pdf_path"] = SUBLIST_PDF_PATH
RUN_SUMMARY["sublist_pdf_error"] = SUBLIST_PDF_ERROR
print("RUN_SUMMARY_JSON=" + json.dumps(RUN_SUMMARY))
