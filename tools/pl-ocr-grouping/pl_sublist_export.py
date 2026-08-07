#!/usr/bin/env python3
"""
pl_sublist_export.py
=====================
Generates the per-carton "Sublist" workbook (topologie standard sublist.xlsx
style) from the same `packages` list run_pipeline() already produced.
Self-contained -- no import from pl_ocr_core.py -- mirrors the existing
pl_group_export.py pattern (`import pl_sublist_export as ...`) so app.html
only needs to fetch one more small source file.

How the layout constants below were derived
---------------------------------------------
The real "topologie standard sublist.xlsx" template was inspected directly
with openpyxl (not guessed from a description) before writing this module:
  - Sheet1, 69 rows x 12 cols, NO merged cells, portrait, paper size 11.
  - 4 carton blocks per page, laid out horizontally in 3-column groups:
    A:C, D:F, G:I, J:L.
  - Metadata label sits in the 2nd column of each 3-col block, value in the
    3rd column; the block's 1st column is left blank for metadata rows
    (e.g. block 1: label in B, value in C, A blank) -- confirmed from the
    real cell coordinates, not assumed.
  - Item grid (Item No. / EAN / QTY) is THIN-BORDERED for 18 rows per block
    even though the one real filled example in the template only had 12
    actual item VALUES in it -- confirmed by reading the real cell border
    styles row by row, not by counting how many rows happened to have data.
    SUBLIST_ITEM_CAPACITY_PER_BLOCK is therefore 18, not 12.
  - The Total QTY cell holds a live `=SUM(range)` formula, with the range
    scoped to however many item rows the example actually used (C8:C19 for
    12 items) rather than the full C8:C25 capacity -- this module reproduces
    that convention: sum only over the rows actually written.
  - Page cycle is 42 rows in the original 5-metadata-row template
    (Carton# / OR# / SO Order# / GW / Packing Code#). Per the business spec
    this module adds a 6th metadata row (Shipping Mark, right after
    Carton#) -- the preferred option from the spec ("tang chieu cao page
    block ... khong nen giam capacity duoi 12") is used: the page cycle
    grows by exactly 1 row (43, not 42) so the 18-row item capacity is
    never reduced to make room.
  - Column widths and an internal row-height jump observed in the real file
    (36.6pt for the first 14 item rows, 33.6pt for the last 4 + total row;
    Item No./EAN/QTY column widths that differ block-to-block for the exact
    same column role, e.g. block 1's Item No. column is 46.3 wide vs block
    3's 22.7) look like incidental auto-fit/manual-resize artifacts of that
    ONE real file rather than an intentional design -- both are normalized
    to one consistent value per column-role/row-role across all 4 blocks
    here instead of reproducing the inconsistency. Documented explicitly,
    not silently decided -- flag if the real production template disagrees.
"""
from __future__ import annotations

import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

log = logging.getLogger("pl_sublist_export")
if not log.handlers:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

# =========================================================================
# 1) Layout constants (see module docstring for how each was derived)
# =========================================================================
BLOCKS_PER_PAGE = 4
BLOCK_WIDTH_COLS = 3
SUBLIST_ITEM_CAPACITY_PER_BLOCK = 18  # real border extent, NOT the 12-item sample count

# Row offsets from a page's first (1-based) row -- add to page_start_row.
OFF_CARTON = 0
OFF_SHIPPING_MARK = 1
OFF_OR = 2
OFF_SO = 3
OFF_GW = 4
OFF_PACKING_CODE = 5
# offset 6 = blank separator row (matches the template's own blank row
# between the metadata block and the item header)
OFF_ITEM_HEADER = 7
OFF_ITEM_FIRST = 8
OFF_ITEM_LAST = OFF_ITEM_FIRST + SUBLIST_ITEM_CAPACITY_PER_BLOCK - 1   # 25
OFF_TOTAL = OFF_ITEM_LAST + 1                                          # 26
# Original template: content offsets 0-25 (26 rows), then a 16-row blank
# gap (offsets 26-41) before the next page at offset 42 -- cycle=42. This
# build's content is one row taller (offsets 0-26, 27 rows) because of the
# added Shipping Mark row; the same 16-row blank gap is preserved, so the
# cycle grows by exactly 1 row.
PAGE_CYCLE_ROWS = OFF_TOTAL + 1 + 16   # 27 + 16 = 43

METADATA_LABELS = [
    (OFF_CARTON, "Carton #"),
    (OFF_SHIPPING_MARK, "Shipping Mark"),
    (OFF_OR, "OR #"),
    (OFF_SO, "SO Order #"),
    (OFF_GW, "GW"),
    (OFF_PACKING_CODE, "Packing Code #"),
]
ITEM_HEADERS = ["Item No.", "EAN", "QTY"]

# Normalized (see docstring) column widths, applied identically to all 4 blocks.
COL_WIDTH_ITEM_NO = 42
COL_WIDTH_EAN = 34
COL_WIDTH_QTY = 16
ROW_HEIGHT_METADATA = 33.6
ROW_HEIGHT_SPACER = 31.2
ROW_HEIGHT_ITEM = 36.6

FONT_LABEL = Font(name="Calibri", size=26, bold=True)
FONT_VALUE = Font(name="Calibri", size=26, bold=False)
FONT_ITEM_HEADER = Font(name="Calibri", size=26, bold=True)
FONT_ITEM = Font(name="Calibri", size=26, bold=False)
FONT_TOTAL = Font(name="Calibri", size=26, bold=True)
FONT_CONTINUED = Font(name="Calibri", size=18, bold=True, italic=True)

ALIGN_LABEL = Alignment(horizontal="right", vertical="center", wrap_text=True)
ALIGN_VALUE = Alignment(horizontal="right", vertical="center", wrap_text=True)
ALIGN_ITEM = Alignment(horizontal="center", vertical="center", wrap_text=True)

THIN = Side(style="thin", color="000000")
BOX_BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

TEXT_FORMAT = "@"  # forces text storage -- no scientific notation, no leading-zero loss


# =========================================================================
# 2) Pure data model (spec section 11) -- never mutates the source package
# =========================================================================
def _apply_requested_excel_view_preferences(workbook) -> None:
    from openpyxl.styles import Alignment

    target_headers = {"SKU#", "SHIPPING MARK"}

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = None

        scan_rows = min(max(worksheet.max_row, 1), 30)
        for row in worksheet.iter_rows(min_row=1, max_row=scan_rows):
            for header_cell in row:
                header = str(header_cell.value or "").strip().upper()
                if header not in target_headers:
                    continue

                col_idx = header_cell.column
                for col_cells in worksheet.iter_cols(
                    min_col=col_idx,
                    max_col=col_idx,
                    min_row=header_cell.row,
                    max_row=max(worksheet.max_row, header_cell.row),
                ):
                    for cell in col_cells:
                        old = cell.alignment
                        cell.alignment = Alignment(
                            horizontal="left",
                            vertical=old.vertical or "center",
                            wrap_text=old.wrap_text,
                            shrink_to_fit=old.shrink_to_fit,
                            text_rotation=old.text_rotation,
                        )


class SublistItemRow:
    item_no: str
    ean: str
    qty: int


@dataclass
class SublistCartonModel:
    carton_sequence: int
    carton_total: int
    carton_display: str
    shipping_mark: str
    shipping_mark_source: str
    or_number: str
    or_source: str
    so_number: str
    so_source: str
    gross_weight: Optional[float]
    gross_weight_display: str
    gross_weight_source: str
    packing_code: str
    items: List[SublistItemRow] = field(default_factory=list)
    total_qty: int = 0
    source_file: str = ""
    reference_code: str = ""
    package_code: str = ""
    pdf_package_seq: str = ""
    # identity used by validate_sublist() to count UNIQUE cartons -- distinct
    # from block count (a >18-item carton spans multiple blocks but is still
    # exactly one carton).
    carton_identity: str = ""
    # Per-Store/shipment numbering scope (pl_ocr_core.compute_counting_scope_key)
    # -- carton_sequence is only unique WITHIN this scope, never globally.
    # Kerry's carton 1/6 and Hangzhou's carton 1/4 legitimately share
    # carton_sequence=1; validate_sublist() must key uniqueness on
    # (counting_scope_key, carton_sequence), never on carton_sequence alone.
    counting_scope_key: str = ""


def resolve_sublist_metadata(package) -> dict:
    """Pull the package-level Sublist metadata fields off a duck-typed
    package object (works against pl_ocr_core.Package or any test double
    exposing the same attribute names). Never mutates `package`.

    GW priority (spec: "GW (from PL, kept separate from existing DIM-
    sourced weight)"): `pl_gross_weight` (captured straight from the PL
    PDF's own text/table, e.g. "35.68 KG") always wins when present --
    it's the field the Sublist's GW column is actually documented against.
    Only when the PL text never gave us a gross weight at all does this
    fall back to the DIM-lookup `weight` field, so the GW cell still shows
    *something* useful rather than going blank. The two are never merged
    or overwritten into one another -- pl_gross_weight is untouched either
    way, this function just decides which one is DISPLAYED here."""
    pl_gw = (getattr(package, "pl_gross_weight", "") or "").strip()
    if pl_gw:
        gw = pl_gw
        # pl_gross_weight is captured raw text and may or may not already
        # carry a unit (e.g. "35.68 KG" vs a bare "35.68") -- only append
        # "KG" when the captured text has no letters of its own, so a
        # differently-labelled unit in the source PDF is never overwritten.
        gw_display = pl_gw if any(ch.isalpha() for ch in pl_gw) else f"{pl_gw} KG"
        gw_source = "PL_TEXT"
    else:
        dim_gw = getattr(package, "weight", None)
        gw = dim_gw
        gw_display = f"{dim_gw:.2f} KG" if isinstance(dim_gw, (int, float)) else ""
        gw_source = "DIM_FALLBACK" if dim_gw not in (None, "", 0) else ""
    shipping_mark = getattr(package, "shipping_mark", "") or getattr(package, "reference_code", "") or ""
    shipping_mark_source = getattr(package, "shipping_mark_source", "") or (
        "FILENAME_REFERENCE_CODE" if shipping_mark == getattr(package, "reference_code", "") else "")
    return {
        "shipping_mark": shipping_mark,
        "shipping_mark_source": shipping_mark_source,
        "or_number": getattr(package, "or_number", "") or "",
        "or_source": getattr(package, "or_source", "") or "",
        "so_number": getattr(package, "so_number", "") or "",
        "so_source": getattr(package, "so_source", "") or "",
        "gross_weight": gw,
        "gross_weight_display": gw_display,
        "gross_weight_source": gw_source,
        "packing_code": getattr(package, "package_code", "") or "",
    }


def calculate_carton_total_qty(package) -> int:
    return sum(int(getattr(it, "quantity", 0) or 0) for it in getattr(package, "items", []))


def build_sublist_carton_model(package, *, carton_display_mode: str = "current_total") -> SublistCartonModel:
    """Pure helper (spec section 10): package -> SublistCartonModel. Does
    NOT mutate `package`."""
    meta = resolve_sublist_metadata(package)
    items = [
        SublistItemRow(
            item_no=str(getattr(it, "product_code", "") or ""),
            ean=str(getattr(it, "barcode", "") or ""),
            qty=int(getattr(it, "quantity", 0) or 0),
        )
        for it in getattr(package, "items", [])
    ]
    carton_sequence = int(getattr(package, "carton_sequence", 0) or 0)
    carton_total = int(getattr(package, "carton_total", 0) or 0)
    if carton_display_mode == "current_only":
        carton_display = str(carton_sequence) if carton_sequence else ""
    else:
        carton_display = getattr(package, "carton_display", "") or getattr(package, "global_carton_num", "")
    identity = f"{getattr(package, 'source_file', '')}|{getattr(package, 'package_code', '')}|{getattr(package, 'reference_code', '')}"
    return SublistCartonModel(
        carton_sequence=carton_sequence,
        carton_total=carton_total,
        carton_display=carton_display,
        shipping_mark=meta["shipping_mark"],
        shipping_mark_source=meta["shipping_mark_source"],
        or_number=meta["or_number"],
        or_source=meta["or_source"],
        so_number=meta["so_number"],
        so_source=meta["so_source"],
        gross_weight=meta["gross_weight"],
        gross_weight_display=meta["gross_weight_display"],
        gross_weight_source=meta["gross_weight_source"],
        packing_code=meta["packing_code"],
        items=items,
        total_qty=sum(r.qty for r in items),
        source_file=getattr(package, "source_file", ""),
        reference_code=getattr(package, "reference_code", ""),
        package_code=getattr(package, "package_code", ""),
        pdf_package_seq=str(getattr(package, "pdf_package_seq", "") or ""),
        carton_identity=identity,
        # Falls back to "" (one implicit shared scope, today's flat 1/N..N/N
        # behaviour) for any caller/test that predates counting_scope_key --
        # same backward-compat default pl_ocr_core.assign_global_numbers()
        # itself uses.
        counting_scope_key=getattr(package, "counting_scope_key", "") or "",
    )


@dataclass
class SublistBlock:
    carton: SublistCartonModel
    block_index: int          # 0 = first/only block, 1+ = continuation
    block_count: int          # total blocks this carton was split into
    items: List[SublistItemRow]
    is_last_block: bool
    block_carton_label: str
    block_total_qty: int      # subtotal for non-last blocks, GRAND TOTAL on the last block


def paginate_carton_blocks(cartons: List[SublistCartonModel]) -> List[SublistBlock]:
    """One package == one carton, ALWAYS -- a carton with more than
    SUBLIST_ITEM_CAPACITY_PER_BLOCK items is split into several Excel
    blocks ("Continued 1", "Continued 2", ...) so no item is ever dropped,
    but block_count > 1 must never be mistaken for more than one carton
    (validate_sublist() counts unique carton_identity, not block count)."""
    blocks: List[SublistBlock] = []
    cap = SUBLIST_ITEM_CAPACITY_PER_BLOCK
    for carton in cartons:
        n = len(carton.items)
        chunks = [carton.items[i:i + cap] for i in range(0, n, cap)] or [[]]
        block_count = len(chunks)
        for idx, chunk in enumerate(chunks):
            is_last = idx == block_count - 1
            if block_count == 1:
                label = carton.carton_display
            elif idx == 0:
                label = carton.carton_display
            else:
                label = f"{carton.carton_display} - Continued {idx}"
            block_total = carton.total_qty if is_last else sum(r.qty for r in chunk)
            blocks.append(SublistBlock(
                carton=carton, block_index=idx, block_count=block_count,
                items=chunk, is_last_block=is_last,
                block_carton_label=label, block_total_qty=block_total,
            ))
    return blocks


# =========================================================================
# 3) Excel writer
# =========================================================================
def _set_text_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=(str(value) if value not in (None, "") else None))
    cell.number_format = TEXT_FORMAT
    return cell


def _apply_col_widths(ws, block_start_col: int):
    ws.column_dimensions[get_column_letter(block_start_col)].width = COL_WIDTH_ITEM_NO
    ws.column_dimensions[get_column_letter(block_start_col + 1)].width = COL_WIDTH_EAN
    ws.column_dimensions[get_column_letter(block_start_col + 2)].width = COL_WIDTH_QTY


def write_carton_block(ws, block: SublistBlock, page_start_row: int, block_start_col: int):
    """Writes one carton block (metadata + item grid + total) into `ws` at
    the given page/column position. Never mutates `block`/`block.carton`."""
    label_col = block_start_col + 1
    value_col = block_start_col + 2
    _apply_col_widths(ws, block_start_col)

    carton = block.carton
    metadata_values = {
        OFF_CARTON: block.block_carton_label,
        OFF_SHIPPING_MARK: carton.shipping_mark,
        OFF_OR: carton.or_number,
        OFF_SO: carton.so_number,
        OFF_GW: carton.gross_weight_display,
        OFF_PACKING_CODE: carton.packing_code,
    }
    for off, label in METADATA_LABELS:
        r = page_start_row + off
        ws.row_dimensions[r].height = ROW_HEIGHT_METADATA
        lbl_cell = ws.cell(row=r, column=label_col, value=label)
        lbl_cell.font = FONT_LABEL
        lbl_cell.alignment = ALIGN_LABEL
        val_cell = _set_text_cell(ws, r, value_col, metadata_values[off])
        val_cell.font = FONT_CONTINUED if (off == OFF_CARTON and block.block_index > 0) else FONT_VALUE
        val_cell.alignment = ALIGN_VALUE

    spacer_row = page_start_row + OFF_ITEM_HEADER - 1
    ws.row_dimensions[spacer_row].height = ROW_HEIGHT_SPACER

    header_row = page_start_row + OFF_ITEM_HEADER
    ws.row_dimensions[header_row].height = ROW_HEIGHT_ITEM
    for i, htext in enumerate(ITEM_HEADERS):
        c = ws.cell(row=header_row, column=block_start_col + i, value=htext)
        c.font = FONT_ITEM_HEADER
        c.alignment = ALIGN_ITEM
        c.border = BOX_BORDER

    first_row = page_start_row + OFF_ITEM_FIRST
    last_row = page_start_row + OFF_ITEM_LAST
    for offset in range(SUBLIST_ITEM_CAPACITY_PER_BLOCK):
        r = first_row + offset
        ws.row_dimensions[r].height = ROW_HEIGHT_ITEM
        item = block.items[offset] if offset < len(block.items) else None
        item_no_cell = _set_text_cell(ws, r, block_start_col, item.item_no if item else "")
        ean_cell = _set_text_cell(ws, r, block_start_col + 1, item.ean if item else "")
        qty_cell = ws.cell(row=r, column=block_start_col + 2, value=(item.qty if item else None))
        for c in (item_no_cell, ean_cell, qty_cell):
            c.font = FONT_ITEM
            c.alignment = ALIGN_ITEM
            c.border = BOX_BORDER

    n_written = min(len(block.items), SUBLIST_ITEM_CAPACITY_PER_BLOCK)
    total_row = page_start_row + OFF_TOTAL
    qty_col_letter = get_column_letter(block_start_col + 2)
    if n_written > 0:
        formula = f"=SUM({qty_col_letter}{first_row}:{qty_col_letter}{first_row + n_written - 1})"
        # Formula stays scoped to the rows actually written (matches the
        # real template's own convention) -- but for a non-last continuation
        # block the DISPLAYED number must be that block's subtotal, and for
        # the last block it must be the grand total across every block of
        # this carton, so the live formula (subtotal of THIS block only)
        # is correct for non-last blocks; the last block instead gets the
        # grand total written as a literal value (its own item rows plus
        # every earlier block's items are not all in one contiguous range
        # any live SUM formula could cover across multiple side-by-side
        # blocks / previous pages).
        if block.is_last_block and block.block_count > 1:
            total_cell = ws.cell(row=total_row, column=block_start_col + 2, value=block.block_total_qty)
        else:
            total_cell = ws.cell(row=total_row, column=block_start_col + 2, value=formula)
    else:
        total_cell = ws.cell(row=total_row, column=block_start_col + 2, value=0)
    total_cell.font = FONT_TOTAL
    total_cell.alignment = ALIGN_ITEM


def generate_sublist_workbook(
    packages: List,
    output_path,
    *,
    template_path=None,
    carton_display_mode: str = "current_total",
) -> "SublistBuildResult":
    """Build the Sublist workbook for `packages`, in the EXACT SAME ORDER
    they were passed in (caller is responsible for passing them in PL_TOTAL
    order -- spec: "SUBLIST_TOTAL.xlsx phai co carton order giong hoan toan
    01_PL_TOTAL"). `template_path` is accepted for interface compatibility
    but unused -- this build recreates the template's measured style from
    constants (see module docstring) rather than opening/mutating the
    original binary file, the same architecture pl_ocr_core.write_workbook()
    already uses for the Packing List sheet.
    """
    cartons = [build_sublist_carton_model(p, carton_display_mode=carton_display_mode) for p in packages]
    blocks = paginate_carton_blocks(cartons)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = 11

    for i, block in enumerate(blocks):
        page_idx = i // BLOCKS_PER_PAGE
        slot = i % BLOCKS_PER_PAGE
        page_start_row = 1 + page_idx * PAGE_CYCLE_ROWS
        block_start_col = 1 + BLOCK_WIDTH_COLS * slot
        write_carton_block(ws, block, page_start_row, block_start_col)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = output_path.with_suffix(output_path.suffix + ".tmp")
    try:
        _apply_requested_excel_view_preferences(wb)
        wb.save(str(tmp_path))
    except PermissionError as e:
        raise PermissionError(
            f"Cannot create temp file for '{output_path.name}': {e}. "
            f"Close any program locking that folder and re-run."
        ) from e
    try:
        if output_path.exists():
            output_path.unlink()
        tmp_path.replace(output_path)
    except PermissionError as e:
        try:
            tmp_path.unlink(missing_ok=True)
        except Exception:
            pass
        raise PermissionError(
            f"Cannot overwrite '{output_path}': the file appears to be open in Excel. "
            f"Close it and re-run."
        ) from e

    log.info(f"Sublist: {len(cartons)} carton(s), {len(blocks)} block(s) "
             f"({sum(1 for b in blocks if b.block_index > 0)} continuation block(s)) -> {output_path}")
    return SublistBuildResult(cartons=cartons, blocks=blocks, output_path=output_path)


@dataclass
class SublistBuildResult:
    cartons: List[SublistCartonModel]
    blocks: List[SublistBlock]
    output_path: Path


# =========================================================================
# 4) Validation (spec section 13) -- never silently drop/duplicate anything
# =========================================================================
def validate_sublist(packages: List, result: "SublistBuildResult"):
    """Reconcile the generated Sublist against the source `packages`.
    Returns (ok: bool, report_text: str). Raises nothing itself -- caller
    decides whether a failed reconciliation should abort the run (mirrors
    pl_group_export.export_grouped_pl's own validate-then-raise pattern)."""
    lines: List[str] = []
    ok = True

    def check(label: str, cond: bool, detail: str = ""):
        nonlocal ok
        status = "OK" if cond else "FAIL"
        if not cond:
            ok = False
        lines.append(f"[{status}] {label}" + (f" -- {detail}" if detail else ""))
        return cond

    cartons = result.cartons
    blocks = result.blocks

    check("Sublist carton count == package count", len(cartons) == len(packages),
          f"sublist={len(cartons)} packages={len(packages)}")

    # unique carton identity count -- continuation blocks must NOT inflate this
    unique_identities = {b.carton.carton_identity for b in blocks}
    check("Unique carton identities in blocks == package count",
          len(unique_identities) == len(packages),
          f"unique={len(unique_identities)} packages={len(packages)}")

    # carton_sequence is only unique WITHIN a counting_scope_key (per-Store/
    # shipment numbering scope, spec section on per-Store carton numbering)
    # -- e.g. Kerry's "1/6" and Hangzhou's "1/4" legitimately share
    # carton_sequence=1 because they're two different Stores/scopes. Identity
    # for BOTH duplicate- and missing-sequence checks must therefore be
    # (counting_scope_key, carton_sequence), never carton_sequence alone --
    # checking global carton_sequence uniqueness across scopes was the exact
    # bug that made a correct, per-Store 1/2 2/2 / 1/2 2/2 layout look like a
    # validation failure ("duplicates=[1, 2]"). carton_display (e.g. "1/6")
    # is used only for human-readable reporting below, never as an identity
    # key -- two different scopes' cartons can share a carton_display too.
    scopes: Dict[str, List[SublistCartonModel]] = {}
    for c in cartons:
        scopes.setdefault(c.counting_scope_key, []).append(c)

    all_dup_pairs: List[tuple] = []
    all_missing_pairs: List[tuple] = []
    for scope_key, scope_cartons in scopes.items():
        scope_seqs = [c.carton_sequence for c in scope_cartons if c.carton_sequence]
        dup_seqs = sorted({s for s in scope_seqs if scope_seqs.count(s) > 1})
        if dup_seqs:
            all_dup_pairs.append((scope_key or "(default scope)", dup_seqs))
        if scope_seqs:
            # Each scope has its own carton_total (e.g. Kerry=6, Hangzhou=4)
            # -- never borrow another scope's total the way a single global
            # cartons[0].carton_total did before.
            scope_total = scope_cartons[0].carton_total or len(scope_cartons)
            expected = set(range(1, scope_total + 1))
            missing = sorted(expected - set(scope_seqs))
            if missing:
                all_missing_pairs.append((scope_key or "(default scope)", missing))

    check("No duplicate carton_sequence within any counting_scope_key", not all_dup_pairs,
          f"duplicates_by_scope={all_dup_pairs}" if all_dup_pairs else "")
    check("No missing carton_sequence within any counting_scope_key", not all_missing_pairs,
          f"missing_by_scope={all_missing_pairs}" if all_missing_pairs else "")

    for pkg, carton in zip(packages, cartons):
        pkg_items = getattr(pkg, "items", [])
        blocks_for_carton = [b for b in blocks if b.carton.carton_identity == carton.carton_identity]
        sublist_item_count = sum(len(b.items) for b in blocks_for_carton)
        check(f"Item count matches for carton {carton.carton_display or carton.package_code}",
              sublist_item_count == len(pkg_items),
              f"sublist={sublist_item_count} package={len(pkg_items)}")
        sublist_qty = sum(r.qty for b in blocks_for_carton for r in b.items)
        pkg_qty = sum(int(getattr(it, "quantity", 0) or 0) for it in pkg_items)
        check(f"Total QTY matches for carton {carton.carton_display or carton.package_code}",
              sublist_qty == pkg_qty, f"sublist={sublist_qty} package={pkg_qty}")
        sublist_item_nos = [r.item_no for b in blocks_for_carton for r in b.items]
        pkg_item_nos = [str(getattr(it, "product_code", "") or "") for it in pkg_items]
        check(f"Item No. sequence matches for carton {carton.carton_display or carton.package_code}",
              sublist_item_nos == pkg_item_nos)
        sublist_eans = [r.ean for b in blocks_for_carton for r in b.items]
        pkg_eans = [str(getattr(it, "barcode", "") or "") for it in pkg_items]
        check(f"EAN sequence matches for carton {carton.carton_display or carton.package_code}",
              sublist_eans == pkg_eans)
        check(f"Packing Code matches for carton {carton.carton_display or carton.package_code}",
              carton.packing_code == (getattr(pkg, "package_code", "") or ""))

    return ok, "\n".join(lines)
