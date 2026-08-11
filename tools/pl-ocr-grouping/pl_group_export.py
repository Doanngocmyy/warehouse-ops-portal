#!/usr/bin/env python3
"""
pl_group_export.py
===================
Splits the packages produced by OCR_Packing List.ipynb (v7 pipeline) into a
grouped set of workbooks:

    PL_SPLIT_OUTPUT/
    ├── 01_PL_TOTAL/PL_TOTAL.xlsx
    ├── 02_BY_FACTORY/PL_FACTORY_{CN,POP,SBGEAR,QIFENG,JION}.xlsx
    ├── 03_CN_BY_PORT/PL_CN_PORT_{PVG,SZX,TFU,PEK}.xlsx
    ├── 04_CN_BY_STORE/PL_CN_STORE_{...9 stores...}.xlsx
    └── PL_SPLIT_CONTROL.csv

Design notes
------------
* This module is intentionally self-contained (no import from the notebook).
  It only needs a `write_workbook(path, packages)` callable — the one already
  defined in the notebook — and a list of `Package`-like objects exposing:
  .package_code .source_file .reference_code .pdf_package_seq .items
  .global_carton_num .calc_qty (property) .declared_total_qty
  It never mutates package_code / pdf_package_seq / items — only
  `global_carton_num` is temporarily rewritten per sub-file and always
  restored afterwards (see _write_group).
* Every workbook write goes through a temp-file + os.replace() pattern so a
  file that is currently open in Excel raises a clear PermissionError instead
  of silently corrupting the target or crashing mid-write.
* Store/port classification only ever runs for packages whose FACTORY is CN.
  It never guesses: if confidence is low, or the top-2 candidates are too
  close, the package is marked REVIEW (store=REVIEW, port=REVIEW) and is
  EXCLUDED from the 03_CN_BY_PORT / 04_CN_BY_STORE files (it is still fully
  visible in PL_SPLIT_CONTROL.csv and in the validation report).
* v8: STORE_MASTER now also carries the full Notify Party / Delivery Address
  block (receiver company, full street address, contact name + phone) for
  each of the 9 CN retail stores, supplied directly by the warehouse team.
  notify_party_block() formats that into the exact "NOTIFY PARTY: / DELIVERY
  ADDRESS: / ..." text used on the real PL template, so pl_ocr_core.py's
  Packing List sheet can auto-fill Notify Party ONLY when every package in
  that file belongs to the same single store (e.g. the 04_CN_BY_STORE
  split) — never guessed/blended when a file mixes stores.
"""
from __future__ import annotations

import csv
import difflib
import logging
import re
import shutil
import unicodedata
from collections import defaultdict, OrderedDict
from pathlib import Path
from typing import Callable, Dict, List, Optional, Tuple

log = logging.getLogger("pl_group_export")
if not log.handlers:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

try:
    import pdfplumber
except ImportError:  # pragma: no cover - pdfplumber is a hard dependency of the notebook already
    pdfplumber = None


# =========================================================================
# 1) Normalization helpers (self-contained — no dependency on the notebook)
# =========================================================================

def _apply_requested_excel_view_preferences(workbook) -> None:
    from openpyxl.styles import Alignment

    target_headers = {"SKU#", "SHIPPING MARK"}

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = None

        scan_rows = min(max(worksheet.max_row, 1), 30)
        for row in worksheet.iter_rows(min_row=1, max_row=scan_rows):
            for header_cell in row:
                header = str(header_cell.value or "").strip().upper()
                if header not in target_headers:
                    continue

                col_idx = header_cell.column
                for col_cells in worksheet.iter_cols(
                    min_col=col_idx,
                    max_col=col_idx,
                    min_row=header_cell.row,
                    max_row=max(worksheet.max_row, header_cell.row),
                ):
                    for cell in col_cells:
                        old = cell.alignment
                        cell.alignment = Alignment(
                            horizontal="left",
                            vertical=old.vertical or "center",
                            wrap_text=old.wrap_text,
                            shrink_to_fit=old.shrink_to_fit,
                            text_rotation=old.text_rotation,
                        )


def _strip_accents(s: str) -> str:
    nfd = unicodedata.normalize("NFD", str(s or "").strip())
    return "".join(c for c in nfd if unicodedata.category(c) != "Mn").upper()


def _norm_text(s: str) -> str:
    """Uppercase, strip accents/diacritics, unicode-normalize, collapse
    punctuation/whitespace into single spaces. Used for fuzzy text compare."""
    if s is None:
        return ""
    s = unicodedata.normalize("NFKC", str(s))
    s = _strip_accents(s)
    s = re.sub(r"[^A-Z0-9]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


_RE_SPACED_SINGLE_LETTERS = re.compile(
    r'(?<![A-Z0-9])(?:[A-Z0-9][\s\-_]+){1,}[A-Z0-9](?![A-Z0-9])'
)
# NOTE: custom lookaround boundaries instead of \b -- \b treats underscore as a
# "word" character, which would block matching across the underscore-delimited
# segments real Shipmark codes actually use (e.g. "..._I A P M_..."). The boundary
# used here matches _tokens()'s own final split class ([^A-Z0-9], which DOES
# treat underscore as a separator).


def _tokens(code: str) -> List[str]:
    text = _strip_accents(code)
    # v14 (spec section 4): collapse a run of >=2 space/hyphen/underscore-
    # separated SINGLE characters into one glued token first -- e.g.
    # "I A P M" -> "IAPM" -- so a spelled-out Store code tokenizes and
    # matches identically on both the alias side (STORE_MASTER aliases) and
    # the live signal side (Shipmark/filename/receiver text), since both
    # go through this same function. Deliberately narrow (only fires on
    # RUNS of lone single characters, never on ordinary multi-letter
    # words) to avoid gluing unrelated short words together.
    text = _RE_SPACED_SINGLE_LETTERS.sub(lambda m: re.sub(r"[\s\-_]+", "", m.group(0)), text)
    return [t for t in re.split(r"[^A-Z0-9]+", text) if t]


# =========================================================================
# 2) Factory detection — suffix based, last token wins
# =========================================================================
# v12: "VN" is a distinct, literal factory token -- used when a shipment's
# Vietnam factory is named generically ("KR100_VN") rather than by its
# specific name (POP/SBGEAR/QIFENG/JION). Spec section 8 explicitly warns
# these must NOT be assumed equivalent -- VN is its own category, never
# silently folded into one of the other four.
FACTORY_KEYWORDS = ["SBGEAR", "QIFENG", "POP", "JION", "VN", "CN"]  # longest-first for the flat-suffix fallback
_KNOWN_FACTORY_TOKENS = {"CN", "POP", "SBGEAR", "QIFENG", "JION", "VN"}

# v11/v12: carton-NUMBERING business order (spec section 7.2/8) --
# deliberately separate from FACTORY_KEYWORDS above, which is only a
# longest-first matching-priority list for detect_factory() and has nothing
# to do with the order cartons should be numbered in. Centralised here (not
# scattered across pl_ocr_core.py) so there is exactly one place to change
# if the warehouse's factory order ever changes.
#   CARTON_FACTORY_ORDER_WITH_CO -- specific named factories, one Store
#     spanning several of them (spec 7.2/8, "CO" case).
#   CARTON_FACTORY_ORDER_NO_CO -- the literal VN/CN case (spec 7.3/8, "no
#     CO"): VN always before CN, whether or not a specific VN sub-factory
#     name is present.
# carton_factory_rank_table() merges the two into one effective order so
# they can never silently disagree with each other -- see its docstring.
CARTON_FACTORY_ORDER_WITH_CO = ["POP", "SBGEAR", "QIFENG", "JION", "CN"]
CARTON_FACTORY_ORDER_NO_CO = ["VN", "CN"]


def carton_factory_rank_table() -> List[str]:
    """Single effective carton-numbering factory order, derived from the two
    named constants above so they cannot silently diverge: every specific
    factory from CARTON_FACTORY_ORDER_WITH_CO except its trailing "CN"
    ranks first (in that order), then CARTON_FACTORY_ORDER_NO_CO's own order
    (VN, then CN) as the tail -- i.e. POP -> SBGEAR -> QIFENG -> JION -> VN
    -> CN. A shipment with only literal VN/CN packages (no specific factory
    names) gets exactly the CARTON_FACTORY_ORDER_NO_CO order (VN before CN)
    from this same table; a shipment with specific factory names gets
    CARTON_FACTORY_ORDER_WITH_CO's order, with any literal-VN packages
    sorting alongside them (after the named ones, still before CN)."""
    specific = [f for f in CARTON_FACTORY_ORDER_WITH_CO if f != "CN"]
    return specific + list(CARTON_FACTORY_ORDER_NO_CO)


# v12: trailing copy-suffix (_1 / -2 / (3) / COPY 4 / COPY_5 / COPY-6),
# anchored to the END of the string only -- same semantics as
# pl_ocr_core.parse_document_sequence's _SEQ_SUFFIX_RE (kept as an
# independent, self-contained copy here per this module's own "no import
# from pl_ocr_core" architecture -- see test_pl_group_export.py for a
# cross-consistency test guarding the two against silently diverging).
_COPY_SUFFIX_RE = re.compile(
    r'^(?P<base>.+?)(?:[_\-]\s*\d{1,4}|\(\d{1,4}\)|\s+COPY[_\-\s]\d{1,4})$',
    re.IGNORECASE)


# v12 (Turn 12 bug report): trailing version/revision marker -- "CN v",
# "CN v1", "CN rev2" -- appended when a corrected Packing List is re-
# uploaded under the same base name. Conservative on purpose: only a BARE
# "V"/"V<digits>" or "REV<digits>", separated from what precedes it by
# "_"/"-"/space, anchored to the very END of the string. This can never
# match the real "VN" factory token (VN is a literal 2-letter token, not
# "V" followed only by digits, so "V\d*$" never consumes the trailing "N")
# and never touches a mid-string occurrence -- e.g. "CN-1529_SH-Airport_PVG"
# on its own is untouched, only an actual trailing " v"/" v1"/" rev2" marker
# is stripped. Same "flag explicitly, never guess" philosophy as
# _strip_trailing_copy_suffix() below, and applied for the identical reason:
# without this, "CN-1529_SH-Airport_PVG_CN v.pdf" reads its last token as
# the literal string "V" (not a known factory token) and falls through to
# REVIEW, even though the real, intended factory suffix "CN" is right there.
_VERSION_SUFFIX_RE = re.compile(
    r'^(?P<base>.+?)[_\-\s](?:V\d*|REV\d*)$',
    re.IGNORECASE)


def _strip_trailing_version_suffix(code: str) -> str:
    """"...CN v" / "...CN v1" / "...CN rev2" -> "...CN" -- see
    _VERSION_SUFFIX_RE's module-level comment for why this is conservative
    and cannot mis-strip the real "VN" factory token or a mid-string value."""
    m = _VERSION_SUFFIX_RE.match(code or "")
    return m.group("base") if m and m.group("base") else (code or "")


def _strip_trailing_copy_suffix(code: str) -> str:
    """"Kerry_POP_1" -> "Kerry_POP" (so factory detection sees the real
    "POP" suffix instead of the copy marker "1"); "Kerry_POP" (no suffix)
    and "CN-1666-PVG-KERRY-POP" (mid-string number, not a trailing copy
    marker) are returned unchanged. Found while testing business_sort_
    packages(): calling detect_factory() on a copy-suffixed filename without
    this step misreads every _1/_2/... copy as factory=REVIEW (the literal
    last token is "1", not "POP"), which would sort a copy PDF to the very
    end of the whole shipment instead of next to its original."""
    m = _COPY_SUFFIX_RE.match(code or "")
    return m.group("base") if m and m.group("base") else (code or "")


def _detect_factory_from_code(code: str) -> Optional[str]:
    # Strip a version/revision marker BEFORE the copy-suffix marker so a
    # name combining both (e.g. "Kerry_POP_v2_1") still resolves correctly
    # -- order matters here because _strip_trailing_copy_suffix's own regex
    # requires a digit directly after "_"/"-", which a trailing "_1" copy
    # marker satisfies but a "_v2" version marker does not (no digit
    # immediately after the separator, "v" is there instead), so running
    # copy-suffix-strip first would leave the version marker behind.
    code = _strip_trailing_version_suffix(code)
    code = _strip_trailing_copy_suffix(code)
    toks = _tokens(code)
    if not toks:
        return None

    last = toks[-1]
    if last in _KNOWN_FACTORY_TOKENS:
        return last

    # Handle split spellings such as "..._SB_GEAR" or "..._QI_FENG"
    if len(toks) >= 2:
        merged2 = toks[-2] + toks[-1]
        if merged2 == "SBGEAR":
            return "SBGEAR"
        if merged2 == "QIFENG":
            return "QIFENG"

    # Fallback: flat (no-separator) suffix match, e.g. filename "TW8785SBGEAR"
    flat = "".join(toks)
    for kw in FACTORY_KEYWORDS:
        if flat.endswith(kw):
            return kw
    return None


def detect_factory(reference_code: str, source_file: str) -> str:
    """Return one of CN / POP / SBGEAR / QIFENG / JION / VN / REVIEW.

    Priority: the LAST suffix of reference_code, then the LAST suffix of the
    PDF filename (stem). A leading 'CN' (e.g. CN-2569_SH_PVG_POP) must NOT be
    read as factory=CN — only the final suffix counts. A trailing copy-
    sequence marker (_1/_2/(3)/COPY 4/...) is stripped first, so
    "Kerry_POP_1" still detects factory=POP (see _strip_trailing_copy_suffix)
    -- this is now the SINGLE source of truth for factory classification:
    every caller (classify_packages_for_port, export_grouped_pl's grouping,
    _is_all_cn_factory, business_sort_packages) goes through this same
    function and therefore gets the same, consistent answer.
    """
    f = _detect_factory_from_code(reference_code or "")
    if f:
        return f
    stem = Path(source_file).stem if source_file else ""
    f = _detect_factory_from_code(stem)
    if f:
        return f
    return "REVIEW"


# =========================================================================
# 3) Master store <-> port mapping (CN retail only)
# =========================================================================
STORE_MASTER: Dict[str, Dict[str, object]] = {
    "CHENGDU": {
        "port": "TFU",
        "receiver": "Topologie CN - Chengdu Taikooli",
        "aliases": ["Chengdu", "Chengdu Taikooli", "Taikoo Li Chengdu", "M060"],
        "address": "Shop M060, Taikoo Li Chengdu, 8 Zhongshamao Street, Jinjiang District, Chengdu City, Sichuan, CN 610021",
        "contact_name": "June",
        "contact_phone": "18084829907",
    },
    "SHENZHEN": {
        "shipping_mark_tokens": ["SZ"],
        "port": "SZX",
        "receiver": "CN - Shenzhen Mixc City (Shop T228)",
        "aliases": ["Shenzhen", "Shenzen", "Shenzhen MixC", "Mixc City", "T228"],
        "address": "Shop T228, Tower 3, Vientiane City (MixC), No. 1881 Baoan South Road, Luohu District, Shenzhen, Guangdong, CN 518000",
        "contact_name": "Ben",
        "contact_phone": "18565775002",
    },
    "GUANGZHOU": {
        "shipping_mark_tokens": ["GZ"],
        "port": "SZX",
        "receiver": "Topologie CN - Guangzhou Central Parc",
        "aliases": ["Guangzhou", "Guangzou", "Guangzhou Central Parc", "Guangzhou Parc Central", "B262-1"],
        "address": "Shop B262-1, B2/F, Parc Central, No.218 Tianhe Road, Tianhe District, Guangzhou City, Guangdong, CN 510620",
        "contact_name": "Zhang Xiaojie",
        "contact_phone": "13662343374",
    },
    "HANGZHOU": {
        "shipping_mark_tokens": ["HZ"],
        "port": "PVG",
        "receiver": "Topologie CN - Hangzhou Mixc",
        "aliases": ["Hangzhou", "Hang Zhou", "Hangzhou MixC", "B1C03"],
        "address": "B1C03, Hangzhou MixC Mall, 701 Fuchun Rd, Jianggan District, Hangzhou, Zhejiang, CN 310008",
        "contact_name": "Su Su",
        "contact_phone": "15606539115",
    },
    "IAPM": {
        "shipping_mark_tokens": ["IAPM"],
        "port": "PVG",
        "receiver": "Topologie CN - Iapm",
        "aliases": ["IAPM", "I A P M", "IAPM Mall", "L4-426"],
        "address": "L4-426, IAPM Mall, 999 Huaihai Rd (M), Xuhui District, Shanghai, Shanghai, CN 200020",
        "contact_name": "Shi Wei Yi",
        "contact_phone": "13621647004",
    },
    "KERRY": {
        # v14 (spec section 4): KR/KRY/KRYY/KER -- explicit short codes only
        # (never fuzzy, per the "short aliases <=3 chars -> exact token
        # only" rule), all resolve unambiguously to KERRY.
        "shipping_mark_tokens": ["KR", "KRY", "KRYY", "KER"],
        "port": "PVG",
        "receiver": "Topologie CN - Kerry Center flagship",
        "aliases": ["Kerry", "Kerry Center", "Kerry Centre", "NB1-23B"],
        "address": "NB1-23B shop, B1 floor, Jing'an Kerry Centre, Jing'an District, Shanghai, Shanghai, CN 200040",
        "contact_name": "Ning Ning",
        "contact_phone": "17602197790",
    },
    "SHANGHAI_TAIKOOLI": {
        "shipping_mark_tokens": ["SHTAIKOOLI"],
        "port": "PVG",
        "receiver": "CN - Shanghai Taikooli (Shop B1-07b)",
        # NOTE (spec section 4): "SH-Taikooli"/"SHai Taikooli" spelling
        # variants are intentionally NOT added as free-text aliases here --
        # both would tokenize down to a bare "SH" unigram shared with
        # SHANGHAI_HONGQIAO's own "SH-Airport" alias, creating exactly the
        # kind of cross-store ambiguity _GENERIC_ALIAS_STOPWORDS exists to
        # prevent. The compound form is already resolved unambiguously via
        # shipping_mark_tokens=["SHTAIKOOLI"] + the bigram-joining exact
        # matcher (see _tokens_with_bigrams) -- "SH"+"Taikooli" in a real
        # Shipmark joins to "SHTAIKOOLI" and matches that token directly,
        # no separate alias needed.
        "aliases": ["Shanghai Taikooli", "Shanghai Taikoo Li", "B1-07b", "S-B1-07b"],
        "address": "Shop S-B1-07b, B/F, No.1-9, 500 Dongyu Road, Pudong, Shanghai, Shanghai, CN 200127",
        "contact_name": "Bobo Shi",
        "contact_phone": "13621647004",
    },
    "SHANGHAI_HONGQIAO": {
        "shipping_mark_tokens": ["SHAIRPORT"],
        "port": "PVG",
        "receiver": "CN - Shanghai Hongqiao Airport",
        # NOTE (spec section 4): see the matching NOTE on SHANGHAI_TAIKOOLI
        # above -- "SH Airport"/"SH-Airport" deliberately NOT added as a
        # free-text alias (bare "SH" collision); already resolved via
        # shipping_mark_tokens=["SHAIRPORT"] + bigram-joining.
        "aliases": ["Shanghai Hongqiao", "Hongqiao Airport", "D60-6"],
        "address": "Shop D60-6, Shanghai Hongqiao International Airport Terminal 2 (Departure Restricted Area), Changning District, Shanghai, Shanghai, China 200335",
        "contact_name": "Bobo Shi",
        "contact_phone": "13621647004",
    },
    "CHINA_WORLD": {
        "port": "PEK",
        "receiver": "China World NB1026",
        "aliases": ["China World", "China World Mall", "NB1026", "CNWORLD", "CNWorld", "CN World", "ChinaWorld", "China World NB1026", "CNWorld NB1026"],
        "address": "Shop NB1026, B1 Floor, China World Mall, No. 1 Jianguomenwai Avenue, Chaoyang District, Beijing, Beijing, China 100004",
        "contact_name": "Bobo Shi",
        "contact_phone": "+86 13621647004",
    },
}

FACTORY_FILE_MAP = {
    "CN": "PL_FACTORY_CN.xlsx",
    "POP": "PL_FACTORY_POP.xlsx",
    "SBGEAR": "PL_FACTORY_SBGEAR.xlsx",
    "QIFENG": "PL_FACTORY_QIFENG.xlsx",
    "JION": "PL_FACTORY_JION.xlsx",
}
PORT_FILE_MAP = {
    "PVG": "PL_CN_PORT_PVG.xlsx",
    "SZX": "PL_CN_PORT_SZX.xlsx",
    "TFU": "PL_CN_PORT_TFU.xlsx",
    "PEK": "PL_CN_PORT_PEK.xlsx",
}
STORE_FILE_MAP = {k: f"PL_CN_STORE_{k}.xlsx" for k in STORE_MASTER}


def _dynamic_store_filename(label: str) -> str:
    """Filename for a Store that resolved via the OR List but isn't one of
    the 9 named STORE_MASTER keys (spec: OR List Store column is dynamic,
    not limited to a fixed enum) -- sanitize into a safe filename instead of
    silently dropping the group."""
    safe = _strip_accents(str(label or "STORE")).upper()
    safe = re.sub(r"[^A-Z0-9]+", "_", safe).strip("_") or "STORE"
    return f"PL_CN_STORE_{safe}.xlsx"


# v18: _NON_CN_COUNTRIES (a finite 5-country blocklist) was removed here
# -- both is_cn_port_eligible() and _resolved_store_for_split() below now
# use a positive allowlist (country in ("", "CN")) instead, see their own
# docstrings for the root-cause writeup (SG-533-TEST regression report).

# v16 (correction -- an earlier v15 draft of this fix wrongly restricted
# PORT/Store eligibility to factory in {"CN", "VN"}, which is JUST AS WRONG
# as the original bug: Store/Port belong to the Shipping Mark BODY and are
# COMPLETELY INDEPENDENT of the trailing factory/origin suffix. CN-1529_
# HZ_PVG_POP / _SBGEAR / _QIFENG / _JION / _VN / _CN must ALL resolve to
# the identical HANGZHOU/PVG -- only the factory/origin dimension differs.
# This already matches 04_CN_BY_STORE's existing behavior (which combines
# a Store's POP+SBGEAR/QIFENG/JION+VN+CN cartons into ONE file -- see
# t_export_grouped_pl_store_split_combines_pop_vn_cn_and_keeps_scope_
# numbering in tests/test_pl_ocr_core.py); 03_CN_BY_PORT/PL_Total's PORT
# must not disagree with that by using a narrower, factory-based gate.
# The ONLY thing that determines eligibility is destination COUNTRY.
def is_cn_port_eligible(pkg, factory: str = None) -> bool:
    """True when `pkg` should go through the CN Store/Port resolver
    (classify_packages_for_port() in pl_ocr_core.py) at all -- i.e. its
    destination country is CN (or not yet/never determined -- "" is
    treated as "don't know, don't exclude", matching every other v14
    country gate in this codebase, e.g. _resolved_store_for_split()
    below). Store/Port are resolved from the Shipping Mark BODY and are
    NEVER gated on the trailing factory/origin suffix -- `factory` is
    accepted only for call-signature stability (some callers already have
    it computed for the unrelated 02_BY_FACTORY grouping) and is not
    consulted here.

    This is the ONE structural gate BOTH classify_packages_for_port()
    (fills pkg.port/pkg.store -- the canonical PL_Total PORT source) and
    export_grouped_pl()'s 03_CN_BY_PORT split / PL_SPLIT_VALIDATION use --
    it decides ELIGIBILITY only, never the resolved Store/Port VALUE
    itself (that always comes from pkg.port/pkg.store, set exactly once by
    classify_packages_for_port() -- exporters must never re-run
    match_store() independently, see export_grouped_pl()'s classify loop).

    v18 (SG-533-TEST real-fixture regression report, root cause): this
    USED TO be `country not in _NON_CN_COUNTRIES` -- a finite 5-country
    BLOCKLIST (KR/JP/BE/US/TW). That is backwards from the documented
    intent above ("destination country is CN, or unknown"): any country
    code NOT in that small hardcoded set -- e.g. SG, TH, PH, AU, MY, or
    any future/typo'd code -- fell through as ELIGIBLE, the exact opposite
    of "only CN". A v18 fix widened this to a positive allowlist that
    still treated "" (unresolved) as eligible.

    v19 (SG-533-TEST consolidation report, requirement 2): now that
    detect_shipment_country() (see pl_ocr_core.py) is a generic structural
    rule instead of a fixed 6-code list, a genuinely non-CN Shipmark/
    filename prefix is detected correctly far more often, so a residual
    "" no longer means "probably CN, just undetected" as often as before
    -- it means "no usable country signal at all". Per explicit
    instruction ("blank/unresolved country must NOT be treated as China-
    eligible... do not attempt CN resolution and hope it fails"), this is
    now STRICT: eligible only when country is literally "CN". "" and
    every other country, known or not, are excluded, no exceptions."""
    return getattr(pkg, "country", "") == "CN"


def _resolved_store_for_split(pkg) -> str:
    """Cross-factory Store identity key for the 04_CN_BY_STORE split (spec
    sections 9-12: a Store's POP/SBGEAR/QIFENG/JION/CN cartons all belong in
    ONE file, sharing the ONE denominator already computed by pl_ocr_core.py's
    counting_scope_key/assign_global_numbers()). Priority mirrors
    compute_counting_scope_key() exactly, INCLUDING its section 3/11 country
    gate: non-China countries (any code other than "CN"/"") are SINGLE_DESTINATION and
    never get a China Store split, even if an OR List/CN classifier would
    otherwise resolve one. Then: an OK OR List match first, else the
    CN-only classify_packages_for_port() result, else "" (excluded on
    purpose -- never silently guessed)."""
    # v19: see is_cn_port_eligible()'s docstring -- same root cause,
    # now STRICT: only country == "CN" is eligible, "" included in the
    # exclusion (was previously also treated as eligible).
    if getattr(pkg, "country", "") != "CN":
        return ""
    if getattr(pkg, "or_list_match_status", "") == "OK" and getattr(pkg, "or_list_store", ""):
        return store_identity(pkg.or_list_store)
    if getattr(pkg, "store", "") and pkg.store != "REVIEW":
        return store_identity(pkg.store)
    return ""


def store_display_name(store_key: str) -> str:
    """Human, customer-facing Store name for a resolved STORE_MASTER key
    (spec sections 8/12/16/17: "Store" is now a fixed, always-shown
    business-backbone field on the Packing List and Sublist, alongside OR
    No./Ref No.). Derived from STORE_MASTER's own `receiver` field with
    the "Topologie CN - " / "CN - " boilerplate prefix stripped, rather
    than a second parallel alias/display table -- e.g. "Topologie CN -
    Hangzhou Mixc" -> "Hangzhou Mixc", "China World NB1026" (no prefix at
    all) -> unchanged. Falls back to the raw `store_key` (e.g. "REVIEW",
    or any non-STORE_MASTER value) when there's no STORE_MASTER entry to
    look up -- never invents a display name. Blank in, blank out."""
    if not store_key:
        return ""
    info = STORE_MASTER.get(store_key)
    if not info:
        return store_key
    receiver = str(info.get("receiver", "") or "").strip()
    if not receiver:
        return store_key
    for prefix in ("Topologie CN - ", "CN - "):
        if receiver.startswith(prefix):
            return receiver[len(prefix):]
    return receiver


def _store_display_label(pkg, store_key: str) -> str:
    """Human-readable label for a resolved store key, preferring the raw OR
    List text (e.g. "Kerry") over the bare identity key, for filenames of
    stores outside the 9 named STORE_MASTER keys."""
    if getattr(pkg, "or_list_match_status", "") == "OK" and getattr(pkg, "or_list_store", ""):
        return pkg.or_list_store
    return getattr(pkg, "store", "") or store_key


def _store_file_key(pkg, store_key: str) -> str:
    """Best-effort CANONICAL STORE_MASTER key, for filename purposes only
    (e.g. "KERRY" -> PL_CN_STORE_KERRY.xlsx instead of a sanitized dump of
    the OR List's full free-text description). Grouping itself always uses
    the raw `store_key` from _resolved_store_for_split() so file contents
    never disagree with pl_ocr_core.compute_counting_scope_key()'s
    numbering scope -- this only affects the filename that scope is
    written under."""
    if getattr(pkg, "or_list_match_status", "") == "OK" and getattr(pkg, "or_list_store", ""):
        # _canonical_store_identity_for_or_row() returns space-separated
        # identity form (e.g. "SHANGHAI HONGQIAO") for multi-word keys --
        # STORE_MASTER itself is keyed with underscores ("SHANGHAI_HONGQIAO").
        canon = _canonical_store_identity_for_or_row(pkg.or_list_store).replace(" ", "_")
        if canon in STORE_MASTER:
            return canon
    if getattr(pkg, "store", "") in STORE_MASTER:
        return pkg.store
    return store_key


def notify_party_block(store_key: str) -> str:
    """Format the exact 'NOTIFY PARTY: / DELIVERY ADDRESS: / ...' text block
    used on the real PL template, from STORE_MASTER's contact info. Returns
    "" if store_key isn't a known store (caller should leave the cell blank
    in that case rather than guessing)."""
    info = STORE_MASTER.get(store_key)
    if not info:
        return ""
    lines = ["NOTIFY PARTY:", "DELIVERY ADDRESS:", str(info.get("receiver", "")), str(info.get("address", ""))]
    contact = str(info.get("contact_name", "") or "")
    phone = str(info.get("contact_phone", "") or "")
    if contact:
        lines.append(contact)
    if phone:
        lines.append(f"Tel: {phone}")
    return "\n".join(lines)


def _build_alias_lookup() -> List[Tuple[str, str]]:
    lookup: List[Tuple[str, str]] = []
    for store_key, info in STORE_MASTER.items():
        cand_texts = [store_key.replace("_", " "), str(info["receiver"])] + list(info["aliases"])
        for t in cand_texts:
            n = _norm_text(t)
            if n:
                lookup.append((store_key, n))
    return lookup


_ALIAS_LOOKUP = _build_alias_lookup()


def _token_overlap_score(signal: str, alias: str) -> float:
    ali_tokens = set(alias.split())
    if not ali_tokens:
        return 0.0
    sig_tokens = set(signal.split())
    inter = sig_tokens & ali_tokens
    return len(inter) / len(ali_tokens)


def match_store(signal_text: str, threshold: float = 0.55, margin: float = 0.08) -> Tuple[str, float, str]:
    """Fuzzy-match free text against STORE_MASTER.

    Returns (store_key_or_REVIEW, confidence[0..1], suggested_store_if_review).
    Never guesses: low confidence or a too-close runner-up both yield REVIEW.
    """
    norm_signal = _norm_text(signal_text)
    if not norm_signal:
        return "REVIEW", 0.0, ""

    # 1) exact / substring alias hits (shop number, distinctive alias) -> high confidence
    exact_hits = set()
    for store_key, alias_norm in _ALIAS_LOOKUP:
        if not alias_norm:
            continue
        if alias_norm == norm_signal or alias_norm in norm_signal or norm_signal in alias_norm:
            exact_hits.add(store_key)
    if len(exact_hits) == 1:
        return next(iter(exact_hits)), 1.0, ""
    if len(exact_hits) > 1:
        return "REVIEW", 0.5, "/".join(sorted(exact_hits))

    # 2) fuzzy scoring — best per store across all its aliases
    scores: Dict[str, float] = {}
    for store_key, alias_norm in _ALIAS_LOOKUP:
        ratio = difflib.SequenceMatcher(None, norm_signal, alias_norm).ratio()
        overlap = _token_overlap_score(norm_signal, alias_norm)
        s = max(ratio, overlap)
        if s > scores.get(store_key, 0.0):
            scores[store_key] = s

    if not scores:
        return "REVIEW", 0.0, ""

    ranked = sorted(scores.items(), key=lambda kv: kv[1], reverse=True)
    best_store, best_score = ranked[0]
    second_score = ranked[1][1] if len(ranked) > 1 else 0.0

    if best_score < threshold:
        return "REVIEW", round(best_score, 3), best_store
    if (best_score - second_score) < margin:
        return "REVIEW", round(best_score, 3), best_store
    return best_store, round(best_score, 3), ""


# =========================================================================
# v12) Store / OR / SO matching against an (optional) OR List (spec section
#      5) -- separate from match_store() above (which is CN-STORE_MASTER-
#      specific, used for Notify Party/address lookup). This matches
#      against the OR List's OWN store vocabulary, which can name ANY
#      store across ANY factory, not just the 9 CN retail stores.
# =========================================================================
from dataclasses import dataclass as _dc_dataclass, field as _dc_field


@_dc_dataclass
class StoreOrMatchResult:
    matched_store: str = ""
    matched_or: str = ""
    matched_so: str = ""
    # v14: the full dynamic business_fields dict for the matched row
    # (display_header -> value, in upload column order -- spec section 6).
    # matched_or/matched_so above are kept as backward-compatible aliases
    # for the 1st/2nd business field (many existing callers still read
    # them generically that way) -- matched_business_fields is the
    # complete, order-preserving source of truth for every field.
    matched_business_fields: "OrderedDict" = _dc_field(default_factory=OrderedDict)
    # v17 (doc correction): the ACTUAL runtime values are SHIPMARK_SAFE_
    # ALIAS | FILENAME_SAFE_ALIAS | RECEIVER_SAFE_ALIAS | "" (one unified
    # "safe" tier per signal source -- covers both literal shipping_mark_
    # tokens hits and safe-typo-tolerant hits under the same label; see
    # match_store_and_or()). The previous SHIPMARK_TOKEN_EXACT/FUZZY names
    # in this comment were stale leftover docstring-era terminology that
    # never matched what the code actually returned -- corrected here
    # rather than perpetuated, per root-cause audit of test_pl_group_
    # export.py's 3 stale match_source assertions.
    match_source: str = ""   # SHIPMARK_SAFE_ALIAS | FILENAME_SAFE_ALIAS | RECEIVER_SAFE_ALIAS | ""
    candidate_store: str = ""
    candidate_score: float = 0.0
    status: str = "REVIEW"   # OK | REVIEW | NO_OR_LIST
    review_reason: str = ""


def store_identity(store_text: str) -> str:
    """Case/whitespace-insensitive identity key so "Kerry" (OR List's own
    casing) and "KERRY" (STORE_MASTER's enum-style key) are recognised as
    the SAME store instead of a false ambiguity between two differently-
    cased spellings of one store. Public (no leading underscore) because
    pl_ocr_core.py's counting-scope-key computation reuses it too (spec
    section 9: counting_scope_key = shipment_key + "|" + normalized_store)."""
    return _strip_accents(store_text).strip()


_store_identity = store_identity  # backward-compat alias for internal callers


_GENERIC_ALIAS_STOPWORDS = {
    "TAIKOOLI", "TAIKOO", "MIXC", "CENTRAL", "PARC", "CENTER", "CENTRE",
    "MALL", "CITY", "SHOP", "PLAZA", "TOWER", "FLOOR",
}


_STORE_MASTER_IDENTITIES = {_store_identity(k.replace("_", " ")) for k in STORE_MASTER}


def _store_alias_token_index(or_list_store_values) -> Dict[str, set]:
    """token(uppercased, >=2 chars) -> set of distinct store IDENTITIES
    (see _store_identity above) it could mean. Built from BOTH
    STORE_MASTER's own alias config (helps match a CN store even when the
    OR List just says "Kerry") AND the OR List's own STORE column values
    (works for factories STORE_MASTER doesn't cover at all, e.g.
    POP/SBGEAR/QIFENG stores). A token that maps to more than one DISTINCT
    store identity is ambiguous by construction -- never used for an exact
    match (spec: "RY must not automatically match KERRY unless RY is
    explicitly configured as a Store alias" -- since tokens are never split
    below word boundaries, "RY" and "KERRY" are simply different tokens;
    this only becomes relevant if two DIFFERENT stores happen to share a
    configured alias token, which is flagged as ambiguous rather than
    picked)."""
    idx: Dict[str, set] = {}

    def add(token, store_identity):
        if len(token) >= 2:
            idx.setdefault(token, set()).add(store_identity)

    for store_key, info in STORE_MASTER.items():
        identity = _store_identity(store_key.replace("_", " "))
        # NOTE: deliberately NOT tokenizing info["receiver"] (a full postal/
        # company description, e.g. "CN - Shenzhen Mixc City (Shop T228)")
        # -- splitting that into individual words would flood the index
        # with generic English tokens ("Shop", "City", "Tower", ...) that
        # collide across many stores and make everything "ambiguous". Only
        # the store's own short/distinctive name + explicitly configured
        # aliases (e.g. "Kerry", "Kerry Center", "NB1-23B") + explicit
        # shipping_mark_tokens (e.g. "GZ", "SHAIRPORT" -- v13) are tokenized.
        alias_texts = [store_key.replace("_", " ")] + list(info["aliases"]) + list(info.get("shipping_mark_tokens", []))
        for alias in alias_texts:
            for tok in _tokens(alias):
                if tok in _GENERIC_ALIAS_STOPWORDS:
                    continue
                add(tok, identity)
    for store_raw in or_list_store_values:
        identity = _canonical_store_identity_for_or_row(store_raw)
        if identity in _STORE_MASTER_IDENTITIES:
            # v13 bugfix: this store already got its curated, collision-safe
            # aliases indexed above (via STORE_MASTER's own alias +
            # shipping_mark_tokens loop) -- do NOT also tokenize its raw OR
            # List free-text description here. Real OR List descriptions
            # share boilerplate words across EVERY row ("CN -", a leading
            # date stamp, "Replen") -- indexing those would make e.g. the
            # literal token "CN" resolve to all 7 stores at once, which then
            # poisons matching for a real Shipping Mark like
            # "CN-1529_HZ_PVG_POP" (its own "CN" token would collide with
            # that same ambiguous entry even though "HZ" alone is a clean,
            # unambiguous hit). Only stores STORE_MASTER doesn't know about
            # at all still need their raw description tokenized here, since
            # that free text is the ONLY signal available for them.
            continue
        for tok in _tokens(store_raw):
            add(tok, identity)
    return idx


def _canonical_store_identity_for_or_row(store_raw: str) -> str:
    """v13 (FIX2): resolve an OR List's free-text Store description (e.g.
    "20260609 CN - Guangzhou Parc Central Replen") to a CANONICAL
    STORE_MASTER identity ("GUANGZHOU") by reusing match_store()'s existing
    alias/fuzzy logic against STORE_MASTER -- never a bespoke new matcher.
    Falls back to the old literal-text identity (_store_identity(store_raw))
    when the description doesn't resolve to any STORE_MASTER store (e.g. a
    POP/SBGEAR/QIFENG-only store STORE_MASTER doesn't know about at all) --
    so non-CN-retail stores keep working exactly as before, just keyed by
    their own literal text instead of a canonical enum key that doesn't
    exist for them."""
    store_key, _score, _suggestion = match_store(store_raw)
    if store_key and store_key != "REVIEW":
        # NOTE: _store_alias_token_index() (and match_store_and_or's own
        # STORE_MASTER-anchored identities) use _store_identity(store_key.
        # replace("_", " ")) as the canonical identity form for multi-word
        # keys like "SHANGHAI_HONGQIAO" -> "SHANGHAI HONGQIAO" -- returning
        # the raw underscored key here instead would silently create a
        # SECOND, incompatible identity space for exactly those stores,
        # breaking identity_to_raw lookups after an otherwise-successful
        # exact-token Shipmark match. Must stay in lockstep with that form.
        return _store_identity(store_key.replace("_", " "))
    return _store_identity(store_raw)


def _tokens_with_bigrams(text: str) -> List[str]:
    """v13 (FIX3): unigram tokens PLUS every adjacent pair joined with no
    separator -- e.g. "CN-1529_SH-Airport_PVG_POP" tokenizes to
    ["CN","1529","SH","AIRPORT","PVG","POP"] and this adds "CN1529",
    "1529SH", "SHAIRPORT", "AIRPORTPVG", "PVGPOP". Needed because a real
    Shipping Mark can carry a compound short code split across a hyphen
    ("SH-Airport", "SH-Taikooli") that would otherwise tokenize into two
    generic, individually-ambiguous pieces ("SH" alone means nothing on its
    own -- it's a prefix shared by both Shanghai stores in this shipment).
    Joining "SH"+"AIRPORT" -> "SHAIRPORT" lets it match the single explicit
    shipping_mark_tokens alias configured for that store (never a fuzzy
    guess -- still an exact-string index lookup, just against a 2-token
    joined key instead of a 1-token one)."""
    toks = _tokens(text)
    bigrams = [toks[i] + toks[i + 1] for i in range(len(toks) - 1)]
    return toks + bigrams


def _exact_token_store_match(text: str, token_index: Dict[str, set]):
    """Returns (store_or_None, ambiguous_candidates_or_None)."""
    hits: set = set()
    ambiguous: set = set()
    for tok in _tokens_with_bigrams(text):
        stores = token_index.get(tok)
        if not stores:
            continue
        if len(stores) == 1:
            hits |= stores
        else:
            ambiguous |= stores
    if len(hits) == 1 and not (ambiguous - hits):
        return next(iter(hits)), None
    if hits or ambiguous:
        return None, sorted(hits | ambiguous)
    return None, None


def _fuzzy_match_against_candidates(signal_text: str, candidates, threshold: float = 0.6, margin: float = 0.1):
    """Generic version of match_store()'s tier-2 fuzzy scoring, parameterized
    over an arbitrary candidate list instead of hardcoded STORE_MASTER --
    used as the conservative last-resort fallback for OR List store names
    that aren't in STORE_MASTER at all. Same "REVIEW beats a wrong guess"
    philosophy: low confidence or a too-close runner-up both yield REVIEW."""
    norm_signal = _norm_text(signal_text)
    if not norm_signal or not candidates:
        return "REVIEW", 0.0, ""
    scores: Dict[str, float] = {}
    for cand in candidates:
        norm_cand = _norm_text(cand)
        if not norm_cand:
            continue
        ratio = difflib.SequenceMatcher(None, norm_signal, norm_cand).ratio()
        overlap = _token_overlap_score(norm_signal, norm_cand)
        s = max(ratio, overlap)
        if s > scores.get(cand, 0.0):
            scores[cand] = s
    if not scores:
        return "REVIEW", 0.0, ""
    ranked = sorted(scores.items(), key=lambda kv: kv[1], reverse=True)
    best_cand, best_score = ranked[0]
    second_score = ranked[1][1] if len(ranked) > 1 else 0.0
    if best_score < threshold:
        return "REVIEW", round(best_score, 3), best_cand
    if (best_score - second_score) < margin:
        return "REVIEW", round(best_score, 3), best_cand
    return best_cand, round(best_score, 3), ""



def _store_name_from_or_memo(raw_value: str, canonical_identity: str = "") -> str:
    """Extract a human-readable Store name from the OR List first/Memo column."""
    raw = str(raw_value or "").strip()
    if raw:
        hits = list(re.finditer(r"(?i)\bCN\s*-\s*", raw))
        if hits:
            raw = raw[hits[-1].end():].strip()
        raw = re.sub(r"^\s*\d{6,8}\s+", "", raw).strip()
        raw = re.split(r"(?i)\s+(?:REPLEN\b|REP(?=\d{6,8}\b))", raw, maxsplit=1)[0].strip()
        raw = re.sub(r"\s+", " ", raw).strip(" -_/")
        if raw:
            return raw
    return str(canonical_identity or "").strip()


def _or_row_kind(row) -> int:
    """Normal Replen first, then Material, then Extra Straps."""
    memo = str(getattr(row, "store_raw", "") or "").upper()
    if "MATERIAL" in memo:
        return 1
    if "EXTRA STRAP" in memo:
        return 2
    return 0





# === KEC_SAFE_STORE_RESOLVER_V4 START ===

_KEC_STORE_DISPLAY = {
    "CHINA_WORLD": "China World NB1026",
    "GUANGZHOU": "Guangzhou Central Parc",
    "HANGZHOU": "Hangzhou Mixc",
    "IAPM": "Iapm",
    "KERRY": "Kerry Center flagship",
    "SHANGHAI_HONGQIAO": "Shanghai Hongqiao Airport",
    "SHANGHAI_TAIKOOLI": "Shanghai Taikooli (Shop B1-07b)",
    "SHENZHEN": "Shenzhen Mixc City (Shop T228)",
}

_KEC_STORE_PORT = {
    "CHINA_WORLD": "PEK",
    "GUANGZHOU": "SZX",
    "HANGZHOU": "PVG",
    "IAPM": "PVG",
    "KERRY": "PVG",
    "SHANGHAI_HONGQIAO": "PVG",
    "SHANGHAI_TAIKOOLI": "PVG",
    "SHENZHEN": "SZX",
}

_KEC_STORE_ALIASES = {
    "CHINA_WORLD": (
        "CNWORLD", "CN WORLD", "CHINAWORLD", "CHINA WORLD",
        "CHINA WORLD NB1026", "CNWORLD NB1026", "NB1026",
    ),
    "GUANGZHOU": (
        "GZ", "GUANGZHOU", "GUANG ZHOU",
        "GUANGZHOU CENTRAL PARC", "GUANGZHOU PARC CENTRAL",
        "CENTRAL PARC", "PARC CENTRAL",
    ),
    "HANGZHOU": (
        "HZ", "HANGZHOU", "HANG ZHOU",
        "HANGZHOU MIXC", "HANGZHOU MIX C", "MIXC HANGZHOU",
    ),
    "IAPM": ("IAPM", "IAPM SHANGHAI"),
    "KERRY": (
        # v17 (real production bug fix, spec sections 9/32): KRY/KRYY/KER
        # were MISSING here even though STORE_MASTER["KERRY"]["shipping_
        # mark_tokens"] already listed all 4 (KR/KRY/KRYY/KER) -- this
        # table (_KEC_STORE_ALIASES) is a SEPARATE alias index consulted
        # by match_store_and_or()/_kec_resolve_store_identity(), which had
        # silently drifted out of sync with STORE_MASTER's own config, so
        # a real "CN-xxxx_KRY_PVG_POP"-style Shipping Mark would fall
        # through to REVIEW instead of resolving to Kerry. Root-caused via
        # test_pl_group_export.py's Kerry regression test; fixed here by
        # bringing this table back in sync rather than adding a second
        # parallel fix elsewhere.
        "KR", "KRY", "KRYY", "KER", "KERY",
        "KERRY CENTER", "KERRY CENTRE",
        "KERRY CENTER FLAGSHIP", "KERRY CENTRE FLAGSHIP",
    ),
    "SHANGHAI_HONGQIAO": (
        "SH AIRPORT", "SHAIRPORT", "SH HONGQIAO", "SHHONGQIAO",
        "HONGQIAO", "HONGQIAO AIRPORT",
        "SHANGHAI HONGQIAO", "SHANGHAI HONGQIAO AIRPORT",
    ),
    "SHANGHAI_TAIKOOLI": (
        "SH TAIKOOLI", "SHTAIKOOLI", "SH TAIKOO LI",
        "TAIKOOLI", "TAIKOO LI", "TAIKOLI",
        "SHANGHAI TAIKOOLI",
    ),
    "SHENZHEN": (
        "SZ", "SHENZHEN", "SHEN ZHEN",
        "SHENZHEN MIXC", "SHENZHEN MIXC CITY",
        "MIXC CITY SHENZHEN", "T228",
    ),
}

# v17: KRY/KRYY/KER added alongside KR (spec section 10: short aliases must
# be explicit/exact-token matches, never fuzzy-substring/typo-tolerant --
# these are short enough that the generic substring-containment scoring
# path a few lines down would be looser than intended for them).
_KEC_SHORT_ALIASES = {"GZ", "HZ", "KR", "SZ", "KRY", "KRYY", "KER"}
_KEC_PORTS = {"PEK", "PVG", "SZX", "TFU"}
_KEC_GENERIC_TOKENS = {
    "CN", "TOPOLOGIE", "SHOP", "MALL", "REPLEN", "MATERIAL",
    "EXTRA", "STRAPS", "STRAP", "FLAGSHIP",
    "PEK", "PVG", "SZX", "TFU", "VN", "POP",
}

def _kec_words(value: str):
    raw = str(value or "").upper()
    raw = re.sub(r"(?<![A-Z0-9])KERY(?![A-Z0-9])", "KERRY", raw)
    return re.findall(r"[A-Z0-9]+", raw)

def _kec_is_cn_signal(value: str) -> bool:
    return "CN" in set(_kec_words(value))

def _kec_signal_port(value: str):
    ports = [t for t in _kec_words(value) if t in _KEC_PORTS]
    return ports[0] if len(set(ports)) == 1 else None

def _kec_port_ok(store_key: str, signal: str) -> bool:
    seen = _kec_signal_port(signal)
    return True if not seen else seen == _KEC_STORE_PORT.get(store_key)

def _kec_damerau_le1(a: str, b: str) -> bool:
    a = str(a or "").upper()
    b = str(b or "").upper()
    if a == b:
        return True
    if abs(len(a) - len(b)) > 1:
        return False

    if len(a) == len(b):
        diffs = [i for i, (x, y) in enumerate(zip(a, b)) if x != y]
        if len(diffs) == 1:
            return True
        if len(diffs) == 2:
            i, j = diffs
            return j == i + 1 and a[i] == b[j] and a[j] == b[i]
        return False

    short, long = (a, b) if len(a) < len(b) else (b, a)
    i = j = edits = 0
    while i < len(short) and j < len(long):
        if short[i] == long[j]:
            i += 1
            j += 1
        else:
            edits += 1
            j += 1
            if edits > 1:
                return False
    return True

def _kec_exact_alias_score(store_key: str, signal: str, alias: str) -> int:
    if not _kec_port_ok(store_key, signal):
        return 0

    words = _kec_words(signal)
    aw = _kec_words(alias)
    if not words or not aw:
        return 0

    word_set = set(words)
    compact = "".join(words)
    acompact = "".join(aw)

    if len(aw) == 1 and aw[0] in _KEC_SHORT_ALIASES:
        if not _kec_is_cn_signal(signal):
            return 0
        return 5000 if aw[0] in word_set else 0

    if len(aw) == 1 and aw[0] in {"NB1026", "T228"}:
        return 4800 if aw[0] in word_set else 0

    if acompact and acompact in compact:
        return 4000 + len(acompact)

    meaningful = [x for x in aw if x not in _KEC_GENERIC_TOKENS]
    if meaningful and all(x in word_set for x in meaningful):
        return 3000 + sum(len(x) for x in meaningful)

    return 0

def _kec_typo_score(store_key: str, signal: str) -> int:
    if not _kec_is_cn_signal(signal):
        return 0
    if not _kec_port_ok(store_key, signal):
        return 0

    signal_tokens = [
        t for t in _kec_words(signal)
        if len(t) >= 4 and t not in _KEC_GENERIC_TOKENS and not t.isdigit()
    ]
    if not signal_tokens:
        return 0

    alias_tokens = set()
    for alias in _KEC_STORE_ALIASES[store_key]:
        aw = _kec_words(alias)
        if len(aw) == 1 and aw[0] in _KEC_SHORT_ALIASES:
            continue
        for t in aw:
            if len(t) >= 4 and t not in _KEC_GENERIC_TOKENS and not t.isdigit():
                alias_tokens.add(t)

    hits = []
    for st in signal_tokens:
        for at in alias_tokens:
            if _kec_damerau_le1(st, at):
                hits.append((st, at))

    if not hits:
        return 0
    return 1000 + max(max(len(a), len(b)) for a, b in hits)

def _kec_resolve_store_identity(value: str):
    if not value:
        return None

    exact_scores = {}
    for key, aliases in _KEC_STORE_ALIASES.items():
        best = max((_kec_exact_alias_score(key, value, a) for a in aliases), default=0)
        if best:
            exact_scores[key] = best

    if exact_scores:
        top = max(exact_scores.values())
        winners = [k for k, v in exact_scores.items() if v == top]
        return winners[0] if len(winners) == 1 else None

    typo_scores = {}
    for key in _KEC_STORE_ALIASES:
        score = _kec_typo_score(key, value)
        if score:
            typo_scores[key] = score

    if not typo_scores:
        return None

    top = max(typo_scores.values())
    winners = [k for k, v in typo_scores.items() if v == top]
    return winners[0] if len(winners) == 1 else None

def _kec_store_name_from_memo(raw_value: str) -> str:
    raw = str(raw_value or "").strip()
    if not raw:
        return ""

    hits = list(re.finditer(r"(?i)\bCN\s*-\s*", raw))
    if hits:
        raw = raw[hits[-1].end():].strip()

    raw = re.sub(r"^\s*\d{6,8}\s+", "", raw).strip()
    raw = re.split(
        r"(?i)\s+(?:REPLEN\b|REP(?=\d{6,8}\b))",
        raw,
        maxsplit=1,
    )[0]
    return re.sub(r"\s+", " ", raw).strip(" -_/")

def _kec_or_row_kind(row) -> int:
    memo = str(getattr(row, "store_raw", "") or "").upper()
    if "EXTRA STRAP" in memo:
        return 2
    if "MATERIAL" in memo:
        return 1
    return 0

# === KEC_SAFE_STORE_RESOLVER_V4 END ===



def _match_store_and_or_flat_no_store(all_rows) -> Optional["StoreOrMatchResult"]:
    """v18 (SG-533-TEST fix): fallback used ONLY from match_store_and_or()
    when identity_to_rows is globally empty (see the call site's long
    comment for the full rationale). Restricted to rows whose
    detection_source is POSITIONAL_FALLBACK -- the one OR-List header tier
    where "store_raw" is structurally ambiguous (spec section 5: "only
    assumption -- first column = Store", which is routinely NOT a real
    Store name at all, e.g. a bare "OR | SO" sheet). A LITERAL_HEADER or
    SEMANTIC_FALLBACK_DUPLICATE_OR_HEADER sheet has an intentionally-
    identified Store column -- if ITS values don't resolve to a known CN
    Store alias, that is a genuine "no match" and must stay REVIEW exactly
    as before, never be reinterpreted as an OR value here.

    Returns None when this fallback does not apply at all (caller keeps
    its existing REVIEW result), or a fully-populated StoreOrMatchResult
    (status OK with matched_store="" -- deliberately never a China Store,
    so this can never grant CN Port/Store eligibility -- or status REVIEW
    with an explanatory reason) when it does."""
    candidate_rows = [r for r in all_rows if getattr(r, "detection_source", "") == "POSITIONAL_FALLBACK"]
    if not candidate_rows:
        return None

    def _flat_record(row) -> "OrderedDict[str, str]":
        rec: "OrderedDict[str, str]" = OrderedDict()
        label = getattr(row, "store_header", "") or "OR"
        rec[label] = str(getattr(row, "store_raw", "") or "").strip()
        for k, v in (getattr(row, "business_fields", {}) or {}).items():
            rec[k] = str(v or "").strip()
        return rec

    records = [_flat_record(r) for r in candidate_rows]
    fingerprints = {tuple(_norm_text(v) for v in rec.values()) for rec in records}

    result = StoreOrMatchResult()
    if len(fingerprints) > 1:
        result.status = "REVIEW"
        result.review_reason = (
            "OR List has no Store/Shop column (bare positional shape) and its rows "
            f"resolve to {len(fingerprints)} different records -- cannot pick one "
            "without a Store or routing dimension to disambiguate."
        )
        return result

    matched_fields = records[0]
    result.matched_store = ""
    result.match_source = "NO_STORE_DIMENSION_FLAT_MATCH"
    result.matched_business_fields = matched_fields
    vals = list(matched_fields.values())
    result.matched_or = vals[0] if len(vals) >= 1 else ""
    result.matched_so = vals[1] if len(vals) >= 2 else ""
    result.status = "OK"
    return result


def match_store_and_or(pkg, or_index: Dict[str, list], receiver_text: str = "") -> StoreOrMatchResult:
    result = StoreOrMatchResult()
    if not or_index:
        result.status = "NO_OR_LIST"
        return result

    all_rows = [r for rows in or_index.values() for r in rows]
    if not all_rows:
        result.status = "NO_OR_LIST"
        return result

    identity_to_rows: Dict[str, list] = {}
    for row in all_rows:
        memo_store = _kec_store_name_from_memo(getattr(row, "store_raw", ""))
        identity = _kec_resolve_store_identity(memo_store)
        if identity:
            identity_to_rows.setdefault(identity, []).append(row)

    store_identity = None
    match_source = ""
    for signal, source in (
        (getattr(pkg, "shipping_mark", ""), "SHIPMARK_SAFE_ALIAS"),
        (getattr(pkg, "reference_code", ""), "FILENAME_SAFE_ALIAS"),
        (receiver_text, "RECEIVER_SAFE_ALIAS"),
    ):
        identity = _kec_resolve_store_identity(signal)
        if identity and identity in identity_to_rows:
            store_identity = identity
            match_source = source
            break

    if not store_identity:
        # v18 (SG-533-TEST real-fixture fix -- spec sections 3/5/6/7 of the
        # non-CN regression report): identity_to_rows is built ONLY from
        # rows whose store_raw resolves to a KNOWN CN Store alias (China
        # World / Guangzhou / Hangzhou / IAPM / Kerry / Hongqiao / Taikooli
        # / Shenzhen / ...). A real production OR List can legitimately
        # have NO Store dimension at all -- e.g. a bare two-column "OR |
        # SO" sheet for a single non-CN shipment (confirmed against the
        # real SG-533-TEST/OR.xlsx: header ["OR","SO"], both data rows
        # identical). pl_or_list_import.py's positional fallback still
        # (by design, spec section 5: "first column = Store") stores that
        # OR value under store_raw/store_header -- it is NEVER a real CN
        # Store name, so it will never resolve via _kec_resolve_store_
        # identity() above, and identity_to_rows stays empty for the WHOLE
        # OR List (not just this one package).
        #
        # This is fundamentally different from "no match for THIS
        # package" (a normal CN OR List where this package's Shipmark/
        # filename/receiver just doesn't hit any of the real Store rows
        # present) -- that case must keep returning REVIEW exactly as
        # before. The distinguishing signal is identity_to_rows being
        # GLOBALLY empty: not one single row anywhere in the uploaded OR
        # List resolved to a real CN Store, so there is structurally no
        # Store dimension to match THIS or ANY OTHER package against.
        #
        # In that situation, per spec section 5 ("non-CN does NOT need a
        # fake China Store... but if OR List matches the destination/
        # routing key, the output still needs OR No./Ref No.") and section
        # 6 ("the matcher must support this without invoking China Store
        # matching"), fall back to a flat, Store-less match: pool every
        # row in the OR List, recover each row's true OR value (store_raw,
        # labelled with its own store_header -- e.g. "OR" -- so it is
        # never lost), and if the WHOLE pool reduces to exactly one
        # distinct business-field record, apply it to every package
        # regardless of destination Store/country (Store/Port stay
        # whatever classify_packages_for_port() already decided --
        # matched_store is deliberately left "" here, never a China Store,
        # so this can never grant CN Port/Store eligibility). If the pool
        # has more than one distinct record, there is no way to safely
        # pick one without a Store/routing dimension -- REVIEW, never a
        # guess.
        if not identity_to_rows:
            flat_result = _match_store_and_or_flat_no_store(all_rows)
            if flat_result is not None:
                return flat_result
        result.status = "REVIEW"
        result.review_reason = "No unique safe Store match from Shipping Mark / filename / receiver."
        return result

    store_rows = identity_to_rows.get(store_identity, [])
    if not store_rows:
        result.status = "REVIEW"
        result.review_reason = f"Store {store_identity!r} resolved but no OR-list rows were indexed."
        result.candidate_store = _KEC_STORE_DISPLAY.get(store_identity, store_identity)
        return result

    ordered_rows = sorted(
        store_rows,
        key=lambda r: (_kec_or_row_kind(r), getattr(r, "row_number", 0)),
    )
    base_row = ordered_rows[0]

    result.matched_store = _KEC_STORE_DISPLAY.get(store_identity, store_identity)
    result.match_source = match_source

    labels = []
    for r in ordered_rows:
        for label in (getattr(r, "business_fields", {}) or {}).keys():
            if label not in labels:
                labels.append(label)

    try:
        from collections import OrderedDict
        resolved_fields = OrderedDict()
    except Exception:
        resolved_fields = {}

    for label in labels:
        base_value = str(
            (getattr(base_row, "business_fields", {}) or {}).get(label, "") or ""
        ).strip()

        unique = []
        for r in ordered_rows:
            v = str(
                (getattr(r, "business_fields", {}) or {}).get(label, "") or ""
            ).strip()
            if v and v not in unique:
                unique.append(v)

        if len(unique) == 1:
            resolved_fields[label] = unique[0]
        elif base_value:
            resolved_fields[label] = base_value
        else:
            resolved_fields[label] = unique[0] if unique else ""

    result.matched_business_fields = resolved_fields
    vals = list(resolved_fields.values())

    result.matched_or = (
        vals[0] if len(vals) >= 1
        else str(getattr(base_row, "or_raw", "") or "").strip()
    )
    result.matched_so = (
        vals[1] if len(vals) >= 2
        else str(getattr(base_row, "so_raw", "") or "").strip()
    )
    result.status = "OK"
    return result






# =========================================================================
# 4) PDF receiver-text extraction (CN packages only)
# =========================================================================
_RECEIVER_LABEL_RE = re.compile(r"receiver\s*(?:company)?\s*[:\-]?\s*(.+)", re.IGNORECASE)
_STOP_LABEL_RE = re.compile(r"^(sender|shipper|consignee|invoice|date|packing|page|notify)\b", re.IGNORECASE)


def _extract_receiver_block(full_text: str) -> str:
    if not full_text:
        return ""
    lines = full_text.splitlines()
    for i, line in enumerate(lines):
        m = _RECEIVER_LABEL_RE.search(line)
        if m:
            block = [m.group(1).strip()]
            for j in range(i + 1, min(i + 4, len(lines))):
                nxt = lines[j].strip()
                if not nxt or _STOP_LABEL_RE.match(nxt):
                    break
                block.append(nxt)
            return " ".join(p for p in block if p)
    return ""


def _find_pdf_path(source_file: str, pdf_folder: Optional[Path], recursive: bool) -> Optional[Path]:
    if not pdf_folder or not source_file:
        return None
    pdf_folder = Path(pdf_folder)
    if not pdf_folder.exists():
        return None
    direct = pdf_folder / source_file
    if direct.exists():
        return direct
    pattern = "**/*" if recursive else "*"
    try:
        for p in pdf_folder.glob(pattern):
            if p.is_file() and p.name == source_file:
                return p
    except Exception as e:  # pragma: no cover
        log.warning(f"  glob error while looking for {source_file}: {e}")
    return None


def _extract_pdf_text_cached(pdf_path: Path, cache: Dict[str, str]) -> str:
    key = str(pdf_path)
    if key in cache:
        return cache[key]
    text = ""
    if pdfplumber is None:
        cache[key] = text
        return text
    try:
        with pdfplumber.open(str(pdf_path)) as pdf:
            pages_text = []
            for page in pdf.pages[:3]:  # receiver block is always on the cover page(s)
                pages_text.append(page.extract_text() or "")
            text = "\n".join(pages_text)
    except Exception as e:
        log.warning(f"  cannot read PDF for store detection: {pdf_path.name} ({e})")
        text = ""
    cache[key] = text
    return text


def _collect_cn_signal(pkg, pdf_folder, recursive, cache: Dict[str, str]) -> str:
    """Build the text blob used for store fuzzy-matching: PDF receiver block
    (if available) + reference_code + source_file, so matching still has a
    chance even when the PDF cannot be located."""
    parts = [pkg.reference_code or "", pkg.source_file or ""]
    pdf_path = _find_pdf_path(pkg.source_file, pdf_folder, recursive)
    if pdf_path:
        full_text = _extract_pdf_text_cached(pdf_path, cache)
        block = _extract_receiver_block(full_text)
        parts.append(block or full_text[:800])
    else:
        log.warning(f"  PDF not found for store detection: {pkg.source_file} (ref={pkg.reference_code})")
    return " ".join(p for p in parts if p)


# =========================================================================
# 5) Safe workbook writer (temp file + replace, never corrupts an open file)
# =========================================================================
def _safe_write(path: Path, write_fn: Callable[[Path], None]):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_suffix(path.suffix + ".tmp")
    try:
        write_fn(tmp_path)
    except PermissionError as e:
        raise PermissionError(
            f"Cannot create temp file for '{path.name}': {e}. "
            f"Close any program locking that folder and re-run."
        ) from e
    try:
        if path.exists():
            path.unlink()
        tmp_path.replace(path)
    except PermissionError as e:
        # cleanup the tmp file so re-runs don't pile up .tmp files
        try:
            tmp_path.unlink(missing_ok=True)
        except Exception:
            pass
        raise PermissionError(
            f"Cannot overwrite '{path}': the file appears to be open in Excel. "
            f"Close it and re-run the export."
        ) from e


def _write_group(path: Path, pkgs: List, write_workbook: Callable, renumber: bool,
                  optional_business_field_labels=None):
    """Write one grouped workbook, optionally renumbering `global_carton_num`
    to be local to the group (1/N .. N/N), then ALWAYS restoring the original
    values afterwards — success or failure — so later groups are unaffected.

    optional_business_field_labels (v16): forwarded to write_workbook() so a
    grouped file's B/C headers match PL_Total's -- see export_grouped_pl()'s
    docstring for why this was previously silently dropped."""
    if not pkgs:
        return None
    saved = [p.global_carton_num for p in pkgs]
    try:
        if renumber:
            def _sort_key(p):
                m = re.match(r"^\s*(\d+)", p.global_carton_num or "")
                return int(m.group(1)) if m else 0
            ordered = sorted(pkgs, key=_sort_key)
            n = len(ordered)
            for i, p in enumerate(ordered, start=1):
                p.global_carton_num = f"{i}/{n}"
        _safe_write(path, lambda tmp: write_workbook(tmp, pkgs, optional_business_field_labels=optional_business_field_labels))
        log.info(f"  wrote {path.name}  ({len(pkgs)} cartons)")
    finally:
        for p, orig in zip(pkgs, saved):
            p.global_carton_num = orig
    return path


def _write_total(dir_total: Path, packages: List, write_workbook: Callable,
                  total_workbook: Optional[Path], optional_business_field_labels=None) -> Path:
    target = dir_total / "PL_TOTAL.xlsx"
    if total_workbook and Path(total_workbook).exists():
        _safe_write(target, lambda tmp: shutil.copyfile(str(total_workbook), str(tmp)))
    else:
        _safe_write(target, lambda tmp: write_workbook(tmp, packages, optional_business_field_labels=optional_business_field_labels))
    log.info(f"  wrote {target.name}  ({len(packages)} cartons)")
    return target


# =========================================================================
# 6) Control CSV
# =========================================================================
CONTROL_FIELDS = [
    "source_file", "reference_code", "package_code",
    "factory", "port", "store", "store_confidence", "suggested_store_if_review",
    "resolved_store_split", "carton_display", "global_carton_display",
]


def _write_control_csv(path: Path, rows: List[dict]):
    def _do(tmp):
        with open(tmp, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=CONTROL_FIELDS)
            w.writeheader()
            for r in rows:
                w.writerow({k: r.get(k, "") for k in CONTROL_FIELDS})
    _safe_write(path, _do)


# =========================================================================
# 7) Validation / reconciliation
# =========================================================================
def _read_match_status_rowcount(path: Path) -> Optional[int]:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        ws = wb["Match_Status"]
        n = max(ws.max_row - 1, 0)
        wb.close()
        return n
    except Exception as e:  # pragma: no cover
        log.warning(f"  could not read back {path.name} for validation: {e}")
        return None


def _validate(packages, classified, factory_groups, cn_by_port, by_store,
              written_paths: Dict[Path, int]) -> Tuple[bool, str]:
    lines: List[str] = []
    ok = True

    def check(label: str, cond: bool, detail: str = ""):
        nonlocal ok
        status = "OK" if cond else "FAIL"
        if not cond:
            ok = False
        lines.append(f"[{status}] {label}" + (f" — {detail}" if detail else ""))
        return cond

    total_cartons = len(packages)
    total_qty = sum(p.calc_qty for p in packages)
    lines.append(f"PL TOTAL: {total_cartons} cartons, {total_qty} qty")

    # -- no duplicate package_code across the whole shipment --
    codes = [p.package_code for p in packages]
    dup_codes = sorted({c for c in codes if codes.count(c) > 1})
    check("No duplicate package_code across full shipment", not dup_codes,
          f"duplicates={dup_codes}" if dup_codes else "")

    # -- factory totals (REVIEW bucket counted explicitly, nothing silently dropped) --
    factory_cartons = sum(len(v) for v in factory_groups.values())
    factory_qty = sum(sum(p.calc_qty for p in v) for v in factory_groups.values())
    check("SUM(factory groups incl. REVIEW) cartons == PL_TOTAL cartons",
          factory_cartons == total_cartons, f"sum={factory_cartons} total={total_cartons}")
    check("SUM(factory groups incl. REVIEW) quantity == PL_TOTAL quantity",
          factory_qty == total_qty, f"sum={factory_qty} total={total_qty}")

    review_factory = factory_groups.get("REVIEW", [])
    if review_factory:
        lines.append(f"[WARN] {len(review_factory)} package(s) could not be classified to a factory (REVIEW):")
        for p in review_factory:
            lines.append(f"    REVIEW-FACTORY  source_file={p.source_file}  reference_code={p.reference_code}  package_code={p.package_code}")

    # -- CN port totals (v16 bug fix, corrected: basis is every DESTINATION-
    # CN carton, regardless of factory/origin suffix -- see
    # is_cn_port_eligible() -- never "factory=='CN'" alone, and never
    # restricted to just the CN/VN pair either (an earlier draft did that
    # and was still wrong: POP/SBGEAR/QIFENG/JION cartons resolve Store/
    # Port from the Shipping Mark BODY exactly like CN/VN do). The old
    # basis (factory_groups.get("CN")) silently excluded every VN-suffix
    # carton from its own expected count, so a VN carton losing its PORT
    # never tripped this check -- see PHASE-1 audit / CN-6557 real-file
    # reconciliation for the concrete regression this fixes: PEK/PVG/SZX
    # previously validated against 5/25/9 CN-suffix-only cartons instead
    # of the true 6/31/11 totals.) --
    cn_port_eligible = [c for c in classified if c["cn_port_eligible"]]
    cn_cartons = len(cn_port_eligible)
    cn_review = [c for c in cn_port_eligible if not c["store"] or c["store"] == "REVIEW"]
    expected_cn_classified = cn_cartons - len(cn_review)

    port_cartons = sum(len(v) for v in cn_by_port.values())

    if cn_port_eligible:
        check("SUM(CN port groups) cartons == all destination-CN-eligible cartons minus REVIEW",
              port_cartons == expected_cn_classified,
              f"port_sum={port_cartons} expected={expected_cn_classified} (CN_eligible_total={cn_cartons}, review={len(cn_review)})")

    # -- cross-output consistency (spec: "a package cannot be PORT blank in
    # PL_Total while simultaneously present in a PEK/PVG/SZX grouped
    # output" and the reverse -- a resolved Store must never leave PORT
    # blank) -- both pkg.port and c["port"] here are the SAME canonical
    # value PL_Total's Packing List sheet already wrote, so this also
    # guards PL_Total/Raw_Data/PL_SPLIT_CONTROL/03_CN_BY_PORT from ever
    # silently disagreeing with each other again. --
    resolved_store_blank_port = [c for c in cn_port_eligible
                                  if c["store"] and c["store"] != "REVIEW" and not c["port"]]
    check("No destination-CN package has a resolved Store but a blank PORT",
          not resolved_store_blank_port,
          f"packages={[c['pkg'].package_code for c in resolved_store_blank_port]}" if resolved_store_blank_port else "")

    # -- BY STORE totals (v14: cross-factory -- a store's POP/VN/CN cartons
    # all count here, so this is NOT compared against the CN-only port sum
    # any more; the only invariants that still hold are "no duplicates" and
    # "never more than PL_TOTAL") --
    store_cartons = sum(len(v) for v in by_store.values())
    check("SUM(store groups) cartons <= PL_TOTAL cartons (unresolved-store packages excluded on purpose)",
          store_cartons <= total_cartons, f"store_sum={store_cartons} total={total_cartons}")
    unresolved_store = total_cartons - store_cartons
    if unresolved_store:
        lines.append(f"[INFO] {unresolved_store} package(s) excluded from 04_CN_BY_STORE "
                      f"(no Store resolved -- OR List didn't match and/or not a CN-classified package).")

    if cn_review:
        lines.append(f"[WARN] {len(cn_review)} CN package(s) could not be confidently mapped to a store (REVIEW):")
        for c in cn_review:
            p = c["pkg"]
            lines.append(f"    REVIEW-STORE  source_file={p.source_file}  reference_code={p.reference_code}  "
                          f"package_code={p.package_code}  confidence={c['confidence']}  suggested={c['suggestion']}")

    # -- no duplicate package_code within any single exported group --
    def _dup_within(name: str, groups: Dict[str, List]):
        for key, plist in groups.items():
            pcodes = [p.package_code for p in plist]
            dups = sorted({c for c in pcodes if pcodes.count(c) > 1})
            check(f"No duplicate package_code within {name}={key}", not dups,
                  f"duplicates={dups}" if dups else "")

    _dup_within("factory", factory_groups)
    _dup_within("cn_port", cn_by_port)
    _dup_within("store", by_store)

    # -- no package lost: every package_code appears in exactly one factory bucket --
    all_grouped_codes = [p.package_code for v in factory_groups.values() for p in v]
    check("No package lost between PL_TOTAL and factory groups",
          sorted(all_grouped_codes) == sorted(codes),
          f"grouped_count={len(all_grouped_codes)} total_count={len(codes)}")

    # -- cross-check the actual files written to disk (catches writer bugs) --
    for path, expected_n in written_paths.items():
        n = _read_match_status_rowcount(path)
        if n is None:
            lines.append(f"[WARN] Could not verify {path.name} on disk (read-back failed)")
            continue
        check(f"{path.name}: Match_Status row count == expected cartons",
              n == expected_n, f"on_disk={n} expected={expected_n}")

    return ok, "\n".join(lines)


# =========================================================================
# 8) Main entry point
# =========================================================================
def export_grouped_pl(
    packages: List,
    output_dir,
    write_workbook: Callable,
    total_workbook=None,
    pdf_folder=None,
    recursive: bool = False,
    store_threshold: float = 0.55,
    store_margin: float = 0.08,
    optional_business_field_labels=None,
) -> Path:
    """Split `packages` (as produced by run_pipeline) into the grouped
    PL_SPLIT_OUTPUT folder tree and return the path to PL_SPLIT_CONTROL.csv.

    v15: store_threshold/store_margin are kept for call-signature backward
    compatibility only -- they are no longer consulted here. Store/Port are
    now read from pkg.store/pkg.port, resolved exactly once by
    pl_ocr_core.classify_packages_for_port() (which is where store_threshold/
    store_margin's real match_store() call now lives) -- see
    is_cn_port_eligible()'s docstring for why a second, independent
    match_store() call here was the root cause of the PORT/03_CN_BY_PORT/
    PL_SPLIT_VALIDATION inconsistency this version fixes.

    optional_business_field_labels (v16, point-3 correction): the SAME canonicalized
    OR List business-field labels (e.g. ["OR No.", "Ref No."]) PL_Total's
    Packing List sheet used -- threaded through to every grouped workbook
    this function writes (02_BY_FACTORY, 03_CN_BY_PORT, 04_CN_BY_STORE,
    01_PL_TOTAL) so a grouped Packing List's B/C headers can never disagree
    with PL_Total's. Before this fix, export_grouped_pl() called
    write_workbook(tmp, pkgs) with no optional_business_field_labels argument at
    all, so every grouped file silently fell back to the "OR No."/"SO No."
    default regardless of what the OR List actually said -- caught during
    review of the point-3 dynamic-business-field test.

    Raises RuntimeError if reconciliation fails after writing everything —
    never just prints "Completed" while data is actually missing/duplicated.
    """
    if not packages:
        raise ValueError(
            "export_grouped_pl: `packages` is empty. Run the OCR pipeline "
            "(run_pipeline(...)) first and pass its return value here."
        )

    output_dir = Path(output_dir)
    dir_total = output_dir / "01_PL_TOTAL"
    dir_factory = output_dir / "02_BY_FACTORY"
    dir_cn_port = output_dir / "03_CN_BY_PORT"
    dir_cn_store = output_dir / "04_CN_BY_STORE"
    for d in (output_dir, dir_total, dir_factory, dir_cn_port, dir_cn_store):
        d.mkdir(parents=True, exist_ok=True)

    # ---- classify every package exactly once ----
    # v15 (bug fix -- CN-6557 PORT regression): Store/Port are CONSUMED from
    # pkg.store/pkg.port -- the ONE canonical resolution already computed by
    # pl_ocr_core.classify_packages_for_port() (the same call that fills
    # PL_Total's PORT column) -- NEVER re-detected here with a separate
    # match_store() call. Two independent classifications is exactly what
    # produced the original bug (PL_Total PORT blank for VN-suffix cartons
    # while 03_CN_BY_PORT/PL_SPLIT_VALIDATION used a different, factory==CN-
    # only definition) -- see is_cn_port_eligible()'s docstring. `factory`
    # itself is still detected locally (a different, legitimate concern --
    # 02_BY_FACTORY grouping and carton ordering -- untouched by this fix).
    receiver_cache: Dict[str, str] = {}
    classified: List[dict] = []
    control_rows: List[dict] = []

    for pkg in packages:
        factory = detect_factory(pkg.reference_code, pkg.source_file)
        eligible = is_cn_port_eligible(pkg, factory)
        store = getattr(pkg, "store", "") or ""
        port = getattr(pkg, "port", "") or ""
        confidence = getattr(pkg, "store_confidence", "") if eligible else ""
        suggestion = getattr(pkg, "store_suggestion", "") if eligible else ""

        classified.append({"pkg": pkg, "factory": factory, "cn_port_eligible": eligible,
                            "store": store, "port": port, "confidence": confidence,
                            "suggestion": suggestion})
        control_rows.append({
            "source_file": pkg.source_file,
            "reference_code": pkg.reference_code,
            "package_code": pkg.package_code,
            "factory": factory,
            "cn_port_eligible": eligible,
            "port": port,
            "store": store,
            "store_confidence": confidence,
            "suggested_store_if_review": suggestion,
            # v14 diagnostics (spec section 14): the ACTUAL cross-factory
            # Store identity used for 04_CN_BY_STORE grouping (OR List
            # match first, else this same canonical `store` above) -- shown
            # separately from `store` because the two can legitimately
            # differ (e.g. a POP-factory package has no CN-port-eligible
            # `store` at all, but resolves here via the OR List).
            "resolved_store_split": _resolved_store_for_split(pkg),
            "carton_display": getattr(pkg, "carton_display", "") or getattr(pkg, "global_carton_num", ""),
            "global_carton_display": getattr(pkg, "global_carton_display", ""),
        })

    written_paths: Dict[Path, int] = {}

    # ---- 1) TOTAL ----
    log.info("Writing 01_PL_TOTAL ...")
    total_path = _write_total(dir_total, packages, write_workbook, total_workbook,
                               optional_business_field_labels=optional_business_field_labels)
    written_paths[total_path] = len(packages)

    # ---- 2) BY FACTORY ----
    log.info("Writing 02_BY_FACTORY ...")
    factory_groups: Dict[str, List] = defaultdict(list)
    for c in classified:
        factory_groups[c["factory"]].append(c["pkg"])

    for factory, pkgs in factory_groups.items():
        fname = FACTORY_FILE_MAP.get(factory)
        if not fname or not pkgs:
            if factory == "REVIEW":
                log.warning(f"  {len(pkgs)} package(s) left unclassified (factory=REVIEW) — see control CSV")
            continue
        p = _write_group(dir_factory / fname, pkgs, write_workbook, renumber=True,
                          optional_business_field_labels=optional_business_field_labels)
        if p:
            written_paths[p] = len(pkgs)

    # ---- 3) CN BY PORT (v16 bug fix, corrected: destination-CN-eligible --
    # ALL destination-CN cartons regardless of factory/origin suffix, see
    # is_cn_port_eligible(); membership + PORT value both come straight
    # off `classified`, which itself only ever reads the single canonical
    # pkg.port/pkg.store -- never re-detected here) ----
    log.info("Writing 03_CN_BY_PORT ...")
    cn_by_port: Dict[str, List] = defaultdict(list)
    for c in classified:
        if not c["cn_port_eligible"]:
            continue
        if not c["store"] or c["store"] == "REVIEW":
            continue  # excluded on purpose — never silently guess
        cn_by_port[c["port"]].append(c["pkg"])

    for port, pkgs in cn_by_port.items():
        fname = PORT_FILE_MAP.get(port)
        if fname and pkgs:
            p = _write_group(dir_cn_port / fname, pkgs, write_workbook, renumber=True,
                              optional_business_field_labels=optional_business_field_labels)
            if p:
                written_paths[p] = len(pkgs)

    # ---- 4) BY STORE (v14 rework, spec sections 9-12) ----
    # Cross-factory, NOT CN-factory-only any more: a Store's POP + SBGEAR/
    # QIFENG/JION + CN cartons must all land in the SAME per-store file,
    # sharing ONE denominator -- exactly what pl_ocr_core.py's v12/v13/v14
    # counting_scope_key / assign_global_numbers() already computed for
    # every package, BEFORE export_grouped_pl() was ever called (see
    # run_pipeline()). So this step does NOT re-classify from scratch with
    # the old CN-only match_store() and does NOT locally renumber -- it only
    # GROUPS by the already-resolved Store identity and writes each group's
    # packages with their EXISTING carton_display/global_carton_num (Store-
    # scoped numbering) untouched (renumber=False). See
    # _resolved_store_for_split() for the exact resolution priority.
    log.info("Writing 04_CN_BY_STORE ...")
    by_store: Dict[str, List] = defaultdict(list)
    store_display_name: Dict[str, str] = {}
    for pkg in packages:
        store_key = _resolved_store_for_split(pkg)
        if not store_key:
            continue  # excluded on purpose — never silently guess
        by_store[store_key].append(pkg)
        store_display_name.setdefault(store_key, _store_display_label(pkg, store_key))

    for store_key, pkgs in by_store.items():
        file_key = _store_file_key(pkgs[0], store_key)
        fname = STORE_FILE_MAP.get(file_key) or _dynamic_store_filename(store_display_name.get(store_key, store_key))
        if pkgs:
            p = _write_group(dir_cn_store / fname, pkgs, write_workbook, renumber=False,
                          optional_business_field_labels=optional_business_field_labels)
            if p:
                written_paths[p] = len(pkgs)

    # ---- safety net: restore original carton numbers on every package ----
    # (each _write_group call already restores in its own finally-block; this
    # loop is a defensive no-op unless an exception skipped that restore.)

    # ---- control CSV ----
    control_path = output_dir / "PL_SPLIT_CONTROL.csv"
    _write_control_csv(control_path, control_rows)
    log.info(f"  wrote {control_path.name}  ({len(control_rows)} rows)")

    # ---- validation / reconciliation ----
    ok, report_text = _validate(packages, classified, factory_groups, cn_by_port, by_store, written_paths)
    report_path = output_dir / "PL_SPLIT_VALIDATION.txt"
    report_path.write_text(report_text, encoding="utf-8")
    print("\n" + "=" * 70)
    print("PL SPLIT VALIDATION REPORT")
    print("=" * 70)
    print(report_text)
    print("=" * 70)

    if not ok:
        raise RuntimeError(
            "PL split reconciliation FAILED — see PL_SPLIT_VALIDATION.txt "
            f"({report_path}) for the full list of mismatches."
        )

    log.info("Reconciliation PASSED.")
    return control_path
