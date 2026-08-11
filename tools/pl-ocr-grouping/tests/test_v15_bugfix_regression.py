#!/usr/bin/env python3
"""
tests/test_v15_bugfix_regression.py
====================================
Permanent regression tests for the v15 CN-6557 production bug-fix pass:

  BUG1/2 -- CN Port detection wrongly gated on factory=="CN" instead of
            destination country, silently dropping PORT for every VN-
            suffix carton and letting 03_CN_BY_PORT / PL_SPLIT_VALIDATION
            use a different (CN-factory-only) definition than PL_Total.
  BUG3   -- OR List "Ref#" business field mislabeled "SO No." on both the
            Packing List sheet and the Sublist PDF metadata block.
  BUG4   -- 4 confirmed material SKU<->EAN pairs missing from the known-
            material enrichment mapping.

No test-framework dependency, by design -- matches this repo's existing
convention (see test_pl_ocr_core.py / test_pl_group_export.py). Run with:

    python3 tools/pl-ocr-grouping/tests/test_v15_bugfix_regression.py

NOTE ON DATA: only literal Shipping Mark / SKU / EAN TEXT STRINGS from the
real CN-6557 production ticket are used below (no customer PDFs, DIM
files, addresses, or phone numbers) -- see the 2026-08-03 incident note in
test_pl_ocr_core.py for why real customer files must never be committed
to this repo.
"""
from __future__ import annotations
import sys, types, tempfile, shutil, csv
from pathlib import Path
from collections import Counter

HERE = Path(__file__).resolve().parent
TOOL_DIR = HERE.parent
CORE_PY = TOOL_DIR / "pl_ocr_core.py"
if str(TOOL_DIR) not in sys.path:
    sys.path.insert(0, str(TOOL_DIR))

import pl_group_export as pge
import pl_or_list_import as oli

_ENTRY_MARKER = "# ── Entry point ────────────────────────────────────────────────────────────"


def _substitute_placeholders(src: str) -> str:
    def lit(v):
        return "None" if v is None else repr(v)
    return (src
            .replace("__DIM_WEIGHT_SHEET__", lit(None))
            .replace("__MASTER_DATA_SHEET__", lit(None))
            .replace("__RECURSIVE__", "False")
            .replace("__MANUAL_CONSIGNEE__", lit(None))
            .replace("__MANUAL_NOTIFY_PARTY__", lit(None))
            .replace("__GENERATE_SUBLIST__", "True")
            .replace("__GENERATE_SUBLIST_PDF__", "True")
            .replace("__OR_LIST_FILE__", "None")
            .replace("__GIT_COMMIT__", lit("test-v15")))


_module_counter = 0


def _exec_module(src: str):
    global _module_counter
    _module_counter += 1
    modname = f"pl_ocr_core_v15test_{_module_counter}"
    mod = types.ModuleType(modname)
    mod.__file__ = str(CORE_PY)
    sys.modules[modname] = mod
    exec(compile(src, str(CORE_PY), "exec"), mod.__dict__)
    return mod.__dict__


def load_core_defs():
    """Load every class/function/constant in pl_ocr_core.py WITHOUT running
    the pipeline -- same technique as test_pl_ocr_core.py's own loader."""
    src = CORE_PY.read_text(encoding="utf-8")
    idx = src.index(_ENTRY_MARKER)
    src = _substitute_placeholders(src[:idx])
    return _exec_module(src)


D = load_core_defs()
Package = D["Package"]
classify_packages_for_port = D["classify_packages_for_port"]
write_workbook = D["write_workbook"]
enrich_known_material_ean = D["enrich_known_material_ean"]
KNOWN_MATERIAL_EAN = D["KNOWN_MATERIAL_EAN"]
canonicalize_business_field_label = D["canonicalize_business_field_label"]
Item = D["Item"]
# v20 (SG-533-TEST final cleanup, requirement 1): bind the REAL production
# GTIN validator directly -- the whole point of this requirement is that
# checksum behavior must be proven against THIS function, not a duplicate
# written inside the test file.
is_valid_gtin13 = D["is_valid_gtin13"]

# ── tiny test runner (mirrors every other test_*.py in this directory) ─────
_passed = 0
_failed = 0


def test(name, fn):
    global _passed, _failed
    try:
        fn()
        _passed += 1
        print(f"  ok  - {name}")
    except AssertionError as e:
        _failed += 1
        print(f"FAIL  - {name}\n        {e}")
    except Exception as e:
        _failed += 1
        print(f"ERROR - {name}\n        {type(e).__name__}: {e}")


def _pkg(mark: str, seq: int) -> "Package":
    p = Package(package_code=f"PKG{seq:03d}", source_file=f"{mark}.pdf",
                reference_code=mark, pdf_package_seq=0)
    p.shipping_mark = mark
    p.country = "CN"
    return p


# =========================================================================
# BUG1/2 -- PORT must resolve from the Shipping Mark BODY (Store), never
# from the final factory/origin suffix (VN vs CN) or from the destination
# country being CN itself.
# =========================================================================
print("== BUG1/2: PORT resolution (CN-6557 real Shipping Marks) ==")


def t_port_real_shipmark_pairs_vn_and_cn_suffix():
    cases = [
        ("CN-6557-CNWORLD_PEK-VN", "PEK"),
        ("CN-6557_CNWORLD_PEK_CN", "PEK"),
        ("CN-6557-GUANGZHOU_SZX-VN", "SZX"),
        ("CN-6557-GUANGZHOU_SZX_CN", "SZX"),
        ("CN-6557-HANGZHOU_PVG-VN", "PVG"),
        ("CN-6557-HANGZHOU_PVG_CN", "PVG"),
    ]
    for i, (mark, expected_port) in enumerate(cases):
        pkg = _pkg(mark, i)
        classify_packages_for_port([pkg], None, False)
        assert pkg.port == expected_port, (
            f"{mark}: expected port={expected_port!r}, got port={pkg.port!r} store={pkg.store!r}"
        )


def t_port_never_blank_solely_because_suffix_is_vn():
    """The exact regression: a VN-suffix carton must never lose its PORT
    just because factory-suffix != 'CN'."""
    pkg = _pkg("CN-6557-Kerry_PVG-VN", 0)
    classify_packages_for_port([pkg], None, False)
    assert pkg.port == "PVG", f"VN-suffix PORT wrongly blank/wrong: {pkg.port!r}"
    assert pkg.store == "KERRY"


def t_port_separator_variants_produce_the_same_port():
    """"-" / "_" / a leading "_" before the store token, mixed with either
    factory suffix -- all four must resolve to the identical PORT."""
    variants = [
        "CN-6557-CNWORLD-PEK-VN",
        "CN-6557-CNWORLD_PEK_VN",
        "CN-6557_CNWORLD_PEK-VN",
        "CN-6557_CNWORLD_PEK_CN",
    ]
    ports = set()
    for i, mark in enumerate(variants):
        pkg = _pkg(mark, i)
        classify_packages_for_port([pkg], None, False)
        assert pkg.port, f"{mark}: PORT unexpectedly blank"
        ports.add(pkg.port)
    assert ports == {"PEK"}, f"separator variants disagree on PORT: {ports}"


def t_port_non_cn_destination_never_runs_resolver():
    """A non-CN destination (KR) must never run the CN Store/Port resolver
    even if its trailing suffix happens to be a CN-network factory token
    (spec: 'KR-1000-POP -> destination country KR -> NOT Kerry')."""
    pkg = Package(package_code="P1", source_file="x.pdf", reference_code="KR-1000-POP", pdf_package_seq=0)
    pkg.shipping_mark = "KR-1000-POP"
    pkg.country = "KR"
    classify_packages_for_port([pkg], None, False)
    assert pkg.port == "" and pkg.store == "", (pkg.port, pkg.store)


def t_port_resolves_from_body_for_every_factory_suffix_not_just_cn_vn():
    """CORRECTED (post-review): an earlier draft of this fix wrongly
    restricted PORT/Store eligibility to factory in {"CN", "VN"} -- that
    was just as wrong as the original "factory=='CN'" bug. Store/Port
    belong to the Shipping Mark BODY and are COMPLETELY INDEPENDENT of the
    trailing factory/origin suffix: CN-1529_HANGZHOU_PVG_POP/_SBGEAR/
    _QIFENG/_JION/_VN/_CN must ALL resolve to the identical HANGZHOU/PVG.
    (Uses the full "HANGZHOU" spelling, not the "HZ" short code, because
    match_store() -- the signal-only matcher classify_packages_for_port()
    uses when no PDF/OR List is available -- only resolves 2-letter codes
    confidently via a PDF receiver block or an OR List's explicit
    shipping_mark_tokens; "HZ" alone is a separate, pre-existing
    match_store_and_or-only capability, not a regression from this fix.)"""
    for suffix in ("POP", "SBGEAR", "QIFENG", "JION", "VN", "CN"):
        pkg = _pkg(f"CN-1529_HANGZHOU_PVG_{suffix}", 0)
        classify_packages_for_port([pkg], None, False)
        assert pkg.store == "HANGZHOU", f"suffix={suffix}: store={pkg.store!r}"
        assert pkg.port == "PVG", f"suffix={suffix}: port={pkg.port!r}"


# =========================================================================
# BUG1/2 -- real CN-6557 production counts: PEK=6, PVG=31, SZX=11, TOTAL=48
# (8 stores, VN+CN cartons per store, matching the real PL_SPLIT_CONTROL.csv
# breakdown from the production fixture) -- zero carton loss, zero
# duplicates across port groups, and full cross-output consistency between
# pkg.port, PL_SPLIT_CONTROL's port column, and 03_CN_BY_PORT membership.
# =========================================================================
print("== BUG1/2: CN-6557 real production counts (PEK=6 / PVG=31 / SZX=11 / TOTAL=48) ==")

# (store token, port, n_cn_suffix, n_vn_suffix) -- exactly the real CN-6557
# breakdown from PL_SPLIT_CONTROL.csv / the production ticket's own numbers.
_CN6557_STORES = [
    ("CNWORLD", "PEK", 5, 1),
    ("GUANGZHOU", "SZX", 4, 1),
    ("HANGZHOU", "PVG", 4, 1),
    ("IAPM", "PVG", 5, 1),
    ("KERRY", "PVG", 5, 2),
    ("SHANGHAI_HONGQIAO", "PVG", 7, 1),
    ("SHANGHAI_TAIKOOLI", "PVG", 4, 1),
    ("SHENZHEN", "SZX", 5, 1),
]


def _build_cn6557_packages():
    packages = []
    seq = 0
    for store_token, port, n_cn, n_vn in _CN6557_STORES:
        for _ in range(n_cn):
            seq += 1
            packages.append(_pkg(f"CN-6557-{store_token}_{port}_CN", seq))
        for _ in range(n_vn):
            seq += 1
            packages.append(_pkg(f"CN-6557-{store_token}_{port}-VN", seq))
    return packages


def t_cn6557_production_port_counts_and_vn_distribution():
    packages = _build_cn6557_packages()
    assert len(packages) == 48, len(packages)
    classify_packages_for_port(packages, None, False)

    port_counts = Counter(p.port for p in packages)
    assert port_counts == {"PEK": 6, "PVG": 31, "SZX": 11}, dict(port_counts)

    blank_port = [p for p in packages if not p.port]
    assert not blank_port, f"{len(blank_port)} package(s) have a blank PORT: " \
        f"{[p.reference_code for p in blank_port]}"

    vn_packages = [p for p in packages if p.reference_code.upper().endswith("VN")]
    assert len(vn_packages) == 9, len(vn_packages)
    vn_port_counts = Counter(p.port for p in vn_packages)
    assert vn_port_counts == {"PEK": 1, "PVG": 6, "SZX": 2}, dict(vn_port_counts)


def t_cn6557_export_grouped_pl_zero_loss_zero_duplicates_cross_output_consistent():
    packages = _build_cn6557_packages()
    classify_packages_for_port(packages, None, False)

    out_dir = Path(tempfile.mkdtemp(prefix="pl_v15_cn6557_"))
    try:
        control_path = pge.export_grouped_pl(
            packages=packages, output_dir=out_dir, write_workbook=write_workbook,
        )
        with open(control_path, newline="", encoding="utf-8") as f:
            rows = list(csv.DictReader(f))
        assert len(rows) == 48, len(rows)

        # cross-output consistency: PL_SPLIT_CONTROL's port column must
        # exactly match pkg.port (the same canonical value PL_Total used).
        by_code = {p.package_code: p for p in packages}
        for row in rows:
            pkg = by_code[row["package_code"]]
            assert row["port"] == pkg.port, (row["package_code"], row["port"], pkg.port)

        ctrl_port_counts = Counter(r["port"] for r in rows if r["port"])
        assert ctrl_port_counts == {"PEK": 6, "PVG": 31, "SZX": 11}, dict(ctrl_port_counts)

        # 03_CN_BY_PORT membership: zero carton loss, zero duplicates.
        import openpyxl
        seen_codes = []
        for fname, expected_n in (("PL_CN_PORT_PEK.xlsx", 6),
                                   ("PL_CN_PORT_PVG.xlsx", 31),
                                   ("PL_CN_PORT_SZX.xlsx", 11)):
            fp = out_dir / "03_CN_BY_PORT" / fname
            assert fp.exists(), f"{fname} was not written"
            wb = openpyxl.load_workbook(str(fp), read_only=True, data_only=True)
            ws = wb["Match_Status"]
            header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            code_col = header.index("package_code")
            codes = [row[code_col] for row in ws.iter_rows(min_row=2, values_only=True) if row[code_col]]
            wb.close()
            assert len(codes) == expected_n, f"{fname}: {len(codes)} != {expected_n}"
            seen_codes.extend(codes)
        assert len(seen_codes) == len(set(seen_codes)) == 48, (
            f"carton loss/duplication across port groups: {len(seen_codes)} rows, "
            f"{len(set(seen_codes))} unique"
        )
    finally:
        shutil.rmtree(out_dir, ignore_errors=True)


for _name, _fn in [
    ("PORT: real pairs (CNWORLD/GUANGZHOU/HANGZHOU, VN and CN suffix) resolve to PEK/SZX/PVG", t_port_real_shipmark_pairs_vn_and_cn_suffix),
    ("PORT: never blank solely because factory suffix is VN (KERRY_PVG-VN)", t_port_never_blank_solely_because_suffix_is_vn),
    ("PORT: '-'/'_' separator variants (incl. leading '_') all resolve to the same PORT", t_port_separator_variants_produce_the_same_port),
    ("PORT: non-CN destination (KR) never runs the CN Store/Port resolver", t_port_non_cn_destination_never_runs_resolver),
    ("PORT: resolves from Shipping Mark BODY for EVERY factory suffix (POP/SBGEAR/QIFENG/JION/VN/CN), not just CN/VN", t_port_resolves_from_body_for_every_factory_suffix_not_just_cn_vn),
    ("CN-6557: PEK=6/PVG=31/SZX=11/TOTAL=48, VN distribution PEK=1/PVG=6/SZX=2, zero blank PORT", t_cn6557_production_port_counts_and_vn_distribution),
    ("CN-6557: export_grouped_pl zero carton loss/duplication, PL_SPLIT_CONTROL port == pkg.port", t_cn6557_export_grouped_pl_zero_loss_zero_duplicates_cross_output_consistent),
]:
    test(_name, _fn)


# =========================================================================
# BUG5 (v18, SG-533-TEST real-fixture regression report) -- Port/Store must
# NEVER populate for a non-CN destination, regardless of factory/origin
# suffix, and this must be proven by driving REAL country values (SG/JP/
# KR/US) through classify_packages_for_port() -- never only by a helper
# that silently hardcodes pkg.country = "CN" (see _pkg() above, which
# does exactly that and must stay untouched for the CN-only tests that
# rely on it). Paired against CN control cases so the same test file
# proves both halves of the ONE invariant: destination CN -> Port/Store
# resolver runs; destination != CN -> it never does, independent of
# suffix. See the real SG-533-TEST/PKL/SG-553_CN.pdf + SG-553_VN.pdf
# archive reconciliation for the end-to-end proof against production
# files (this section covers the permanent unit-level regression net).
# =========================================================================
print("== BUG5: non-CN destination never gets a China Port/Store, for ANY suffix ==")


def _pkg_with_country(mark: str, country: str, seq: int = 0) -> "Package":
    """Unlike _pkg() (which hardcodes country='CN' -- correct for the
    existing CN-only tests above), this constructs a Package with an
    EXPLICIT, caller-chosen destination country, so non-CN tests drive the
    real production code path (classify_packages_for_port ->
    is_cn_port_eligible) with a genuine country value instead of relying
    on a helper that would silently make everything CN."""
    p = Package(package_code=f"PKGX{seq:03d}", source_file=f"{mark}.pdf",
                reference_code=mark, pdf_package_seq=0)
    p.shipping_mark = mark
    p.country = country
    return p


def t_non_cn_destination_port_and_store_blank_for_every_suffix():
    """spec (SG-533-TEST consolidation report) sections 2-4: destination !=
    CN -> PORT and Store MUST stay blank, independent of the trailing
    factory/origin suffix -- tested across the full required non-CN
    destination matrix (SG/JP/TW/KR/US/EU/BE/PH/TH, PLUS two markets never
    named in any earlier pass of this engagement -- AU/MY -- to prove the
    gate is genuinely generic and not a slightly-longer hardcoded list)
    and every suffix named in the report (CN/VN/POP/QF)."""
    countries = ("SG", "JP", "TW", "KR", "US", "EU", "BE", "PH", "TH", "AU", "MY")
    for country in countries:
        for suffix in ("CN", "VN", "POP", "QF"):
            mark = f"{country}-9001_HANGZHOU_PVG_{suffix}"
            pkg = _pkg_with_country(mark, country)
            classify_packages_for_port([pkg], None, False)
            assert pkg.port == "" and pkg.store == "" and pkg.store_display == "", (
                f"destination={country} suffix={suffix}: expected Port/Store both blank, "
                f"got port={pkg.port!r} store={pkg.store!r} store_display={pkg.store_display!r}"
            )
            assert not pge.is_cn_port_eligible(pkg), f"destination={country}: must never be CN-port-eligible"


def t_cn_destination_control_still_resolves_for_every_suffix():
    """Control half of the same invariant: destination CN must KEEP
    resolving PEK/PVG regardless of suffix (CN/VN/POP) -- proves BUG5's
    fix (or any future change near it) never accidentally makes ALL
    packages Port-blank, only non-CN ones."""
    cases = [
        ("CN-9002_CNWORLD_PEK_CN", "PEK"),
        ("CN-9002_CNWORLD_PEK_VN", "PEK"),
        ("CN-9002_HANGZHOU_PVG_POP", "PVG"),
    ]
    for mark, expected_port in cases:
        pkg = _pkg_with_country(mark, "CN")
        classify_packages_for_port([pkg], None, False)
        assert pkg.port == expected_port, (mark, pkg.port, expected_port)


def t_unresolved_country_never_leaks_a_fabricated_store_or_port_onto_real_outputs():
    """v19 (SG-533-TEST consolidation report, requirement 2 -- supersedes
    the v18 version of this test): the CN eligibility gate is now STRICT
    -- "" (unresolved/unknown destination) is EXCLUDED, exactly like every
    other non-CN country, not "eligible, don't exclude, try anyway" as a
    v18 draft of this fix had it. is_cn_port_eligible() returns False for
    "", so classify_packages_for_port() never even calls match_store() for
    an unresolved-country package -- store_confidence/store_suggestion
    stay at their untouched defaults too, not just port/store_display."""
    pkg = _pkg_with_country("SG-553_CN", "", seq=1)
    assert not pge.is_cn_port_eligible(pkg)
    classify_packages_for_port([pkg], None, False)
    assert pkg.port == "", pkg.port
    assert pkg.store == "", pkg.store
    assert pkg.store_display == "", pkg.store_display
    assert pkg.store_confidence == "", (
        "an unresolved-country package must never even be OFFERED to match_store() -- "
        f"store_confidence should stay at its untouched default, got {pkg.store_confidence!r}"
    )


for _name, _fn in [
    ("BUG5: non-CN destination (SG/JP/TW/KR/US/EU/BE/PH/TH/AU/MY) x suffix (CN/VN/POP/QF) -> Port and Store always blank, never CN-eligible", t_non_cn_destination_port_and_store_blank_for_every_suffix),
    ("BUG5 control: CN destination still resolves PEK/PVG for every suffix (CN/VN/POP)", t_cn_destination_control_still_resolves_for_every_suffix),
    ("BUG5: unresolved (blank) country is STRICTLY excluded, not offered to match_store() at all", t_unresolved_country_never_leaks_a_fabricated_store_or_port_onto_real_outputs),
]:
    test(_name, _fn)


# =========================================================================
# BUG6 (v18, SG-533-TEST real-fixture regression report) -- an OR List
# with NO Store/Shop column at all (a bare "OR | SO" 2-column sheet, the
# REAL shape of SG-533-TEST/OR.xlsx) must still propagate OR No./Ref No.
# to every package it applies to, WITHOUT ever inventing a China Store
# (spec section 5/6/7: "non-CN does NOT need a fake China Store... but if
# OR List matches the destination/routing key, the output still needs OR
# No./Ref No." and "the matcher must support this without invoking China
# Store matching"). Exercises the real match_store_and_or() function
# directly (not a hand-set Package.business_fields stand-in), same
# principle as BUG5 above.
# =========================================================================
print("== BUG6: OR List with no Store column still propagates OR/Ref (never invents a Store) ==")


def t_bare_or_so_two_column_or_list_propagates_unambiguously_without_a_store():
    """The real SG-533-TEST/OR.xlsx shape: header ['OR','SO'], every row
    resolving to the SAME OR/SO pair (duplicate rows are how the real file
    is shaped) -- unambiguous, so it must propagate OR No./Ref No. to a
    package with NO resolvable CN Store identity, while matched_store
    stays "" (never a China Store)."""
    import pl_or_list_import as oli
    import pl_group_export as pge
    import openpyxl, tempfile, shutil
    out_dir = Path(tempfile.mkdtemp(prefix="bug6_bare_or_so_"))
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["OR", "SO"])
        ws.append(["OR1159", "po38515"])
        ws.append(["OR1159", "po38515"])
        p = out_dir / "or_list.xlsx"
        wb.save(str(p))

        or_result = oli.load_or_list(p)
        assert or_result.status == "OK", or_result.errors
        assert or_result.detection_source == "POSITIONAL_FALLBACK"
        or_index = oli.build_or_index(or_result)

        # v20 (SG-533-TEST final cleanup, requirement 3): country="SG" --
        # the FINAL real production state for this exact archive shape (see
        # t_resolve_package_country_real_sg533_test_archive_shape in
        # test_pl_ocr_core.py), not "" -- this regression now represents
        # the real end-to-end pipeline, not a hypothetical unresolved-
        # country stand-in (that case has its own dedicated test above).
        pkg = _pkg_with_country("SG-553_CN", "SG", seq=2)
        m = pge.match_store_and_or(pkg, or_index)
        assert m.status == "OK", m.review_reason
        assert m.matched_store == "", f"must never invent a China Store, got {m.matched_store!r}"
        assert m.match_source == "NO_STORE_DIMENSION_FLAT_MATCH", m.match_source
        assert m.matched_or == "OR1159", m.matched_or
        assert m.matched_so == "po38515", m.matched_so
        assert dict(m.matched_business_fields) == {"OR": "OR1159", "SO": "po38515"}, dict(m.matched_business_fields)
    finally:
        shutil.rmtree(out_dir, ignore_errors=True)


def t_bare_or_so_two_column_or_list_stays_review_when_genuinely_ambiguous():
    """Same no-Store shape, but with TWO genuinely different OR/SO records
    and no Store/routing dimension to pick between them -- must stay
    REVIEW, never guess which one applies."""
    import pl_or_list_import as oli
    import pl_group_export as pge
    import openpyxl, tempfile, shutil
    out_dir = Path(tempfile.mkdtemp(prefix="bug6_ambiguous_"))
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["OR", "SO"])
        ws.append(["OR1159", "po38515"])
        ws.append(["OR2200", "po90099"])
        p = out_dir / "or_list.xlsx"
        wb.save(str(p))

        or_result = oli.load_or_list(p)
        assert or_result.status == "OK", or_result.errors
        or_index = oli.build_or_index(or_result)

        # v20: country="SG" -- see the note in the previous test.
        pkg = _pkg_with_country("SG-553_CN", "SG", seq=3)
        m = pge.match_store_and_or(pkg, or_index)
        assert m.status == "REVIEW", (m.status, m.matched_business_fields)
        assert m.matched_store == ""
    finally:
        shutil.rmtree(out_dir, ignore_errors=True)


def t_real_store_labeled_or_list_that_fails_to_match_still_stays_review():
    """Guard: a GENUINE Store/Shop-labeled OR List (LITERAL_HEADER tier)
    whose store text simply doesn't match any known CN store alias must
    NOT be reinterpreted by the new no-Store fallback -- that fallback is
    restricted to POSITIONAL_FALLBACK rows only (see
    _match_store_and_or_flat_no_store's docstring). A real Store column
    with unrecognized store names is a genuine "no match", not a signal
    that there's no Store dimension at all."""
    import pl_or_list_import as oli
    import pl_group_export as pge
    import openpyxl, tempfile, shutil
    out_dir = Path(tempfile.mkdtemp(prefix="bug6_real_store_no_match_"))
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Store", "OR#", "Ref#"])
        ws.append(["Some Unrecognized Boutique", "OR7000", "po1"])
        p = out_dir / "or_list.xlsx"
        wb.save(str(p))

        or_result = oli.load_or_list(p)
        assert or_result.status == "OK", or_result.errors
        assert or_result.detection_source == "LITERAL_HEADER"
        or_index = oli.build_or_index(or_result)

        # v20: country="SG" -- see the note above.
        pkg = _pkg_with_country("SG-553_CN", "SG", seq=4)
        m = pge.match_store_and_or(pkg, or_index)
        assert m.status == "REVIEW", (
            "a real Store column with an unrecognized store name must stay REVIEW, "
            f"never be reinterpreted as a flat OR/Ref record -- got status={m.status!r}"
        )
    finally:
        shutil.rmtree(out_dir, ignore_errors=True)


def t_modern_store_or_ref_so_schema_keeps_so_distinct_from_ref():
    """spec (SG-533-TEST consolidation report) requirement 5: the legacy
    bare-OR|SO compatibility fallback (BUG6 above) must be ISOLATED to
    the specific no-Store schema it targets -- a MODERN OR List with a
    real Store column, "Store | OR | Ref | SO | PO", must keep Ref and SO
    as two distinct fields (never merge SO into a "Ref No. compatibility
    value", never let the legacy fallback touch this shape at all). This
    positively proves the SUCCESS path (a real Store DOES match), as a
    companion to t_real_store_labeled_or_list_that_fails_to_match_still_
    stays_review below (which only proves the FAILURE path)."""
    import pl_or_list_import as oli
    import pl_group_export as pge
    import openpyxl, tempfile, shutil
    out_dir = Path(tempfile.mkdtemp(prefix="bug6_modern_schema_"))
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Store", "OR", "Ref", "SO", "PO"])
        ws.append(["Kerry Center flagship", "OR7001", "po70001", "SO7002", "PO7003"])
        p = out_dir / "or_list.xlsx"
        wb.save(str(p))

        or_result = oli.load_or_list(p)
        assert or_result.status == "OK", or_result.errors
        assert or_result.detection_source == "LITERAL_HEADER"
        or_index = oli.build_or_index(or_result)

        pkg = _pkg_with_country("CN-9099_KERRY_PVG_POP", "CN", seq=5)
        m = pge.match_store_and_or(pkg, or_index)
        assert m.status == "OK", m.review_reason
        assert m.match_source != "NO_STORE_DIMENSION_FLAT_MATCH", (
            "a real Store-matched OR List must never go through the legacy no-Store fallback"
        )
        assert dict(m.matched_business_fields) == {
            "OR": "OR7001", "Ref": "po70001", "SO": "SO7002", "PO": "PO7003",
        }, dict(m.matched_business_fields)
        assert m.matched_or == "OR7001", m.matched_or
        assert m.matched_so == "po70001", (
            "position 1 (Ref) is the backward-compat 'matched_so' value -- "
            f"must never become the SO field's value, got {m.matched_so!r}"
        )
    finally:
        shutil.rmtree(out_dir, ignore_errors=True)


for _name, _fn in [
    ("BUG6: bare 'OR | SO' OR List (no Store column, real SG-533-TEST shape) propagates OR/Ref unambiguously, matched_store stays ''", t_bare_or_so_two_column_or_list_propagates_unambiguously_without_a_store),
    ("BUG6: bare 'OR | SO' OR List with 2 genuinely different records stays REVIEW (no Store to disambiguate)", t_bare_or_so_two_column_or_list_stays_review_when_genuinely_ambiguous),
    ("BUG6 guard: a real Store-labeled OR List that fails to match stays REVIEW, never reinterpreted as no-Store", t_real_store_labeled_or_list_that_fails_to_match_still_stays_review),
    ("BUG6 modern-schema guard: Store|OR|Ref|SO|PO keeps Ref and SO as two distinct fields (legacy fallback never touches it)", t_modern_store_or_ref_so_schema_keeps_so_distinct_from_ref),
]:
    test(_name, _fn)


# =========================================================================
# BUG3 -- OR List "Shop | OR# | Ref#" -> Packing List "OR No." / "Ref No."
# (not "SO No."), and the Sublist PDF metadata block agrees.
# =========================================================================
print("== BUG3: OR List 'Ref#' -> 'Ref No.' (not 'SO No.') ==")


def t_or_ref_business_field_labels_canonicalize_correctly():
    assert canonicalize_business_field_label("OR#") == "OR No."
    assert canonicalize_business_field_label("OR No") == "OR No."
    assert canonicalize_business_field_label("Ref#") == "Ref No."
    assert canonicalize_business_field_label("Ref No") == "Ref No."
    assert canonicalize_business_field_label("Reference No") == "Ref No."
    # unrecognized headers preserved verbatim (spec: OR List business
    # fields stay fully dynamic, never hardcoded to just OR/Ref).
    assert canonicalize_business_field_label("Fulfillment No.") == "Fulfillment No."


def t_packing_list_headers_or_no_ref_no_for_shop_or_ref_or_list():
    """v17 (spec sections 12/16): Store/OR No./Ref No. are FIXED columns
    2/3/4 -- ALWAYS these exact labels, never renamed by the OR List's own
    header text (a "Shop|OR#|Ref#" OR List, with no columns beyond OR/Ref,
    produces no optional columns at all -- Product Name stays at column 5,
    matching PL_HEADERS_EN's original column 4 shifted by exactly Store's
    +1)."""
    import openpyxl
    out_dir = Path(tempfile.mkdtemp(prefix="pl_v15_orref_"))
    try:
        p1 = _pkg("CN-6557-CNWORLD_PEK-VN", 0)
        p1.or_number, p1.so_number = "OR1172", "po38533"
        p1.store_display = "China World NB1026"
        p1.global_carton_num = "1/1"
        out_path = out_dir / "PL_Total.xlsx"
        write_workbook(out_path, [p1])

        wb = openpyxl.load_workbook(str(out_path))
        ws = wb["Packing List"]
        header_row = D["TABLE_HDR_ROW1"]
        assert ws.cell(row=header_row, column=2).value == "Store", ws.cell(row=header_row, column=2).value
        assert ws.cell(row=header_row, column=3).value == "OR No.", ws.cell(row=header_row, column=3).value
        assert ws.cell(row=header_row, column=4).value == "Ref No.", ws.cell(row=header_row, column=4).value
        assert ws.cell(row=header_row, column=5).value == "Product Name\nin English"
        first_item_row = D["FIRST_ITEM_ROW"]
        assert ws.cell(row=first_item_row, column=2).value == "China World NB1026"
        assert ws.cell(row=first_item_row, column=3).value == "OR1172"
        assert ws.cell(row=first_item_row, column=4).value == "po38533"
        wb.close()
    finally:
        shutil.rmtree(out_dir, ignore_errors=True)


def t_sublist_pdf_metadata_shows_or_no_ref_no():
    try:
        import pl_sublist_pdf_export as ppe
        import pdfplumber
    except ImportError as e:
        print(f"        (skipped -- {e})")
        return
    p1 = _pkg("CN-6557-CNWORLD_PEK-VN", 0)
    p1.or_number, p1.so_number = "OR1172", "po38533"
    p1.store_display = "China World NB1026"
    p1.global_carton_num = "1/1"
    p1.package_code = "PGKEC0001"
    with tempfile.TemporaryDirectory() as td:
        out = Path(td) / "SUBLIST_TOTAL.pdf"
        result = ppe.generate_sublist_pdf([p1], out)
        assert result.status == "SUCCESS", result.error
        with pdfplumber.open(str(out)) as pdf:
            text = pdf.pages[0].extract_text() or ""
        assert "Store" in text and "China World NB1026" in text, text
        assert "OR No." in text, text
        assert "Ref No." in text, text
        assert "SO Order #" not in text, text
        assert "OR #" not in text, text


for _name, _fn in [
    ("canonicalize_business_field_label: OR#/OR No -> 'OR No.', Ref#/Ref No/Reference No -> 'Ref No.'", t_or_ref_business_field_labels_canonicalize_correctly),
    ("Packing List sheet: Shop|OR#|Ref# OR List -> headers 'OR No.'/'Ref No.' (not 'SO No.')", t_packing_list_headers_or_no_ref_no_for_shop_or_ref_or_list),
    ("Sublist PDF metadata: shows 'OR No.'/'Ref No.', never 'SO Order #'", t_sublist_pdf_metadata_shows_or_no_ref_no),
]:
    test(_name, _fn)


# =========================================================================
# BUG4 -- confirmed material SKU<->EAN pairs, enrichment-only.
# =========================================================================
print("== BUG4: confirmed material SKU/EAN pairs (enrichment-only) ==")

_CONFIRMED_PAIRS = [
    ("TP-PLB-GRY-L-01", "4895227932472"),
    ("TP-PLB-GRY-M-01", "4895227933912"),
    ("TP-PKG-POLYD-PE", "4894961086113"),
    ("TP-PKG-POLYE-R1-PE", "4894961086120"),
]


def _item(sku, barcode, qty=1):
    return Item(no="1", product_name="Material", product_code=sku, barcode=barcode,
                unit="PCS", quantity=qty)


def t_all_four_confirmed_pairs_are_in_the_known_mapping():
    for sku, ean in _CONFIRMED_PAIRS:
        assert KNOWN_MATERIAL_EAN.get(sku) == ean, (sku, KNOWN_MATERIAL_EAN.get(sku))


def t_blank_barcode_gets_enriched_from_known_mapping():
    for sku, ean in _CONFIRMED_PAIRS:
        it = _item(sku, "", qty=2)
        enrich_known_material_ean(it)
        assert it.barcode == ean, (sku, it.barcode)
        assert it.gtin_valid is True
        assert it.quantity == 2  # Qty is never touched


def t_sku_text_as_barcode_gets_enriched_not_dropped():
    """The real EAN List 6587.xlsx row for TP-PLB-GRY-M-01 has the SKU
    TEXT itself in the barcode cell instead of a real EAN -- confirmed
    production evidence for this exact case."""
    it = _item("TP-PLB-GRY-M-01", "TP-PLB-GRY-M-01", qty=1)
    enrich_known_material_ean(it)
    assert it.barcode == "4895227933912", it.barcode
    assert it.gtin_valid is True


def t_conflicting_valid_gtin_is_never_silently_overwritten():
    """4006381333931 is a genuinely valid GTIN-13 (real GS1 check-digit
    arithmetic -- see t_production_gtin_validator_matches_confirmed_gs1_
    fixtures below, which proves this against the ACTUAL production
    validator) and differs from the confirmed mapping for TP-PLB-GRY-L-01
    (4895227932472).

    v20 (SG-533-TEST final cleanup, requirement 1): this docstring
    previously claimed production's is_valid_gtin13() only checked digit
    count and did NOT implement GS1 check-digit validation. That claim is
    now FALSE and has been removed -- production has done real GS1/EAN-13
    modulo-10 checksum validation since an earlier pass of this
    engagement (see gtin13_checksum_is_valid() in pl_ocr_core.py). This
    test still uses a genuinely valid GTIN-13 (rather than any other
    13-digit string) so its intent stays unambiguous regardless."""
    it = _item("TP-PLB-GRY-L-01", "4006381333931", qty=1)  # a DIFFERENT, genuinely valid GTIN-13
    enrich_known_material_ean(it)
    assert it.barcode == "4006381333931", "a different valid parsed GTIN must never be overwritten"
    assert "MATERIAL_EAN_REVIEW" in it.remark, it.remark


def t_production_gtin_validator_matches_confirmed_gs1_fixtures():
    """v20 (SG-533-TEST final cleanup, requirement 1 -- supersedes the
    prior t_gtin13_test_fixture_is_actually_a_valid_checksum, which
    deliberately proved the fixture against a SEPARATE validator written
    inside the test file rather than production's own. That was correct
    at the time (production genuinely only checked digit count then), but
    is no longer the right test now that production implements real GS1
    checksum validation -- per explicit instruction, there must be ONE
    production validator as source of truth, not a duplicate. Calls
    is_valid_gtin13() (bound directly from pl_ocr_core.py above, not
    reimplemented) for every required fixture."""
    assert is_valid_gtin13("4006381333931") is True, "genuine GS1 GTIN-13 must validate"
    assert is_valid_gtin13("1111111111111") is False, "13 digits but fails the GS1 checksum -- must NOT validate"
    # the 4 confirmed material EANs (spec requirement) must all validate
    # through this SAME production function.
    material_eans = {
        "TP-PLB-GRY-L-01": "4895227932472",
        "TP-PLB-GRY-M-01": "4895227933912",
        "TP-PKG-POLYD-PE": "4894961086113",
        "TP-PKG-POLYE-R1-PE": "4894961086120",
    }
    for sku, ean in material_eans.items():
        assert is_valid_gtin13(ean) is True, f"{sku} ({ean}) must validate via production is_valid_gtin13()"


def t_matching_barcode_left_untouched_and_marked_valid():
    it = _item("TP-PKG-POLYD-PE", "4894961086113", qty=1)
    enrich_known_material_ean(it)
    assert it.barcode == "4894961086113"
    assert it.gtin_valid is True
    assert it.remark == ""


def t_unknown_material_sku_is_a_no_op():
    it = _item("TP-PLB-SOMETHING-ELSE", "", qty=1)
    enrich_known_material_ean(it)
    assert it.barcode == ""  # never invented for an unmapped material SKU


# =========================================================================
# BUG3 (point-3 correction) -- FULL end-to-end dynamic business-field
# propagation, not just relabeling columns B/C. v17 (master prompt
# correction -- supersedes the "only first 2 fields physically display"
# design): Store/OR No./Ref No. are the fixed business backbone (spec
# section 12); everything an OR List has BEYOND OR/Ref is a dynamic
# OPTIONAL field that must appear PHYSICALLY -- not just in Raw_Data --
# in every one of Packing List / grouped Packing List / Sublist Excel /
# Sublist PDF, in the OR List's own original order, with real values (not
# hand-set stand-ins). A minimal OR List (no columns beyond OR/Ref) must
# show NO invented blank optional columns/rows anywhere.
# =========================================================================
print("== BUG3 (point 3): full dynamic business-field propagation (7-column OR List) ==")


def t_seven_field_or_list_propagates_end_to_end():
    """v17 (master prompt correction): an 8-column OR List (Shop|OR#|
    Ref#|SO|PO|Invoice|Fulfillment No.|Buyer) -- Store/OR No./Ref No.
    fixed, PLUS 5 distinct optional fields (SO is now its own optional
    column, separate from Ref No.) -- must propagate ALL of Store/OR No./
    Ref No./SO/PO/Invoice/Fulfillment No./Buyer, real labels AND real
    values, PHYSICALLY into PL_Total, grouped Packing List, Sublist Excel,
    AND Sublist PDF (not just Raw_Data)."""
    import openpyxl
    out_dir = Path(tempfile.mkdtemp(prefix="pl_v15_dynamic_fields_"))
    try:
        # ---- 1) a REAL 8-column OR List, parsed through the real loader ----
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Shop", "OR#", "Ref#", "SO", "PO", "Invoice", "Fulfillment No.", "Buyer"])
        ws.append(["CN - Hangzhou Mixc", "OR9001", "po90001", "SO123", "PO-A1", "INV-778", "FUL-55", "Alice"])
        or_list_path = out_dir / "or_list_8field.xlsx"
        wb.save(str(or_list_path))

        or_result = oli.load_or_list(or_list_path)
        assert or_result.status == "OK", or_result.errors
        assert or_result.business_field_labels == [
            "OR#", "Ref#", "SO", "PO", "Invoice", "Fulfillment No.", "Buyer"
        ], or_result.business_field_labels

        or_index = {}
        for r in or_result.rows:
            or_index.setdefault(r.or_norm, []).append(r)

        # ---- 2) match a real Shipping Mark against it (pl_group_export,
        #         the exact function run_pipeline() itself calls) ----
        pkg = _pkg("CN-1529_HANGZHOU_PVG_POP", 0)
        m = pge.match_store_and_or(pkg, or_index)
        assert m.status == "OK", m.review_reason
        assert dict(m.matched_business_fields) == {
            "OR#": "OR9001", "Ref#": "po90001", "SO": "SO123", "PO": "PO-A1",
            "Invoice": "INV-778", "Fulfillment No.": "FUL-55", "Buyer": "Alice",
        }, dict(m.matched_business_fields)

        # ---- 3) Package.business_fields / store_display -- mirroring
        #         exactly what run_pipeline() itself now does ----
        classify_packages_for_port([pkg], None, False)
        assert pkg.store == "HANGZHOU" and pkg.store_display == "Hangzhou Mixc", (pkg.store, pkg.store_display)
        pkg.business_fields = dict(m.matched_business_fields)
        pkg.or_number, pkg.so_number = m.matched_or, m.matched_so
        pkg.global_carton_num = "1/1"
        assert len(pkg.business_fields) == 7
        assert pkg.business_fields["Fulfillment No."] == "FUL-55"
        assert pkg.business_fields["Buyer"] == "Alice"

        optional_labels = list(or_result.business_field_labels[2:])
        assert optional_labels == ["SO", "PO", "Invoice", "Fulfillment No.", "Buyer"], optional_labels

        expected_fields = [
            ("Store", "Hangzhou Mixc"), ("OR No.", "OR9001"), ("Ref No.", "po90001"),
            ("SO", "SO123"), ("PO", "PO-A1"), ("Invoice", "INV-778"),
            ("Fulfillment No.", "FUL-55"), ("Buyer", "Alice"),
        ]

        # ---- 4) Packing List (PL_Total) -- Store/OR/Ref fixed + ALL 5
        #         optional fields, real values, physically present ----
        pl_total_path = out_dir / "PL_Total.xlsx"
        write_workbook(pl_total_path, [pkg], optional_business_field_labels=optional_labels)
        wb_pl = openpyxl.load_workbook(str(pl_total_path))
        ws_pl = wb_pl["Packing List"]
        header_row = D["TABLE_HDR_ROW1"]
        first_item_row = D["FIRST_ITEM_ROW"]
        for i, (label, value) in enumerate(expected_fields, start=2):
            assert ws_pl.cell(row=header_row, column=i).value == label, (
                i, label, ws_pl.cell(row=header_row, column=i).value)
            assert ws_pl.cell(row=first_item_row, column=i).value == value, (
                i, value, ws_pl.cell(row=first_item_row, column=i).value)
        # Product Name (originally column 4) is now column 2+8=10 (Store/OR/
        # Ref/5-optional = 8 business columns after Item#).
        assert ws_pl.cell(row=header_row, column=10).value == "Product Name\nin English"
        wb_pl.close()

        # ---- 5) grouped Packing List (export_grouped_pl -> 02_BY_FACTORY)
        #         -- the exact propagation gap point 3 originally found:
        #         export_grouped_pl() must use the SAME optional_labels,
        #         never silently fall back to a shorter/default set. ----
        split_dir = out_dir / "PL_SPLIT_OUTPUT"
        pge.export_grouped_pl(
            packages=[pkg], output_dir=split_dir, write_workbook=write_workbook,
            optional_business_field_labels=optional_labels,
        )
        grouped_path = split_dir / "02_BY_FACTORY" / "PL_FACTORY_POP.xlsx"
        assert grouped_path.exists(), "PL_FACTORY_POP.xlsx was not written"
        wb_g = openpyxl.load_workbook(str(grouped_path))
        ws_g = wb_g["Packing List"]
        for i, (label, value) in enumerate(expected_fields, start=2):
            assert ws_g.cell(row=header_row, column=i).value == label, (
                "grouped Packing List header did not inherit the OR List's own "
                f"optional fields -- column {i}: expected {label!r}, got "
                f"{ws_g.cell(row=header_row, column=i).value!r}"
            )
            assert ws_g.cell(row=first_item_row, column=i).value == value
        wb_g.close()

        # ---- 6) Sublist Excel -- Store/OR/Ref + ALL 5 optional fields ----
        import pl_sublist_export as pse
        sublist_xlsx_path = out_dir / "SUBLIST_TOTAL.xlsx"
        pse.generate_sublist_workbook([pkg], sublist_xlsx_path, optional_business_field_labels=optional_labels)
        wb_s = openpyxl.load_workbook(str(sublist_xlsx_path))
        ws_s = wb_s["Sheet1"]
        off = pse._sublist_offsets(len(optional_labels))
        sublist_rows = [
            (off["carton"], "Carton #", "1/1"),
            (off["store"], "Store", "Hangzhou Mixc"),
            (off["or"], "OR No.", "OR9001"),
            (off["ref"], "Ref No.", "po90001"),
            (off["optional_first"] + 0, "SO", "SO123"),
            (off["optional_first"] + 1, "PO", "PO-A1"),
            (off["optional_first"] + 2, "Invoice", "INV-778"),
            (off["optional_first"] + 3, "Fulfillment No.", "FUL-55"),
            (off["optional_first"] + 4, "Buyer", "Alice"),
            (off["gw"], "GW", None),
            (off["packing_code"], "Packing Code #", None),
        ]
        for offset, label, value in sublist_rows:
            r = 1 + offset
            assert ws_s[f"B{r}"].value == label, (r, label, ws_s[f"B{r}"].value)
            if value is not None:
                assert ws_s[f"C{r}"].value == value, (r, value, ws_s[f"C{r}"].value)
        wb_s.close()

        # ---- 7) Sublist PDF -- Store/OR/Ref + ALL 5 optional fields
        #         rendered, real values, no clipping/overlap (page still
        #         builds successfully with the taller metadata block) ----
        try:
            import pl_sublist_pdf_export as ppe
            import pdfplumber
            sublist_pdf_path = out_dir / "SUBLIST_TOTAL.pdf"
            pdf_result = ppe.generate_sublist_pdf(
                [pkg], sublist_pdf_path, optional_business_field_labels=optional_labels)
            assert pdf_result.status == "SUCCESS", pdf_result.error
            with pdfplumber.open(str(sublist_pdf_path)) as pdf:
                assert len(pdf.pages) == 1, "8-field metadata block must still fit a single page for a 0-item carton"
                pdf_text = pdf.pages[0].extract_text() or ""
            for label, value in expected_fields:
                assert label in pdf_text, f"{label!r} missing from Sublist PDF text:\n{pdf_text}"
                assert value in pdf_text, f"{value!r} missing from Sublist PDF text:\n{pdf_text}"
        except ImportError as e:
            print(f"        (Sublist PDF check skipped -- {e})")

        # ---- 8) Raw_Data -- the FULL 7-field business_fields record still
        #         auditable, len(row) == len(header) (structural safety) ----
        wb_raw = openpyxl.load_workbook(str(pl_total_path))
        ws_raw = wb_raw["Raw_Data"]
        header = [c.value for c in next(ws_raw.iter_rows(min_row=1, max_row=1))]
        assert "business_fields" in header
        col = header.index("business_fields")
        row = next(ws_raw.iter_rows(min_row=2, max_row=2, values_only=True))
        assert len(row) == len(header), (
            f"Raw_Data row/header length mismatch: row={len(row)} header={len(header)}"
        )
        raw_bf = row[col]
        for label, value in [("OR#", "OR9001"), ("Ref#", "po90001"), ("SO", "SO123"),
                              ("PO", "PO-A1"), ("Invoice", "INV-778"),
                              ("Fulfillment No.", "FUL-55"), ("Buyer", "Alice")]:
            assert f"{label}: {value}" in raw_bf, f"{label}: {value} missing from Raw_Data.business_fields={raw_bf!r}"
        wb_raw.close()
    finally:
        shutil.rmtree(out_dir, ignore_errors=True)


def t_minimal_or_list_never_invents_optional_columns():
    """spec section 33 point A: a minimal Shop|OR#|Ref# OR List (no
    columns beyond OR/Ref) must show Store/OR No./Ref No. ONLY -- no
    blank invented SO/PO/etc columns/rows anywhere (Packing List, Sublist
    Excel, Sublist PDF)."""
    import openpyxl
    out_dir = Path(tempfile.mkdtemp(prefix="pl_v15_minimal_or_list_"))
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Shop", "OR#", "Ref#"])
        ws.append(["CN - China World NB1026", "OR1172", "po38533"])
        or_list_path = out_dir / "or_list_minimal.xlsx"
        wb.save(str(or_list_path))

        or_result = oli.load_or_list(or_list_path)
        assert or_result.status == "OK", or_result.errors
        optional_labels = list(or_result.business_field_labels[2:])
        assert optional_labels == [], optional_labels

        or_index = {}
        for r in or_result.rows:
            or_index.setdefault(r.or_norm, []).append(r)
        pkg = _pkg("CN-6557-CNWORLD_PEK-VN", 0)
        classify_packages_for_port([pkg], None, False)
        m = pge.match_store_and_or(pkg, or_index)
        assert m.status == "OK", m.review_reason
        pkg.business_fields = dict(m.matched_business_fields)
        pkg.or_number, pkg.so_number = m.matched_or, m.matched_so
        pkg.global_carton_num = "1/1"

        pl_total_path = out_dir / "PL_Total.xlsx"
        write_workbook(pl_total_path, [pkg], optional_business_field_labels=optional_labels)
        wb_pl = openpyxl.load_workbook(str(pl_total_path))
        ws_pl = wb_pl["Packing List"]
        header_row = D["TABLE_HDR_ROW1"]
        assert ws_pl.cell(row=header_row, column=2).value == "Store"
        assert ws_pl.cell(row=header_row, column=3).value == "OR No."
        assert ws_pl.cell(row=header_row, column=4).value == "Ref No."
        # column 5 must be Product Name -- no invented blank optional column.
        assert ws_pl.cell(row=header_row, column=5).value == "Product Name\nin English"
        wb_pl.close()

        import pl_sublist_export as pse
        sublist_xlsx_path = out_dir / "SUBLIST_TOTAL.xlsx"
        pse.generate_sublist_workbook([pkg], sublist_xlsx_path, optional_business_field_labels=optional_labels)
        wb_s = openpyxl.load_workbook(str(sublist_xlsx_path))
        ws_s = wb_s["Sheet1"]
        off = pse._sublist_offsets(0)
        assert ws_s[f"B{1 + off['ref']}"].value == "Ref No."
        assert ws_s[f"B{1 + off['gw']}"].value == "GW", (
            "GW row must immediately follow Ref No. -- no blank optional row in between"
        )
        wb_s.close()
    finally:
        shutil.rmtree(out_dir, ignore_errors=True)


def t_raw_data_row_length_matches_header_for_every_branch():
    """spec section 26: structural safety net -- Raw_Data's row length
    must equal its header length for BOTH the "package has items" and
    "package has zero items" writer branches, for every package shape."""
    import openpyxl
    out_dir = Path(tempfile.mkdtemp(prefix="pl_v15_rawdata_structural_"))
    try:
        p_with_items = _pkg("CN-6557-CNWORLD_PEK-VN", 0)
        p_with_items.items = [Item(no="1", product_name="Bungee Strap", product_code="TP-A-1",
                                     barcode="4894961069222", unit="PCS", quantity=10)]
        p_with_items.declared_total_qty = 10
        p_with_items.global_carton_num = "1/2"

        p_zero_items = _pkg("CN-6557-GUANGZHOU_SZX_CN", 1)
        p_zero_items.items = []
        p_zero_items.global_carton_num = "2/2"

        out_path = out_dir / "PL_Total.xlsx"
        write_workbook(out_path, [p_with_items, p_zero_items])
        wb = openpyxl.load_workbook(str(out_path))
        ws = wb["Raw_Data"]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            assert len(row) == len(header), (
                f"Raw_Data row/header length mismatch: row_len={len(row)} header_len={len(header)} row={row!r}"
            )
        wb.close()
    finally:
        shutil.rmtree(out_dir, ignore_errors=True)


for _name, _fn in [
    ("8-field OR List (Shop|OR#|Ref#|SO|PO|Invoice|Fulfillment No.|Buyer): full propagation end-to-end into every physical output", t_seven_field_or_list_propagates_end_to_end),
    ("Minimal Shop|OR#|Ref# OR List never invents blank optional columns/rows", t_minimal_or_list_never_invents_optional_columns),
    ("Raw_Data row length == header length for both the has-items and zero-items writer branches", t_raw_data_row_length_matches_header_for_every_branch),
]:
    test(_name, _fn)


for _name, _fn in [
    ("All 4 confirmed SKU/EAN pairs are present in KNOWN_MATERIAL_EAN", t_all_four_confirmed_pairs_are_in_the_known_mapping),
    ("Blank material barcode is enriched from the confirmed mapping, Qty untouched", t_blank_barcode_gets_enriched_from_known_mapping),
    ("SKU-text-as-barcode (real EAN List defect) is enriched, row never dropped", t_sku_text_as_barcode_gets_enriched_not_dropped),
    ("A DIFFERENT already-valid parsed GTIN is never silently overwritten (REVIEW flagged)", t_conflicting_valid_gtin_is_never_silently_overwritten),
    ("Production is_valid_gtin13(): 4006381333931 valid, 1111111111111 invalid, all 4 confirmed material EANs valid (real function, not a duplicate)", t_production_gtin_validator_matches_confirmed_gs1_fixtures),
    ("An already-matching barcode is left untouched and marked gtin_valid", t_matching_barcode_left_untouched_and_marked_valid),
    ("An unmapped TP-PLB-* material SKU is a no-op (never invents an EAN)", t_unknown_material_sku_is_a_no_op),
]:
    test(_name, _fn)


print(f"\n{_passed} passed, {_failed} failed")
if _failed:
    sys.exit(1)
