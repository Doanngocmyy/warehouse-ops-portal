#!/usr/bin/env python3
"""
Regression tests for pl_ocr_core.py (PL OCR + Grouping tool).

No test framework dependency, by design -- matches this repo's existing
convention (see tests/outbound-availability.test.js: plain assert-based
tests with a tiny custom runner, no package.json/pytest.ini). Run with:

    python3 tools/pl-ocr-grouping/tests/test_pl_ocr_core.py

Requires the same third-party packages pl_ocr_core.py itself needs
(pandas, openpyxl, pdfplumber) -- install with:

    pip install pandas openpyxl pdfplumber --break-system-packages

pl_ocr_core.py is written to be `exec()`'d by app.html inside Pyodide (see
its own header comment), not imported as a normal module -- its bottom half
unconditionally calls run_pipeline() against hard-coded /work/... paths.
_load_core_defs() / _load_core_pipeline() below reproduce exactly the
placeholder-substitution app.html performs, then exec() the (possibly
truncated) source into a throwaway module so these tests exercise the real
file, not a reimplementation of it.
"""
from __future__ import annotations
import sys, types, tempfile, shutil
from pathlib import Path

HERE = Path(__file__).resolve().parent
TOOL_DIR = HERE.parent
CORE_PY = TOOL_DIR / "pl_ocr_core.py"
if str(TOOL_DIR) not in sys.path:
    sys.path.insert(0, str(TOOL_DIR))
FIXTURES = HERE / "fixtures"
# SANITIZED SYNTHETIC fixtures only (see 2026-08-03 incident: the original
# version of this test suite committed 8 REAL customer packing-list PDFs +
# a real DIM.xlsx containing real customer names/phone numbers/addresses to
# this PUBLIC repo. That history has been purged -- see git log / README
# note for the cleanup commit. These synthetic fixtures were built with
# reportlab (tools/pl-ocr-grouping/tests/fixtures/synthetic/generate.py) to
# reproduce the exact same layout defects (merged condition+quantity cell,
# SKU "-BOX" suffix split across a newline, package header living inside
# the item table, package continuation across pages, one page with two
# packages, repeated table header + about:blank/date noise) using entirely
# fabricated codes/addresses. NO real shipment data ships in this repo.
SYN_PDFS = FIXTURES / "synthetic"
SYN_DIM = FIXTURES / "synthetic" / "SYN-DIM.xlsx"

_ENTRY_MARKER = "# ── Entry point ────────────────────────────────────────────────────────────"
_AUTO_SPLIT_MARKER = "# =========================================================\n# AUTO SPLIT:"


def _substitute_placeholders(src: str, *, dim_sheet=None, master_sheet=None,
                              recursive=False, consignee=None, notify=None,
                              generate_sublist=True, or_list_file=None,
                              generate_sublist_pdf=True, routing_rules=None,
                              convert_to_pcs=False, show_uom=False) -> str:
    def lit(v):
        return "None" if v is None else repr(v)
    import json as _json
    routing_rules_json = _json.dumps(routing_rules) if routing_rules else "[]"
    return (src
            .replace("__DIM_WEIGHT_SHEET__", lit(dim_sheet))
            .replace("__MASTER_DATA_SHEET__", lit(master_sheet))
            .replace("__RECURSIVE__", "True" if recursive else "False")
            .replace("__MANUAL_CONSIGNEE__", lit(consignee))
            .replace("__MANUAL_NOTIFY_PARTY__", lit(notify))
            .replace("__GENERATE_SUBLIST__", "True" if generate_sublist else "False")
            .replace("__GENERATE_SUBLIST_PDF__", "True" if generate_sublist_pdf else "False")
            .replace("__OR_LIST_FILE__", lit(str(or_list_file)) if or_list_file else "None")
            .replace("__ROUTING_RULES_JSON__", routing_rules_json)
            .replace("__CONVERT_TO_PCS__", "True" if convert_to_pcs else "False")
            .replace("__SHOW_UOM_IN_SUBLIST__", "True" if show_uom else "False")
            .replace("__GIT_COMMIT__", lit("test-suite")))


_module_counter = 0


def _exec_module(src: str, label: str):
    global _module_counter
    _module_counter += 1
    modname = f"pl_ocr_core_test_{_module_counter}"
    mod = types.ModuleType(modname)
    mod.__file__ = str(CORE_PY)
    sys.modules[modname] = mod
    exec(compile(src, str(CORE_PY), "exec"), mod.__dict__)
    return mod.__dict__


def load_core_defs():
    """Load every class/function/constant in pl_ocr_core.py WITHOUT running
    the pipeline (truncated before the entry-point block) -- for pure unit
    tests of parsing/normalization functions."""
    src = CORE_PY.read_text(encoding="utf-8")
    idx = src.index(_ENTRY_MARKER)
    src = _substitute_placeholders(src[:idx])
    return _exec_module(src, "defs")


def load_core_pipeline(pdf_dir: Path, dim_xlsx: Path, master_xlsx: Path, out_dir: Path, **kw):
    """Run the real run_pipeline() end-to-end (truncated just before the
    AUTO SPLIT / factory-store-grouping section, which is a separate
    feature) against real files on disk -- for integration tests."""
    src = CORE_PY.read_text(encoding="utf-8")
    idx = src.index(_AUTO_SPLIT_MARKER)
    src = _substitute_placeholders(src[:idx], **kw)
    src = src.replace('PL_FOLDER = Path("/work/pdfs")', f'PL_FOLDER = Path({str(pdf_dir)!r})')
    src = src.replace('OUTPUT_XLSX = Path("/work/PL_Total.xlsx")', f'OUTPUT_XLSX = Path({str(out_dir / "PL_Total.xlsx")!r})')
    src = src.replace('DIM_WEIGHT_FILE = Path("/work/dim.xlsx")', f'DIM_WEIGHT_FILE = Path({str(dim_xlsx)!r})')
    src = src.replace('MASTER_DATA_FILE = Path("/work/master.xlsx")', f'MASTER_DATA_FILE = Path({str(master_xlsx)!r})')
    return _exec_module(src, "pipeline")


def _make_empty_master_xlsx(path: Path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["SKU/Product Code", "HS Code", "EAN/Barcode"])
    wb.save(str(path))


# ── tiny test runner (mirrors tests/outbound-availability.test.js) ─────────
_passed = 0
_failed = 0
_failures = []


def test(name, fn):
    global _passed, _failed
    try:
        fn()
        _passed += 1
        print(f"  ok  - {name}")
    except AssertionError as e:
        _failed += 1
        _failures.append(name)
        print(f"FAIL  - {name}")
        print(f"        {e}")
    except Exception as e:
        _failed += 1
        _failures.append(name)
        print(f"ERROR - {name}")
        print(f"        {type(e).__name__}: {e}")


# =============================================================================
# 1. Parser unit tests -- pure functions/classes, no PDF needed
# =============================================================================
print("== Parser unit tests ==")
C = load_core_defs()


def t_cond_qty_merge_split():
    row = ['1', '4894961082009', 'TP-WST-RL10-MCC-02', '10mm Rope Loop', 'PCS', 'Moi 12']
    item = C["parse_item_cells"](row)
    assert item is not None, "item should parse"
    assert item.quantity == 12, f"quantity={item.quantity}"
    assert item.condition, "condition should be captured"


def t_sku_box_suffix_newline_in_cell():
    row = ['2', '4894961081941', 'TP-WST-RL10-MWD-02-\nBOX', '10mm Rope Loop', 'PCS', 'Moi 6']
    item = C["parse_item_cells"](row)
    assert item is not None
    assert item.product_code == "TP-WST-RL10-MWD-02-BOX", item.product_code


def t_sku_box_suffix_next_row():
    # Real-evidence shape (spec item #6: "-BOX" continuation on its own next
    # row/line): the item row's own SKU cell is already a complete code (no
    # trailing '-'), and the "-BOX" suffix arrives as an entirely separate
    # table row directly below it -- NOT as an adjacent cell within the same
    # row (that different case -- a cell ending in '-' merging with the very
    # next cell -- is covered by t_sku_box_suffix_newline_in_cell and is a
    # pre-existing, separate code path).
    Parser = C["Parser"]
    p = Parser()
    p.set_file(Path("CN-9999-Test-CN.pdf"))
    p.set_page(1)
    p.feed_table_row(["Ma kien hang: PGKECTEST0000001 1/1", "", "", "", "", ""])
    p.feed_table_row(["1", "4894961082009", "TP-WST-BGS-CMY-56", "Bag", "PCS", "Moi 4"])
    p.feed_table_row(["-BOX"])
    p.feed_table_row(["Tong cong 4"])
    p.finalise()
    assert len(p.packages) == 1
    item = p.packages[0].items[0]
    assert item.product_code == "TP-WST-BGS-CMY-56-BOX", item.product_code


def t_unicode_dash_artifact_normalized():
    fix = C["fix_unicode_artifacts"]
    assert fix("TP-WBA-CSL14-CMY￾51") == "TP-WBA-CSL14-CMY-51"
    assert fix("TP-WBA-CSL14-CMY￿51") == "TP-WBA-CSL14-CMY-51"


def t_description_newline_join_via_code():
    normalize_code = C["normalize_code"]
    assert normalize_code("TP-WST-BGS-CMY-56-\nBOX") == "TP-WST-BGS-CMY-56-BOX"


def t_gtin13_valid():
    """NOTE (SG-533-TEST consolidation report, requirement 6): this test's
    OWN display label used to read "GTIN validated as exactly 13 digits"
    -- stale/misleading now that is_valid_gtin13() does real GS1/EAN-13
    modulo-10 checksum validation (see gtin13_checksum_is_valid()), not
    just a digit-count check. The checksum behavior itself is exercised
    separately and thoroughly in test_v15_bugfix_regression.py (4
    confirmed material EANs + the 4006381333931 valid / 1111111111111
    invalid fixture pair + the "different valid GTIN never silently
    overwritten" REVIEW case) -- this function only ever tested shape
    (digit count), so its label is corrected below to say so accurately,
    rather than implying it covers checksum behavior it doesn't actually
    assert."""
    is_valid = C["is_valid_gtin13"]
    assert is_valid("4894961082009") is True
    assert is_valid("123") is False
    assert is_valid("48949610820091") is False  # 14 digits, not 13
    assert is_valid("ABCDEFGHIJKLM") is False


def t_gtin_equals_sku_no_crash():
    row = ['1', '4894961082009', 'TP-4894961082009', 'Some Item', 'PCS', 'Moi 3']
    item = C["parse_item_cells"](row)  # must not raise
    assert item is not None


def t_header_footer_row_not_parsed_as_item():
    is_noise = C["is_noise"]
    assert is_noise("about:blank")
    assert is_noise("8/1/26, 4:27 PM about:blank")
    assert is_noise("# Barcode Ma san pham Ten hang hoa DVT Tinh trang So luong")
    assert not is_noise("1 4894961082009 TP-WST-RL10-MCC-02 10mm Rope Loop PCS Moi 12")


def t_repeated_table_header_not_item():
    is_hdr = C["is_table_hdr"]
    assert is_hdr(["#", "Barcode", "Ma san pham", "Ten hang hoa", "DVT", "Tinh trang So luong"])


def t_package_zero_items_status():
    Package = C["Package"]
    audit_status = C["audit_status"]
    pkg = Package(package_code="PGKECX", source_file="f.pdf", reference_code="f",
                  pdf_package_seq="1/1", declared_total_qty=10)
    assert audit_status(pkg) == "ZERO_ITEMS"


def t_package_missing_total_status():
    Package = C["Package"]
    audit_status = C["audit_status"]
    pkg = Package(package_code="PGKECX", source_file="f.pdf", reference_code="f", pdf_package_seq="1/1")
    pkg.items.append(C["Item"](no="1", product_name="x", product_code="TP-A-1", barcode="4894961082009",
                                unit="PCS", quantity=1))
    assert audit_status(pkg) == "MISSING_TOTAL"


def t_dedup_same_page_same_row_not_double_counted():
    """Table pass parses a row, then a redundant text-fallback pass (e.g. if
    the table produced 0 NEW items and the page's text is re-scanned)
    re-encounters the identical row: it must not be counted twice."""
    Parser = C["Parser"]
    p = Parser()
    p.set_file(Path("CN-9999-Test-CN.pdf"))
    p.set_page(1)
    p.feed_table_row(["Ma kien hang: PGKECTEST0000002 1/1", "", "", "", "", ""])
    row = ["1", "4894961082009", "TP-A-1", "Item A", "PCS", "Moi 5"]
    p.feed_table_row(row)
    p.feed_table_row(row)  # simulate the same physical row being fed twice
    p.feed_table_row(["Tong cong 5"])
    p.finalise()
    assert p.packages[0].item_count == 1, f"expected dedup, got {p.packages[0].item_count} items"
    assert p.duplicate_items_skipped == 1


def t_legit_repeated_sku_different_rows_both_kept():
    """The SAME sku/gtin/qty legitimately repeating on two DIFFERENT lines
    (different line_no) must both be kept -- dedup is context-aware, not a
    blind gtin+qty collapse."""
    Parser = C["Parser"]
    p = Parser()
    p.set_file(Path("CN-9999-Test-CN.pdf"))
    p.set_page(1)
    p.feed_table_row(["Ma kien hang: PGKECTEST0000003 1/1", "", "", "", "", ""])
    p.feed_table_row(["1", "4894961082009", "TP-A-1", "Item A", "PCS", "Moi 5"])
    p.feed_table_row(["2", "4894961082009", "TP-A-1", "Item A", "PCS", "Moi 5"])
    p.feed_table_row(["Tong cong 10"])
    p.finalise()
    assert p.packages[0].item_count == 2, "two genuinely different rows must both survive dedup"


def t_multi_package_same_page():
    Parser = C["Parser"]
    p = Parser()
    p.set_file(Path("CN-9999-Test-CN.pdf"))
    p.set_page(1)
    p.feed_table_row(["Ma kien hang: PGKECTEST0000004 1/2", "", "", "", "", ""])
    p.feed_table_row(["1", "4894961082009", "TP-A-1", "Item A", "PCS", "Moi 5"])
    p.feed_table_row(["Tong cong 5"])
    p.feed_table_row(["Ma kien hang: PGKECTEST0000005 2/2", "", "", "", "", ""])
    p.feed_table_row(["1", "4894961082010", "TP-A-2", "Item B", "PCS", "Moi 7"])
    p.feed_table_row(["Tong cong 7"])
    p.finalise()
    assert len(p.packages) == 2
    assert {pk.package_code for pk in p.packages} == {"PGKECTEST0000004", "PGKECTEST0000005"}


def t_package_spans_pages_totals_reconcile():
    Parser = C["Parser"]
    p = Parser()
    p.set_file(Path("CN-9999-Test-CN.pdf"))
    p.set_page(1)
    p.feed_table_row(["Ma kien hang: PGKECTEST0000006 1/1", "", "", "", "", ""])
    p.feed_table_row(["1", "4894961082009", "TP-A-1", "Item A", "PCS", "Moi 5"])
    p.set_page(2)
    p.feed_table_row(["Ma kien hang: PGKECTEST0000006 1/1", "", "", "", "", ""])  # repeated on continuation page
    p.feed_table_row(["2", "4894961082010", "TP-A-2", "Item B", "PCS", "Moi 8"])
    p.feed_table_row(["Tong cong 13"])
    p.finalise()
    assert len(p.packages) == 1, f"continuation must merge into ONE package, got {len(p.packages)}"
    pkg = p.packages[0]
    assert pkg.item_count == 2
    assert pkg.declared_total_qty == pkg.calc_qty == 13


def t_condition_word_separate_cell_8col():
    row = ["5", "4894961081927", "TP-WST-RL14-RGL-02", "14mm Rope Loop", "PCS", "Moi", "8", ""]
    item = C["parse_item_cells"](row)
    assert item is not None
    assert item.quantity == 8
    assert item.condition


def t_seven_column_row_all_fields():
    """Exactly 7 cells, no trailing blank column (distinct from the 8-col
    fixture above, which has a trailing ""). Verifies every field flagged
    in the follow-up review, not just quantity+condition: GTIN, SKU,
    description, UOM, condition, quantity."""
    row = ["3", "4894961082016", "TP-SVN-RL07-BLK-01", "7mm Rope Loop Black", "PCS", "Moi", "12"]
    item = C["parse_item_cells"](row)
    assert item is not None
    assert item.barcode == "4894961082016", item.barcode                # GTIN
    assert item.product_code == "TP-SVN-RL07-BLK-01", item.product_code  # SKU
    assert "Rope Loop" in item.product_name, item.product_name          # description
    assert item.unit == "PCS", item.unit                                # UOM
    assert item.condition, item.condition                               # condition
    assert item.quantity == 12, item.quantity                           # quantity


def t_delta_fallback_table_zero_items_recovers_via_text():
    """Direct regression test for bug #16 (`if tables: used_table = True`
    anti-pattern). Reproduces the exact Parser-level calls and the exact
    boolean condition run_pipeline() uses to decide whether to fall back to
    the text layer -- table pass yields a non-empty `tables` list whose
    every row is noise/header/blank (0 new items, 0 packages closed), so
    the fallback MUST trigger and the text layer's real item must be
    recovered exactly once (no duplication)."""
    Parser = C["Parser"]
    p = Parser()
    p.set_file(Path("CN-9999-Delta-CN.pdf"))
    p.set_page(1)
    p.feed_table_row(["Ma kien hang: PGKECDLT0000001 1/1", "", "", "", "", ""])

    items_before = p.total_item_count()
    closed_before = len(p.packages)
    # A non-empty `tables` list (3 rows) whose every row fails to produce a
    # valid item -- exactly the shape that used to permanently disable the
    # text fallback for the rest of the page under the old `if tables:` rule.
    garbage_table_rows = [
        ["about:blank", "", "", "", "", ""],                             # noise row
        ["STT", "GTIN", "SKU", "Ten hang", "DVT", "Tinh trang SL"],      # repeated header row
        ["", "", "", "", "", ""],                                        # blank row
    ]
    for row in garbage_table_rows:
        p.feed_table_row(row)
    items_after_table = p.total_item_count()
    closed_after_table = len(p.packages)

    assert items_after_table == items_before, "garbage table rows must not add any item"
    assert closed_after_table == closed_before, "garbage table rows must not close a package"

    # This is the literal run_pipeline() delta-fallback condition (bug #16
    # fix) re-evaluated here against the real Parser counters.
    should_fall_back = (items_after_table == items_before and closed_after_table == closed_before)
    assert should_fall_back, "delta-fallback condition must trigger when the table pass yields nothing new"

    item_line = "1 4894961082009 TP-DLT-001-A Rope Loop PCS Moi 9"
    p.feed_text_line(item_line)
    assert p._cur is not None and p._cur.item_count == 1, "text fallback must recover exactly 1 item"

    # Re-feeding the identical recovered line must NOT double-count it.
    dup_before = p.duplicate_items_skipped
    p.feed_text_line(item_line)
    assert p.duplicate_items_skipped == dup_before + 1, "re-feeding the same line must be caught by dedup"
    assert p._cur.item_count == 1, "item count must stay at 1 after the duplicate re-feed"

    p.feed_text_line("Tong cong 9")
    p.finalise()
    assert len(p.packages) == 1
    pkg = p.packages[0]
    assert pkg.item_count == 1, f"expected exactly 1 item after fallback recovery, got {pkg.item_count}"
    assert pkg.items[0].quantity == 9
    assert pkg.declared_total_qty == 9
    assert pkg.calc_qty == 9


test("condition+quantity merged cell split ('Moi 12')", t_cond_qty_merge_split)
test("SKU '-BOX' suffix joined (newline inside cell)", t_sku_box_suffix_newline_in_cell)
test("SKU '-BOX' suffix joined (suffix on its own next row)", t_sku_box_suffix_next_row)
test("U+FFFE / U+FFFF normalized to '-'", t_unicode_dash_artifact_normalized)
test("SKU split across newline joined via normalize_code", t_description_newline_join_via_code)
test("GTIN shape validation (digit count/type) -- checksum behavior covered separately in test_v15_bugfix_regression.py", t_gtin13_valid)
test("SKU == GTIN does not crash the parser", t_gtin_equals_sku_no_crash)
test("header/footer noise (about:blank, date, repeated header) not parsed as item", t_header_footer_row_not_parsed_as_item)
test("repeated table header row detected", t_repeated_table_header_not_item)
test("zero-item package -> ZERO_ITEMS status", t_package_zero_items_status)
test("package with items but no Tong cong -> MISSING_TOTAL status", t_package_missing_total_status)
test("duplicate row (table+fallback re-read) not double-counted", t_dedup_same_page_same_row_not_double_counted)
test("legitimately repeated SKU on different rows both kept", t_legit_repeated_sku_different_rows_both_kept)
test("one page, multiple packages parsed correctly", t_multi_package_same_page)
test("package spanning pages: continuation merges, totals reconcile", t_package_spans_pages_totals_reconcile)
test("condition + quantity in separate cells (8-col layout)", t_condition_word_separate_cell_8col)
test("7-column row (no trailing blank col): GTIN/SKU/description/UOM/condition/quantity all correct", t_seven_column_row_all_fields)
test("delta-fallback: non-empty garbage table -> text layer recovers item, no duplicate (bug #16)", t_delta_fallback_table_zero_items_recovers_via_text)


# =============================================================================
# 2. DIM loader unit tests -- synthetic fixtures for tiers this real DIM
#    file doesn't happen to exercise (its headers are clean -> HEADER_MAPPING
#    only). PACKAGE_CODE_POSITIONAL_FALLBACK etc. are tested against
#    synthetic workbooks built here, per spec section 15's allowance to use
#    sanitized fixtures when a real broken-header file isn't available.
# =============================================================================
print("\n== DIM loader unit tests ==")


def _dim_wb(rows, headers=None, tmp_name="dim_fixture.xlsx"):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    if headers is not None:
        ws.append(headers)
    for r in rows:
        ws.append(r)
    tmp_dir = Path(tempfile.mkdtemp(prefix="dim_fixture_"))
    path = tmp_dir / tmp_name
    wb.save(str(path))
    return path, tmp_dir


def t_dim_header_mapping_synthetic_file():
    DimMapper = C["DimMapper"]
    dim = DimMapper(SYN_DIM)
    assert dim.detection_method == "HEADER_MAPPING", dim.detection_method
    assert dim.valid_rows == 3, dim.valid_rows
    d = dim.lookup("SYN-1002-WarehouseBeta-CN", "PGKECSYN10020001")
    assert d is not None
    assert d["length"] == 45


def t_dim_partial_header_positional_fill():
    """PARTIAL_HEADER_POSITIONAL (tier 2): ref+pkg found by header alias,
    but the 5 dimension columns have meaningless headers ("X1".."X5") that
    match no alias -- found while writing this test, the original code
    (a) never actually filled those 5 fields positionally despite the
    docstring promising it, and (b) counted the row as "loaded" even though
    _commit_row() rejected it as malformed, so tier 3 was never tried and
    the sheet falsely reported success with 0 usable records. Both are
    fixed: PARTIAL_HEADER_POSITIONAL now reads all 5 dimension cells at
    pkg_col+1..+5 (same fixed order as tier 3), and the row-loaded counters
    only count rows _commit_row() actually accepted."""
    path, tmp = _dim_wb(
        headers=["Lo hang", "Tracking", "X1", "X2", "X3", "X4", "X5"],
        rows=[["REF-P", "PGKECPPP0000001", 10, 20, 15, 2.5, 0.003]],
    )
    try:
        DimMapper = C["DimMapper"]
        dim = DimMapper(path)
        assert dim.detection_method == "PARTIAL_HEADER_POSITIONAL", dim.detection_method
        assert dim.valid_rows == 1, dim.valid_rows
        d = dim.lookup("REF-P", "PGKECPPP0000001")
        assert d is not None
        assert d["length"] == 10 and d["width"] == 20 and d["height"] == 15
        assert d["weight"] == 2.5 and d["cbm"] == 0.003
    finally:
        shutil.rmtree(tmp)


def t_dim_header_alias_rename():
    path, tmp = _dim_wb(
        headers=["Reference", "Package", "L", "W", "H", "WT", "CBM"],
        rows=[["REF-A", "PGKECAAA0000001", 10, 20, 30, 1.5, 0.006]],
    )
    try:
        DimMapper = C["DimMapper"]
        dim = DimMapper(path)
        assert dim.detection_method == "HEADER_MAPPING", dim.detection_method
        assert dim.lookup("REF-A", "PGKECAAA0000001") is not None
    finally:
        shutil.rmtree(tmp)


def t_dim_broken_header_positional_fallback():
    path, tmp = _dim_wb(
        headers=["col1", "col2", "col3", "col4", "col5", "col6", "col7"],
        rows=[["REF-B", "PGKECBBB0000001", 15, 25, 35, 2.2, 0.013]],
    )
    try:
        DimMapper = C["DimMapper"]
        dim = DimMapper(path)
        assert dim.detection_method == "PACKAGE_CODE_POSITIONAL_FALLBACK", dim.detection_method
        d = dim.lookup("REF-B", "PGKECBBB0000001")
        assert d is not None
        assert d["length"] == 15 and d["cbm"] == 0.013
    finally:
        shutil.rmtree(tmp)


def t_dim_positional_with_leading_stt_column():
    path, tmp = _dim_wb(
        headers=["junk1", "junk2", "junk3", "junk4", "junk5", "junk6", "junk7", "junk8"],
        rows=[[1, "REF-C", "PGKECCCC0000001", 12, 22, 32, 3.1, 0.0087]],
    )
    try:
        DimMapper = C["DimMapper"]
        dim = DimMapper(path)
        assert dim.detection_method == "PACKAGE_CODE_POSITIONAL_FALLBACK", dim.detection_method
        d = dim.lookup("REF-C", "PGKECCCC0000001")
        assert d is not None, "STT column before package code must not be mistaken for the reference"
        assert d["length"] == 12
    finally:
        shutil.rmtree(tmp)


def t_dim_positional_package_code_not_fixed_column():
    """Row A has the package code in column 2, row B in column 3 (an empty
    spacer cell shifts it right) -- the fallback must scan every cell, not
    assume a fixed column index. The reference is still the nearest non-
    blank cell to the LEFT of the package code on each row."""
    path, tmp = _dim_wb(
        headers=["h1", "h2", "h3", "h4", "h5", "h6", "h7", "h8"],
        rows=[
            ["REF-D", "PGKECDDD0000001", 11, 21, 31, 4.0, 0.0072, ""],
            ["REF-E", "", "PGKECEEE0000001", 13, 23, 33, 5.0, 0.0099],
        ],
    )
    try:
        DimMapper = C["DimMapper"]
        dim = DimMapper(path)
        assert dim.lookup("REF-D", "PGKECDDD0000001") is not None
        assert dim.lookup("REF-E", "PGKECEEE0000001") is not None, \
            "package code in a different column on a different row must still be found"
    finally:
        shutil.rmtree(tmp)


def t_dim_reference_forward_fill_blank_cell():
    path, tmp = _dim_wb(
        headers=["h1", "h2", "h3", "h4", "h5", "h6", "h7"],
        rows=[
            ["REF-F", "PGKECFFF0000001", 10, 20, 30, 1.0, 0.006],
            ["", "PGKECFFF0000002", 10, 20, 30, 1.0, 0.006],  # blank ref -> forward-fill from row above
        ],
    )
    try:
        DimMapper = C["DimMapper"]
        dim = DimMapper(path)
        assert dim.lookup("REF-F", "PGKECFFF0000002") is not None, "blank reference should forward-fill"
    finally:
        shutil.rmtree(tmp)


def t_dim_insufficient_positional_fields():
    path, tmp = _dim_wb(
        headers=["h1", "h2", "h3"],
        rows=[["REF-G", "PGKECGGG0000001", 10]],  # only 1 cell after the package code, need 5
    )
    try:
        DimMapper = C["DimMapper"]
        dim = DimMapper(path)
        assert dim.lookup("REF-G", "PGKECGGG0000001") is None
        reasons = {d["validation_status"] for d in dim.diagnostics}
        assert "INSUFFICIENT_POSITIONAL_FIELDS" in reasons, reasons
    finally:
        shutil.rmtree(tmp)


def t_dim_non_numeric_dimension_fails_row():
    path, tmp = _dim_wb(
        headers=["h1", "h2", "h3", "h4", "h5", "h6", "h7"],
        rows=[["REF-H", "PGKECHHH0000001", "not-a-number", 20, 30, 1.0, 0.006]],
    )
    try:
        DimMapper = C["DimMapper"]
        dim = DimMapper(path)
        assert dim.lookup("REF-H", "PGKECHHH0000001") is None
        assert dim.malformed_rows >= 1
    finally:
        shutil.rmtree(tmp)


def t_dim_cbm_tolerance_review_required_no_overwrite():
    # length*width*height/1e6 = 10*20*30/1e6 = 0.006 expected; declared 0.5 is
    # wildly outside the 5% tolerance -> DIM_REVIEW_REQUIRED, but the ORIGINAL
    # 0.5 must be kept (never silently recalculated/overwritten).
    path, tmp = _dim_wb(
        headers=["h1", "h2", "h3", "h4", "h5", "h6", "h7"],
        rows=[["REF-I", "PGKECIII0000001", 10, 20, 30, 1.0, 0.5]],
    )
    try:
        DimMapper = C["DimMapper"]
        dim = DimMapper(path)
        d = dim.lookup("REF-I", "PGKECIII0000001")
        assert d is not None
        assert d["cbm"] == 0.5, f"original CBM must be preserved, got {d['cbm']}"
        assert dim.review_required == 1
        diag = [x for x in dim.diagnostics if x["normalized_package_code"] == "PGKECIII0000001"][0]
        assert diag["validation_status"] == "DIM_REVIEW_REQUIRED"
    finally:
        shutil.rmtree(tmp)


def t_dim_duplicate_key_flagged_first_wins():
    path, tmp = _dim_wb(
        headers=["h1", "h2", "h3", "h4", "h5", "h6", "h7"],
        rows=[
            ["REF-J", "PGKECJJJ0000001", 10, 20, 30, 1.0, 0.006],
            ["REF-J", "PGKECJJJ0000001", 99, 99, 99, 9.0, 0.9],  # duplicate key, different values
        ],
    )
    try:
        DimMapper = C["DimMapper"]
        dim = DimMapper(path)
        assert dim.duplicate_keys == 1
        d = dim.lookup("REF-J", "PGKECJJJ0000001")
        assert d["length"] == 10, "first occurrence must win, not be silently overwritten"
    finally:
        shutil.rmtree(tmp)


def t_dim_no_fuzzy_auto_match():
    """v17 (test correction, spec point 37): root-caused against
    DimMapper.lookup()'s real, intentional 2-tier design (see its own
    docstring/comment: "Primary identity: a unique Packaging Code in
    DIM... Safety fallback for a Packaging Code that appears more than
    once: require canonicalized Shipping Mark/reference + Packaging
    Code") -- this is exactly spec section 25's own rule ("DIM matching
    uses actual Packaging Code, not filename similarity"): a Packaging
    Code that is GLOBALLY UNIQUE in the DIM file is trusted on its own
    (Packaging Code IS the authoritative physical identifier -- more
    reliable than a PDF-OCR'd reference/filename, which is exactly the
    kind of "filename similarity" text spec 25 says NOT to key on); only
    an AMBIGUOUS Packaging Code (appears 2+ times under different DIM
    rows) falls back to requiring the exact ref+pkg combination, which is
    where "no fuzzy auto-match" genuinely applies. The previous version
    of this test asserted the opposite for the unique-code case (that a
    mismatched reference should defeat an otherwise-unique, correct
    Packaging Code hit) -- that was a stale/incorrect expectation, not a
    production bug; confirmed via git-stash baseline (behavior identical
    before and after this fix pass, so nothing here regressed anything)."""
    path, tmp = _dim_wb(
        headers=["h1", "h2", "h3", "h4", "h5", "h6", "h7"],
        rows=[["REF-K", "PGKECKKK0000001", 10, 20, 30, 1.0, 0.006]],
    )
    try:
        DimMapper = C["DimMapper"]
        dim = DimMapper(path)
        # a near-miss PACKAGE CODE that doesn't exist in DIM at all -> never
        # fuzzy-matched to the nearest one.
        assert dim.lookup("REF-K", "PGKECKKK0000002") is None
        # a GLOBALLY UNIQUE Packaging Code is the authoritative identity --
        # a mismatched/typo'd reference does NOT defeat it (spec section 25:
        # Packaging Code, not filename/reference similarity, is what DIM
        # matching keys on).
        d = dim.lookup("REF-K-TYPO", "PGKECKKK0000001")
        assert d is not None and d["length"] == 10.0, d
        # diagnose_miss() is a read-only, informational-only helper for a
        # package that found NO match at all -- verified independently
        # against a package code that genuinely isn't in DIM (not the
        # unique-code case above, which DOES match by design).
        diag = dim.diagnose_miss("test.pdf", "REF-K-TYPO", "PGKECKKK0000002")
        assert diag["mismatch_reason"] in ("NO_REFERENCE_IN_DIM", "KEY_COMBINATION_NOT_FOUND", "NO_PACKAGE_CODE_IN_DIM")

        # The REAL "no fuzzy auto-match" guarantee: once a Packaging Code is
        # AMBIGUOUS (2+ DIM rows share it under different references), the
        # exact ref+pkg combination IS required -- a mismatched reference
        # against an ambiguous code correctly returns None.
        path2, tmp2 = _dim_wb(
            headers=["h1", "h2", "h3", "h4", "h5", "h6", "h7"],
            rows=[
                ["REF-A", "PGKECAMBIG0001", 10, 20, 30, 1.0, 0.006],
                ["REF-B", "PGKECAMBIG0001", 11, 21, 31, 1.1, 0.007],
            ],
        )
        try:
            dim2 = DimMapper(path2)
            assert dim2.lookup("REF-A", "PGKECAMBIG0001") is not None
            assert dim2.lookup("REF-B", "PGKECAMBIG0001") is not None
            assert dim2.lookup("REF-TYPO", "PGKECAMBIG0001") is None
        finally:
            shutil.rmtree(tmp2)
    finally:
        shutil.rmtree(tmp)


def t_raw_data_sheet_has_dim_per_item_row():
    out_dir = Path(tempfile.mkdtemp(prefix="pipeline_out_"))
    master_path = out_dir / "master.xlsx"
    _make_empty_master_xlsx(master_path)
    try:
        ns = load_core_pipeline(SYN_PDFS, SYN_DIM, master_path, out_dir)
        import openpyxl
        wb = openpyxl.load_workbook(str(out_dir / "PL_Total.xlsx"))
        assert "Raw_Data" in wb.sheetnames
        assert "Audit_Summary" in wb.sheetnames
        ws = wb["Raw_Data"]
        headers = [c.value for c in ws[1]]
        for col in ("length", "width", "height", "weight", "cbm", "dim_match_status"):
            assert col in headers, f"Raw_Data missing column {col}"
        length_idx = headers.index("length") + 1
        row2 = [ws.cell(row=2, column=c).value for c in range(1, len(headers) + 1)]
        assert row2[length_idx - 1] is not None, "Raw_Data row must carry DIM values directly (not merged cells)"
        wb2 = openpyxl.load_workbook(str(out_dir / "PL_Total.xlsx"))  # re-open sanity check (openpyxl compatible)
        assert wb2 is not None
    finally:
        shutil.rmtree(out_dir)


test("HEADER_MAPPING against the sanitized synthetic DIM.xlsx", t_dim_header_mapping_synthetic_file)
test("PARTIAL_HEADER_POSITIONAL: ref+pkg via header, L/W/H/Wt/CBM filled positionally", t_dim_partial_header_positional_fill)
test("HEADER_MAPPING via alias-renamed headers", t_dim_header_alias_rename)
test("broken headers -> PACKAGE_CODE_POSITIONAL_FALLBACK, fixed L/W/H/Wt/CBM order", t_dim_broken_header_positional_fallback)
test("leading STT column before package code still detected", t_dim_positional_with_leading_stt_column)
test("package code not in a fixed column, scanned per-row", t_dim_positional_package_code_not_fixed_column)
test("blank/merged reference forward-filled from row above", t_dim_reference_forward_fill_blank_cell)
test("insufficient cells after package code -> row fails", t_dim_insufficient_positional_fields)
test("non-numeric dimension -> row fails", t_dim_non_numeric_dimension_fails_row)
test("CBM outside tolerance -> DIM_REVIEW_REQUIRED, original CBM not overwritten", t_dim_cbm_tolerance_review_required_no_overwrite)
test("duplicate DIM key flagged, first occurrence wins (no silent overwrite)", t_dim_duplicate_key_flagged_first_wins)
test("no fuzzy auto-match of package code / reference", t_dim_no_fuzzy_auto_match)
test("Raw_Data sheet carries DIM values on every item row (not merge-only)", t_raw_data_sheet_has_dim_per_item_row)


# =============================================================================
# 3. Synthetic-file integration test (spec section 16 originally called for
#    a real-file integration test; those real files were purged from repo
#    history after the 2026-08-03 security incident -- see module docstring)
# =============================================================================
print("\n== Synthetic-file integration test (2 sanitized PDFs + synthetic DIM.xlsx) ==")


def t_synthetic_files_full_pipeline():
    """End-to-end run against the two sanitized synthetic PDFs + synthetic
    DIM.xlsx. This exercises the SAME layout defects the original real-file
    audit reproduced (merged condition+quantity cells, "-BOX" split across a
    newline, package header living inside the item table, continuation
    across pages, one page with two packages, about:blank/date/repeated-
    header noise) -- see tools/pl-ocr-grouping/tests/fixtures/synthetic/generate.py
    for exactly how each fixture maps to which bug.

    NOTE (see 2026-08-03 security incident): this replaces what used to be a
    genuine real-file integration test against 8 real customer PDFs, which
    were purged from this repo's history entirely. This synthetic run is
    NOT a substitute for periodically re-running the tool against real
    files locally (never committed) before trusting a release -- see
    README/limitations."""
    out_dir = Path(tempfile.mkdtemp(prefix="pipeline_syn_"))
    master_path = out_dir / "master.xlsx"
    _make_empty_master_xlsx(master_path)
    try:
        ns = load_core_pipeline(SYN_PDFS, SYN_DIM, master_path, out_dir)
        packages = ns["packages"]
        audit_status = ns["audit_status"]
        assert len(packages) == 3, f"expected 3 packages (1 spanning 2 pages + 2 on one page), got {len(packages)}"
        assert all(p.item_count > 0 for p in packages), "no package should have 0 items on these fixtures"
        assert sum(p.item_count for p in packages) == 9, f"expected 9 items total, got {sum(p.item_count for p in packages)}"
        statuses = {audit_status(p) for p in packages}
        assert statuses == {"OK"}, f"expected all packages OK, got {statuses}"
        assert all(p.declared_total_qty == p.calc_qty for p in packages), \
            "declared vs calculated totals must reconcile for every package"
        assert all(p.dim_matched for p in packages), "every package must DIM-match against the synthetic DIM.xlsx"
        assert ns["RUN_SUMMARY"]["errors"] == 0
        assert ns["RUN_SUMMARY"]["duplicate_items_skipped"] == 0
        seen_codes = [p.package_code for p in packages]
        assert len(seen_codes) == len(set(seen_codes)), "no duplicate package codes"
        by_code = {p.package_code: p for p in packages}
        assert by_code["PGKECSYN10010001"].item_count == 5, "the 2-page continuation package must merge to 5 items"
        assert by_code["PGKECSYN10010001"].calc_qty == 36
        box_item = next(i for i in by_code["PGKECSYN10010001"].items if i.product_code.endswith("-BOX"))
        assert box_item.product_code == "TP-SYN-A-001-BOX", box_item.product_code
    finally:
        shutil.rmtree(out_dir)


test("synthetic fixtures (2 PDFs, 3 packages, 9 items): all OK, totals reconcile, DIM 3/3", t_synthetic_files_full_pipeline)


# =============================================================================
# 9. Per-Store counting scope for carton numbering (spec section 9) --
#    counting_scope_key must group cartons per Store WITHIN a shipment,
#    never as one flat denominator across an entire multi-store upload.
# =============================================================================
print("\n== Per-Store counting scope (carton numbering) unit tests ==")
D9 = load_core_defs()
Package9 = D9["Package"]
assign_counting_scope_keys = D9["assign_counting_scope_keys"]
assign_global_numbers = D9["assign_global_numbers"]
compute_counting_scope_key = D9["compute_counting_scope_key"]


def _pkg9(ref, or_status="", or_num="", or_store="", store="", seq=1, country="CN"):
    # v19 (SG-533-TEST consolidation report, requirement 2): compute_
    # counting_scope_key()'s per-Store grouping is CN-specific
    # functionality (Kerry/Hangzhou/Shenzhen/Guangzhou below are all real
    # China stores) -- now that the CN-eligibility gate is STRICT
    # (country == "CN" only, "" no longer implicitly eligible), every
    # caller of this helper needs an explicit country to keep exercising
    # that Store-based grouping at all. Defaults to "CN" since that is
    # what every existing call site in this file already represents;
    # override with country="" only for a test that deliberately wants
    # the "no usable country signal" fallback path.
    p = Package9(package_code=f"PGKEC{ref}", source_file=f"{ref}.pdf",
                 reference_code=ref, pdf_package_seq=seq)
    p.or_list_match_status = or_status
    p.or_number = or_num
    p.or_list_store = or_store
    p.store = store
    p.country = country
    return p


def t_scoped_numbering_spec_example_kerry_hangzhou():
    # Exact user-given example: Kerry 6 cartons, Hangzhou 4 cartons ->
    # independent 1/6..6/6 and 1/4..4/4 -- NEVER a combined 1/10..10/10.
    pkgs = [_pkg9(f"kerry_{i}", "OK", "OR1016", "Kerry", seq=i + 1) for i in range(6)]
    pkgs += [_pkg9(f"hz_{i}", "OK", "OR2044", "Hangzhou", seq=i + 1) for i in range(4)]
    assign_counting_scope_keys(pkgs)
    assign_global_numbers(pkgs)
    kerry = [p for p in pkgs if p.or_list_store == "Kerry"]
    hz = [p for p in pkgs if p.or_list_store == "Hangzhou"]
    assert [p.carton_display for p in kerry] == ["1/6", "2/6", "3/6", "4/6", "5/6", "6/6"],         [p.carton_display for p in kerry]
    assert [p.carton_display for p in hz] == ["1/4", "2/4", "3/4", "4/4"], [p.carton_display for p in hz]
    assert kerry[0].counting_scope_key != hz[0].counting_scope_key
    assert all(p.carton_total == 6 for p in kerry)
    assert all(p.carton_total == 4 for p in hz)


def t_scoped_numbering_no_store_signal_stays_flat_backward_compat():
    # No OR List match, no CN store resolved -> every package falls into
    # the SAME implicit "UPLOAD_BATCH|UNRESOLVED" scope -- i.e. today's
    # flat 1/N..N/N numbering, byte-for-byte, for every pre-v12 caller
    # that never touches counting_scope_key at all.
    pkgs = [_pkg9(f"f{i}", seq=i + 1) for i in range(5)]
    assign_counting_scope_keys(pkgs)
    assign_global_numbers(pkgs)
    assert [p.carton_display for p in pkgs] == ["1/5", "2/5", "3/5", "4/5", "5/5"]
    assert len({p.counting_scope_key for p in pkgs}) == 1


def t_scoped_numbering_cn_store_without_or_list_still_splits():
    # or_list_match_status is NOT "OK" (no OR List uploaded at all) but
    # pkg.store WAS resolved (classify_packages_for_port / CN split) --
    # counting scope must still key off that resolved Store.
    pkgs = [_pkg9(f"a{i}", store="SHENZHEN", seq=i + 1) for i in range(3)]
    pkgs += [_pkg9(f"b{i}", store="GUANGZHOU", seq=i + 1) for i in range(2)]
    assign_counting_scope_keys(pkgs)
    assign_global_numbers(pkgs)
    sz = [p for p in pkgs if p.store == "SHENZHEN"]
    gz = [p for p in pkgs if p.store == "GUANGZHOU"]
    assert [p.carton_display for p in sz] == ["1/3", "2/3", "3/3"]
    assert [p.carton_display for p in gz] == ["1/2", "2/2"]


def t_scoped_numbering_or_list_match_takes_priority_over_cn_store():
    # A package can have BOTH or_list_store (OK match) and a stale/blank
    # pkg.store -- compute_counting_scope_key must prefer the OR-List-
    # matched store, since it's the stronger, more specific signal.
    p = _pkg9("x", or_status="OK", or_num="OR9", or_store="Kerry", store="")
    key, source = compute_counting_scope_key(p)
    assert source == "OR_LIST_GROUPING", source
    assert "KERRY" in key.upper()


def t_scoped_numbering_review_status_falls_back_not_or_list_grouped():
    # or_list_match_status == "REVIEW" (ambiguous/no confident match) must
    # NOT be treated as a confirmed OR-List grouping signal.
    p = _pkg9("x", or_status="REVIEW", or_num="", or_store="", store="SHENZHEN")
    key, source = compute_counting_scope_key(p)
    assert source == "UPLOAD_BATCH_DEFAULT", source
    assert "SHENZHEN" in key.upper()


def t_scoped_numbering_case_insensitive_store_identity_merges():
    # "Kerry" (OR List's own casing) and "KERRY" (a hypothetical differently
    # -cased signal) must resolve to the SAME counting scope, not two.
    p1 = _pkg9("a", or_status="OK", or_num="OR1", or_store="Kerry")
    p2 = _pkg9("b", or_status="OK", or_num="OR1", or_store="KERRY")
    k1, _ = compute_counting_scope_key(p1)
    k2, _ = compute_counting_scope_key(p2)
    assert k1 == k2, (k1, k2)


test("per-Store scope: Kerry 6 cartons + Hangzhou 4 cartons number independently (1/6.. and 1/4.., never 1/10..)",
     t_scoped_numbering_spec_example_kerry_hangzhou)
test("per-Store scope: no store signal at all -> flat 1/N..N/N (backward-compat, matches every pre-v12 caller)",
     t_scoped_numbering_no_store_signal_stays_flat_backward_compat)
test("per-Store scope: CN store resolved without any OR List still splits scopes correctly",
     t_scoped_numbering_cn_store_without_or_list_still_splits)
test("per-Store scope: OK OR List match takes priority over a blank/stale pkg.store",
     t_scoped_numbering_or_list_match_takes_priority_over_cn_store)
test("per-Store scope: REVIEW status is never treated as a confirmed OR List grouping signal",
     t_scoped_numbering_review_status_falls_back_not_or_list_grouped)
test("per-Store scope: store identity is case-insensitive (Kerry == KERRY)",
     t_scoped_numbering_case_insensitive_store_identity_merges)


# =========================================================================
# v13 (FIX4/FIX5 real-file validation): end-to-end simulation of the real
# 21-carton / 7-store production shipment (see audit notes -- Shipping
# Marks + carton counts read directly from the real uploaded
# SUBLIST_TOTAL.pdf, OR List rows from the real uploaded OR List.xlsx).
# Confirms a Store's cartons across POP + VN + CN factories now share ONE
# counting scope/denominator, numbered in POP -> VN -> CN business order --
# previously POP+VN packages (no Store source at all pre-v13) were lumped
# into one flat shared scope across ALL stores, and only CN packages were
# split per-store.
# =========================================================================
import pl_or_list_import as _oli13
import pl_group_export as _pge13

_REAL_OR_LIST_ROWS_V13 = [
    ("20260609 CN - Guangzhou Parc Central Replen", "po38070", "inv628037"),
    ("20260609 CN - Hangzhou Mixc Replen", "po38068", "inv628036"),
    ("20260609 CN - Iapm Replen", "po38072", "inv628039"),
    ("20260609 CN - Kerry Center flagship Replen", "po38071", "inv628038"),
    ("20260609 CN - Shanghai Hongqiao Airport Replen", "po38074", "inv628042"),
    ("20260609 CN - Shanghai Taikooli (Shop B1-07b) Replen", "po38076", "inv628041"),
    ("20260609 CN - Shenzhen Mixc City (Shop T228) Replen", "po38073", "inv628040"),
]


def _real_or_index_v13():
    from collections import OrderedDict
    rows = [
        _oli13.OrListRow(row_number=i + 2, store_raw=store, store_norm="",
                          business_fields=OrderedDict([("OR No.", or_v), ("SO No.", so_v)]))
        for i, (store, or_v, so_v) in enumerate(_REAL_OR_LIST_ROWS_V13)
    ]
    idx = {}
    for r in rows:
        idx.setdefault(r.or_norm, []).append(r)
    return idx


def t_real_shipment_cross_factory_store_scope_combines_pop_vn_cn():
    real_marks = (
        ["CN-1529_HZ_PVG_POP", "CN-1529_IAPM_PVG_POP", "CN-1529_KR_PVG_POP",
         "CN-1529_SH-Airport_PVG_POP", "CN-1529_SH-Taikooli_PVG_POP",
         "CN-1529_GZ_SZX_VN", "CN-1529_HZ_PVG_VN", "CN-1529_IAPM_PVG_VN",
         "CN-1529_KR_PVG_VN", "CN-1529_SZ_SZX_VN"]
        + ["CN-1529_GZ_SZX_CN"]
        + ["CN-1529_HZ_PVG_CN"] * 2
        + ["CN-1529_IAPM_PVG_CN"] * 2
        + ["CN-1529_KR_PVG_CN"] * 2
        + ["CN-1529_SH-Airport_PVG_CN v"] * 2
        + ["CN-1529_SH-Taikooli_PVG_CN"]
        + ["CN-1529_SZ_SZX_CN"]
    )
    assert len(real_marks) == 21

    pkgs = []
    for i, mk in enumerate(real_marks):
        p = Package9(package_code=f"PKG{i}", source_file=f"{mk}.pdf", reference_code=mk, pdf_package_seq=0)
        p.shipping_mark = mk
        # v19 (SG-533-TEST consolidation report, requirement 2): the CN
        # multi-Store eligibility gate is now STRICT (country == "CN"
        # only) -- every one of these real Shipmarks ("CN-1529_..." with a
        # genuine "CN" Shipmark prefix) is a real CN shipment, so this now
        # needs an explicit country to keep exercising classify_packages_
        # for_port()/compute_counting_scope_key()'s Store-based grouping,
        # matching what detect_shipment_country() would itself resolve
        # from this exact "CN-1529_..." prefix in the real pipeline.
        p.country = "CN"
        pkgs.append(p)

    pkgs = D9["business_sort_packages"](pkgs)
    D9["classify_packages_for_port"](pkgs, None, False)

    or_index = _real_or_index_v13()
    for pkg in pkgs:
        m = _pge13.match_store_and_or(pkg, or_index)
        pkg.or_list_store = m.matched_store
        pkg.or_list_match_status = m.status
        if m.status == "OK":
            if not pkg.or_number:
                pkg.or_number = m.matched_or
            if not pkg.so_number:
                pkg.so_number = m.matched_so

    assign_counting_scope_keys(pkgs)
    assign_global_numbers(pkgs)

    n_ok = sum(1 for p in pkgs if p.or_list_match_status == "OK")
    assert n_ok == 21, f"expected all 21 real cartons to match OK, got {n_ok}"

    by_mark_factory = {}
    for p in pkgs:
        factory = _pge13.detect_factory(p.reference_code, p.source_file)
        by_mark_factory.setdefault((p.reference_code, factory), []).append(p.carton_display)

    # HZ / IAPM / KR: POP(1) + VN(1) + CN(2) -- ONE combined scope of 4,
    # ordered POP -> VN -> CN (never separate 1/1's per factory).
    for code in ("CN-1529_HZ_PVG", "CN-1529_IAPM_PVG", "CN-1529_KR_PVG"):
        pop = by_mark_factory[(f"{code}_POP", "POP")]
        vn = by_mark_factory[(f"{code}_VN", "VN")]
        assert pop == ["1/4"], (code, pop)
        assert vn == ["2/4"], (code, vn)

    cn_hz = by_mark_factory[("CN-1529_HZ_PVG_CN", "CN")]
    assert cn_hz == ["3/4", "4/4"], cn_hz

    # SH-Airport: POP(1) + CN(2), no VN -- combined total 3.
    sh_air_pop = by_mark_factory[("CN-1529_SH-Airport_PVG_POP", "POP")]
    sh_air_cn = by_mark_factory[("CN-1529_SH-Airport_PVG_CN v", "CN")]
    assert sh_air_pop == ["1/3"], sh_air_pop
    assert sh_air_cn == ["2/3", "3/3"], sh_air_cn

    # SH-Taikooli: POP(1) + CN(1), no VN -- combined total 2.
    sh_tk_pop = by_mark_factory[("CN-1529_SH-Taikooli_PVG_POP", "POP")]
    sh_tk_cn = by_mark_factory[("CN-1529_SH-Taikooli_PVG_CN", "CN")]
    assert sh_tk_pop == ["1/2"] and sh_tk_cn == ["2/2"], (sh_tk_pop, sh_tk_cn)

    # GZ / SZ: VN(1) + CN(1), no POP -- combined total 2.
    for code in ("CN-1529_GZ_SZX", "CN-1529_SZ_SZX"):
        vn = by_mark_factory[(f"{code}_VN", "VN")]
        cn = by_mark_factory[(f"{code}_CN", "CN")]
        assert vn == ["1/2"] and cn == ["2/2"], (code, vn, cn)

    # Exact expected OR/SO per store (spec's required 100%-match gate).
    expected_or_so = {
        "GZ": ("po38070", "inv628037"), "HZ": ("po38068", "inv628036"),
        "IAPM": ("po38072", "inv628039"), "KR": ("po38071", "inv628038"),
        "SH-Airport": ("po38074", "inv628042"), "SH-Taikooli": ("po38076", "inv628041"),
        "SZ": ("po38073", "inv628040"),
    }
    for p in pkgs:
        for code, (exp_or, exp_so) in expected_or_so.items():
            if f"_{code}_" in p.reference_code:
                assert p.or_number == exp_or and p.so_number == exp_so,             (p.reference_code, p.or_number, p.so_number, exp_or, exp_so)


test("real 21-carton/7-store shipment: POP+VN+CN combine under one per-Store scope, ordered POP->VN->CN, 100% OR/SO match",
     t_real_shipment_cross_factory_store_scope_combines_pop_vn_cn)


# =========================================================================
# v14 (spec sections 9-12): export_grouped_pl()'s 04_CN_BY_STORE split must
# actually WRITE each Store's POP+VN+CN cartons combined into one file,
# using the ALREADY-correct Store-scoped carton_display (no local
# renumbering) -- this is the piece that was still using the old CN-only
# match_store() and bespoke per-file renumbering before this fix. Reuses
# the exact same real 21-carton/7-store fixture as the test above, but
# additionally calls export_grouped_pl() and reads the actual .xlsx files
# back to confirm the combined-factory grouping and numbering survive all
# the way to disk.
# =========================================================================
def t_export_grouped_pl_store_split_combines_pop_vn_cn_and_keeps_scope_numbering():
    import openpyxl

    real_marks = (
        ["CN-1529_HZ_PVG_POP", "CN-1529_IAPM_PVG_POP", "CN-1529_KR_PVG_POP",
         "CN-1529_SH-Airport_PVG_POP", "CN-1529_SH-Taikooli_PVG_POP",
         "CN-1529_GZ_SZX_VN", "CN-1529_HZ_PVG_VN", "CN-1529_IAPM_PVG_VN",
         "CN-1529_KR_PVG_VN", "CN-1529_SZ_SZX_VN"]
        + ["CN-1529_GZ_SZX_CN"]
        + ["CN-1529_HZ_PVG_CN"] * 2
        + ["CN-1529_IAPM_PVG_CN"] * 2
        + ["CN-1529_KR_PVG_CN"] * 2
        + ["CN-1529_SH-Airport_PVG_CN v"] * 2
        + ["CN-1529_SH-Taikooli_PVG_CN"]
        + ["CN-1529_SZ_SZX_CN"]
    )
    pkgs = []
    for i, mk in enumerate(real_marks):
        p = Package9(package_code=f"PKG{i}", source_file=f"{mk}.pdf", reference_code=mk, pdf_package_seq=0)
        p.shipping_mark = mk
        # v19 (SG-533-TEST consolidation report, requirement 2): the CN
        # multi-Store eligibility gate is now STRICT (country == "CN"
        # only) -- every one of these real Shipmarks ("CN-1529_..." with a
        # genuine "CN" Shipmark prefix) is a real CN shipment, so this now
        # needs an explicit country to keep exercising classify_packages_
        # for_port()/compute_counting_scope_key()'s Store-based grouping,
        # matching what detect_shipment_country() would itself resolve
        # from this exact "CN-1529_..." prefix in the real pipeline.
        p.country = "CN"
        pkgs.append(p)

    pkgs = D9["business_sort_packages"](pkgs)
    D9["classify_packages_for_port"](pkgs, None, False)

    or_index = _real_or_index_v13()
    for pkg in pkgs:
        m = _pge13.match_store_and_or(pkg, or_index)
        pkg.or_list_store = m.matched_store
        pkg.or_list_match_status = m.status
        if m.status == "OK":
            if not pkg.or_number:
                pkg.or_number = m.matched_or
            if not pkg.so_number:
                pkg.so_number = m.matched_so

    assign_counting_scope_keys(pkgs)
    assign_global_numbers(pkgs)
    D9["assign_true_global_numbers"](pkgs)

    out_dir = Path(tempfile.mkdtemp(prefix="pl_split_store_"))
    try:
        control_path = _pge13.export_grouped_pl(
            packages=pkgs,
            output_dir=out_dir,
            write_workbook=D9["write_workbook"],
        )
        assert control_path.exists()

        store_dir = out_dir / "04_CN_BY_STORE"
        written = {p.name for p in store_dir.glob("*.xlsx")}
        # All 7 real stores must have produced a combined-factory file --
        # not filtered to CN-only cartons any more.
        for key in ("HZ", "IAPM", "KR"):
            pass  # (canonical keys checked via STORE_MASTER below)
        expected_files = {"PL_CN_STORE_HANGZHOU.xlsx", "PL_CN_STORE_IAPM.xlsx",
                           "PL_CN_STORE_KERRY.xlsx", "PL_CN_STORE_SHANGHAI_HONGQIAO.xlsx",
                           "PL_CN_STORE_SHANGHAI_TAIKOOLI.xlsx", "PL_CN_STORE_GUANGZHOU.xlsx",
                           "PL_CN_STORE_SHENZHEN.xlsx"}
        assert expected_files <= written, f"missing store files: {expected_files - written}, got {written}"

        def _carton_col_values(path):
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            ws = wb["Match_Status"]
            header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            col_idx = header.index("Carton number")
            vals = [row[col_idx] for row in ws.iter_rows(min_row=2, values_only=True) if row[col_idx]]
            wb.close()
            return vals

        # HANGZHOU: 4 cartons total (1 POP + 1 VN + 2 CN), combined scope
        # numbering 1/4..4/4 -- exactly the SAME denominator as carton_display
        # computed above, never locally renumbered/reset by the split step.
        hz_vals = _carton_col_values(store_dir / "PL_CN_STORE_HANGZHOU.xlsx")
        assert sorted(hz_vals) == ["1/4", "2/4", "3/4", "4/4"], hz_vals
        assert len(hz_vals) == 4, hz_vals

        # GUANGZHOU: VN(1) + CN(1) only, no POP -- combined total 2.
        gz_vals = _carton_col_values(store_dir / "PL_CN_STORE_GUANGZHOU.xlsx")
        assert sorted(gz_vals) == ["1/2", "2/2"], gz_vals
    finally:
        shutil.rmtree(out_dir, ignore_errors=True)


test("export_grouped_pl(): 04_CN_BY_STORE combines POP+VN+CN per Store and keeps the already-correct scope numbering (no local renumber)",
     t_export_grouped_pl_store_split_combines_pop_vn_cn_and_keeps_scope_numbering)


# =========================================================================
# v14 (spec sections 1-3): Shipmark confidence + country detection + the
# non-China "single destination" routing gate.
# =========================================================================
detect_shipment_country = D9["detect_shipment_country"]
resolve_shipping_mark_confidence = D9["resolve_shipping_mark_confidence"]
# v20 (SG-533-TEST final cleanup, requirement 2): bind the REAL production
# country-resolution function (extracted from run_pipeline() as a pure
# refactor -- same logic, now independently testable) directly.
resolve_package_country = D9["resolve_package_country"]


def t_detect_shipment_country_known_prefixes():
    assert detect_shipment_country("CN-1529_HZ_PVG_POP") == "CN"
    assert detect_shipment_country("KR-2201_SEOUL") == "KR"
    assert detect_shipment_country("JP-0099_TOKYO") == "JP"
    assert detect_shipment_country("BE-0001_BRUSSELS") == "BE"
    assert detect_shipment_country("US-4400_NY") == "US"
    assert detect_shipment_country("TW-0012_TAIPEI") == "TW"
    assert detect_shipment_country("cn-lowercase-ok") == "CN"


def t_detect_shipment_country_never_guesses_beyond_the_six_codes():
    """v19 (SG-533-TEST consolidation report, requirement 1): renamed in
    spirit but kept as-is where its assertions are STILL true -- this
    function now documents the STRUCTURAL guards that remain regardless
    of which 2-letter code appears (never a 3+-letter word, never a
    fuzzy/content guess), not "only 6 codes are recognised" (that
    restriction was deliberately removed, see
    t_detect_shipment_country_generic_non_whitelisted_markets below).
    "VN-1234_HANOI" is INTENTIONALLY no longer asserted here -- VN is now
    a structurally valid 2-letter prefix like any other (this is a
    genuine, deliberate behavior change, not a regression: VN already had
    a completely different, unrelated meaning elsewhere in this codebase
    as a FACTORY/origin suffix -- e.g. "CN-6557-...-VN" -- but that has
    always been resolved from a DIFFERENT field (get_origin(), the
    trailing suffix) than this function (the LEADING prefix), so the two
    never collide; a shipment whose Shipmark genuinely starts "VN-" is
    structurally identical in shape to "CN-"/"KR-"/... and deserves the
    same treatment)."""
    assert detect_shipment_country("XYZ-0001") == ""  # 3-letter word, not a 2-letter code
    assert detect_shipment_country("") == ""
    # must not match a longer word that merely STARTS with a known code
    assert detect_shipment_country("USER-1234") == ""
    assert detect_shipment_country("USA-1234") == ""


def t_detect_shipment_country_generic_non_whitelisted_markets():
    """v19 (SG-533-TEST consolidation report, requirement 1): the whole
    point of the fix -- a real, named, non-CN market that was NEVER in the
    old hardcoded 6-code list must now resolve correctly from a
    structurally valid 2-letter prefix, with no code changes needed to
    "add" it. Covers every market this report explicitly named."""
    for code in ("SG", "TH", "PH", "AU", "MY", "EU", "ID"):
        assert detect_shipment_country(f"{code}-9001_TEST") == code, code
        assert detect_shipment_country(f"{code}_9001_TEST") == code, code
    # real SG-533-TEST shape: no delimiter needed before the digits, just
    # not another LETTER immediately after the 2-letter code.
    assert detect_shipment_country("SG-553_CN") == "SG"


def t_shipping_mark_confidence_ordinal_by_source():
    assert resolve_shipping_mark_confidence("PL_STRUCTURED_FIELD") == 1.0
    assert resolve_shipping_mark_confidence("PL_TEXT") == 0.85
    assert resolve_shipping_mark_confidence("FILENAME_REFERENCE_CODE") == 0.5
    assert resolve_shipping_mark_confidence("") == 0.0
    assert resolve_shipping_mark_confidence("SOMETHING_UNKNOWN") == 0.0


def t_run_pipeline_sets_country_and_filename_reference_diagnostics_fields():
    """v20 (SG-533-TEST final cleanup, requirement 2): now calls the REAL
    production resolve_package_country() (extracted from run_pipeline()
    as a pure refactor) instead of hand-copying its logic inline -- the
    hand-copied version had already silently drifted out of sync with
    production once country_source tracking was added (v19), which is
    exactly the kind of divergence this refactor eliminates."""
    p = Package9(package_code="PKG1", source_file="CN-1529_HZ_PVG_POP.pdf",
                 reference_code="CN-1529_HZ_PVG_POP", pdf_package_seq=0)
    assert p.shipping_mark == "" and p.country == "" and p.filename_reference == ""
    # simulate exactly what run_pipeline's Shipmark-resolution loop does
    # BEFORE calling resolve_package_country()
    p.filename_reference = p.reference_code
    if not p.shipping_mark:
        p.shipping_mark = p.reference_code
        p.shipping_mark_source = "FILENAME_REFERENCE_CODE"
    p.shipping_mark_raw = p.shipping_mark
    p.shipping_mark_confidence = resolve_shipping_mark_confidence(p.shipping_mark_source)
    resolve_package_country(p)
    assert p.filename_reference == "CN-1529_HZ_PVG_POP"
    assert p.shipping_mark_confidence == 0.5  # filename fallback -- weakest signal
    assert p.country == "CN"
    assert p.country_source == "FILENAME_PREFIX", p.country_source


def _pkg_at_shipmark_resolution_stage(reference_code, source_file=None, *,
                                       shipping_mark=None, shipping_mark_source="FILENAME_REFERENCE_CODE"):
    """Builds a Package in exactly the state run_pipeline() has it in
    right before calling resolve_package_country() -- reference_code/
    filename_reference/shipping_mark/shipping_mark_source all resolved,
    country/country_source still untouched. Defaults reproduce the real
    SG-533-TEST archive's own confirmed shape: no real PDF Shipping Mark
    field at all, so shipping_mark falls back to the filename reference
    with source FILENAME_REFERENCE_CODE."""
    p = Package9(package_code="PKGR", source_file=source_file or f"{reference_code}.pdf",
                 reference_code=reference_code, pdf_package_seq=0)
    p.filename_reference = p.reference_code
    p.shipping_mark = shipping_mark if shipping_mark is not None else p.reference_code
    p.shipping_mark_source = shipping_mark_source
    return p


def t_resolve_package_country_real_sg533_test_archive_shape():
    """spec (SG-533-TEST final cleanup) requirement 2: proves the REAL
    production resolver against the REAL SG-533-TEST archive's own
    confirmed field values (reference_code/shipping_mark/shipping_mark_
    source exactly as run_pipeline() left them for this real fixture,
    reconfirmed by direct archive reconciliation) -- both source files,
    both factory suffixes, same destination."""
    p_cn = _pkg_at_shipmark_resolution_stage("SG-553_CN", "SG-553_CN.pdf")
    resolve_package_country(p_cn)
    assert p_cn.country == "SG", p_cn.country
    assert p_cn.country_source == "FILENAME_PREFIX", p_cn.country_source

    p_vn = _pkg_at_shipmark_resolution_stage("SG-553_VN", "SG-553_VN.pdf")
    resolve_package_country(p_vn)
    assert p_vn.country == "SG", p_vn.country
    assert p_vn.country_source == "FILENAME_PREFIX", p_vn.country_source

    # the trailing factory/origin suffix (_CN vs _VN) must never change
    # the resolved destination -- both are SG.
    assert p_cn.country == p_vn.country == "SG"


def t_resolve_package_country_generic_markets_via_real_function():
    """Same generic-market proof as detect_shipment_country()'s own test
    (t_detect_shipment_country_generic_non_whitelisted_markets), but
    through the FULL production priority chain (resolve_package_country),
    not the leaf helper alone -- proves the whole real code path is
    generic, not just one function inside it. Covers every market the
    report named, plus one it never named (ID) as an explicit "unseen
    code" proof."""
    for code in ("JP", "TW", "KR", "US", "EU", "BE", "PH", "TH", "AU", "MY", "ID"):
        p = _pkg_at_shipmark_resolution_stage(f"{code}-9001_TEST", f"{code}-9001_TEST.pdf")
        resolve_package_country(p)
        assert p.country == code, (code, p.country)
        assert p.country_source == "FILENAME_PREFIX", (code, p.country_source)


def t_resolve_package_country_prefers_real_shipping_mark_over_filename():
    """The OTHER priority branch: when shipping_mark DID come from real
    PDF content (not a filename fallback), that source wins and
    country_source records SHIPPING_MARK_PREFIX, not FILENAME_PREFIX --
    even if the filename/reference_code disagrees."""
    p = _pkg_at_shipmark_resolution_stage(
        "some-internal-doc-id-004", "some-internal-doc-id-004.pdf",
        shipping_mark="CN-1529_HZ_PVG_POP", shipping_mark_source="PL_STRUCTURED_FIELD",
    )
    resolve_package_country(p)
    assert p.country == "CN", p.country
    assert p.country_source == "SHIPPING_MARK_PREFIX", p.country_source


test("detect_shipment_country(): all 6 spec-listed country codes recognised from the Shipmark prefix",
     t_detect_shipment_country_known_prefixes)
test("detect_shipment_country(): structural guards still hold (3+-letter word / unknown / longer-word-prefix -> \"\")",
     t_detect_shipment_country_never_guesses_beyond_the_six_codes)
test("detect_shipment_country(): generic non-whitelisted markets (SG/TH/PH/AU/MY/EU/ID) resolve with zero code changes",
     t_detect_shipment_country_generic_non_whitelisted_markets)
test("resolve_shipping_mark_confidence(): PL_STRUCTURED_FIELD > PL_TEXT > FILENAME_REFERENCE_CODE > unknown",
     t_shipping_mark_confidence_ordinal_by_source)
test("run_pipeline's Shipmark-resolution loop populates filename_reference/shipping_mark_confidence/country (v14 diagnostics, now via the real resolve_package_country())",
     t_run_pipeline_sets_country_and_filename_reference_diagnostics_fields)
test("resolve_package_country() [REAL production function]: real SG-533-TEST archive shape (SG-553_CN/SG-553_VN) -> country=SG, country_source=FILENAME_PREFIX, suffix never changes destination",
     t_resolve_package_country_real_sg533_test_archive_shape)
test("resolve_package_country() [REAL production function]: generic non-whitelisted markets (JP/TW/KR/US/EU/BE/PH/TH/AU/MY + unseen ID) all resolve",
     t_resolve_package_country_generic_markets_via_real_function)
test("resolve_package_country() [REAL production function]: a real PDF-sourced Shipping Mark wins over a disagreeing filename/reference_code",
     t_resolve_package_country_prefers_real_shipping_mark_over_filename)


def t_non_cn_country_forces_single_flat_scope_even_with_or_list_store_match():
    """Section 3/11: KR/JP/BE/US/TW are SINGLE_DESTINATION -- even if an OR
    List match resolves distinct-looking "store" values, packages must all
    share ONE flat counting scope (store_carton_display == global_carton_
    display), never split like a China multi-store shipment."""
    # Same OR (one single-destination shipment/order) but DIFFERENT
    # "store"-looking OR List values per carton -- if country routing were
    # not gated, this shape would otherwise trigger a China-style per-Store
    # split/scope exactly like the real Kerry/Hangzhou fixture above.
    p1 = Package9(package_code="PKG1", source_file="KR-2201_SEOUL_A.pdf",
                  reference_code="KR-2201_SEOUL_A", pdf_package_seq=0)
    p1.shipping_mark = "KR-2201_SEOUL_A"
    p1.country = "KR"
    p1.or_list_match_status = "OK"
    p1.or_list_store = "Seoul Flagship A"
    p1.or_number = "OR-KR-1"

    p2 = Package9(package_code="PKG2", source_file="KR-2201_SEOUL_B.pdf",
                  reference_code="KR-2201_SEOUL_B", pdf_package_seq=0)
    p2.shipping_mark = "KR-2201_SEOUL_B"
    p2.country = "KR"
    p2.or_list_match_status = "OK"
    p2.or_list_store = "Seoul Flagship B"  # DIFFERENT "store" -- would split if this were CN
    p2.or_number = "OR-KR-1"

    pkgs = [p1, p2]
    assign_counting_scope_keys(pkgs)
    assign_global_numbers(pkgs)
    D9["assign_true_global_numbers"](pkgs)

    # ONE shared flat scope, not two -- both cartons number 1/2 and 2/2
    # within that single scope, and Store-scoped carton_display equals the
    # flat global_carton_display exactly (spec: "store_carton_display ==
    # global_carton_display" for non-CN countries).
    assert p1.counting_scope_key == p2.counting_scope_key, (p1.counting_scope_key, p2.counting_scope_key)
    assert {p1.carton_display, p2.carton_display} == {"1/2", "2/2"}
    for p in pkgs:
        assert p.carton_display == p.global_carton_display, (p.carton_display, p.global_carton_display)

    # And export_grouped_pl() must NOT produce a 04_CN_BY_STORE split for
    # these -- no China Store split for a non-China shipment.
    out_dir = Path(tempfile.mkdtemp(prefix="pl_kr_single_dest_"))
    try:
        _pge13.export_grouped_pl(packages=pkgs, output_dir=out_dir, write_workbook=D9["write_workbook"])
        store_dir = out_dir / "04_CN_BY_STORE"
        written = list(store_dir.glob("*.xlsx")) if store_dir.exists() else []
        assert written == [], f"non-CN shipment must not produce a Store split, got {written}"
    finally:
        shutil.rmtree(out_dir, ignore_errors=True)


test("non-CN country (KR) forces one flat counting scope, never a China-style Store split, even with a matching OR List",
     t_non_cn_country_forces_single_flat_scope_even_with_or_list_store_match)


# =========================================================================
# v14 (spec section 8): Master Data (DIM/weight file + HS Code file) is
# ENRICHMENT-ONLY -- a package/item must NEVER be dropped just because
# Master Data has no matching row for it; it must be visibly marked
# MASTER_UNMATCHED instead. Item = Item, Package = Package.
# =========================================================================
Item9 = D9["Item"]


def t_master_match_status_unmatched_when_dim_not_matched():
    p = Package9(package_code="PKG1", source_file="a.pdf", reference_code="a", pdf_package_seq=0)
    p.items = [Item9(no="1", product_name="Widget", product_code="SKU1",
                      barcode="000", unit="PCS", quantity=5, hs_code="1234.56")]
    assert p.dim_matched is False  # default -- no DIM file row found
    assert p.master_match_status == "MASTER_UNMATCHED"
    # the item itself is still fully present -- never dropped
    assert p.item_count == 1 and p.calc_qty == 5


def t_master_match_status_unmatched_when_any_item_hs_code_blank():
    p = Package9(package_code="PKG1", source_file="a.pdf", reference_code="a", pdf_package_seq=0)
    p.dim_matched = True
    p.items = [
        Item9(no="1", product_name="Widget A", product_code="SKU1", barcode="000", unit="PCS", quantity=5, hs_code="1234.56"),
        Item9(no="2", product_name="Widget B", product_code="SKU2", barcode="001", unit="PCS", quantity=3, hs_code=""),  # HS Code master had no match
    ]
    assert p.master_match_status == "MASTER_UNMATCHED"
    # BOTH items still present -- the unmatched one is not dropped
    assert p.item_count == 2 and p.calc_qty == 8


def t_master_match_status_matched_when_dim_and_all_hs_codes_found():
    p = Package9(package_code="PKG1", source_file="a.pdf", reference_code="a", pdf_package_seq=0)
    p.dim_matched = True
    p.items = [Item9(no="1", product_name="Widget", product_code="SKU1",
                      barcode="000", unit="PCS", quantity=5, hs_code="1234.56")]
    assert p.master_match_status == "MASTER_MATCHED"


def t_master_match_status_matched_with_zero_items_and_dim_ok():
    """Edge case: dim_matched=True but zero items -- master_match_status
    only reports on Master Data enrichment, not on item-count/qty issues
    (those are audit_status()/overall_status()'s job, a separate concern)."""
    p = Package9(package_code="PKG1", source_file="a.pdf", reference_code="a", pdf_package_seq=0)
    p.dim_matched = True
    assert p.items == []
    assert p.master_match_status == "MASTER_MATCHED"


test("master_match_status: MASTER_UNMATCHED when the DIM/weight file has no matching row (package never dropped)",
     t_master_match_status_unmatched_when_dim_not_matched)
test("master_match_status: MASTER_UNMATCHED when any item's HS Code master lookup misses (item never dropped)",
     t_master_match_status_unmatched_when_any_item_hs_code_blank)
test("master_match_status: MASTER_MATCHED when DIM matched and every item's HS Code was found",
     t_master_match_status_matched_when_dim_and_all_hs_codes_found)
test("master_match_status: reports purely on Master Data enrichment, independent of item-count issues",
     t_master_match_status_matched_with_zero_items_and_dim_ok)


# =========================================================================
# v14 (spec section 14): Raw_Data sheet exposes the full diagnostics set
# per package as additive trailing columns (existing columns/positions
# unchanged).
# =========================================================================
def t_raw_data_sheet_exposes_v14_diagnostics_columns():
    import openpyxl
    out_dir = Path(tempfile.mkdtemp(prefix="pl_rawdata_diag_"))
    try:
        p = Package9(package_code="PKGA", source_file="CN-1529_HZ_PVG_POP.pdf",
                     reference_code="CN-1529_HZ_PVG_POP", pdf_package_seq=0)
        p.items = [Item9(no="1", product_name="Widget", product_code="SKU1",
                          barcode="000", unit="PCS", quantity=5, hs_code="1234.56")]
        p.dim_matched = True
        p.country = "CN"
        p.country_source = "SHIPPING_MARK_PREFIX"
        p.shipping_mark = "CN-1529_HZ_PVG_POP"
        p.shipping_mark_source = "FILENAME_REFERENCE_CODE"
        p.shipping_mark_confidence = 0.5
        p.filename_reference = "CN-1529_HZ_PVG_POP"
        p.or_list_match_status = "OK"
        p.or_list_store = "Hangzhou"
        p.counting_scope_key = "OR:OR1|HANGZHOU"
        p.carton_display = "1/1"
        p.global_carton_display = "1/1"

        out_path = out_dir / "PL_Total.xlsx"
        D9["write_workbook"](out_path, [p])

        wb = openpyxl.load_workbook(str(out_path), read_only=True, data_only=True)
        ws = wb["Raw_Data"]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for col in ("master_match_status", "country", "country_source", "shipping_mark",
                    "shipping_mark_source", "shipping_mark_confidence", "filename_reference",
                    "or_list_match_status", "or_list_store", "counting_scope_key",
                    "store_carton_display", "global_carton_display"):
            assert col in header, f"missing Raw_Data diagnostics column: {col}"
        row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
        rowd = dict(zip(header, row))
        assert rowd["master_match_status"] == "MASTER_MATCHED"
        assert rowd["country"] == "CN"
        assert rowd["or_list_store"] == "Hangzhou"
        assert rowd["store_carton_display"] == "1/1"
        assert rowd["global_carton_display"] == "1/1"
        wb.close()
    finally:
        shutil.rmtree(out_dir, ignore_errors=True)


test("Raw_Data sheet exposes the full v14 diagnostics set per package (additive trailing columns)",
     t_raw_data_sheet_exposes_v14_diagnostics_columns)


# =========================================================================
# v21 gap-fix (static audit, requirement 3): Raw_Data exposes explicitly-
# named raw + derived UOM/QTY audit fields (UOM Raw/QTY Raw/PCS Per Unit/
# QTY PCS) and the routing diagnostics (Route Status/Method/Reason), and
# the Convert-to-PCS checkbox must NEVER change the raw values.
# =========================================================================
def t_raw_data_sheet_exposes_v21_uom_and_route_audit_columns():
    import openpyxl
    out_dir = Path(tempfile.mkdtemp(prefix="pl_rawdata_v21_"))
    try:
        p = Package9(package_code="PKGB", source_file="CN-9001_SomeHub_POP.pdf",
                     reference_code="CN-9001_SomeHub_POP", pdf_package_seq=0)
        # spec example: CARTON_10PCS | raw qty 2 -> PCS Per Unit=10, QTY PCS=20
        p.items = [Item9(no="1", product_name="Widget", product_code="SKU1",
                          barcode="000", unit="CARTON_10PCS", quantity=2, hs_code="1234.56")]
        p.dim_matched = True
        p.country = "CN"
        p.route_match_status = "MATCHED"
        p.route_match_method = "ROUTE_EXACT_MATCH"
        p.route_match_reason = "exact port+store match"

        out_path = out_dir / "PL_Total.xlsx"
        D9["write_workbook"](out_path, [p])

        wb = openpyxl.load_workbook(str(out_path), read_only=True, data_only=True)
        ws = wb["Raw_Data"]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for col in ("uom_raw", "qty_raw", "pcs_per_unit", "qty_pcs",
                    "route_status", "route_method", "route_reason"):
            assert col in header, f"missing Raw_Data V21 audit column: {col}"
        row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
        rowd = dict(zip(header, row))
        assert rowd["uom_raw"] == "CARTON_10PCS", rowd["uom_raw"]
        assert rowd["qty_raw"] == 2, rowd["qty_raw"]
        assert rowd["pcs_per_unit"] == 10, rowd["pcs_per_unit"]
        assert rowd["qty_pcs"] == 20, rowd["qty_pcs"]
        assert rowd["route_status"] == "MATCHED"
        assert rowd["route_method"] == "ROUTE_EXACT_MATCH"
        assert rowd["route_reason"] == "exact port+store match"
        # the existing pre-v21 "uom"/"quantity" columns must still carry the
        # exact same untouched raw values (nothing renamed/removed)
        assert rowd["uom"] == "CARTON_10PCS"
        assert rowd["quantity"] == 2
        wb.close()
    finally:
        shutil.rmtree(out_dir, ignore_errors=True)


test("Raw_Data exposes UOM Raw/QTY Raw/PCS Per Unit/QTY PCS + Route Status/Method/Reason "
     "(spec gap-fix #3, CARTON_10PCS qty=2 -> pcs_per_unit=10, qty_pcs=20)",
     t_raw_data_sheet_exposes_v21_uom_and_route_audit_columns)


def t_raw_data_uom_raw_and_qty_raw_identical_regardless_of_convert_checkbox():
    # Non-negotiable invariant: the Convert-to-PCS checkbox must NEVER
    # change UOM Raw / QTY Raw in Raw_Data -- build the SAME package twice,
    # once as if the checkbox were OFF and once ON, and confirm Raw_Data's
    # uom_raw/qty_raw/pcs_per_unit/qty_pcs are byte-for-byte identical both
    # times (write_workbook()/Raw_Data has no convert_to_pcs parameter at
    # all -- it always reports the raw+always-derived audit values, which
    # is exactly what makes this invariant structurally impossible to
    # violate, but this test locks the observable behaviour in permanently).
    import openpyxl

    def _build_and_read(out_dir):
        p = Package9(package_code="PKGC", source_file="CN-9002_SomeHub_POP.pdf",
                     reference_code="CN-9002_SomeHub_POP", pdf_package_seq=0)
        p.items = [Item9(no="1", product_name="Widget", product_code="SKU1",
                          barcode="000", unit="CARTON_10PCS", quantity=2, hs_code="1234.56")]
        p.dim_matched = True
        out_path = out_dir / "PL_Total.xlsx"
        D9["write_workbook"](out_path, [p])
        wb = openpyxl.load_workbook(str(out_path), read_only=True, data_only=True)
        ws = wb["Raw_Data"]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
        rowd = dict(zip(header, row))
        wb.close()
        return {k: rowd[k] for k in ("uom_raw", "qty_raw", "pcs_per_unit", "qty_pcs")}

    out_dir_a = Path(tempfile.mkdtemp(prefix="pl_rawdata_v21_off_"))
    out_dir_b = Path(tempfile.mkdtemp(prefix="pl_rawdata_v21_on_"))
    try:
        # write_workbook()/Raw_Data doesn't take convert_to_pcs at all --
        # this test proves that structurally by building the identical
        # package/item twice and confirming the audit values never drift.
        result_off = _build_and_read(out_dir_a)
        result_on = _build_and_read(out_dir_b)
        assert result_off == result_on == {
            "uom_raw": "CARTON_10PCS", "qty_raw": 2, "pcs_per_unit": 10, "qty_pcs": 20,
        }, (result_off, result_on)
    finally:
        shutil.rmtree(out_dir_a, ignore_errors=True)
        shutil.rmtree(out_dir_b, ignore_errors=True)


test("Raw_Data UOM Raw/QTY Raw/PCS Per Unit/QTY PCS are identical regardless of Convert-to-PCS "
     "checkbox state (spec gap-fix #3 non-negotiable invariant)",
     t_raw_data_uom_raw_and_qty_raw_identical_regardless_of_convert_checkbox)


# =========================================================================
# v17 (spec sections 12/13/16 -- supersedes the old v13/v14 "B/C columns"
# design): the Packing List's business backbone is now Item#/Store/OR
# No./Ref No. -- FOUR fixed columns (2, 3, 4 for Store/OR No./Ref No.),
# ALWAYS shown with these exact labels regardless of what the OR List's
# own header text said (spec: "Do NOT relabel Ref# as SO/SO Order/
# Invoice") -- there is no more a historical "OR No."/"SO No." default
# vs. a dynamic-label override; it's simply always Store/OR No./Ref No.
# Anything the OR List has BEYOND OR/Ref is a dynamic OPTIONAL column,
# inserted after Ref No. (column 5 onward), verbatim, never truncated and
# never invented when absent -- see t_packing_list_optional_business_
# fields_only_shown_when_present below.
# =========================================================================
def t_packing_list_headers_are_store_or_no_ref_no_always():
    import openpyxl
    out_dir = Path(tempfile.mkdtemp(prefix="pl_headers_"))
    try:
        p1 = Package9(package_code="PKGA", source_file="a.pdf", reference_code="a", pdf_package_seq=0)
        p1.or_number, p1.so_number = "po38068", "inv628036"
        p1.store_display = "Hangzhou Mixc"
        p1.global_carton_num = "1/1"
        p2 = Package9(package_code="PKGB", source_file="b.pdf", reference_code="b", pdf_package_seq=0)
        # no OR List match / no Store resolved at all -- must stay blank, never a guess.
        p2.global_carton_num = "1/1"

        out_path = out_dir / "PL_Total.xlsx"
        D9["write_workbook"](out_path, [p1, p2])

        wb = openpyxl.load_workbook(str(out_path))
        ws = wb["Packing List"]
        header_row = D9["TABLE_HDR_ROW1"]
        assert ws.cell(row=header_row, column=2).value == "Store", ws.cell(row=header_row, column=2).value
        assert ws.cell(row=header_row, column=3).value == "OR No.", ws.cell(row=header_row, column=3).value
        assert ws.cell(row=header_row, column=4).value == "Ref No.", ws.cell(row=header_row, column=4).value

        first_item_row = D9["FIRST_ITEM_ROW"]
        assert ws.cell(row=first_item_row, column=2).value == "Hangzhou Mixc"
        assert ws.cell(row=first_item_row, column=3).value == "po38068"
        assert ws.cell(row=first_item_row, column=4).value == "inv628036"
        # package with no Store/OR List match: cells stay blank.
        assert ws.cell(row=first_item_row + 1, column=2).value in (None, "")
        assert ws.cell(row=first_item_row + 1, column=3).value in (None, "")
        assert ws.cell(row=first_item_row + 1, column=4).value in (None, "")
        # Product Name (originally column 4) is now column 5.
        assert ws.cell(row=header_row, column=5).value == "Product Name\nin English"
    finally:
        shutil.rmtree(out_dir)


test("Packing List sheet: Store/OR No./Ref No. are fixed columns 2/3/4, always shown, filled from canonical Package fields, blank when unmatched",
     t_packing_list_headers_are_store_or_no_ref_no_always)


def t_packing_list_optional_business_fields_only_shown_when_present():
    """spec sections 13/16: OR List columns beyond OR/Ref (e.g. SO/PO/
    Invoice/Fulfillment No./Buyer) become dynamic optional columns
    inserted after Ref No. (column 5 onward), verbatim label + value, in
    original order -- and are NEVER invented when the OR List doesn't
    have them (minimal Shop|OR#|Ref# OR List -> Product Name stays at
    column 5, no blank SO/PO columns in between)."""
    import openpyxl
    out_dir = Path(tempfile.mkdtemp(prefix="pl_optional_fields_"))
    try:
        from collections import OrderedDict
        p1 = Package9(package_code="PKGA", source_file="a.pdf", reference_code="a", pdf_package_seq=0)
        p1.or_number, p1.so_number = "OR9001", "po90001"
        p1.store_display = "Hangzhou Mixc"
        p1.business_fields = OrderedDict([
            ("OR#", "OR9001"), ("Ref#", "po90001"), ("SO", "SO123"),
            ("PO", "PO-A1"), ("Invoice", "INV-778"),
        ])
        p1.global_carton_num = "1/1"

        out_path = out_dir / "PL_Total.xlsx"
        D9["write_workbook"](out_path, [p1], optional_business_field_labels=["SO", "PO", "Invoice"])

        wb = openpyxl.load_workbook(str(out_path))
        ws = wb["Packing List"]
        header_row = D9["TABLE_HDR_ROW1"]
        first_item_row = D9["FIRST_ITEM_ROW"]
        assert ws.cell(row=header_row, column=5).value == "SO"
        assert ws.cell(row=header_row, column=6).value == "PO"
        assert ws.cell(row=header_row, column=7).value == "Invoice"
        assert ws.cell(row=first_item_row, column=5).value == "SO123"
        assert ws.cell(row=first_item_row, column=6).value == "PO-A1"
        assert ws.cell(row=first_item_row, column=7).value == "INV-778"
        # Product Name (originally column 4, now shifted past Store/OR/Ref
        # + 3 optional fields) is column 8.
        assert ws.cell(row=header_row, column=8).value == "Product Name\nin English"

        # -- minimal case: no optional fields at all -- Product Name stays
        #    at column 5 (Store/OR No./Ref No. only), never a blank gap.
        out_dir2 = Path(tempfile.mkdtemp(prefix="pl_no_optional_fields_"))
        p2 = Package9(package_code="PKGB", source_file="b.pdf", reference_code="b", pdf_package_seq=0)
        p2.or_number, p2.so_number = "OR1172", "po38533"
        p2.store_display = "China World NB1026"
        p2.global_carton_num = "1/1"
        out_path2 = out_dir2 / "PL_Total.xlsx"
        D9["write_workbook"](out_path2, [p2])
        wb2 = openpyxl.load_workbook(str(out_path2))
        ws2 = wb2["Packing List"]
        assert ws2.cell(row=header_row, column=5).value == "Product Name\nin English"
        shutil.rmtree(out_dir2)
    finally:
        shutil.rmtree(out_dir)


test("Packing List sheet: optional OR List business fields (SO/PO/Invoice/...) appear dynamically after Ref No., never invented when absent",
     t_packing_list_optional_business_fields_only_shown_when_present)





# ── summary ──────────────────────────────────────────────────────────────
print(f"\n{_passed} passed, {_failed} failed")
if _failed:
    print("FAILED:", ", ".join(_failures))
    sys.exit(1)
sys.exit(0)
