#!/usr/bin/env python3
"""
Regression tests for:
  - pl_sublist_export.py (new in this change -- Sublist / carton-block Excel
    generation, layout math derived from the real "topologie standard
    sublist.xlsx" template -- see that module's docstring)
  - the v11 additions inside pl_ocr_core.py: parse_document_sequence(),
    business_sort_packages(), OR#/SO#/Shipping Mark label parsing, and the
    carton_sequence/carton_total/carton_display fields on Package.

Same no-framework convention as test_pl_ocr_core.py / test_pl_group_export.py.
Run with: python3 tools/pl-ocr-grouping/tests/test_pl_sublist_export.py
"""
from __future__ import annotations
import sys, types
from pathlib import Path

HERE = Path(__file__).resolve().parent
TOOL_DIR = HERE.parent
CORE_PY = TOOL_DIR / "pl_ocr_core.py"
if str(TOOL_DIR) not in sys.path:
    sys.path.insert(0, str(TOOL_DIR))

import pl_sublist_export as pse

_ENTRY_MARKER = "# ── Entry point ────────────────────────────────────────────────────────────"


def _substitute(src: str) -> str:
    return (src
            .replace("__DIM_WEIGHT_SHEET__", "None")
            .replace("__MASTER_DATA_SHEET__", "None")
            .replace("__RECURSIVE__", "False")
            .replace("__MANUAL_CONSIGNEE__", "None")
            .replace("__MANUAL_NOTIFY_PARTY__", "None")
            .replace("__GENERATE_SUBLIST__", "True")
            .replace("__GENERATE_SUBLIST_PDF__", "True")
            .replace("__OR_LIST_FILE__", "None")
            .replace("__GIT_COMMIT__", repr("test-suite")))


_counter = 0


def load_core_defs():
    global _counter
    _counter += 1
    src = CORE_PY.read_text(encoding="utf-8")
    src = _substitute(src[:src.index(_ENTRY_MARKER)])
    modname = f"pl_ocr_core_sublist_test_{_counter}"
    mod = types.ModuleType(modname)
    mod.__file__ = str(CORE_PY)
    sys.modules[modname] = mod
    exec(compile(src, str(CORE_PY), "exec"), mod.__dict__)
    return mod.__dict__


C = load_core_defs()

_passed = 0
_failed = 0


def test(name, fn):
    global _passed, _failed
    try:
        fn()
        _passed += 1
        print(f"  ok  - {name}")
    except AssertionError as e:
        _failed += 1
        print(f"FAIL  - {name}\n        {e}")
    except Exception as e:
        _failed += 1
        print(f"ERROR - {name}\n        {type(e).__name__}: {e}")


# =============================================================================
# 1. parse_document_sequence -- pure function, natural (not lexical) order
# =============================================================================
print("== parse_document_sequence ==")
pds = C["parse_document_sequence"]


def t_original_has_sequence_zero():
    r = pds("Kerry_POP")
    assert r.is_original and r.document_sequence == 0 and r.base_document_key == "Kerry_POP", r


def t_underscore_1_sequence_one():
    r = pds("Kerry_POP_1")
    assert not r.is_original and r.document_sequence == 1 and r.base_document_key == "Kerry_POP", r


def t_zero_padded_suffix():
    r = pds("Kerry_POP_02")
    assert r.document_sequence == 2 and r.base_document_key == "Kerry_POP", r


def t_two_digit_suffix_not_lexical_trap():
    r = pds("Kerry_POP_10")
    assert r.document_sequence == 10 and r.base_document_key == "Kerry_POP", r


def t_shipment_number_not_mistaken_for_sequence():
    # "1666" is a shipment/reference number, NOT a trailing copy suffix --
    # there is more text (PVG-KERRY-POP) after it, so it must not match.
    r = pds("CN-1666-PVG-KERRY-POP")
    assert r.is_original and r.document_sequence == 0, r
    assert r.base_document_key == "CN-1666-PVG-KERRY-POP", r


def t_paren_suffix():
    r = pds("Kerry_POP(2)")
    assert r.document_sequence == 2 and r.base_document_key == "Kerry_POP", r


def t_copy_word_suffix():
    r = pds("Kerry_POP COPY 3")
    assert r.document_sequence == 3 and r.base_document_key == "Kerry_POP", r


def t_natural_sort_order_2_before_10():
    docs = ["Kerry_POP_10", "Kerry_POP", "Kerry_POP_2", "Kerry_POP_1"]
    ordered = sorted(docs, key=lambda d: pds(d).document_sequence)
    assert ordered == ["Kerry_POP", "Kerry_POP_1", "Kerry_POP_2", "Kerry_POP_10"], ordered
    lexical = sorted(docs)
    assert lexical != ordered, "sanity check: lexical order really is different from natural order here"


test("no suffix -> sequence 0, is_original=True", t_original_has_sequence_zero)
test("'_1' suffix -> sequence 1, base key strips the suffix", t_underscore_1_sequence_one)
test("zero-padded '_02' -> sequence 2 (int, not string-compared)", t_zero_padded_suffix)
test("'_10' -> sequence 10 (not lexically before '_2')", t_two_digit_suffix_not_lexical_trap)
test("embedded shipment number '1666' is NOT read as a copy sequence", t_shipment_number_not_mistaken_for_sequence)
test("'(2)' suffix pattern recognised", t_paren_suffix)
test("'COPY 3' suffix pattern recognised", t_copy_word_suffix)
test("natural order (0,1,2,10) != lexical order for the same set", t_natural_sort_order_2_before_10)


# =============================================================================
# 2. OR# / SO# / Shipping Mark label parsing (Parser-level, per-package scoped)
# =============================================================================
print("== OR# / SO# / Shipping Mark label parsing ==")
Parser = C["Parser"]


def _run_labelled_pdf(rows_or_lines, use_table=True):
    p = Parser()
    p.set_file(Path("Kerry_POP.pdf"))
    p.set_page(1)
    if use_table:
        for row in rows_or_lines:
            p.feed_table_row(row)
    else:
        for line in rows_or_lines:
            p.feed_text_line(line)
    p.finalise()
    return p


def t_or_so_shipping_mark_two_cell_rows():
    p = _run_labelled_pdf([
        ["Ma kien hang: PGKECO377R7J0320001", "", "", "", "", ""],
        ["OR #", "OR1016"],
        ["SO Order #", "so402064"],
        ["Shipping Mark", "CN-1666-PVG-KERRY-POP"],
        ["1", "4894961082009", "TP-WST-RL10-MCC-02", "10mm Rope Loop", "PCS", "Moi 12"],
        ["Tong cong 12"],
    ])
    assert len(p.packages) == 1
    pkg = p.packages[0]
    assert pkg.or_number == "OR1016", pkg.or_number
    assert pkg.or_source == "PL_STRUCTURED_FIELD", pkg.or_source
    assert pkg.so_number == "so402064", pkg.so_number
    assert pkg.shipping_mark == "CN-1666-PVG-KERRY-POP", pkg.shipping_mark
    assert pkg.item_count == 1


def t_or_so_text_line_form():
    p = _run_labelled_pdf([
        "Ma kien hang: PGKECO377R7J0320001",
        "OR #: OR1016",
        "SO Order #  so402064",
        "1 4894961082009 TP-WST-RL10-MCC-02 10mm Rope Loop PCS Moi 12",
        "Tong cong 12",
    ], use_table=False)
    assert len(p.packages) == 1
    pkg = p.packages[0]
    assert pkg.or_number == "OR1016", pkg.or_number
    assert pkg.or_source == "PL_TEXT", pkg.or_source
    assert pkg.so_number == "so402064", pkg.so_number


def t_or_first_match_wins_not_overwritten():
    p = _run_labelled_pdf([
        ["OR #", "OR1016"],
        ["OR #", "OR9999"],  # must NOT overwrite the first value
    ])
    p._cur = None  # no total line closes it in this micro-test; force-inspect directly
    # re-run properly instead (package never closes without Tong cong) --
    # just check the in-progress package via the parser's internal state.
    p2 = Parser()
    p2.set_file(Path("Kerry_POP.pdf"))
    p2.feed_table_row(["Ma kien hang: PGKECTEST0000009", "", "", "", "", ""])
    p2.feed_table_row(["OR #", "OR1016"])
    p2.feed_table_row(["OR #", "OR9999"])
    p2.feed_table_row(["Tong cong 0"])
    assert p2.packages[0].or_number == "OR1016", p2.packages[0].or_number


def t_or_so_scoped_to_open_package_only_not_forward_filled():
    p = _run_labelled_pdf([
        ["Ma kien hang: PGKECAAA0000001", "", "", "", "", ""],
        ["OR #", "OR-FIRST"],
        ["1", "4894961082009", "TP-A", "Name", "PCS", "Moi 5"],
        ["Tong cong 5"],
        ["Ma kien hang: PGKECBBB0000002", "", "", "", "", ""],
        # no OR# line for the second package
        ["1", "4894961082010", "TP-B", "Name", "PCS", "Moi 3"],
        ["Tong cong 3"],
    ])
    assert len(p.packages) == 2
    assert p.packages[0].or_number == "OR-FIRST"
    assert p.packages[1].or_number == "", \
        "OR# must not forward-fill into a package that never had its own OR# line"


def t_bare_or_so_words_not_false_matched():
    # "Or"/"So" as ordinary English sentence words must NEVER be captured.
    p = _run_labelled_pdf([
        "Ma kien hang: PGKECTEST0000003",
        "Or something completely unrelated to this field",
        "So happy to see this order shipped",
        "1 4894961082009 TP-A Name PCS Moi 5",
        "Tong cong 5",
    ], use_table=False)
    assert p.packages[0].or_number == "", p.packages[0].or_number
    assert p.packages[0].so_number == "", p.packages[0].so_number


def t_no_or_so_in_pdf_still_parses_fine():
    # Backward-compat / "OR List not required" equivalent at the parser
    # level: a PL with no OR#/SO#/Shipping-Mark label lines at all must
    # still parse items/totals exactly as before -- fields just stay "".
    p = _run_labelled_pdf([
        ["Ma kien hang: PGKECPLAIN0000001", "", "", "", "", ""],
        ["1", "4894961082009", "TP-A", "Name", "PCS", "Moi 5"],
        ["Tong cong 5"],
    ])
    pkg = p.packages[0]
    assert pkg.or_number == "" and pkg.so_number == "" and pkg.shipping_mark == ""
    assert pkg.item_count == 1 and pkg.declared_total_qty == 5


test("2-cell table rows: OR#/SO#/Shipping Mark captured, source=PL_STRUCTURED_FIELD",
     t_or_so_shipping_mark_two_cell_rows)
test("text-line form: 'LABEL: value' captured, source=PL_TEXT", t_or_so_text_line_form)
test("first OR# match wins, later duplicate line never overwrites it", t_or_first_match_wins_not_overwritten)
test("OR# scoped to the package it appeared in -- never forward-filled into the next package",
     t_or_so_scoped_to_open_package_only_not_forward_filled)
test("bare English sentences starting with 'Or'/'So' are never false-matched as labels",
     t_bare_or_so_words_not_false_matched)
test("PL with no OR#/SO#/Shipping-Mark lines at all still parses items/totals correctly (no OR List / no labels)",
     t_no_or_so_in_pdf_still_parses_fine)


# =============================================================================
# 3. carton_sequence / carton_total / carton_display + business_sort_packages
# =============================================================================
print("== carton numbering fields + business_sort_packages ==")
Package = C["Package"]
assign_global_numbers = C["assign_global_numbers"]
business_sort_packages = C["business_sort_packages"]


def _mk_pkg(reference_code, source_file=None, package_code="PGKECX0000001"):
    return Package(package_code=package_code, source_file=source_file or (reference_code + ".pdf"),
                    reference_code=reference_code, pdf_package_seq="1")


def t_assign_global_numbers_sets_structured_fields():
    pkgs = [_mk_pkg("A"), _mk_pkg("B"), _mk_pkg("C")]
    assign_global_numbers(pkgs)
    assert [p.carton_sequence for p in pkgs] == [1, 2, 3]
    assert all(p.carton_total == 3 for p in pkgs)
    assert [p.carton_display for p in pkgs] == ["1/3", "2/3", "3/3"]
    assert [p.global_carton_num for p in pkgs] == [p.carton_display for p in pkgs], \
        "global_carton_num must stay in sync with carton_display (backward compat)"


def t_business_sort_natural_order_within_factory():
    # All same factory (POP) -- must end up in natural document order, not
    # lexical (Kerry_POP, _1, _2, _10 -- not Kerry_POP, _1, _10, _2).
    pkgs = [_mk_pkg("Kerry_POP_10"), _mk_pkg("Kerry_POP"), _mk_pkg("Kerry_POP_2"), _mk_pkg("Kerry_POP_1")]
    ordered = business_sort_packages(pkgs)
    assert [p.reference_code for p in ordered] == ["Kerry_POP", "Kerry_POP_1", "Kerry_POP_2", "Kerry_POP_10"], \
        [p.reference_code for p in ordered]


def t_business_sort_factory_order_pop_before_cn():
    pkgs = [_mk_pkg("Kerry_CN"), _mk_pkg("Kerry_POP")]
    ordered = business_sort_packages(pkgs)
    assert [p.reference_code for p in ordered] == ["Kerry_POP", "Kerry_CN"], [p.reference_code for p in ordered]


def t_business_sort_full_factory_business_order():
    pkgs = [_mk_pkg("X_CN"), _mk_pkg("X_JION"), _mk_pkg("X_QIFENG"), _mk_pkg("X_SBGEAR"), _mk_pkg("X_POP")]
    ordered = business_sort_packages(pkgs)
    assert [p.reference_code for p in ordered] == ["X_POP", "X_SBGEAR", "X_QIFENG", "X_JION", "X_CN"], \
        [p.reference_code for p in ordered]


def t_business_sort_does_not_interleave_different_base_documents():
    # Two independent shipments under the same factory must not interleave
    # just because they share a factory suffix.
    pkgs = [_mk_pkg("Alpha_POP_1"), _mk_pkg("Beta_POP"), _mk_pkg("Alpha_POP"), _mk_pkg("Beta_POP_1")]
    ordered = business_sort_packages(pkgs)
    refs = [p.reference_code for p in ordered]
    assert refs.index("Alpha_POP") < refs.index("Alpha_POP_1"), refs
    assert refs.index("Beta_POP") < refs.index("Beta_POP_1"), refs
    # each base-document's own packages stay contiguous (not interleaved)
    alpha_positions = [i for i, r in enumerate(refs) if r.startswith("Alpha")]
    beta_positions = [i for i, r in enumerate(refs) if r.startswith("Beta")]
    assert alpha_positions == list(range(min(alpha_positions), max(alpha_positions) + 1)), refs
    assert beta_positions == list(range(min(beta_positions), max(beta_positions) + 1)), refs


def t_business_sort_stable_within_same_pdf():
    # Two packages parsed from the SAME pdf_package_seq context (same base
    # document, sequence 0) must keep their original relative (parse) order.
    p1 = _mk_pkg("SamePdf", package_code="PGKEC0001")
    p2 = _mk_pkg("SamePdf", package_code="PGKEC0002")
    ordered = business_sort_packages([p1, p2])
    assert [p.package_code for p in ordered] == ["PGKEC0001", "PGKEC0002"]


def t_carton_numbering_continuous_no_reset_across_factories():
    pkgs = [_mk_pkg("X_POP"), _mk_pkg("X_POP_1"), _mk_pkg("X_SBGEAR"), _mk_pkg("X_CN"), _mk_pkg("X_CN_1")]
    ordered = business_sort_packages(pkgs)
    assign_global_numbers(ordered)
    assert [p.carton_display for p in ordered] == ["1/5", "2/5", "3/5", "4/5", "5/5"], \
        [p.carton_display for p in ordered]


test("assign_global_numbers sets carton_sequence/carton_total/carton_display, "
     "global_carton_num stays in sync", t_assign_global_numbers_sets_structured_fields)
test("business_sort_packages: natural order within one factory (not lexical _1/_10/_2)",
     t_business_sort_natural_order_within_factory)
test("business_sort_packages: POP before CN (factory business order)",
     t_business_sort_factory_order_pop_before_cn)
test("business_sort_packages: full order POP -> SBGEAR -> QIFENG -> JION -> CN",
     t_business_sort_full_factory_business_order)
test("business_sort_packages: two independent shipments under the same factory are not interleaved",
     t_business_sort_does_not_interleave_different_base_documents)
test("business_sort_packages: stable sort preserves in-PDF parse order for ties",
     t_business_sort_stable_within_same_pdf)
test("carton numbering is continuous/global -- no reset between factories/PDFs",
     t_carton_numbering_continuous_no_reset_across_factories)


# =============================================================================
# 4. pl_sublist_export -- carton model, pagination, Excel layout, validation
# =============================================================================
print("== pl_sublist_export ==")
from types import SimpleNamespace


def _mk_item(sku, ean, qty):
    return SimpleNamespace(product_code=sku, barcode=ean, quantity=qty)


def _mk_full_pkg(seq, total, items, shipping_mark="CN-1666-PVG-KERRY-POP",
                  or_number="OR1016", so_number="so402064", weight=35.68,
                  package_code="PGKECO377R7J0320001", reference_code="Kerry_POP",
                  counting_scope_key="", store_display="Kerry Center flagship",
                  business_fields=None):
    return SimpleNamespace(
        carton_sequence=seq, carton_total=total, carton_display=f"{seq}/{total}",
        global_carton_num=f"{seq}/{total}",
        shipping_mark=shipping_mark, shipping_mark_source="FILENAME_REFERENCE_CODE",
        or_number=or_number, or_source="PL_TEXT", so_number=so_number, so_source="PL_TEXT",
        weight=weight, package_code=package_code, items=items,
        source_file=reference_code + ".pdf", reference_code=reference_code, pdf_package_seq="1",
        counting_scope_key=counting_scope_key, store_display=store_display,
        business_fields=business_fields or {},
    )


def t_build_carton_model_basic_fields():
    items = [_mk_item("SKU1", "EAN1", 3)]
    pkg = _mk_full_pkg(1, 1, items)
    model = pse.build_sublist_carton_model(pkg)
    assert model.carton_display == "1/1"
    assert model.shipping_mark == "CN-1666-PVG-KERRY-POP"
    assert model.or_number == "OR1016" and model.so_number == "so402064"
    assert model.gross_weight_display == "35.68 KG"
    assert model.packing_code == "PGKECO377R7J0320001"
    assert model.total_qty == 3


def t_build_carton_model_does_not_mutate_package():
    items = [_mk_item("SKU1", "EAN1", 3)]
    pkg = _mk_full_pkg(1, 1, items)
    original_items = list(pkg.items)
    pse.build_sublist_carton_model(pkg)
    assert pkg.items == original_items, "package.items must be untouched"


def t_current_only_display_mode():
    pkg = _mk_full_pkg(3, 10, [_mk_item("S", "E", 1)])
    model = pse.build_sublist_carton_model(pkg, carton_display_mode="current_only")
    assert model.carton_display == "3", model.carton_display


def t_paginate_no_continuation_when_under_capacity():
    items = [_mk_item(f"S{i}", f"E{i}", 1) for i in range(5)]
    pkg = _mk_full_pkg(1, 1, items)
    model = pse.build_sublist_carton_model(pkg)
    blocks = pse.paginate_carton_blocks([model])
    assert len(blocks) == 1 and blocks[0].block_count == 1


def t_paginate_exact_capacity_no_continuation():
    items = [_mk_item(f"S{i}", f"E{i}", 1) for i in range(pse.SUBLIST_ITEM_CAPACITY_PER_BLOCK)]
    pkg = _mk_full_pkg(1, 1, items)
    model = pse.build_sublist_carton_model(pkg)
    blocks = pse.paginate_carton_blocks([model])
    assert len(blocks) == 1, "exactly at capacity must NOT trigger a continuation block"


def t_paginate_over_capacity_splits_no_item_lost():
    n = pse.SUBLIST_ITEM_CAPACITY_PER_BLOCK + 7
    items = [_mk_item(f"S{i}", f"E{i}", 2) for i in range(n)]
    pkg = _mk_full_pkg(1, 1, items)
    model = pse.build_sublist_carton_model(pkg)
    blocks = pse.paginate_carton_blocks([model])
    assert len(blocks) == 2, len(blocks)
    assert len(blocks[0].items) == pse.SUBLIST_ITEM_CAPACITY_PER_BLOCK
    assert len(blocks[1].items) == 7
    total_items_in_blocks = sum(len(b.items) for b in blocks)
    assert total_items_in_blocks == n, "no item may be dropped across a continuation split"
    assert not blocks[0].is_last_block and blocks[1].is_last_block
    assert "Continued 1" in blocks[1].block_carton_label, blocks[1].block_carton_label
    assert blocks[0].block_total_qty == pse.SUBLIST_ITEM_CAPACITY_PER_BLOCK * 2, \
        "non-last block shows its own subtotal"
    assert blocks[1].block_total_qty == n * 2, "last block shows the GRAND total across all blocks"


def t_continuation_does_not_inflate_carton_count():
    n = pse.SUBLIST_ITEM_CAPACITY_PER_BLOCK + 1
    items = [_mk_item(f"S{i}", f"E{i}", 1) for i in range(n)]
    pkg = _mk_full_pkg(1, 1, items)
    model = pse.build_sublist_carton_model(pkg)
    blocks = pse.paginate_carton_blocks([model])
    unique_cartons = {b.carton.carton_identity for b in blocks}
    assert len(unique_cartons) == 1, "2 blocks, but still exactly 1 carton"


def t_gw_prefers_pl_text_over_dim_weight():
    # spec: "GW (from PL, kept separate from existing DIM-sourced weight)"
    # -- pl_gross_weight (captured from the PL PDF text) must win over the
    # DIM-lookup `weight` field whenever both are present.
    items = [_mk_item("SKU1", "EAN1", 3)]
    pkg = _mk_full_pkg(1, 1, items, weight=99.99)
    pkg.pl_gross_weight = "35.68 KG"
    model = pse.build_sublist_carton_model(pkg)
    assert model.gross_weight_display == "35.68 KG", model.gross_weight_display
    assert model.gross_weight_source == "PL_TEXT"


def t_gw_falls_back_to_dim_weight_when_pl_text_missing():
    items = [_mk_item("SKU1", "EAN1", 3)]
    pkg = _mk_full_pkg(1, 1, items, weight=12.5)
    pkg.pl_gross_weight = ""
    model = pse.build_sublist_carton_model(pkg)
    assert model.gross_weight_display == "12.50 KG", model.gross_weight_display
    assert model.gross_weight_source == "DIM_FALLBACK"


def t_gw_bare_number_without_unit_gets_kg_appended():
    items = [_mk_item("SKU1", "EAN1", 3)]
    pkg = _mk_full_pkg(1, 1, items)
    pkg.pl_gross_weight = "35.68"  # no letters -> unit appended
    model = pse.build_sublist_carton_model(pkg)
    assert model.gross_weight_display == "35.68 KG", model.gross_weight_display


test("GW prefers PL-text pl_gross_weight over DIM-sourced weight when both present", t_gw_prefers_pl_text_over_dim_weight)
test("GW falls back to DIM-sourced weight when pl_gross_weight is empty", t_gw_falls_back_to_dim_weight_when_pl_text_missing)
test("GW: bare numeric pl_gross_weight (no unit letters) gets 'KG' appended", t_gw_bare_number_without_unit_gets_kg_appended)
test("build_sublist_carton_model maps all metadata fields correctly", t_build_carton_model_basic_fields)
test("build_sublist_carton_model does not mutate the source package", t_build_carton_model_does_not_mutate_package)
test("carton_display_mode='current_only' shows just the sequence number", t_current_only_display_mode)
test("<=capacity items -> single block, no continuation", t_paginate_no_continuation_when_under_capacity)
test("exactly at capacity (18 items) -> still a single block", t_paginate_exact_capacity_no_continuation)
test(">capacity items -> splits into continuation block(s), zero items lost, "
     "subtotal on first block, grand total on last block", t_paginate_over_capacity_splits_no_item_lost)
test("continuation blocks never inflate the unique-carton count", t_continuation_does_not_inflate_carton_count)


# --- full workbook generation + real-file layout checks -------------------
import tempfile
import openpyxl


def t_generate_workbook_layout_matches_real_template_measurements():
    """v17 (spec sections 12/16/17): the metadata block gained a fixed
    Store row (Carton#/Shipping Mark/Store/OR No./Ref No./GW/Packing
    Code# = 7 rows, was 6) -- every row below it shifts down by exactly
    1 versus the old v15/v16 layout. v20 (Sublist display-only UI change):
    Store's row was removed again -- Carton#/Shipping Mark/OR No./Ref No./
    GW/Packing Code# = 6 rows, back to the pre-Store geometry -- Store is
    still fully resolved on the package (store_display below) and used
    elsewhere (PL_Total etc.), just never written to a Sublist row. Row
    numbers here are computed from pse._sublist_offsets(0) rather than
    hardcoded, so this test can never silently drift out of sync with the
    production geometry again."""
    with tempfile.TemporaryDirectory() as td:
        items = [_mk_item("TP-A-1", "4894961069222", 10), _mk_item("TP-A-2", "4895227935312", 15)]
        pkg = _mk_full_pkg(1, 1, items)
        out = Path(td) / "sublist.xlsx"
        result = pse.generate_sublist_workbook([pkg], out)
        wb = openpyxl.load_workbook(str(out))
        ws = wb["Sheet1"]
        off = pse._sublist_offsets(0)
        assert "store" not in off, "Sublist offsets must never have a Store row (v20 display-only removal)"
        # metadata block: label col B, value col C (block 1 = A:C)
        assert ws[f"B{1 + off['carton']}"].value == "Carton #" and ws[f"C{1 + off['carton']}"].value == "1/1"
        assert ws[f"B{1 + off['shipping_mark']}"].value == "Shipping Mark" and ws[f"C{1 + off['shipping_mark']}"].value == "CN-1666-PVG-KERRY-POP"
        assert ws[f"B{1 + off['or']}"].value == "OR No." and ws[f"C{1 + off['or']}"].value == "OR1016"
        assert ws[f"B{1 + off['ref']}"].value == "Ref No." and ws[f"C{1 + off['ref']}"].value == "so402064"
        assert ws[f"B{1 + off['gw']}"].value == "GW" and ws[f"C{1 + off['gw']}"].value == "35.68 KG"
        assert ws[f"B{1 + off['packing_code']}"].value == "Packing Code #" and ws[f"C{1 + off['packing_code']}"].value == "PGKECO377R7J0320001"
        header_row = 1 + off["item_header"]
        assert [ws.cell(row=header_row, column=c).value for c in (1, 2, 3)] == ["Item No.", "EAN", "QTY"]
        first_row = 1 + off["item_first"]
        assert ws.cell(row=first_row, column=1).value == "TP-A-1"
        assert ws.cell(row=first_row, column=2).value == "4894961069222"
        assert ws.cell(row=first_row, column=3).value == 10
        total_row = 1 + off["total"]
        assert ws.cell(row=total_row, column=3).value == f"=SUM(C{first_row}:C{first_row + 1})"
        result.output_path.exists()  # sanity
        # v20 regression: "Store" must never appear as a label ANYWHERE in
        # the Sublist Excel sheet, for either block.
        all_labels = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
        assert "Store" not in all_labels, f"Sublist Excel must never show a 'Store' label -- found: {all_labels}"
        assert "Kerry Center flagship" not in all_labels, (
            f"Store's resolved display value must never appear on the Sublist Excel: {all_labels}"
        )


def t_page_two_starts_at_correct_row():
    with tempfile.TemporaryDirectory() as td:
        pkgs = [_mk_full_pkg(i, 5, [_mk_item("S", "E", 1)], reference_code=f"R{i}") for i in range(1, 6)]
        out = Path(td) / "sublist.xlsx"
        pse.generate_sublist_workbook(pkgs, out)
        wb = openpyxl.load_workbook(str(out))
        ws = wb["Sheet1"]
        # page 1 holds blocks 1-4 (cols A:C, D:F, G:I, J:L); block 5 (carton 5) -> page 2, col A:C
        expected_page2_start = 1 + pse.PAGE_CYCLE_ROWS
        assert ws.cell(row=expected_page2_start, column=2).value == "Carton #"
        assert ws.cell(row=expected_page2_start, column=3).value == "5/5"


def t_ean_and_metadata_are_stored_as_text_not_number():
    with tempfile.TemporaryDirectory() as td:
        items = [_mk_item("TP-A-1", "4894961069222", 10)]
        pkg = _mk_full_pkg(1, 1, items, or_number="000123", so_number="000456")
        out = Path(td) / "sublist.xlsx"
        pse.generate_sublist_workbook([pkg], out)
        wb = openpyxl.load_workbook(str(out))
        ws = wb["Sheet1"]
        off = pse._sublist_offsets(0)
        first_row = 1 + off["item_first"]
        assert ws[f"B{first_row}"].data_type == "s" and ws[f"B{first_row}"].value == "4894961069222", \
            "EAN must be stored as text, never converted to a number"
        assert ws[f"C{1 + off['or']}"].value == "000123", "leading zeros in OR# must survive (text, not number)"
        assert ws[f"C{1 + off['ref']}"].value == "000456", "leading zeros in Ref# must survive (text, not number)"
        assert ws["C1"].number_format == "@"


def t_sublist_carton_order_matches_input_order():
    with tempfile.TemporaryDirectory() as td:
        pkgs = [_mk_full_pkg(i, 3, [_mk_item("S", "E", 1)], reference_code=f"R{i}") for i in (1, 2, 3)]
        out = Path(td) / "sublist.xlsx"
        result = pse.generate_sublist_workbook(pkgs, out)
        assert [c.reference_code for c in result.cartons] == ["R1", "R2", "R3"], \
            "Sublist must preserve exactly the order it was given (== PL_TOTAL order)"


def t_validate_sublist_passes_for_well_formed_input():
    pkgs = [_mk_full_pkg(i, 2, [_mk_item(f"S{i}", f"E{i}", i)], reference_code=f"R{i}") for i in (1, 2)]
    with tempfile.TemporaryDirectory() as td:
        result = pse.generate_sublist_workbook(pkgs, Path(td) / "s.xlsx")
        ok, report = pse.validate_sublist(pkgs, result)
        assert ok, report


def t_generate_sublist_handles_zero_item_package():
    pkg = _mk_full_pkg(1, 1, [])
    with tempfile.TemporaryDirectory() as td:
        result = pse.generate_sublist_workbook([pkg], Path(td) / "s.xlsx")
        ok, report = pse.validate_sublist([pkg], result)
        assert ok, report
        assert result.blocks[0].block_total_qty == 0


# --- scope-aware carton_sequence validation (Turn 12: per-Store numbering
#     legitimately repeats 1, 2, 3... across different counting_scope_key
#     values -- validate_sublist() must key uniqueness/completeness on
#     (counting_scope_key, carton_sequence), never carton_sequence alone) --
def t_validate_sublist_two_stores_each_1_of_2_2_of_2_passes():
    # HANGZHOU 1/2, 2/2 and KERRY 1/2, 2/2 -- both scopes independently
    # complete and non-duplicated. Global carton_sequence looks like
    # [1, 2, 1, 2] here -- must NOT be flagged as duplicates.
    pkgs = [
        _mk_full_pkg(1, 2, [_mk_item("H1", "EH1", 1)], reference_code="HZ1",
                     package_code="PKG-HZ-1", counting_scope_key="UPLOAD_BATCH|HANGZHOU"),
        _mk_full_pkg(2, 2, [_mk_item("H2", "EH2", 1)], reference_code="HZ2",
                     package_code="PKG-HZ-2", counting_scope_key="UPLOAD_BATCH|HANGZHOU"),
        _mk_full_pkg(1, 2, [_mk_item("K1", "EK1", 1)], reference_code="KR1",
                     package_code="PKG-KR-1", counting_scope_key="UPLOAD_BATCH|KERRY"),
        _mk_full_pkg(2, 2, [_mk_item("K2", "EK2", 1)], reference_code="KR2",
                     package_code="PKG-KR-2", counting_scope_key="UPLOAD_BATCH|KERRY"),
    ]
    with tempfile.TemporaryDirectory() as td:
        result = pse.generate_sublist_workbook(pkgs, Path(td) / "s.xlsx")
        ok, report = pse.validate_sublist(pkgs, result)
        assert ok, report


def t_validate_sublist_duplicate_within_same_scope_fails():
    # Two DIFFERENT cartons both claiming carton_sequence=1 in the SAME
    # scope (KERRY) is a real bug and must still FAIL.
    pkgs = [
        _mk_full_pkg(1, 2, [_mk_item("K1", "EK1", 1)], reference_code="KR1",
                     package_code="PKG-KR-1", counting_scope_key="UPLOAD_BATCH|KERRY"),
        _mk_full_pkg(1, 2, [_mk_item("K2", "EK2", 1)], reference_code="KR2",
                     package_code="PKG-KR-2", counting_scope_key="UPLOAD_BATCH|KERRY"),
    ]
    with tempfile.TemporaryDirectory() as td:
        result = pse.generate_sublist_workbook(pkgs, Path(td) / "s.xlsx")
        ok, report = pse.validate_sublist(pkgs, result)
        assert not ok, "duplicate carton_sequence WITHIN one scope must fail validation"
        assert "duplicate" in report.lower() and "UPLOAD_BATCH|KERRY" in report


def t_validate_sublist_missing_within_one_scope_fails():
    # KERRY claims carton_total=2 but only carton_sequence=1 exists --
    # carton_sequence=2 is missing WITHIN that scope.
    pkgs = [
        _mk_full_pkg(1, 2, [_mk_item("K1", "EK1", 1)], reference_code="KR1",
                     package_code="PKG-KR-1", counting_scope_key="UPLOAD_BATCH|KERRY"),
    ]
    with tempfile.TemporaryDirectory() as td:
        result = pse.generate_sublist_workbook(pkgs, Path(td) / "s.xlsx")
        ok, report = pse.validate_sublist(pkgs, result)
        assert not ok, "missing carton_sequence WITHIN one scope must fail validation"
        assert "missing" in report.lower()


def t_validate_sublist_repeated_1_across_different_scopes_passes():
    # IAPM 1/1 and a totally separate UNRESOLVED-scope 1/1 both legitimately
    # use carton_sequence=1 -- must PASS (this is the exact "duplicates=[1, 2]"
    # false positive the business owner hit live).
    pkgs = [
        _mk_full_pkg(1, 1, [_mk_item("I1", "EI1", 1)], reference_code="IAPM1",
                     package_code="PKG-IAPM-1", counting_scope_key="UPLOAD_BATCH|IAPM"),
        _mk_full_pkg(1, 1, [_mk_item("U1", "EU1", 1)], reference_code="UN1",
                     package_code="PKG-UN-1", counting_scope_key="UPLOAD_BATCH|UNRESOLVED"),
    ]
    with tempfile.TemporaryDirectory() as td:
        result = pse.generate_sublist_workbook(pkgs, Path(td) / "s.xlsx")
        ok, report = pse.validate_sublist(pkgs, result)
        assert ok, report


test("generated workbook layout matches the real template's measured row/col positions",
     t_generate_workbook_layout_matches_real_template_measurements)
test("5th carton correctly starts page 2 at the measured page-cycle row",
     t_page_two_starts_at_correct_row)
test("EAN / OR# / SO# stored as text (leading zeros survive, no scientific notation)",
     t_ean_and_metadata_are_stored_as_text_not_number)
test("Sublist carton order exactly matches the order packages were passed in (== PL_TOTAL order)",
     t_sublist_carton_order_matches_input_order)
test("validate_sublist passes (OK) for well-formed input, full reconciliation", t_validate_sublist_passes_for_well_formed_input)
test("zero-item package produces a valid (empty) carton block, not a crash", t_generate_sublist_handles_zero_item_package)
test("scope-aware validation: two Stores each with their own 1/2, 2/2 -> PASS", t_validate_sublist_two_stores_each_1_of_2_2_of_2_passes)
test("scope-aware validation: duplicate carton_sequence=1 WITHIN one Store's scope -> FAIL", t_validate_sublist_duplicate_within_same_scope_fails)
test("scope-aware validation: missing carton_sequence=2 WITHIN one Store's scope -> FAIL", t_validate_sublist_missing_within_one_scope_fails)
test("scope-aware validation: carton_sequence=1 repeated across DIFFERENT scopes -> PASS", t_validate_sublist_repeated_1_across_different_scopes_passes)


# =============================================================================
# 5. Full end-to-end pipeline (real synthetic PDFs) through the AUTO SPLIT
#    section -- proves the actual app.html wiring (import pl_group_export,
#    import pl_sublist_export, sys.path setup, GENERATE_SUBLIST flag) works
#    together, not just each piece in isolation.
# =============================================================================
print("== end-to-end: run_pipeline + AUTO SPLIT + Sublist ==")
import shutil, tempfile
import openpyxl as _openpyxl

FIXTURES = HERE / "fixtures" / "synthetic"
SYN_DIM = FIXTURES / "SYN-DIM.xlsx"


def _make_empty_master_xlsx(path: Path):
    wb = _openpyxl.Workbook()
    ws = wb.active
    ws.append(["SKU/Product Code", "HS Code", "EAN/Barcode"])
    wb.save(str(path))


def _run_full_module(pdf_dir, dim_xlsx, master_xlsx, out_dir, or_list_path=None, **kw):
    src = CORE_PY.read_text(encoding="utf-8")
    src = _substitute(src)
    src = src.replace('PL_FOLDER = Path("/work/pdfs")', f'PL_FOLDER = Path({str(pdf_dir)!r})')
    src = src.replace('OUTPUT_XLSX = Path("/work/PL_Total.xlsx")', f'OUTPUT_XLSX = Path({str(out_dir / "PL_Total.xlsx")!r})')
    src = src.replace('DIM_WEIGHT_FILE = Path("/work/dim.xlsx")', f'DIM_WEIGHT_FILE = Path({str(dim_xlsx)!r})')
    src = src.replace('MASTER_DATA_FILE = Path("/work/master.xlsx")', f'MASTER_DATA_FILE = Path({str(master_xlsx)!r})')
    if or_list_path is not None:
        src = src.replace('OR_LIST_FILE = None', f'OR_LIST_FILE = Path({str(or_list_path)!r})')
    modname = "pl_ocr_core_e2e_test"
    mod = types.ModuleType(modname)
    mod.__file__ = str(CORE_PY)
    sys.modules[modname] = mod
    exec(compile(src, str(CORE_PY), "exec"), mod.__dict__)
    return mod.__dict__


def t_end_to_end_generates_sublist_alongside_existing_outputs():
    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        pdf_dir = td / "pdfs"
        pdf_dir.mkdir()
        for f in FIXTURES.glob("*.pdf"):
            shutil.copy(f, pdf_dir / f.name)
        master_xlsx = td / "master.xlsx"
        _make_empty_master_xlsx(master_xlsx)
        out_dir = td
        result_ns = _run_full_module(pdf_dir, SYN_DIM, master_xlsx, out_dir)

        packages = result_ns["packages"]
        assert len(packages) == 3, "existing synthetic fixture behaviour must be unchanged (3 packages)"

        split_dir = pdf_dir / "PL_SPLIT_OUTPUT"
        assert (split_dir / "01_PL_TOTAL" / "PL_TOTAL.xlsx").exists(), "existing grouping output must be unaffected"
        sublist_path = split_dir / "05_SUBLIST" / "SUBLIST_TOTAL.xlsx"
        assert sublist_path.exists(), "SUBLIST_TOTAL.xlsx must be generated by default (GENERATE_SUBLIST=True)"

        wb = _openpyxl.load_workbook(str(sublist_path))
        ws = wb["Sheet1"]
        # 3 cartons, no continuation expected (synthetic fixture items are few) -> 3 blocks on page 1
        assert ws["B1"].value == "Carton #" and ws["C1"].value == "1/3"
        assert ws["E1"].value == "Carton #" and ws["F1"].value == "2/3"

        # v12: the A5 PDF is ALSO generated by default alongside the Excel
        # Sublist (GENERATE_SUBLIST_PDF=True default) -- both are optional/
        # secondary+primary Sublist outputs living side by side, neither
        # one affects the other.
        pdf_path = split_dir / "05_SUBLIST" / "SUBLIST_TOTAL.pdf"
        assert pdf_path.exists(), "SUBLIST_TOTAL.pdf must be generated by default (GENERATE_SUBLIST_PDF=True)"
        assert result_ns["SUBLIST_PDF_STATUS"] == "SUCCESS"
        assert result_ns["SUBLIST_XLSX_STATUS"] == "SUCCESS"
        assert result_ns["RUN_SUMMARY"]["sublist_pdf_status"] == "SUCCESS"


def t_end_to_end_sublist_pdf_failure_never_blocks_legacy_export():
    # spec bug report: "The current patch claims Sublist is optional, but
    # re-raises RuntimeError and PermissionError. Fix this." -- simulate
    # the A5 PDF module being unimportable (e.g. reportlab missing) and
    # confirm the REST of the pipeline (PL_TOTAL, PL_SPLIT_OUTPUT, Excel
    # Sublist) still completes with no exception escaping run_pipeline().
    class _Blocker:
        def find_spec(self, name, path, target=None):
            if name == "pl_sublist_pdf_export":
                raise ImportError("simulated: reportlab not available")
            return None
    blocker = _Blocker()
    real_cached = sys.modules.pop("pl_sublist_pdf_export", None)
    sys.meta_path.insert(0, blocker)
    try:
        with tempfile.TemporaryDirectory() as td:
            td = Path(td)
            pdf_dir = td / "pdfs"
            pdf_dir.mkdir()
            for f in FIXTURES.glob("*.pdf"):
                shutil.copy(f, pdf_dir / f.name)
            master_xlsx = td / "master.xlsx"
            _make_empty_master_xlsx(master_xlsx)
            # Must not raise -- this call itself is the assertion.
            result_ns = _run_full_module(pdf_dir, SYN_DIM, master_xlsx, td)

            assert result_ns["SUBLIST_PDF_STATUS"] == "FAILED"
            assert "pl_sublist_pdf_export" in result_ns["SUBLIST_PDF_ERROR"]
            # Everything else completed anyway:
            assert len(result_ns["packages"]) == 3
            split_dir = pdf_dir / "PL_SPLIT_OUTPUT"
            assert (split_dir / "01_PL_TOTAL" / "PL_TOTAL.xlsx").exists()
            assert result_ns["SUBLIST_XLSX_STATUS"] == "SUCCESS", (
                "Excel Sublist is a separate try/except block and must be unaffected by the PDF one failing")
            pdf_path = split_dir / "05_SUBLIST" / "SUBLIST_TOTAL.pdf"
            assert not pdf_path.exists(), "no partial/corrupt PDF should be left behind on FAILED"
    finally:
        sys.meta_path.remove(blocker)
        if real_cached is not None:
            sys.modules["pl_sublist_pdf_export"] = real_cached


def t_end_to_end_generate_sublist_off_still_produces_everything_else():
    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        pdf_dir = td / "pdfs"
        pdf_dir.mkdir()
        for f in FIXTURES.glob("*.pdf"):
            shutil.copy(f, pdf_dir / f.name)
        master_xlsx = td / "master.xlsx"
        _make_empty_master_xlsx(master_xlsx)
        src = CORE_PY.read_text(encoding="utf-8")
        src = _substitute(src).replace("__GENERATE_SUBLIST__", "True").replace("__GENERATE_SUBLIST_PDF__", "True").replace("__OR_LIST_FILE__", "None")  # placeholder for the next replace
        # re-substitute with GENERATE_SUBLIST=False specifically
        src = CORE_PY.read_text(encoding="utf-8")
        src = (src.replace("__DIM_WEIGHT_SHEET__", "None").replace("__MASTER_DATA_SHEET__", "None")
                  .replace("__RECURSIVE__", "False").replace("__MANUAL_CONSIGNEE__", "None")
                  .replace("__MANUAL_NOTIFY_PARTY__", "None").replace("__GENERATE_SUBLIST__", "False")
                  .replace("__GENERATE_SUBLIST_PDF__", "False")
                  .replace("__OR_LIST_FILE__", "None")
                  .replace("__GIT_COMMIT__", repr("test-suite")))
        src = src.replace('PL_FOLDER = Path("/work/pdfs")', f'PL_FOLDER = Path({str(pdf_dir)!r})')
        src = src.replace('OUTPUT_XLSX = Path("/work/PL_Total.xlsx")', f'OUTPUT_XLSX = Path({str(td / "PL_Total.xlsx")!r})')
        src = src.replace('DIM_WEIGHT_FILE = Path("/work/dim.xlsx")', f'DIM_WEIGHT_FILE = Path({str(SYN_DIM)!r})')
        src = src.replace('MASTER_DATA_FILE = Path("/work/master.xlsx")', f'MASTER_DATA_FILE = Path({str(master_xlsx)!r})')
        mod = types.ModuleType("pl_ocr_core_e2e_off_test")
        mod.__file__ = str(CORE_PY)
        sys.modules["pl_ocr_core_e2e_off_test"] = mod
        exec(compile(src, str(CORE_PY), "exec"), mod.__dict__)

        split_dir = pdf_dir / "PL_SPLIT_OUTPUT"
        assert (split_dir / "01_PL_TOTAL" / "PL_TOTAL.xlsx").exists(),             "legacy grouping output must still be produced when Sublist is turned off"
        assert not (split_dir / "05_SUBLIST").exists(),             "Sublist folder must not appear at all when GENERATE_SUBLIST=False"


def t_end_to_end_or_list_header_not_found_surfaces_distinctly_in_run_summary():
    # Turn 12 bug: a genuinely uploaded-but-unparseable OR List previously
    # looked IDENTICAL to "no file uploaded" (or_index just came back empty
    # either way). RUN_SUMMARY must now say HEADER_NOT_FOUND explicitly so
    # the UI can show "OR List uploaded but header not recognized" instead
    # of silently behaving as if nothing was supplied.
    import openpyxl as _oxl
    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        pdf_dir = td / "pdfs"
        pdf_dir.mkdir()
        for f in FIXTURES.glob("*.pdf"):
            shutil.copy(f, pdf_dir / f.name)
        master_xlsx = td / "master.xlsx"
        _make_empty_master_xlsx(master_xlsx)

        bad_or_list = td / "or_list.xlsx"
        wb = _oxl.Workbook()
        ws = wb.active
        ws.append(["Warehouse", "Reference", "Notes"])  # no STORE/OR aliases at all
        # v14: a blank first cell on the data row denies even the new
        # positional fallback tier (spec section 5) its "data continues
        # under column A" signal -- this now genuinely stays unparseable.
        ws.append(["", "B", "C"])
        wb.save(str(bad_or_list))

        result_ns = _run_full_module(pdf_dir, SYN_DIM, master_xlsx, td, or_list_path=bad_or_list)

        assert result_ns["RUN_SUMMARY"]["or_list_status"] == "HEADER_NOT_FOUND", result_ns["RUN_SUMMARY"]
        assert result_ns["RUN_SUMMARY"]["or_list_error"], "or_list_error must be populated, not blank"
        # Must never be confused with "no file uploaded":
        assert result_ns["RUN_SUMMARY"]["or_list_status"] != "NO_FILE"
        # And the rest of the pipeline must be completely unaffected:
        split_dir = pdf_dir / "PL_SPLIT_OUTPUT"
        assert (split_dir / "01_PL_TOTAL" / "PL_TOTAL.xlsx").exists()


def t_end_to_end_or_list_no_file_reports_no_file_status():
    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        pdf_dir = td / "pdfs"
        pdf_dir.mkdir()
        for f in FIXTURES.glob("*.pdf"):
            shutil.copy(f, pdf_dir / f.name)
        master_xlsx = td / "master.xlsx"
        _make_empty_master_xlsx(master_xlsx)

        result_ns = _run_full_module(pdf_dir, SYN_DIM, master_xlsx, td)  # no or_list_path at all

        assert result_ns["RUN_SUMMARY"]["or_list_status"] == "NO_FILE"


test("full run_pipeline + AUTO SPLIT produces 05_SUBLIST/SUBLIST_TOTAL.xlsx alongside every existing output",
     t_end_to_end_generates_sublist_alongside_existing_outputs)
test("OR List uploaded but header not recognized -> RUN_SUMMARY.or_list_status=HEADER_NOT_FOUND, distinct from NO_FILE",
     t_end_to_end_or_list_header_not_found_surfaces_distinctly_in_run_summary)
test("no OR List uploaded -> RUN_SUMMARY.or_list_status=NO_FILE (the normal/expected case)",
     t_end_to_end_or_list_no_file_reports_no_file_status)
test("GENERATE_SUBLIST=False: legacy grouping output unaffected, no Sublist folder created, no error",
     t_end_to_end_generate_sublist_off_still_produces_everything_else)
test("Sublist PDF module unimportable -> FAILED status, but legacy export + Excel Sublist still complete (non-blocking)",
     t_end_to_end_sublist_pdf_failure_never_blocks_legacy_export)


print(f"\n{_passed} passed, {_failed} failed")
if _failed:
    sys.exit(1)
