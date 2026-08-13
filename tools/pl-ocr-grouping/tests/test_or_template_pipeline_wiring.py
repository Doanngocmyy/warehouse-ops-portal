#!/usr/bin/env python3
"""
tests/test_or_template_pipeline_wiring.py
===========================================
Proves the OR List / Routing Template (pl_or_routing_template.py) is
wired into the REAL execution path -- run_pipeline() itself, exactly as
the browser app calls it -- not just exercised in isolation as a
standalone loader (that narrower coverage already lives in
tests/test_or_routing_template.py, 13 tests).

Every test here runs the actual pl_ocr_core.py source (truncated just
before the AUTO SPLIT section, same technique as tests/test_pl_ocr_core.py
and tests/test_v21_routing_and_uom.py) against a real synthetic PDF on
disk and a real OR-Template xlsx on disk, via or_list_file= -- the SAME
parameter app.html's orListInput wires up. There is no more separate
"routing_rules" JS table: routing now comes exclusively from whatever
was uploaded as the OR List (spec "OR-TEMPLATE VALIDATION CORRECTION",
sections 2/3/9).

Covers the required matrix (spec section 11):
  A. app/pipeline consumes the OR Template rows, not just the standalone loader.
  B. CN/PVG/CNWorld resolves route + OR + Ref.
  C. CN/blank/Tmall resolves with Port blank.
  D. blank OR continues with warning.
  E. blank Ref continues with warning.
  F. both blank continue with warning.
  G. SG blank Port/Store unique fallback.
  H. ambiguous template rows -> REVIEW.
  I. old Routing Rules UI is absent (app.html).
  J. Download OR List Template asset exists and is referenced by app.html.
Plus: legacy OR List (no Country Code column) keeps the pre-existing
compatibility path (spec section 9) -- a permanent regression guard for
the "never silently break the old behavior" requirement.

Same no-framework convention as the rest of this tool's tests. Run with:
    python3 tools/pl-ocr-grouping/tests/test_or_template_pipeline_wiring.py
"""
from __future__ import annotations
import sys, tempfile, shutil, types, logging
from pathlib import Path

HERE = Path(__file__).resolve().parent
TOOL_DIR = HERE.parent
CORE_PY = TOOL_DIR / "pl_ocr_core.py"
APP_HTML = TOOL_DIR / "app.html"
if str(TOOL_DIR) not in sys.path:
    sys.path.insert(0, str(TOOL_DIR))

logging.getLogger().setLevel(logging.ERROR)  # keep test output readable

SYN_PDFS = TOOL_DIR / "tests" / "fixtures" / "synthetic"
SYN_DIM = SYN_PDFS / "SYN-DIM.xlsx"
_AUTO_SPLIT_MARKER = "# =========================================================\n# AUTO SPLIT:"


def _lit(v):
    return "None" if v is None else repr(v)


def _substitute_placeholders(src: str, *, or_list_file=None) -> str:
    return (src
            .replace("__DIM_WEIGHT_SHEET__", "None")
            .replace("__MASTER_DATA_SHEET__", "None")
            .replace("__RECURSIVE__", "False")
            .replace("__MANUAL_CONSIGNEE__", "None")
            .replace("__MANUAL_NOTIFY_PARTY__", "None")
            .replace("__GENERATE_SUBLIST__", "False")
            .replace("__GENERATE_SUBLIST_PDF__", "False")
            .replace("__OR_LIST_FILE__", _lit(str(or_list_file)) if or_list_file else "None")
            .replace("__CONVERT_TO_PCS__", "False")
            .replace("__SHOW_UOM_IN_SUBLIST__", "False")
            .replace("__GIT_COMMIT__", _lit("test-or-template-wiring")))


def load_core_pipeline(pdf_dir: Path, dim_xlsx: Path, master_xlsx: Path, out_dir: Path, or_list_file=None):
    """Runs the REAL run_pipeline() end-to-end (truncated just before the
    AUTO SPLIT / factory-store-grouping section) -- the exact same
    function app.html's Pyodide run calls, with the exact same
    or_list_file parameter app.html wires from the orListInput upload."""
    src = CORE_PY.read_text(encoding="utf-8")
    idx = src.index(_AUTO_SPLIT_MARKER)
    src = _substitute_placeholders(src[:idx], or_list_file=or_list_file)
    src = src.replace('PL_FOLDER = Path("/work/pdfs")', f'PL_FOLDER = Path({str(pdf_dir)!r})')
    src = src.replace('OUTPUT_XLSX = Path("/work/PL_Total.xlsx")', f'OUTPUT_XLSX = Path({str(out_dir / "PL_Total.xlsx")!r})')
    src = src.replace('DIM_WEIGHT_FILE = Path("/work/dim.xlsx")', f'DIM_WEIGHT_FILE = Path({str(dim_xlsx)!r})')
    src = src.replace('MASTER_DATA_FILE = Path("/work/master.xlsx")', f'MASTER_DATA_FILE = Path({str(master_xlsx)!r})')
    mod = types.ModuleType(f"pl_ocr_core_or_template_wiring_{id(src)}")
    mod.__file__ = str(CORE_PY)
    sys.modules[mod.__name__] = mod
    exec(compile(src, str(CORE_PY), "exec"), mod.__dict__)
    return mod.__dict__


def _make_empty_master_xlsx(path: Path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["SKU/Product Code", "HS Code", "EAN/Barcode"])
    wb.save(str(path))


def _make_or_template(path: Path, rows):
    """rows: list of [country, port, store, or_no, ref_no]."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Country Code", "Port", "Store", "OR No.", "Ref No."])
    for r in rows:
        ws.append(r)
    wb.save(str(path))


def _make_legacy_or_list(path: Path, rows):
    """rows: list of [store, or_no, so_no] -- the OLD Store-first shape,
    no Country Code column anywhere."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Store", "OR No.", "SO No."])
    for r in rows:
        ws.append(r)
    wb.save(str(path))


def _single_pkg_run(filename: str, or_template_rows, tmpdir: Path):
    """Copies the 1-package synthetic fixture under `filename` (its
    reference_code -- and therefore its Shipmark BODY tokens, since these
    fixtures have no in-PDF Shipping Mark text -- comes straight from the
    filename), builds an OR-Template xlsx with `or_template_rows`, runs
    the REAL pipeline, and returns (packages, run_summary)."""
    pdf_dir = tmpdir / "pdfs"
    pdf_dir.mkdir(parents=True, exist_ok=True)
    shutil.copy(SYN_PDFS / "SYN-1001-WarehouseAlpha-CN.pdf", pdf_dir / filename)
    out_dir = tmpdir / "out"
    out_dir.mkdir(parents=True, exist_ok=True)
    master = out_dir / "master.xlsx"
    _make_empty_master_xlsx(master)
    or_tpl = out_dir / "or_template.xlsx"
    _make_or_template(or_tpl, or_template_rows)
    ns = load_core_pipeline(pdf_dir, SYN_DIM, master, out_dir, or_list_file=or_tpl)
    return ns["packages"], ns["RUN_SUMMARY"]


# ── tiny test runner (mirrors every other test_*.py in this directory) ─────
_passed = 0
_failed = 0
_failures = []


def test(name, fn):
    global _passed, _failed
    try:
        fn()
        _passed += 1
        print(f"  ok  - {name}")
    except AssertionError as e:
        _failed += 1
        _failures.append(name)
        print(f"FAIL  - {name}\n        {e}")
    except Exception as e:
        _failed += 1
        _failures.append(name)
        print(f"ERROR - {name}\n        {type(e).__name__}: {e}")


# =============================================================================
# A. app/pipeline consumes the OR Template rows, not just the standalone loader
# =============================================================================
print("== A. real run_pipeline() wiring (not just the standalone loader) ==")


def t_a_run_summary_reports_or_template_active_from_real_pipeline():
    with tempfile.TemporaryDirectory() as td:
        packages, summary = _single_pkg_run(
            "CN-9001-WarehouseAlpha.pdf",
            [["CN", "", "Tmall", "OR1075", "to10880"]],
            Path(td))
        assert summary["or_template_active"] is True, "RUN_SUMMARY must show OR-Template mode active for the REAL pipeline run"
        assert summary["or_template_rows_loaded"] == 1
        assert len(packages) == 1
        assert packages[0].or_number == "OR1075"
        assert packages[0].so_number == "to10880"


test("A: real run_pipeline() (not the standalone loader) reports or_template_active + attaches OR/Ref", t_a_run_summary_reports_or_template_active_from_real_pipeline)


# =============================================================================
# B. CN/PVG/CNWorld resolves route + OR + Ref (exact match)
# =============================================================================
def t_b_cn_pvg_cnworld_resolves_route_and_or_and_ref():
    with tempfile.TemporaryDirectory() as td:
        packages, _ = _single_pkg_run(
            "CN-9001-PVG-CNWorld-POP.pdf",
            [["CN", "PVG", "CNWorld", "OR1172", "po38533"]],
            Path(td))
        p = packages[0]
        assert p.route_match_status == "MATCHED", p.route_match_reason
        assert p.route_match_method == "ROUTE_EXACT_MATCH"
        assert p.country == "CN" and p.port == "PVG" and p.store == "CNWorld"
        assert p.or_number == "OR1172" and p.so_number == "po38533"
        assert p.or_ref_warning == ""


test("B: CN/PVG/CNWorld -> exact route match, Port/Store/OR/Ref all resolve", t_b_cn_pvg_cnworld_resolves_route_and_or_and_ref)


# =============================================================================
# C. CN/blank/Tmall resolves with Port blank (unique fallback)
# =============================================================================
def t_c_cn_blank_tmall_resolves_with_port_blank():
    with tempfile.TemporaryDirectory() as td:
        packages, _ = _single_pkg_run(
            "CN-9001-WarehouseAlpha.pdf",
            [["CN", "", "Tmall", "OR1075", "to10880"]],
            Path(td))
        p = packages[0]
        assert p.route_match_status == "MATCHED", p.route_match_reason
        assert p.port == "", f"Port must stay blank, got {p.port!r}"
        assert p.store == "Tmall"
        assert p.or_number == "OR1075" and p.so_number == "to10880"
        assert p.or_ref_warning == ""


test("C: CN/blank/Tmall -> route resolves with Port blank (never guessed)", t_c_cn_blank_tmall_resolves_with_port_blank)


# =============================================================================
# D/E/F. Missing OR/Ref must NEVER block a valid route -- warning only
# =============================================================================
def t_d_blank_or_continues_with_warning():
    with tempfile.TemporaryDirectory() as td:
        packages, _ = _single_pkg_run(
            "CN-9001-WarehouseAlpha.pdf",
            [["CN", "", "Tmall", "", "to10880"]],
            Path(td))
        p = packages[0]
        assert p.route_match_status == "MATCHED", "a blank OR No. must never block the route"
        assert p.store == "Tmall"
        assert p.or_number == ""
        assert p.so_number == "to10880"
        assert p.or_ref_warning == "WARNING_MISSING_OR"


def t_e_blank_ref_continues_with_warning():
    with tempfile.TemporaryDirectory() as td:
        packages, _ = _single_pkg_run(
            "CN-9001-WarehouseAlpha.pdf",
            [["CN", "", "Tmall", "OR1075", ""]],
            Path(td))
        p = packages[0]
        assert p.route_match_status == "MATCHED", "a blank Ref No. must never block the route"
        assert p.store == "Tmall"
        assert p.or_number == "OR1075"
        assert p.so_number == ""
        assert p.or_ref_warning == "WARNING_MISSING_REF"


def t_f_both_blank_continue_with_warning():
    with tempfile.TemporaryDirectory() as td:
        packages, _ = _single_pkg_run(
            "CN-9001-WarehouseAlpha.pdf",
            [["CN", "", "Tmall", "", ""]],
            Path(td))
        p = packages[0]
        assert p.route_match_status == "MATCHED", "both OR and Ref blank must still never block the route"
        assert p.store == "Tmall"
        assert p.or_number == "" and p.so_number == ""
        assert p.or_ref_warning == "WARNING_MISSING_OR_REF"


test("D: CN/blank/Tmall/blank-OR/to10880 -> route OK, WARNING_MISSING_OR", t_d_blank_or_continues_with_warning)
test("E: CN/blank/Tmall/OR1075/blank-Ref -> route OK, WARNING_MISSING_REF", t_e_blank_ref_continues_with_warning)
test("F: CN/blank/Tmall/blank/blank -> route OK, WARNING_MISSING_OR_REF", t_f_both_blank_continue_with_warning)


# =============================================================================
# G. SG blank Port/Store unique fallback
# =============================================================================
def t_g_sg_blank_port_store_unique_fallback():
    with tempfile.TemporaryDirectory() as td:
        packages, _ = _single_pkg_run(
            "SG-9001-WarehouseAlpha.pdf",
            [["SG", "", "", "OR1159", "po38515"]],
            Path(td))
        p = packages[0]
        assert p.route_match_status == "MATCHED", p.route_match_reason
        assert p.country == "SG"
        assert p.port == "" and p.store == "", "Port and Store may both legitimately stay blank"
        assert p.or_number == "OR1159" and p.so_number == "po38515"


test("G: SG with blank Port/Store -> unique-country fallback still resolves the route", t_g_sg_blank_port_store_unique_fallback)


# =============================================================================
# H. Ambiguous template rows -> REVIEW (never a silent guess)
# =============================================================================
def t_h_ambiguous_template_rows_review():
    with tempfile.TemporaryDirectory() as td:
        packages, _ = _single_pkg_run(
            "CN-9001-WarehouseAlpha.pdf",
            [["CN", "", "StoreA", "OR1", "Ref1"], ["CN", "", "StoreB", "OR2", "Ref2"]],
            Path(td))
        p = packages[0]
        assert p.route_match_status == "REVIEW", "2 CN rules with no disambiguating Shipmark signal must REVIEW, never guess"
        assert p.or_list_match_status == "REVIEW"
        assert p.or_number == "" and p.so_number == "", "no business refs attached to an unresolved route"


test("H: 2 conflicting CN rules, no disambiguating signal -> REVIEW, no refs guessed", t_h_ambiguous_template_rows_review)


# =============================================================================
# Legacy compatibility: an OR List with NO Country Code column keeps using
# the pre-existing Store-first compatibility path (spec section 9) --
# routing_rules stays empty, so classify_packages_for_port() falls back to
# the pre-V21 STORE_MASTER resolver exactly as before.
# =============================================================================
def t_legacy_or_list_without_country_code_keeps_compatibility_path():
    with tempfile.TemporaryDirectory() as td:
        tmpdir = Path(td)
        pdf_dir = tmpdir / "pdfs"
        pdf_dir.mkdir(parents=True, exist_ok=True)
        shutil.copy(SYN_PDFS / "SYN-1001-WarehouseAlpha-CN.pdf", pdf_dir / "CN-9001-WarehouseAlpha.pdf")
        out_dir = tmpdir / "out"
        out_dir.mkdir(parents=True, exist_ok=True)
        master = out_dir / "master.xlsx"
        _make_empty_master_xlsx(master)
        legacy = out_dir / "legacy_or_list.xlsx"
        _make_legacy_or_list(legacy, [["Tmall", "OR1075", "to10880"]])
        ns = load_core_pipeline(pdf_dir, SYN_DIM, master, out_dir, or_list_file=legacy)
        assert ns["RUN_SUMMARY"]["or_template_active"] is False, "a legacy (no Country Code) OR List must NOT activate OR-Template routing"
        p = ns["packages"][0]
        # legacy path: route_match_status stays "" (routing_rules empty ->
        # classify_packages_for_port never even enters the routing-rule
        # branch) -- pkg.or_ref_warning also stays "" (only ever set in
        # OR-Template mode).
        assert p.route_match_status == "", "legacy path must never populate route_match_status (that's OR-Template-mode-only)"
        assert p.or_ref_warning == ""


test("legacy OR List (no Country Code column) keeps the pre-existing compatibility path unchanged", t_legacy_or_list_without_country_code_keeps_compatibility_path)


# =============================================================================
# I. Old Routing Rules UI is absent from app.html
# =============================================================================
print("\n== I/J. app.html structural checks ==")
_APP_HTML_TEXT = APP_HTML.read_text(encoding="utf-8")


def t_i_routing_rules_ui_fully_removed():
    forbidden = [
        "routingRulesTable", "routingRulesBody", "routingRulesStatus", "addRoutingRuleBtn",
        "addRoutingRuleRow", "readRoutingRules", "validateRoutingRules",
        "4. (Tuỳ chọn) Routing Rules", "__ROUTING_RULES_JSON__",
    ]
    present = [s for s in forbidden if s in _APP_HTML_TEXT]
    assert not present, f"Routing Rules UI remnants still present in app.html: {present}"


def t_i_sublist_section_renumbered_to_4():
    assert "4. (Tuỳ chọn) Hiển thị Sublist" in _APP_HTML_TEXT
    assert "5. (Tuỳ chọn) Hiển thị Sublist" not in _APP_HTML_TEXT
    # both checkbox groups (Section 1 + Section 4) must still be present,
    # unchanged, with their original ids/defaults (presentation-only
    # renumbering, never a behavior change -- spec section 8).
    for cid in ("recursiveCheck", "generateSublistCheck", "generateSublistPdfCheck",
                "convertToPcsCheck", "showUomInSublistCheck"):
        assert f'id="{cid}"' in _APP_HTML_TEXT, f"checkbox #{cid} must still exist"
    assert 'id="generateSublistCheck" checked' in _APP_HTML_TEXT
    assert 'id="generateSublistPdfCheck" checked' in _APP_HTML_TEXT
    assert 'id="convertToPcsCheck">' in _APP_HTML_TEXT  # unchecked (no "checked")
    assert 'id="showUomInSublistCheck">' in _APP_HTML_TEXT  # unchecked


test("I: Routing Rules UI (table/buttons/JS/state) fully removed from app.html, not just hidden", t_i_routing_rules_ui_fully_removed)
test("I: Sublist Display Options renumbered 5->4, both checkbox groups + defaults preserved", t_i_sublist_section_renumbered_to_4)


# =============================================================================
# J. Download OR List Template asset exists and is referenced by app.html
# =============================================================================
def t_j_template_asset_exists_on_disk():
    asset = TOOL_DIR / "assets" / "OR_List_Template.xlsx"
    assert asset.exists(), f"expected static asset at {asset}"
    import openpyxl
    wb = openpyxl.load_workbook(str(asset))
    ws = wb["OR Template"] if "OR Template" in wb.sheetnames else wb.worksheets[0]
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=10):
        values = [c.value for c in row]
        if values[:1] == ["Country Code"] or (values and "Country Code" in str(values[0] or "")):
            header_row = values
            break
    assert header_row is not None, "template must have a Country Code header row"


def t_j_template_link_referenced_in_app_html():
    assert 'assets/OR_List_Template.xlsx' in _APP_HTML_TEXT
    assert 'download' in _APP_HTML_TEXT.lower()
    assert 'Download OR List Template' in _APP_HTML_TEXT


def t_j_template_colors_country_code_red_others_yellow():
    import openpyxl
    asset = TOOL_DIR / "assets" / "OR_List_Template.xlsx"
    wb = openpyxl.load_workbook(str(asset))
    ws = wb["OR Template"]
    header_cells = {ws.cell(row=4, column=c).value: ws.cell(row=4, column=c) for c in range(1, 6)}
    country_cell = next(v for k, v in header_cells.items() if k and "Country Code" in str(k))
    assert country_cell.fill.fgColor.rgb == "FFF4CCCC", "Country Code header must stay red/required"
    assert "*" in str(country_cell.value), "Country Code must still visually read as required"
    for label in ("Port", "Store", "OR No.", "Ref No."):
        cell = header_cells.get(label)
        assert cell is not None, f"missing {label} header cell"
        assert cell.fill.fgColor.rgb == "FFFFF2CC", f"{label} header must be yellow/may-be-blank, got {cell.fill.fgColor.rgb}"
        assert "*" not in str(cell.value), f"{label} must NOT have an asterisk (not mandatory)"


test("J: assets/OR_List_Template.xlsx exists on disk with a Country Code header", t_j_template_asset_exists_on_disk)
test("J: app.html references the download link (assets/OR_List_Template.xlsx, download attr, label text)", t_j_template_link_referenced_in_app_html)
test("J: template coloring -- Country Code red/required, Port/Store/OR/Ref yellow, no asterisks on OR/Ref", t_j_template_colors_country_code_red_others_yellow)


# ── summary ──────────────────────────────────────────────────────────────
print(f"\n{_passed} passed, {_failed} failed")
if _failed:
    print("FAILED:", ", ".join(_failures))
    sys.exit(1)
sys.exit(0)
